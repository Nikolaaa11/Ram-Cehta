"""Script one-shot — importa proveedores desde Excel a la tabla core.proveedores.

Uso:
    python scripts/import_proveedores_xlsx.py "C:/Users/DELL/Documents/nikolaya/Proveedores.xlsx"

Lee TODAS las hojas (TRONGKAI, EVOQUE, DTE, CLIMATE, CENERGY, AFIS,
Consolida Proveedores), consolida por RUT único, valida modulo 11 y
hace INSERT ... ON CONFLICT (rut) DO NOTHING en la DB.

Conexión: lee DATABASE_URL del .env del backend o del arg --dsn.

Reporta al final:
  - filas leídas por hoja
  - RUTs únicos consolidados
  - RUTs inválidos (descartados)
  - RUTs insertados (nuevos)
  - RUTs ya existentes (skipped)
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

# Path hack para importar el value object Rut del backend
BACKEND_DIR = Path(__file__).resolve().parent.parent / "backend"
sys.path.insert(0, str(BACKEND_DIR))

import openpyxl  # noqa: E402
import psycopg2  # noqa: E402

from app.domain.value_objects.rut import format_rut, validate_rut  # noqa: E402

# DSN sync para psycopg2 (NO el +asyncpg que usa el backend).
# Acepta override por env var o arg posicional --dsn=...
DEFAULT_DSN = (
    "postgresql://postgres.dqwwqfhzejscgcynkbip:088ybpNLf5kWAW3e"
    "@aws-1-us-east-2.pooler.supabase.com:6543/postgres"
)


def parse_args() -> tuple[str, str]:
    xlsx_path = "C:/Users/DELL/Documents/nikolaya/Proveedores.xlsx"
    dsn = DEFAULT_DSN
    for a in sys.argv[1:]:
        if a.startswith("--dsn="):
            dsn = a.split("=", 1)[1]
        elif a.startswith("--"):
            continue
        else:
            xlsx_path = a
    return xlsx_path, dsn


def collect_unique(xlsx_path: str) -> tuple[
    dict[str, tuple[str, str | None]],  # rut_canonical → (razon_social, direccion)
    dict[str, int],  # sheet → rows_read
    list[tuple[str, str, str]],  # invalid: (sheet, rut_raw, razon)
]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    consolidated: dict[str, tuple[str, str | None]] = {}
    rows_per_sheet: dict[str, int] = {}
    invalid: list[tuple[str, str, str]] = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        count = 0
        for r in range(2, ws.max_row + 1):
            rut_raw = ws.cell(r, 1).value
            razon = ws.cell(r, 2).value
            direccion = ws.cell(r, 3).value
            if not rut_raw:
                continue
            count += 1
            rut_str = str(rut_raw).strip()
            if not validate_rut(rut_str):
                invalid.append((sh, rut_str, str(razon or "")))
                continue
            rut_canonical = format_rut(rut_str)
            razon_clean = str(razon).strip() if razon else "(sin razón social)"
            direccion_clean = str(direccion).strip() if direccion else None
            # Si el RUT ya estaba (otra hoja), mantenemos la versión con más
            # información (la que tenga dirección no-vacía gana).
            existing = consolidated.get(rut_canonical)
            if existing is None or (direccion_clean and not existing[1]):
                consolidated[rut_canonical] = (razon_clean[:255], direccion_clean)
        rows_per_sheet[sh] = count

    return consolidated, rows_per_sheet, invalid


def upsert(
    dsn: str, providers: dict[str, tuple[str, str | None]]
) -> tuple[int, int]:
    """INSERT ... ON CONFLICT (rut) DO NOTHING para cada proveedor.

    Devuelve (inserted, skipped).
    """
    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    inserted = 0
    skipped = 0
    try:
        with conn.cursor() as cur:
            for rut, (razon, direccion) in providers.items():
                cur.execute(
                    """
                    INSERT INTO core.proveedores (rut, razon_social, direccion, activo)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (rut) DO NOTHING
                    RETURNING proveedor_id
                    """,
                    (rut, razon, direccion),
                )
                if cur.fetchone() is not None:
                    inserted += 1
                else:
                    skipped += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return inserted, skipped


def main() -> int:
    xlsx_path, dsn = parse_args()
    print(f"[1/3] Reading {xlsx_path}")
    consolidated, rows_per_sheet, invalid = collect_unique(xlsx_path)

    print("\n=== Sumario lectura ===")
    total_rows = 0
    for sh, count in rows_per_sheet.items():
        print(f"  {sh}: {count} filas con RUT")
        total_rows += count
    print(f"  TOTAL filas: {total_rows}")
    print(f"  RUTs únicos válidos: {len(consolidated)}")
    print(f"  RUTs inválidos descartados: {len(invalid)}")
    if invalid:
        print("\n  Primeros 10 inválidos:")
        for sh, rut, razon in invalid[:10]:
            print(f"    [{sh}] {rut!r} → {razon[:50]}")

    print(f"\n[2/3] Conectando a Postgres: {dsn[:60]}…")
    inserted, skipped = upsert(dsn, consolidated)

    print("\n=== Sumario import ===")
    print(f"  INSERTED: {inserted}")
    print(f"  SKIPPED (ya existían): {skipped}")
    print(f"  TOTAL procesados: {inserted + skipped}")

    print(f"\n[3/3] Verificando con SELECT COUNT(*) FROM core.proveedores …")
    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM core.proveedores WHERE activo = TRUE")
            total_activos = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM core.proveedores")
            total = cur.fetchone()[0]
    finally:
        conn.close()
    print(f"  core.proveedores total: {total} ({total_activos} activos)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
