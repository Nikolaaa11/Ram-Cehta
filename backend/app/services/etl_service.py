"""ETL service — sincroniza Data Madre.xlsx desde Dropbox a `core.movimientos`.

Flow (idempotente, con full audit trail):

    1. Buscar el archivo en Dropbox: `/Cehta Capital/00-Inteligencia de Negocios/Data Madre.xlsx`
       (acepta también la variante legacy sin prefijo `00-`).
    2. Calcular SHA256 del file content.
    3. Si el hash coincide con el último etl_run con status='success' → SKIP.
       Devolvemos un `ETLResult` con `status='skipped'` sin escribir filas
       nuevas — el archivo no cambió.
    4. Si hay cambios:
        a. Crear `audit.etl_runs` row con status='running'.
        b. Parsear hoja "Resumen" con openpyxl (read_only + values_only).
        c. Volcar TODAS las filas crudas → `raw.resumen_excel` (preserva
           reproducibilidad).
        d. Validar + transformar → upsert en `core.movimientos` por natural_key.
        e. Filas inválidas → `audit.rejected_rows` (no bloquean el resto).
        f. Guardar snapshot en `Histórico/YYYY-MM-DD-Data-Madre.xlsx`.
        g. Cerrar etl_run con status='success' o 'partial' (si hubo rejected) +
           counts; o 'failed' si la corrida explotó antes de cerrar.

Decisiones técnicas:
- **Idempotencia por hash**: barata (SHA256 64 hex chars en `audit.etl_runs.source_hash`).
  Si el archivo no cambió, el resto del pipeline se evita.
- **natural_key**: hash determinístico de (fecha, descripcion, abono, egreso,
  empresa, banco). Permite UPSERT — si una fila ya existe con esa firma, se
  actualiza; sino se inserta. Re-runs del mismo Excel son idempotentes.
- **rejected_rows en lugar de fail-fast**: cada validación que falla queda
  registrada con razón clara, pero NO aborta el resto del archivo. El admin
  ve `/admin/etl/{run_id}/rejected-rows` y arregla.
- **Soft-fail si Dropbox no conectado**: el endpoint que invoca run_etl decide
  el código HTTP (503 con mensaje útil); el servicio mismo solo necesita un
  `DropboxService` ya construido.
"""
from __future__ import annotations

import contextlib
import hashlib
import io
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import uuid4

import structlog
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.empresa import Empresa
from app.models.etl_run import EtlRun
from app.models.rejected_row import RejectedRow
from app.services.dropbox_service import DropboxNotConfigured, DropboxService
from app.services.event_broadcaster import get_broadcaster

log = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# Constants — paths Dropbox + sheet schema
# ---------------------------------------------------------------------------

CEHTA_ROOT = "Cehta Capital"
"""Carpeta raíz dentro del Dropbox conectado."""

INTELIGENCIA_FOLDER_CANDIDATES: tuple[str, ...] = (
    "00-Inteligencia de Negocios",
    "Inteligencia de Negocios",
)
"""Aceptamos la variante legacy sin prefijo numérico (`Inteligencia de Negocios`)
para compatibilidad con instalaciones que aún no migraron a la convención V3."""

DATA_MADRE_FILENAME_SUBSTR = "data madre"
"""Match case-insensitive: `Data Madre.xlsx`, `data-madre.xlsx`, etc."""

HISTORICO_SUBFOLDER = "Histórico"

RESUMEN_SHEET = "Resumen"
"""Nombre canónico legacy. La hoja real puede llamarse 'CONSOLIDA' (formato
del Excel madre Cehta) o 'Resumen'. La búsqueda en `parse_resumen_sheet`
prueba ambos."""

CONSOLIDA_SHEET_CANDIDATES: tuple[str, ...] = ("CONSOLIDA", "Consolida", "Resumen")
"""Hojas a probar en orden. CONSOLIDA es el formato real del Excel Cehta
(2.700+ filas con todas las empresas)."""

# Mapeo de header del Excel → nombre lógico canónico (snake_case).
# Aceptamos múltiples variantes para tolerar pequeñas diferencias en headers,
# typos, prefijos/sufijos. La key se compara con .strip().casefold().
HEADER_ALIASES: dict[str, str] = {
    # Hipervínculo (link al respaldo) — Excel real usa "Hyper-Vinculo"
    "hipervinculo": "hipervinculo",
    "hipervínculo": "hipervinculo",
    "hyper-vinculo": "hipervinculo",
    "hyper-vínculo": "hipervinculo",
    "hypervinculo": "hipervinculo",
    # Fecha
    "fecha": "fecha",
    # Descripción
    "descripcion": "descripcion",
    "descripción": "descripcion",
    # Abonos / Egreso
    "abonos": "abonos",
    "abono": "abonos",
    "egreso": "egreso",
    "egresos": "egreso",
    # Saldos
    "saldo contable": "saldo_contable",
    "saldo cehta": "saldo_cehta",
    # Typo conocido en hoja TRONGKAI: "ALDO CEHTA" en vez de "SALDO CEHTA"
    "aldo cehta": "saldo_cehta",
    "saldo corfo": "saldo_corfo",
    # Conceptos
    "concepto general": "concepto_general",
    "concepto detallado": "concepto_detallado",
    # Tipo egreso — Excel real usa "TIPO DE EGRESO"
    "tipo egreso": "tipo_egreso",
    "tipo de egreso": "tipo_egreso",
    # Fuentes
    "fuentes": "fuentes",
    "fuente": "fuentes",
    # Proyecto / Banco
    "proyecto": "proyecto",
    "banco": "banco",
    # Real / Proyectado
    "real/proyectado": "real_proyectado",
    "real proyectado": "real_proyectado",
    "real_proyectado": "real_proyectado",
    # Año / Periodo
    "año": "anio",
    "anio": "anio",
    "ano": "anio",
    "periodo": "periodo",
    # Empresa
    "empresa": "empresa",
    # IVA — Excel real usa "IVA CRÉDITO FISCAL" / "IVA DÉBITO FISCAL"
    "iva crédito": "iva_credito",
    "iva credito": "iva_credito",
    "iva crédito fiscal": "iva_credito",
    "iva credito fiscal": "iva_credito",
    "iva débito": "iva_debito",
    "iva debito": "iva_debito",
    "iva débito fiscal": "iva_debito",
    "iva debito fiscal": "iva_debito",
}


# Normalización de nombres de empresa del Excel → código canónico en core.empresas.
# El Excel viene con casing inconsistente y nombres expandidos; la app usa
# códigos UPPER. C&E es el código de la hoja de CENERGY.
EMPRESA_NAME_MAP: dict[str, str] = {
    "csl": "CSL",
    "trongkai": "TRONGKAI",
    "revtech": "REVTECH",
    "rho": "RHO",
    "evoque": "EVOQUE",
    "dte": "DTE",
    "afis": "AFIS",
    "fip_cehta": "FIP_CEHTA",
    "fip cehta": "FIP_CEHTA",
    "fip-cehta": "FIP_CEHTA",
    "cenergy": "CENERGY",
    "c&e": "CENERGY",
    "c & e": "CENERGY",
    "c y e": "CENERGY",
}


def normalize_empresa_codigo(value: Any) -> str | None:
    """Normaliza el campo EMPRESA del Excel a un código canónico (UPPER).

    "Trongkai" → "TRONGKAI"
    "Evoque"   → "EVOQUE"
    "C&E"      → "CENERGY"
    None / ""  → None
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    key = s.casefold()
    if key in EMPRESA_NAME_MAP:
        return EMPRESA_NAME_MAP[key]
    # Default: uppercase + underscores
    return s.upper().replace(" ", "_")


# ---------------------------------------------------------------------------
# Result / typed dataclasses
# ---------------------------------------------------------------------------


@dataclass
class RawRow:
    """Fila tal cual viene del Excel (todos los campos string-coerced)."""

    source_row_num: int
    data: dict[str, str | None]


@dataclass
class RejectedRowDTO:
    source_row_num: int
    reason: str
    raw_data: dict[str, Any]


@dataclass
class ETLResult:
    """Resumen serializable del run para responder por API."""

    run_id: str | None
    status: str  # 'success' | 'partial' | 'failed' | 'skipped'
    source_file: str
    source_hash: str | None
    rows_extracted: int = 0
    rows_loaded: int = 0
    rows_rejected: int = 0
    error_message: str | None = None
    snapshot_path: str | None = None
    triggered_by: str = "manual"
    started_at: datetime | None = None
    finished_at: datetime | None = None
    rejected_samples: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "run_id": self.run_id,
            "status": self.status,
            "source_file": self.source_file,
            "source_hash": self.source_hash,
            "rows_extracted": self.rows_extracted,
            "rows_loaded": self.rows_loaded,
            "rows_rejected": self.rows_rejected,
            "error_message": self.error_message,
            "snapshot_path": self.snapshot_path,
            "triggered_by": self.triggered_by,
            "started_at": self.started_at.isoformat() if self.started_at else None,
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
            "rejected_samples": self.rejected_samples[:10],
        }


# ---------------------------------------------------------------------------
# Pure helpers (no I/O — fáciles de testear)
# ---------------------------------------------------------------------------


def compute_file_hash(content: bytes) -> str:
    """SHA256 hex (64 chars) del contenido binario."""
    return hashlib.sha256(content).hexdigest()


def compute_natural_key(row: dict[str, Any]) -> str:
    """Hash determinístico de (fecha, descripcion, abono, egreso, empresa, banco).

    Sirve como identidad estable de la fila para UPSERT idempotente. Cualquier
    cambio en uno de estos campos genera nueva fila — diseño deliberado: si
    cambia el monto o la fecha, es lógicamente otro movimiento.
    """
    fecha = str(row.get("fecha") or "")
    desc = (row.get("descripcion") or "").strip().casefold()
    abono = str(row.get("abono") or "0")
    egreso = str(row.get("egreso") or "0")
    empresa = (row.get("empresa_codigo") or "").strip().upper()
    banco = (row.get("banco") or "").strip().casefold()
    payload = f"{fecha}|{desc}|{abono}|{egreso}|{empresa}|{banco}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip().casefold()
    return HEADER_ALIASES.get(s, s.replace(" ", "_") if s else None)


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    s = str(value).strip()
    return s if s else None


def _to_decimal(value: Any) -> Decimal | None:
    """Convierte celda a Decimal. Acepta números, strings con coma o punto.
    Devuelve None si no se puede parsear (string vacío, texto no numérico, etc.).
    """
    if value is None:
        return None
    if isinstance(value, int | float | Decimal):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    s = str(value).strip()
    if not s:
        return None
    # Soporta '1.234,56' (es-CL), '1,234.56' (en) y '1234.56' (raw).
    if "," in s and "." in s:
        # Asumir el último separador como decimal, los otros son miles
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_date(value: Any) -> date | None:
    """Convierte celda a `date`. Acepta datetime, date, ISO string, dd/mm/yyyy.

    El parser openpyxl + nuestro `_coerce_str` convierten datetime a
    `2024-05-31T00:00:00` (ISO con T). Tenemos que aceptarlo.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Probamos primero `fromisoformat` que maneja '2024-05-31',
    # '2024-05-31T00:00:00', '2024-05-31 00:00:00' y muchas variantes ISO.
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        pass
    # Fallback a formatos comunes no-ISO (es-CL).
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _is_valid_periodo(value: str | None, anio: int | None) -> bool:
    """Periodo debe ser 'MM_YY' (ej '11_25', '02_26') y consistente con anio."""
    if not value:
        return False
    parts = value.split("_")
    if len(parts) != 2:
        return False
    mm_s, yy_s = parts
    if not (mm_s.isdigit() and yy_s.isdigit()):
        return False
    mm = int(mm_s)
    yy = int(yy_s)
    if not (1 <= mm <= 12):
        return False
    if anio is None:
        return True
    # YY → expandir a YYYY: heurística simple. Si YY < 70 → 20YY, sino 19YY.
    expected_yy = anio % 100
    return yy == expected_yy


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------


def parse_resumen_sheet(
    content: bytes,
) -> tuple[list[RawRow], list[RejectedRowDTO]]:
    """Parsea la hoja 'Resumen' del Excel madre.

    Devuelve `(raw_rows, rejected_at_parse)`. Las rejected_at_parse son filas
    completamente vacías o estructuralmente rotas; el resto se devuelve como
    raw para validación posterior.
    """
    raw_rows: list[RawRow] = []
    rejected: list[RejectedRowDTO] = []

    with io.BytesIO(content) as buf:
        wb = load_workbook(buf, read_only=True, data_only=True)
        try:
            # Buscamos la hoja en orden de prioridad: CONSOLIDA (formato real
            # del Excel Cehta con todas las empresas) → Resumen (legacy) →
            # primera hoja con datos.
            sheet_name: str | None = None
            sheet_names_lower = {s.casefold(): s for s in wb.sheetnames}
            for candidate in CONSOLIDA_SHEET_CANDIDATES:
                if candidate in wb.sheetnames:
                    sheet_name = candidate
                    break
                if candidate.casefold() in sheet_names_lower:
                    sheet_name = sheet_names_lower[candidate.casefold()]
                    break

            if sheet_name is None:
                # Last resort: primera hoja con >100 filas (asumimos data sheet).
                for s in wb.sheetnames:
                    if (wb[s].max_row or 0) > 100:
                        sheet_name = s
                        break
                if sheet_name is None:
                    sheet_name = wb.sheetnames[0]

            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return [], []

            header_map: dict[int, str] = {}
            for idx, cell in enumerate(header_row):
                norm = _normalize_header(cell)
                if norm:
                    header_map[idx] = norm

            if not header_map:
                return [], []

            for row_num, row in enumerate(rows_iter, start=2):
                # Fila completamente vacía → ignorar silencioso (no es rejected).
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                    continue

                raw_data: dict[str, str | None] = {}
                for idx, val in enumerate(row):
                    field_name = header_map.get(idx)
                    if field_name:
                        raw_data[field_name] = _coerce_str(val)

                raw_rows.append(RawRow(source_row_num=row_num, data=raw_data))
        finally:
            wb.close()

    return raw_rows, rejected


# ---------------------------------------------------------------------------
# Validation + transform
# ---------------------------------------------------------------------------


def _validate_and_transform_row(
    raw: RawRow, valid_empresas: set[str]
) -> tuple[dict[str, Any] | None, RejectedRowDTO | None]:
    """Aplica las reglas de negocio. Devuelve (row_dict, None) si OK, o
    (None, rejected) si falla. Las reglas:

    1. fecha no puede estar vacía / inválida.
    2. empresa debe existir en core.empresas.
    3. periodo debe ser 'MM_YY' válido.
    4. periodo (anio, MM_YY) deben ser consistentes.
    5. abono > 0 y egreso > 0 simultáneamente NO permitido.
    """
    d = raw.data
    raw_payload: dict[str, Any] = dict(d)  # snapshot para rejected.raw_data

    # 1) Fecha
    fecha = _to_date(d.get("fecha"))
    if fecha is None:
        return None, RejectedRowDTO(raw.source_row_num, "fecha vacía o inválida", raw_payload)

    # 2) Empresa — normalizar usando el mapeo (Evoque→EVOQUE, C&E→CENERGY, etc.)
    empresa_raw = d.get("empresa")
    empresa_codigo = normalize_empresa_codigo(empresa_raw)
    if not empresa_codigo:
        return None, RejectedRowDTO(
            raw.source_row_num, "empresa vacía", raw_payload
        )
    if empresa_codigo not in valid_empresas:
        return None, RejectedRowDTO(
            raw.source_row_num,
            f"empresa '{empresa_codigo}' (raw: {empresa_raw!r}) no existe en core.empresas",
            raw_payload,
        )

    # 3) Periodo del Excel (puede tener whitespace tras pegar de Excel)
    periodo = (d.get("periodo") or "").strip()

    # 4) Año — derivamos de la fecha (más confiable que la columna AÑO del Excel,
    # que en este dataset tiene typos: filas con fecha=2025 pero AÑO=2024).
    # Si la columna AÑO está vacía, también usamos fecha.year.
    anio_raw = d.get("anio")
    anio_excel: int | None = None
    if anio_raw is not None:
        try:
            anio_excel = int(float(str(anio_raw).strip()))
        except (ValueError, TypeError):
            anio_excel = None

    # Authoritative: año de la fecha. anio_excel solo se usa para warning.
    anio = fecha.year

    # 5) Periodo: si viene vacío o inválido, auto-derivar de la fecha.
    # El periodo en el Excel es el "mes contable" pero en general coincide
    # con el mes de la fecha. Si difiere o falta, usamos la fecha como
    # autoridad — es más confiable que celdas mal pegadas.
    expected_periodo = f"{fecha.month:02d}_{fecha.year % 100:02d}"
    if not periodo or not _is_valid_periodo(periodo, anio):
        periodo = expected_periodo

    # 5) Abono y egreso
    abono = _to_decimal(d.get("abonos")) or Decimal("0")
    egreso = _to_decimal(d.get("egreso")) or Decimal("0")
    if abono > 0 and egreso > 0:
        return None, RejectedRowDTO(
            raw.source_row_num,
            f"abono ({abono}) y egreso ({egreso}) ambos > 0 en misma fila",
            raw_payload,
        )

    # Real/Proyectado — sólo aceptamos esos dos valores; cualquier otra cosa
    # se setea en None (no es rejected — es flag opcional).
    real_proy = (d.get("real_proyectado") or "").strip()
    real_proy_norm: str | None = None
    if real_proy.casefold() == "real":
        real_proy_norm = "Real"
    elif real_proy.casefold() in {"proyectado", "proy"}:
        real_proy_norm = "Proyectado"

    row_dict: dict[str, Any] = {
        "fecha": fecha,
        "descripcion": d.get("descripcion"),
        "abono": abono,
        "egreso": egreso,
        "saldo_contable": _to_decimal(d.get("saldo_contable")),
        "saldo_cehta": _to_decimal(d.get("saldo_cehta")),
        "saldo_corfo": _to_decimal(d.get("saldo_corfo")),
        "concepto_general": d.get("concepto_general"),
        "concepto_detallado": d.get("concepto_detallado"),
        "tipo_egreso": d.get("tipo_egreso"),
        "fuente": d.get("fuentes"),
        "proyecto": d.get("proyecto"),
        "banco": d.get("banco"),
        "real_proyectado": real_proy_norm,
        "anio": anio,
        "periodo": periodo,
        "empresa_codigo": empresa_codigo,
        "iva_credito_fiscal": _to_decimal(d.get("iva_credito")) or Decimal("0"),
        "iva_debito_fiscal": _to_decimal(d.get("iva_debito")) or Decimal("0"),
        "hipervinculo": d.get("hipervinculo"),
    }
    row_dict["natural_key"] = compute_natural_key(row_dict)
    return row_dict, None


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------


async def _fetch_valid_empresa_codigos(db: AsyncSession) -> set[str]:
    rows = (await db.scalars(select(Empresa.codigo))).all()
    return {r for r in rows if r}


async def _last_success_hash(db: AsyncSession) -> str | None:
    q = (
        select(EtlRun.source_hash)
        .where(EtlRun.status == "success")
        .order_by(EtlRun.started_at.desc())
        .limit(1)
    )
    return await db.scalar(q)


async def check_should_run(db: AsyncSession, file_hash: str) -> bool:
    """True si hay que correr (hash distinto al último success)."""
    last = await _last_success_hash(db)
    return last != file_hash


# ---------------------------------------------------------------------------
# Service principal
# ---------------------------------------------------------------------------


class ETLService:
    """Orquestador del ETL Dropbox→Postgres."""

    def __init__(self) -> None:
        pass

    # ----- discovery -----

    def _find_data_madre_path(self, dropbox: DropboxService) -> str | None:
        """Localiza Data Madre.xlsx en el Dropbox conectado.

        Estrategia:
          1. Buscar `Cehta Capital/` en root.
          2. Dentro buscar la primera variante de 'Inteligencia de Negocios'
             (con o sin prefijo `00-`).
          3. Listar y devolver path del archivo cuyo nombre contenga 'data madre'.
        """
        cehta = dropbox.find_folder(CEHTA_ROOT, "")
        search_root = cehta or ""
        ig_path: str | None = None
        for candidate in INTELIGENCIA_FOLDER_CANDIDATES:
            ig_path = dropbox.find_folder(candidate, search_root)
            if ig_path:
                break
        if not ig_path:
            return None
        items = dropbox.list_folder(ig_path)
        for item in items:
            if (
                item["type"] == "file"
                and DATA_MADRE_FILENAME_SUBSTR in item["name"].casefold()
            ):
                path = item["path"]
                return path if isinstance(path, str) else None
        return None

    # ----- raw load -----

    async def _insert_raw_rows(
        self,
        db: AsyncSession,
        run_id: str,
        raw_rows: list[RawRow],
    ) -> None:
        """Inserta filas crudas en raw.resumen_excel para reproducibilidad.
        Usa text() + executemany para performance — bypass del ORM porque
        no necesitamos modelo SQLAlchemy de raw.resumen_excel.
        """
        if not raw_rows:
            return
        sql = text(
            """
            INSERT INTO raw.resumen_excel (
                run_id, source_row_num,
                hipervinculo, fecha, descripcion, abonos, egreso,
                saldo_contable, saldo_cehta, saldo_corfo,
                concepto_general, concepto_detallado, tipo_egreso,
                fuentes, proyecto, banco, real_proyectado,
                anio, periodo, empresa, iva_credito, iva_debito
            ) VALUES (
                :run_id, :source_row_num,
                :hipervinculo, :fecha, :descripcion, :abonos, :egreso,
                :saldo_contable, :saldo_cehta, :saldo_corfo,
                :concepto_general, :concepto_detallado, :tipo_egreso,
                :fuentes, :proyecto, :banco, :real_proyectado,
                :anio, :periodo, :empresa, :iva_credito, :iva_debito
            )
            """
        )
        params: list[dict[str, Any]] = []
        for raw in raw_rows:
            d = raw.data
            params.append(
                {
                    "run_id": run_id,
                    "source_row_num": raw.source_row_num,
                    "hipervinculo": d.get("hipervinculo"),
                    "fecha": d.get("fecha"),
                    "descripcion": d.get("descripcion"),
                    "abonos": d.get("abonos"),
                    "egreso": d.get("egreso"),
                    "saldo_contable": d.get("saldo_contable"),
                    "saldo_cehta": d.get("saldo_cehta"),
                    "saldo_corfo": d.get("saldo_corfo"),
                    "concepto_general": d.get("concepto_general"),
                    "concepto_detallado": d.get("concepto_detallado"),
                    "tipo_egreso": d.get("tipo_egreso"),
                    "fuentes": d.get("fuentes"),
                    "proyecto": d.get("proyecto"),
                    "banco": d.get("banco"),
                    "real_proyectado": d.get("real_proyectado"),
                    "anio": d.get("anio"),
                    "periodo": d.get("periodo"),
                    "empresa": d.get("empresa"),
                    "iva_credito": d.get("iva_credito"),
                    "iva_debito": d.get("iva_debito"),
                }
            )
        await db.execute(sql, params)

    async def _upsert_movimientos(
        self,
        db: AsyncSession,
        run_id: str,
        rows: list[dict[str, Any]],
    ) -> int:
        """UPSERT por natural_key. Devuelve cantidad de filas afectadas."""
        if not rows:
            return 0
        sql = text(
            """
            INSERT INTO core.movimientos (
                natural_key, fecha, descripcion,
                abono, egreso, saldo_contable, saldo_cehta, saldo_corfo,
                concepto_general, concepto_detallado, tipo_egreso,
                fuente, proyecto, banco,
                real_proyectado, anio, periodo, empresa_codigo,
                iva_credito_fiscal, iva_debito_fiscal,
                hipervinculo, run_id
            ) VALUES (
                :natural_key, :fecha, :descripcion,
                :abono, :egreso, :saldo_contable, :saldo_cehta, :saldo_corfo,
                :concepto_general, :concepto_detallado, :tipo_egreso,
                :fuente, :proyecto, :banco,
                :real_proyectado, :anio, :periodo, :empresa_codigo,
                :iva_credito_fiscal, :iva_debito_fiscal,
                :hipervinculo, :run_id
            )
            ON CONFLICT (natural_key) DO UPDATE SET
                fecha = EXCLUDED.fecha,
                descripcion = EXCLUDED.descripcion,
                abono = EXCLUDED.abono,
                egreso = EXCLUDED.egreso,
                saldo_contable = EXCLUDED.saldo_contable,
                saldo_cehta = EXCLUDED.saldo_cehta,
                saldo_corfo = EXCLUDED.saldo_corfo,
                concepto_general = EXCLUDED.concepto_general,
                concepto_detallado = EXCLUDED.concepto_detallado,
                tipo_egreso = EXCLUDED.tipo_egreso,
                fuente = EXCLUDED.fuente,
                proyecto = EXCLUDED.proyecto,
                banco = EXCLUDED.banco,
                real_proyectado = EXCLUDED.real_proyectado,
                anio = EXCLUDED.anio,
                periodo = EXCLUDED.periodo,
                empresa_codigo = EXCLUDED.empresa_codigo,
                iva_credito_fiscal = EXCLUDED.iva_credito_fiscal,
                iva_debito_fiscal = EXCLUDED.iva_debito_fiscal,
                hipervinculo = EXCLUDED.hipervinculo,
                run_id = EXCLUDED.run_id,
                updated_at = now()
            """
        )
        # FK a catálogos (concepto_general, concepto_detallado, tipo_egreso,
        # fuente, proyecto, banco) puede romper UPSERT si la fila no existe
        # todavía. Estrategia pragmática: pre-poblar catálogos para los
        # valores únicos vistos en este batch antes del upsert principal.
        await self._ensure_catalog_values(db, rows)

        params = [
            {**r, "run_id": run_id}
            for r in rows
        ]
        await db.execute(sql, params)
        return len(rows)

    async def _ensure_catalog_values(
        self, db: AsyncSession, rows: list[dict[str, Any]]
    ) -> None:
        """Inserta valores de catálogo que no existan aún. Idempotente.

        Las tablas core.concepto_general/etc. tienen el campo como PK; usamos
        ON CONFLICT DO NOTHING para no romper si el valor ya existe.
        """
        # Recolectar valores únicos no vacíos
        cg = {r["concepto_general"] for r in rows if r.get("concepto_general")}
        cd = {
            (r["concepto_detallado"], r.get("concepto_general"))
            for r in rows
            if r.get("concepto_detallado")
        }
        te = {r["tipo_egreso"] for r in rows if r.get("tipo_egreso")}
        fu = {r["fuente"] for r in rows if r.get("fuente")}
        pr = {r["proyecto"] for r in rows if r.get("proyecto")}
        bc = {r["banco"] for r in rows if r.get("banco")}

        if cg:
            await db.execute(
                text(
                    "INSERT INTO core.concepto_general (concepto_general) "
                    "VALUES (:v) ON CONFLICT DO NOTHING"
                ),
                [{"v": v} for v in cg],
            )
        if cd:
            await db.execute(
                text(
                    "INSERT INTO core.concepto_detallado "
                    "(concepto_detallado, concepto_general) VALUES (:cd, :cg) "
                    "ON CONFLICT DO NOTHING"
                ),
                [{"cd": cd_, "cg": cg_} for cd_, cg_ in cd],
            )
        if te:
            await db.execute(
                text(
                    "INSERT INTO core.tipo_egreso (tipo_egreso) "
                    "VALUES (:v) ON CONFLICT DO NOTHING"
                ),
                [{"v": v} for v in te],
            )
        if fu:
            await db.execute(
                text(
                    "INSERT INTO core.fuente (fuente) "
                    "VALUES (:v) ON CONFLICT DO NOTHING"
                ),
                [{"v": v} for v in fu],
            )
        if pr:
            await db.execute(
                text(
                    "INSERT INTO core.proyecto (proyecto) "
                    "VALUES (:v) ON CONFLICT DO NOTHING"
                ),
                [{"v": v} for v in pr],
            )
        if bc:
            await db.execute(
                text(
                    "INSERT INTO core.banco (banco) "
                    "VALUES (:v) ON CONFLICT DO NOTHING"
                ),
                [{"v": v} for v in bc],
            )

    async def _insert_rejected(
        self,
        db: AsyncSession,
        run_id: str,
        rejected: list[RejectedRowDTO],
        sheet: str = RESUMEN_SHEET,
    ) -> None:
        if not rejected:
            return
        for r in rejected:
            db.add(
                RejectedRow(
                    run_id=run_id,
                    source_sheet=sheet,
                    source_row_num=r.source_row_num,
                    reason=r.reason,
                    raw_data=r.raw_data,
                )
            )
        await db.flush()

    # ----- entrypoint -----

    async def run_etl(
        self,
        db: AsyncSession,
        dropbox: DropboxService,
        triggered_by: str = "manual",
    ) -> ETLResult:
        """Ejecuta el ETL completo. Devuelve `ETLResult` siempre — incluso
        si hubo skip o failure, para que el endpoint pueda responder algo
        coherente.
        """
        started_at = datetime.now(UTC)

        # 1. Localizar archivo
        try:
            file_path = self._find_data_madre_path(dropbox)
        except DropboxNotConfigured:
            raise
        except Exception as exc:  # pragma: no cover — defensivo
            log.exception("etl.find_data_madre_failed")
            return ETLResult(
                run_id=None,
                status="failed",
                source_file="",
                source_hash=None,
                error_message=f"No se pudo localizar Data Madre: {exc}",
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=datetime.now(UTC),
            )
        if not file_path:
            return ETLResult(
                run_id=None,
                status="failed",
                source_file="",
                source_hash=None,
                error_message=(
                    "Data Madre.xlsx no encontrado en Dropbox. Verificá que exista en "
                    f"'{CEHTA_ROOT}/00-Inteligencia de Negocios/'."
                ),
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=datetime.now(UTC),
            )

        # 2. Descargar + hash + idempotencia
        try:
            content = dropbox.download_file(file_path)
        except Exception as exc:
            log.exception("etl.download_failed", path=file_path)
            return ETLResult(
                run_id=None,
                status="failed",
                source_file=file_path,
                source_hash=None,
                error_message=f"Error descargando archivo: {exc}",
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=datetime.now(UTC),
            )

        file_hash = compute_file_hash(content)
        if not await check_should_run(db, file_hash):
            log.info("etl.skipped", reason="hash_unchanged", hash=file_hash[:12])
            return ETLResult(
                run_id=None,
                status="skipped",
                source_file=file_path,
                source_hash=file_hash,
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=datetime.now(UTC),
            )

        # 3. Crear etl_run row
        run_id = str(uuid4())
        run = EtlRun(
            run_id=run_id,
            source_file=file_path,
            source_hash=file_hash,
            status="running",
            triggered_by=triggered_by,
        )
        db.add(run)
        await db.flush()
        await db.commit()  # commit la row de running para que sea visible

        # A partir de acá, cualquier error se loggea en el run con status='failed'.
        try:
            # 4a. Parsear
            raw_rows, parse_rejected = parse_resumen_sheet(content)
            log.info("etl.parsed", rows=len(raw_rows), file=file_path)

            # 4b. Volcar raw
            await self._insert_raw_rows(db, run_id, raw_rows)

            # 4c. Validar + transform
            valid_empresas = await _fetch_valid_empresa_codigos(db)
            transformed: list[dict[str, Any]] = []
            rejected: list[RejectedRowDTO] = list(parse_rejected)
            for raw in raw_rows:
                row_dict, rej = _validate_and_transform_row(raw, valid_empresas)
                if rej is not None:
                    rejected.append(rej)
                elif row_dict is not None:
                    transformed.append(row_dict)

            # 4d. UPSERT a core.movimientos
            loaded = await self._upsert_movimientos(db, run_id, transformed)

            # 4e. Persistir rejected
            await self._insert_rejected(db, run_id, rejected)

            # 4f. Snapshot histórico (best-effort — fallo acá no rompe el run)
            snapshot_path: str | None = None
            try:
                snapshot_path = self._save_historico_snapshot(
                    dropbox, file_path, content
                )
            except Exception as exc:  # pragma: no cover — no debe fallar el run
                log.warning("etl.snapshot_failed", error=str(exc))

            # 4g. Cerrar run con status final
            final_status = "success" if not rejected else "partial"
            run.status = final_status
            run.rows_extracted = len(raw_rows)
            run.rows_loaded = loaded
            run.rows_rejected = len(rejected)
            run.finished_at = datetime.now(UTC)
            await db.flush()
            await db.commit()

            result = ETLResult(
                run_id=run_id,
                status=final_status,
                source_file=file_path,
                source_hash=file_hash,
                rows_extracted=len(raw_rows),
                rows_loaded=loaded,
                rows_rejected=len(rejected),
                snapshot_path=snapshot_path,
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=run.finished_at,
                rejected_samples=[
                    {
                        "row": r.source_row_num,
                        "reason": r.reason,
                    }
                    for r in rejected[:10]
                ],
            )
            log.info(
                "etl.completed",
                run_id=run_id,
                status=final_status,
                loaded=loaded,
                rejected=len(rejected),
            )
            # Real-time push (V4 fase 2): admins viendo /admin/etl reciben
            # el evento sin tener que refrescar.
            try:
                await get_broadcaster().publish(
                    "etl.completed",
                    {
                        "run_id": run_id,
                        "status": final_status,
                        "rows_loaded": loaded,
                        "rows_rejected": len(rejected),
                        "rows_extracted": len(raw_rows),
                        "source_file": file_path,
                        "triggered_by": triggered_by,
                    },
                    role="admin",
                )
            except Exception as pub_exc:  # pragma: no cover — defensivo
                log.warning("etl_sse_publish_failed", error=str(pub_exc))
            return result

        except Exception as exc:
            log.exception("etl.failed", run_id=run_id)
            try:
                # Intentar dejar el run marcado como failed (rollback puede haber
                # invalidado la sesión, así que abrimos otra ruta — usamos text()
                # raw connection-level update vía session si todavía está viva).
                await db.rollback()
                await db.execute(
                    text(
                        "UPDATE audit.etl_runs SET status='failed', "
                        "error_message=:err, finished_at=now() "
                        "WHERE run_id = :rid"
                    ),
                    {"err": str(exc)[:1000], "rid": run_id},
                )
                await db.commit()
            except Exception:  # pragma: no cover — defensa contra DB rota
                log.exception("etl.failed_to_mark_failed")
            # Real-time push: avisar a admins que el ETL falló.
            # Defensivo — el error de SSE no debe ocultar el error original.
            with contextlib.suppress(Exception):
                await get_broadcaster().publish(
                    "etl.failed",
                    {
                        "run_id": run_id,
                        "source_file": file_path,
                        "error_message": str(exc)[:500],
                        "triggered_by": triggered_by,
                    },
                    role="admin",
                )
            return ETLResult(
                run_id=run_id,
                status="failed",
                source_file=file_path,
                source_hash=file_hash,
                error_message=str(exc)[:500],
                triggered_by=triggered_by,
                started_at=started_at,
                finished_at=datetime.now(UTC),
            )

    # ----- snapshot -----

    def _save_historico_snapshot(
        self, dropbox: DropboxService, source_path: str, content: bytes
    ) -> str:
        """Sube una copia del Excel a `Histórico/YYYY-MM-DD-Data-Madre.xlsx`."""
        # Inferir directorio padre del Excel
        parent = source_path.rsplit("/", 1)[0] if "/" in source_path else ""
        target_dir = f"{parent}/{HISTORICO_SUBFOLDER}"
        # `create_folder` ya es idempotente cuando la carpeta existe (chequea
        # 'conflict' en el ApiError). Si falla por otra razón, suprimimos —
        # no queremos romper el snapshot por un edge case de permisos cuando
        # el upload del archivo final probablemente igual va a funcionar.
        with contextlib.suppress(Exception):
            dropbox.create_folder(target_dir)
        today = date.today().isoformat()
        target_path = f"{target_dir}/{today}-Data-Madre.xlsx"
        dropbox.upload_file(target_path, content, overwrite=True)
        return target_path
