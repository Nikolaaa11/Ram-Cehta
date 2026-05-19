"""Round 123 — Parser del Libro de Remuneraciones de Nubox.

Nubox exporta el Libro de Remuneraciones mensual como Excel (.xlsx).
El formato suele tener:
  - 3-5 filas de cabecera (logo, razón social, período, etc.)
  - 1 fila con nombres de columnas
  - N filas: 1 por trabajador
  - 1-2 filas de totales al final

Columnas típicas (varían por configuración de la empresa):
  RUT | Nombre | Cargo | Sueldo Base | Gratificación |
  Horas Extras | Bonos | Colación | Movilización | Otros Haberes |
  TOTAL HABERES |
  AFP | Salud | AFC | Impuesto Único | Otros Desc | TOTAL DESCUENTOS |
  SUELDO LÍQUIDO |
  SIS | AFC Patronal | Mutual

El parser detecta los headers por nombre canónico (tolerante a reorder
y nombres con/sin tildes) e ignora filas que no son trabajadores.

Compatible con xlsx exportado de Nubox y con xlsx que el contador
externo MCG mande en formato similar.
"""
from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


# Mapeo de headers canónicos → posibles aliases en el Excel real
HEADER_ALIASES: dict[str, list[str]] = {
    "trabajador_rut": [
        "rut", "rut trabajador", "rut empleado", "rut/dni",
    ],
    "trabajador_nombre": [
        "nombre", "nombre completo", "trabajador", "empleado",
        "apellido y nombre", "nombre apellido",
    ],
    # Haberes
    "sueldo_base": [
        "sueldo base", "sueldo bruto", "remuneracion base",
        "remuneracion mensual", "sueldo",
    ],
    "gratificacion": ["gratificacion", "gratif", "gratificacion legal"],
    "horas_extras": ["horas extras", "he", "horas extra", "h extras"],
    "bonos": ["bonos", "bono", "incentivos", "bono produccion"],
    "colacion": ["colacion", "colación", "alimentacion"],
    "movilizacion": ["movilizacion", "movilizacion no tributable", "transporte"],
    "otros_haberes": ["otros haberes", "otros", "haberes adicionales"],
    "total_haberes": [
        "total haberes", "tot haberes", "total imponible",
        "total bruto", "total remuneraciones",
    ],
    # Descuentos
    "afp_descuento": ["afp", "descuento afp", "cotiz afp"],
    "salud_descuento": [
        "salud", "isapre", "fonasa", "descuento salud", "cotiz salud",
    ],
    "afc_descuento": ["afc", "seguro cesantia", "descuento afc"],
    "impuesto_unico": [
        "impuesto unico", "impuesto", "imp unico", "iut",
        "impuesto a la renta",
    ],
    "otros_descuentos": ["otros descuentos", "otros desc", "descuentos varios"],
    "total_descuentos": [
        "total descuentos", "tot descuentos", "total descuento",
    ],
    "sueldo_liquido": [
        "sueldo liquido", "líquido", "liquido a pagar",
        "neto a pagar", "sueldo neto", "liquido",
    ],
    # Patronal
    "sis_patronal": ["sis", "sis patronal", "seguro invalidez"],
    "afc_patronal": [
        "afc patronal", "aporte afc patronal", "afc empleador",
    ],
    "mutual_patronal": [
        "mutual", "mutual patronal", "aporte mutual", "achs", "mutuales",
    ],
}


def _norm(s: object) -> str:
    """Normaliza para comparación: minúsculas, sin tildes, sin espacios extras."""
    if s is None:
        return ""
    out = str(s).strip().lower()
    repl = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
    for k, v in repl.items():
        out = out.replace(k, v)
    # Compactar espacios y eliminar caracteres no alfanuméricos al final
    out = re.sub(r"\s+", " ", out).strip()
    return out


def _to_int(v: Any) -> int:
    """'1.234.567' / '$1,234' / Decimal / float → int. Tolerante."""
    if v is None or v == "":
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip()
    # Quitar $ y otros símbolos
    s = re.sub(r"[$\s]", "", s)
    # Detectar negativos (paréntesis o signo menos)
    negative = False
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
        negative = True
    if s.startswith("-"):
        s = s[1:]
        negative = True
    # Eliminar puntos de miles (típico CLP) y dejar coma como decimal
    # En montos CLP no hay decimales, así que sacamos puntos directamente
    if "," in s and s.count(".") >= 1:
        # Formato "1.234.567,00" — sacar puntos, coma → punto decimal
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "")
    if not s or s == "-":
        return 0
    try:
        val = int(float(s))
        return -val if negative else val
    except ValueError:
        return 0


def _resolve_headers(headers_row: list[Any]) -> dict[str, int]:
    """Encuentra qué columna corresponde a cada campo canónico."""
    out: dict[str, int] = {}
    normalized = [_norm(h) for h in headers_row]
    for idx, h_norm in enumerate(normalized):
        if not h_norm:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in out:
                continue
            for alias in aliases:
                if _norm(alias) == h_norm:
                    out[canonical] = idx
                    break
    return out


def _looks_like_header_row(row: list[Any]) -> bool:
    """True si esta fila tiene >=3 nombres de campo canónicos.

    Sirve para encontrar la fila de headers en un xlsx con muchas filas de
    título arriba (logo, razón social, período...).
    """
    matched = 0
    norms = [_norm(c) for c in row]
    all_aliases_norm = {
        _norm(a) for aliases in HEADER_ALIASES.values() for a in aliases
    }
    for n in norms:
        if n in all_aliases_norm:
            matched += 1
            if matched >= 3:
                return True
    return False


def _is_total_row(row: list[Any]) -> bool:
    """Filas con 'TOTAL' en primera celda → ignorar."""
    if not row:
        return False
    first = _norm(row[0])
    return first in {"total", "totales", "total general"} or "total" in first


def parse_libro_remuneraciones(
    xlsx_path_or_bytes: str | bytes,
    periodo: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Parsea el xlsx del Libro de Remuneraciones.

    `periodo` se asigna a todas las filas (Nubox no siempre lo incluye en
    cada fila, viene en el header del documento).

    Returns: (lista de dicts con campos canónicos, lista de warnings/errors)
    """
    errors: list[str] = []

    try:
        if isinstance(xlsx_path_or_bytes, bytes):
            from io import BytesIO
            wb = load_workbook(BytesIO(xlsx_path_or_bytes), data_only=True, read_only=True)
        else:
            wb = load_workbook(xlsx_path_or_bytes, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"No se pudo abrir el xlsx: {exc}"]

    # Usar la primera hoja con datos
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    # Encontrar la fila de headers
    header_idx = -1
    for i, r in enumerate(rows[:30]):  # buscar en las primeras 30 filas
        row_list = list(r) if r else []
        if _looks_like_header_row(row_list):
            header_idx = i
            break

    if header_idx < 0:
        return [], [
            "No se encontró fila de headers reconocibles. "
            "El xlsx debe tener al menos 3 columnas con nombres como "
            "'RUT', 'Sueldo Base', 'AFP', 'Sueldo Líquido'."
        ]

    headers = list(rows[header_idx])
    header_map = _resolve_headers(headers)

    # Campos mínimos requeridos
    required = {"trabajador_rut", "sueldo_liquido"}
    missing = required - set(header_map.keys())
    if missing:
        return [], [
            f"Columnas requeridas no encontradas: {sorted(missing)}. "
            f"Headers vistos: {headers}"
        ]

    data: list[dict[str, Any]] = []
    for line_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        row_list = list(row) if row else []
        if not row_list or all((c is None or str(c).strip() == "") for c in row_list):
            continue
        if _is_total_row(row_list):
            continue

        def cell(field: str) -> Any:
            idx = header_map.get(field)
            if idx is None or idx >= len(row_list):
                return None
            return row_list[idx]

        # RUT vacío → skip silencioso (filas de subtotales por sección)
        rut = cell("trabajador_rut")
        if not rut or not str(rut).strip():
            continue

        try:
            item = {
                "periodo": periodo,
                "trabajador_rut": str(rut).strip(),
                "trabajador_nombre": (
                    str(cell("trabajador_nombre") or "").strip() or None
                ),
                "sueldo_base": _to_int(cell("sueldo_base")),
                "gratificacion": _to_int(cell("gratificacion")),
                "horas_extras": _to_int(cell("horas_extras")),
                "bonos": _to_int(cell("bonos")),
                "colacion": _to_int(cell("colacion")),
                "movilizacion": _to_int(cell("movilizacion")),
                "otros_haberes": _to_int(cell("otros_haberes")),
                "total_haberes": _to_int(cell("total_haberes")),
                "afp_descuento": _to_int(cell("afp_descuento")),
                "salud_descuento": _to_int(cell("salud_descuento")),
                "afc_descuento": _to_int(cell("afc_descuento")),
                "impuesto_unico": _to_int(cell("impuesto_unico")),
                "otros_descuentos": _to_int(cell("otros_descuentos")),
                "total_descuentos": _to_int(cell("total_descuentos")),
                "sueldo_liquido": _to_int(cell("sueldo_liquido")),
                "sis_patronal": _to_int(cell("sis_patronal")),
                "afc_patronal": _to_int(cell("afc_patronal")),
                "mutual_patronal": _to_int(cell("mutual_patronal")),
            }
            # Si no vino total_haberes, calcularlo desde componentes
            if not item["total_haberes"]:
                item["total_haberes"] = (
                    item["sueldo_base"] + item["gratificacion"]
                    + item["horas_extras"] + item["bonos"]
                    + item["colacion"] + item["movilizacion"]
                    + item["otros_haberes"]
                )
            # Idem total_descuentos
            if not item["total_descuentos"]:
                item["total_descuentos"] = (
                    item["afp_descuento"] + item["salud_descuento"]
                    + item["afc_descuento"] + item["impuesto_unico"]
                    + item["otros_descuentos"]
                )
            data.append(item)
        except Exception as exc:  # noqa: BLE001 — defensive
            errors.append(f"Fila {line_idx}: {exc}")

    return data, errors
