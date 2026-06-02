"""R152vvv · Parser de libros de remuneraciones (formato Nubox / estándar Chile).

El Excel típico de Nubox tiene esta estructura:

  Filas 1-7: header empresa (razón social, RUT, dirección, mes)
  Fila 9:    headers de la primera tabla (Cód, RUT, Nombre, DT, S.Base, ...)
  Filas 10+: una fila por empleado, intercaladas con filas-subtotal por área
             (donde columna A está vacía pero columna B tiene el cod área)
  Fila TOTAL GENERAL antes del separador
  ~Fila 18:  headers "Patronales" + "Calculo Imp Unico"
  Fila 19:   headers segunda tabla
  Filas 20+: una fila por empleado con aportes patronales + cálculo IUSC

El parser localiza las dos tablas, las alinea por RUT, y entrega una
lista de dicts con TODOS los campos por empleado.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl


# ─────────────────────────────────────────────────────────────────────
# Tipos
# ─────────────────────────────────────────────────────────────────────


@dataclass
class LibroLinea:
    """Una fila del libro = un empleado-mes."""

    rut: str
    nombre: str
    area: str | None = None
    dias_trabajados: Decimal = Decimal("30")

    # Haberes imponibles
    sueldo_base: Decimal = Decimal("0")
    horas_extras: Decimal = Decimal("0")
    gratificacion_legal: Decimal = Decimal("0")
    otros_imponibles: Decimal = Decimal("0")
    total_imponibles: Decimal = Decimal("0")

    # Haberes no imponibles
    asignacion_familiar: Decimal = Decimal("0")
    otros_no_imponibles: Decimal = Decimal("0")
    total_no_imponibles: Decimal = Decimal("0")

    total_haberes: Decimal = Decimal("0")

    # Descuentos legales (trabajador)
    prevision: Decimal = Decimal("0")
    salud: Decimal = Decimal("0")
    seguro_cesantia_trab: Decimal = Decimal("0")
    otros_descuentos_legales: Decimal = Decimal("0")
    total_descuentos_legales: Decimal = Decimal("0")

    # Descuentos varios
    descuentos_varios: Decimal = Decimal("0")
    total_descuentos: Decimal = Decimal("0")

    liquido_pagado: Decimal = Decimal("0")

    # Aportes patronales (segunda tabla)
    aporte_afp_empleador: Decimal = Decimal("0")
    sis: Decimal = Decimal("0")
    seguro_cesantia_empleador: Decimal = Decimal("0")
    seguro_social: Decimal = Decimal("0")
    mutual: Decimal = Decimal("0")
    total_aportes_patronales: Decimal = Decimal("0")

    # Impuesto único
    base_tributable: Decimal = Decimal("0")
    impuesto_unico: Decimal = Decimal("0")

    # Calculado
    costo_total_empresa: Decimal = Decimal("0")


@dataclass
class LibroParseado:
    """Resultado completo del parser para un Excel."""

    empresa_razon_social: str
    empresa_rut: str
    periodo: str  # 'YYYY-MM'
    mes_label: str  # 'ABRIL DEL 2026'
    lineas: list[LibroLinea] = field(default_factory=list)
    archivo_hash: str = ""
    archivo_origen: str = ""

    @property
    def total_haberes(self) -> Decimal:
        return sum((l.total_haberes for l in self.lineas), Decimal("0"))

    @property
    def total_liquido(self) -> Decimal:
        return sum((l.liquido_pagado for l in self.lineas), Decimal("0"))

    @property
    def total_aportes_patronales(self) -> Decimal:
        return sum(
            (l.total_aportes_patronales for l in self.lineas), Decimal("0")
        )

    @property
    def total_costo_empresa(self) -> Decimal:
        return sum((l.costo_total_empresa for l in self.lineas), Decimal("0"))

    @property
    def total_descuentos_legales(self) -> Decimal:
        return sum(
            (l.total_descuentos_legales for l in self.lineas), Decimal("0")
        )


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────


_MESES = {
    "ENERO": "01", "FEBRERO": "02", "MARZO": "03", "ABRIL": "04",
    "MAYO": "05", "JUNIO": "06", "JULIO": "07", "AGOSTO": "08",
    "SEPTIEMBRE": "09", "OCTUBRE": "10", "NOVIEMBRE": "11", "DICIEMBRE": "12",
}


def _dec(v: Any) -> Decimal:
    """Convierte celda a Decimal seguro. None / '' → 0."""
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float)):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"))
    try:
        s = str(v).strip().replace(".", "").replace(",", ".")
        return Decimal(s).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _periodo_from_label(label: str) -> tuple[str, str]:
    """'MES: ABRIL DEL 2026' → ('2026-04', 'ABRIL DEL 2026')."""
    cleaned = (label or "").replace("MES:", "").strip().upper()
    parts = cleaned.split()
    # Encontrar mes + año
    mes = next((m for m in parts if m in _MESES), None)
    año = next(
        (int(p) for p in parts if p.isdigit() and 2020 <= int(p) <= 2099),
        None,
    )
    if not mes or not año:
        # Default al mes actual si no se pudo parsear
        now = datetime.now()
        return f"{now.year}-{now.month:02d}", cleaned
    return f"{año}-{_MESES[mes]}", cleaned


def _normalize_rut(rut: Any) -> str:
    """'7736580-K' → '7736580-K'. Limpia puntos y mayúscula DV."""
    s = str(rut or "").strip().replace(".", "").replace(" ", "").upper()
    if not s or s == "0":
        return ""
    # Si vino sin guion (parser puede comerse el guion en algunos casos)
    if "-" not in s and len(s) > 1:
        s = f"{s[:-1]}-{s[-1]}"
    return s


def _is_empty_row(ws, row: int, cols: range) -> bool:
    return all(ws.cell(row=row, column=c).value in (None, "") for c in cols)


# ─────────────────────────────────────────────────────────────────────
# Parser principal
# ─────────────────────────────────────────────────────────────────────


# Mapeo conocido para libros estilo Nubox (validado contra
# Reporte Remuneraciones Abril.xlsx · AFIS · Abril 2026)
PRIMERA_TABLA = {
    "codigo": 1,        # A
    "rut": 2,           # B
    "nombre": 3,        # C
    "dt": 4,            # D — días trabajados
    "sueldo_base": 5,   # E
    "horas_extras": 6,  # F
    "gratificacion_legal": 7,    # G
    "otros_imponibles": 8,       # H
    "total_imponibles": 9,       # I
    "asignacion_familiar": 10,   # J
    "otros_no_imponibles": 11,   # K
    "total_no_imponibles": 12,   # L
    "total_haberes": 13,         # M
    "prevision": 14,             # N
    "salud": 15,                 # O
    "seguro_cesantia_trab": 16,  # P
    "otros_descuentos_legales": 17,  # Q
    "total_descuentos_legales": 18,  # R
    "descuentos_varios": 19,     # S
    "total_descuentos": 20,      # T
    "liquido_pagado": 21,        # U
}

SEGUNDA_TABLA = {
    "codigo": 1,                          # A
    "rut": 2,                             # B
    "nombre": 3,                          # C
    "dt": 4,                              # D
    "aporte_afp_empleador": 5,            # E
    "sis": 6,                             # F
    "seguro_cesantia_empleador": 7,       # G
    "seguro_social": 8,                   # H
    "mutual": 9,                          # I
    # Columna J = IDS Patronales total (no la modelamos individual)
    "base_tributable": 11,                # K
    "impuesto_unico": 12,                 # L
}


def parse_libro_remuneraciones(file_path: str | Path) -> LibroParseado:
    """Parsea un libro estilo Nubox/SII. Retorna estructura completa."""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"No existe: {path}")

    # Hash para dedup
    with open(path, "rb") as f:
        archivo_hash = hashlib.sha256(f.read()).hexdigest()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Header (razón social, RUT, mes)
    razon_social = str(ws.cell(row=1, column=1).value or "").strip()
    rut_raw = str(ws.cell(row=2, column=1).value or "")
    m = re.search(r"(\d{1,2}\.?\d{3}\.?\d{3}-[\dkK])|(\d{7,8}-[\dkK])", rut_raw)
    empresa_rut = _normalize_rut(m.group(0)) if m else ""
    mes_cell = str(ws.cell(row=7, column=1).value or "")
    periodo, mes_label = _periodo_from_label(mes_cell)

    # Localizar inicio de primera tabla (header con "C�d" o "Cód")
    fila_header_1 = _find_header_row(ws, "RUT", search_from=8, search_to=15)
    fila_header_2 = _find_header_row(
        ws, "AFP EMP", search_from=fila_header_1 + 1, search_to=ws.max_row
    )

    # Empleados de la primera tabla
    primera_data = _extract_table(
        ws, PRIMERA_TABLA, start_row=fila_header_1 + 1, end_row=fila_header_2 - 2
    )
    segunda_data = _extract_table(
        ws, SEGUNDA_TABLA, start_row=fila_header_2 + 1, end_row=ws.max_row
    )

    # Merge por RUT
    segunda_map = {row["rut"]: row for row in segunda_data if row.get("rut")}

    # Distinguir empleados de subtotales:
    #   - Empleado: col A tiene número correlativo, col B tiene RUT con guion ("21089265-6")
    #   - Subtotal área: col A vacío, col B tiene un número 1-99 que NO es RUT real,
    #     col C tiene el nombre del área ("Gerencia", "Administración y Finanzas")
    #   - TOTAL GENERAL: col A y B vacíos, col C dice "TOTAL GENERAL"
    #
    # Estrategia: separar en empleados (con RUT real "NNNNN-X") + subtotales,
    # luego mapear cada subtotal a los empleados que vienen ANTES de él.

    def _is_real_rut(rut: str | None) -> bool:
        if not rut:
            return False
        # RUT real: contiene guion y la parte numérica tiene >= 5 dígitos
        if "-" not in rut:
            return False
        numero = rut.split("-")[0]
        return numero.isdigit() and len(numero) >= 5

    # Pre-procesar: detectar índice de cada empleado real y agruparlos por área
    empleados_raw: list[dict] = []
    pending_employees: list[dict] = []
    area_map: dict[int, str] = {}  # idx en empleados_raw → area

    for row in primera_data:
        rut = row.get("rut")
        nombre = (row.get("nombre") or "").strip()

        if _is_real_rut(rut):
            empleados_raw.append(row)
            pending_employees.append(row)
            continue

        # No es empleado real → puede ser subtotal de área
        if nombre and "TOTAL" not in nombre.upper():
            # Es un subtotal de área: aplica el nombre a todos los pendientes
            for emp_row in pending_employees:
                idx = empleados_raw.index(emp_row)
                area_map[idx] = nombre
            pending_employees = []

    lineas: list[LibroLinea] = []
    for idx, row in enumerate(empleados_raw):
        rut = row["rut"]
        area_actual = area_map.get(idx)

        # Es una fila de empleado
        s = segunda_map.get(rut, {})
        linea = LibroLinea(
            rut=rut,
            nombre=row.get("nombre", ""),
            area=area_actual,
            dias_trabajados=_dec(row.get("dt")),
            sueldo_base=_dec(row.get("sueldo_base")),
            horas_extras=_dec(row.get("horas_extras")),
            gratificacion_legal=_dec(row.get("gratificacion_legal")),
            otros_imponibles=_dec(row.get("otros_imponibles")),
            total_imponibles=_dec(row.get("total_imponibles")),
            asignacion_familiar=_dec(row.get("asignacion_familiar")),
            otros_no_imponibles=_dec(row.get("otros_no_imponibles")),
            total_no_imponibles=_dec(row.get("total_no_imponibles")),
            total_haberes=_dec(row.get("total_haberes")),
            prevision=_dec(row.get("prevision")),
            salud=_dec(row.get("salud")),
            seguro_cesantia_trab=_dec(row.get("seguro_cesantia_trab")),
            otros_descuentos_legales=_dec(row.get("otros_descuentos_legales")),
            total_descuentos_legales=_dec(row.get("total_descuentos_legales")),
            descuentos_varios=_dec(row.get("descuentos_varios")),
            total_descuentos=_dec(row.get("total_descuentos")),
            liquido_pagado=_dec(row.get("liquido_pagado")),
            aporte_afp_empleador=_dec(s.get("aporte_afp_empleador")),
            sis=_dec(s.get("sis")),
            seguro_cesantia_empleador=_dec(s.get("seguro_cesantia_empleador")),
            seguro_social=_dec(s.get("seguro_social")),
            mutual=_dec(s.get("mutual")),
            base_tributable=_dec(s.get("base_tributable")),
            impuesto_unico=_dec(s.get("impuesto_unico")),
        )
        linea.total_aportes_patronales = (
            linea.aporte_afp_empleador
            + linea.sis
            + linea.seguro_cesantia_empleador
            + linea.seguro_social
            + linea.mutual
        ).quantize(Decimal("0.01"))
        linea.costo_total_empresa = (
            linea.total_haberes + linea.total_aportes_patronales
        ).quantize(Decimal("0.01"))
        lineas.append(linea)

    return LibroParseado(
        empresa_razon_social=razon_social,
        empresa_rut=empresa_rut,
        periodo=periodo,
        mes_label=mes_label,
        lineas=lineas,
        archivo_hash=archivo_hash,
        archivo_origen=path.name,
    )


def _find_header_row(ws, needle: str, *, search_from: int, search_to: int) -> int:
    """Busca la fila que contenga `needle` en cualquier columna.

    Normaliza: upper + quita espacios y puntos (para que 'R.U.T' matchee 'RUT').
    """
    def _norm(s: str) -> str:
        return s.upper().replace(" ", "").replace(".", "")

    needle_u = _norm(needle)
    for r in range(search_from, min(search_to + 1, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and needle_u in _norm(str(v)):
                return r
    raise ValueError(f"No se encontro la fila header con '{needle}'")


def _extract_table(
    ws, mapping: dict[str, int], *, start_row: int, end_row: int
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in range(start_row, end_row + 1):
        row_data: dict[str, Any] = {}
        any_val = False
        for field_name, col in mapping.items():
            v = ws.cell(row=r, column=col).value
            if v not in (None, ""):
                any_val = True
            if field_name == "rut":
                row_data[field_name] = _normalize_rut(v) if v else None
            elif field_name in ("codigo", "nombre"):
                row_data[field_name] = str(v).strip() if v else None
            else:
                row_data[field_name] = v
        if any_val:
            out.append(row_data)
    return out
