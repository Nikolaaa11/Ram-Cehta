"""Parser de Cartas Gantt del portafolio (V4 fase 8).

Tres familias detectadas en los archivos del portafolio:

- **Clásico** (RHO / TRONGKAI / DTE): hoja `'Gantt '` con columnas
  Estado/REAL/PROYECTADO/Codigo/Cliente|Tarea|Actividades/Ubicación/
  Encargado/Estado/Semana|Fecha Crítica/Observaciones + 98 columnas
  semanales. Hoja `Proyectos` con catálogo (Proyecto, Codigo, Estado).
- **EE** (Eco Engineering): hoja `'PROJECT_MANAGEMENT'` con columnas
  ESTADO/REAL/PROYECTADO/CÓDIGO PROYECTO/ACTIVIDAD PRINCIPAL/SUB
  ACTIVIDAD/ENCARGADO/FECHA INICIO/FECHA TÉRMINO/PERIODO/OBSERVACIONES.
- **REVTECH**: hoja `' Gantt_Master'` con columnas Actividades/Fecha
  inicio/Fecha término/Semanas/Encargado/Gasto Proyectado/**Avance**
  (decimal 0-1). Catálogo en hoja `Proyectos`.

Salida unificada: lista de proyectos + sus hitos (actividades), todos en
formato dict listo para inflar `ProyectoCreate` / `HitoCreate`. El parser
es **puro** — no toca DB. El endpoint decide preview vs commit.

Diseño: cada función de parsing devuelve `ParsedGantt` con la lista de
proyectos y, dentro de cada uno, lista de hitos. El dispatcher
`detect_format()` mira los nombres de hojas para decidir parser.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Modelos de salida (DTOs internos del parser, no Pydantic)
# ---------------------------------------------------------------------------


@dataclass
class ParsedHito:
    """Hito (actividad) parseado, listo para inflar `HitoCreate`."""

    nombre: str
    descripcion: str | None = None
    fecha_planificada: date | None = None
    fecha_completado: date | None = None
    estado: str = "pendiente"  # pendiente|en_progreso|completado|cancelado
    progreso_pct: int = 0
    orden: int = 0
    # Campos extra que se persisten en metadata_ del hito (no en columnas)
    encargado: str | None = None
    observaciones: str | None = None
    monto_real: float | None = None
    monto_proyectado: float | None = None
    actividad_principal: str | None = None  # sólo EE
    avance_decimal: float | None = None  # sólo REVTECH (0.0-1.0)


@dataclass
class ParsedProyecto:
    """Proyecto parseado con sus hitos.

    `codigo` es el código que usa la empresa (RHO0001, TKAI002,
    EE.CHO.001, REVTECH0002…). Lo guardamos en `metadata_.codigo_excel`
    para mantener trazabilidad con el archivo de origen.
    """

    codigo: str
    nombre: str
    descripcion: str | None = None
    estado: str = "en_progreso"
    fecha_inicio: date | None = None
    fecha_fin_estimada: date | None = None
    progreso_pct: int = 0
    hitos: list[ParsedHito] = field(default_factory=list)


@dataclass
class ParsedGantt:
    """Resultado del parsing — proyectos con hitos + warnings."""

    formato: str  # "classic" | "ee" | "revtech"
    empresa_codigo: str
    proyectos: list[ParsedProyecto] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    total_hitos: int = 0


# ---------------------------------------------------------------------------
# Utilidades comunes
# ---------------------------------------------------------------------------


# Rango de sanidad para fechas — descarta typos como "2004-03-26" o
# "2099" que aparecen en algunos Excel. Los Gantts del portafolio
# operan típicamente entre 2024 y 2030.
_DATE_MIN_YEAR = 2020
_DATE_MAX_YEAR = 2035


def _to_date(value: Any) -> date | None:
    """Convierte cell value a date. Soporta datetime, date, ISO string.

    Aplica filtro de sanidad: si el año está fuera de
    [_DATE_MIN_YEAR, _DATE_MAX_YEAR] devuelve None — probablemente es
    un typo en el Excel (ej: '2004' en vez de '2024').
    """
    parsed: date | None = None
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                parsed = datetime.strptime(s, fmt).date()
                break
            except ValueError:
                continue

    if parsed is None:
        return None
    if parsed.year < _DATE_MIN_YEAR or parsed.year > _DATE_MAX_YEAR:
        return None
    return parsed


def _normalize_codigo(codigo: str) -> str:
    """Normaliza códigos para matchear typos: EE.RES.001 == EE:RES.001 == ee_res_001.

    Reemplaza separadores comunes (`.`, `:`, `_`, `-`, espacio) con un
    solo `.` y baja a minúsculas. Esto permite reconciliar typos
    en el catálogo (vistos en EE y REVTECH).
    """
    s = codigo.strip().lower()
    for sep in (":", "_", "-", " ", "/"):
        s = s.replace(sep, ".")
    while ".." in s:
        s = s.replace("..", ".")
    return s.strip(".")


def _normalize_codigo_strict(codigo: str) -> str:
    """Normalización agresiva — además de _normalize_codigo, equivale
    leading-zero variants después del prefijo alfa.

    Ejemplos:
      - "REVTECH004"   → "revtech.4"
      - "REVTECH0004"  → "revtech.4"
      - "RHO0001"      → "rho.1"
      - "EE.PAN.001"   → "ee.pan.1"

    Útil cuando un Excel mezcla padding inconsistente (algunos sheets
    con 3 dígitos, otros con 4).
    """
    base = _normalize_codigo(codigo)
    # Quitar leading zeros del último segmento numérico
    parts = base.split(".")
    cleaned: list[str] = []
    for part in parts:
        # Si el segmento es totalmente numérico, sacamos zeros leading
        if part.isdigit():
            cleaned.append(str(int(part)) if int(part) != 0 else "0")
        else:
            # Mixto (letras+números): separar la cola numérica
            # ej: "REVTECH004" → "REVTECH" + "004"
            i = len(part)
            while i > 0 and part[i - 1].isdigit():
                i -= 1
            if i < len(part):
                prefix = part[:i]
                num = part[i:]
                cleaned.append(prefix + (str(int(num)) if int(num) != 0 else "0"))
            else:
                cleaned.append(part)
    return ".".join(cleaned)


# Palabras que NO son códigos de proyecto válidos — aparecen en sub-tablas
# del catálogo (equipo, roles) que el parser solía interpretar mal.
_PALABRAS_NO_PROYECTO = {
    "rol", "ceo", "coo", "cto", "cfo", "cmo", "gm", "director",
    "directora", "gerente", "gerenta", "ingeniero", "ingeniera",
    "administrativo", "administrativa", "asistente", "secretaria",
    "secretario", "equipo", "team", "staff",
}


def _es_codigo_proyecto_valido(codigo: str | None, empresa_codigo: str) -> bool:
    """Heurística para distinguir códigos de proyecto reales vs filas
    de equipo/roles que aparecen en sub-tablas del catálogo.

    Acepta si:
    - Tiene al menos 3 caracteres
    - NO está en la blacklist de palabras (ceo, rol, ingeniero, etc.)
    - Contiene al menos un dígito O empieza con el prefijo de la empresa
    """
    if not codigo:
        return False
    s = codigo.strip()
    if len(s) < 3:
        return False
    s_lower = s.lower()
    if s_lower in _PALABRAS_NO_PROYECTO:
        return False
    # Acepta si contiene dígito (RHO0001, EE.CHO.001, REVTECH0007)
    if any(c.isdigit() for c in s):
        return True
    # Acepta si empieza con prefijo conocido (case-insensitive)
    empresa_lower = empresa_codigo.lower()
    if s_lower.startswith(empresa_lower) or s_lower.startswith(
        empresa_lower[:4]
    ):
        return True
    return False


def _to_float(value: Any) -> float | None:
    """Convierte cell value a float. Acepta None / '' / 'N/A' como None."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if not s or s.lower() in {"n/a", "na", "-", "—", "#error!", "#ref!"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_str(value: Any) -> str | None:
    """Trim + None si vacío."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_estado_texto(raw: str | None) -> str:
    """[DEPRECATED] Mapeo genérico de estado para HITOS.

    Devuelve uno de: pendiente / en_progreso / completado / cancelado.
    Solo válido para Hitos (`EstadoHito`). Para proyectos usar
    `_normalize_estado_proyecto()` que mapea a `planificado` en vez de
    `pendiente` (los enums de DB son distintos).
    """
    return _normalize_estado_hito(raw)


def _normalize_estado_hito(raw: str | None) -> str:
    """Mapea texto de estado a enum HITO (pendiente/en_progreso/completado/cancelado)."""
    if not raw:
        return "pendiente"
    s = raw.strip().lower()
    if s in {"completado", "completada", "terminado", "terminada", "pagado", "finalizado"}:
        return "completado"
    if s in {"cancelado", "cancelada", "anulado"}:
        return "cancelado"
    if s in {"en curso", "en_progreso", "en progreso", "en revisión", "en revision",
             "en revisión cliente", "en revision cliente", "en proceso", "iniciado"}:
        return "en_progreso"
    if s in {"pendiente", "futuro", "no iniciado", "en espera", "planificado"}:
        return "pendiente"
    return "pendiente"


def _normalize_estado_proyecto(raw: str | None) -> str:
    """Mapea texto de estado a enum PROYECTO.

    `EstadoProyecto = planificado | en_progreso | completado | cancelado | pausado`

    NOTA: NO existe "pendiente" en este enum — la DB rechaza el valor
    con check constraint `proyectos_empresa_estado_check`. Por eso
    mapeamos "pendiente"/"futuro"/"no iniciado" → "planificado".
    """
    if not raw:
        return "planificado"
    s = raw.strip().lower()
    if s in {"completado", "completada", "terminado", "terminada", "pagado", "finalizado"}:
        return "completado"
    if s in {"cancelado", "cancelada", "anulado"}:
        return "cancelado"
    if s in {"pausado", "pausada", "en pausa", "en espera"}:
        return "pausado"
    if s in {"en curso", "en_progreso", "en progreso", "en revisión", "en revision",
             "en revisión cliente", "en revision cliente", "en proceso", "iniciado",
             "activo"}:
        return "en_progreso"
    if s in {"pendiente", "futuro", "no iniciado", "planificado", "no aplica",
             "revisión interna", "revision interna"}:
        return "planificado"
    # Default conservador: si no reconocemos el texto, "planificado"
    return "planificado"


def _avance_to_estado(avance: float | None) -> tuple[str, int]:
    """REVTECH usa decimal 0-1. Convierte a (estado, progreso_pct)."""
    if avance is None:
        return ("pendiente", 0)
    pct = max(0, min(100, int(round(avance * 100))))
    if pct == 0:
        return ("pendiente", 0)
    if pct >= 100:
        return ("completado", 100)
    return ("en_progreso", pct)


def _find_sheet(wb: Workbook, candidates: list[str]) -> Worksheet | None:
    """Busca la primera hoja cuyo nombre (trim+lower) coincida con candidatos."""
    norm = {s.strip().lower(): s for s in wb.sheetnames}
    for c in candidates:
        key = c.strip().lower()
        if key in norm:
            return wb[norm[key]]
    return None


# ---------------------------------------------------------------------------
# Detección de formato
# ---------------------------------------------------------------------------


def detect_format(wb: Workbook) -> str:
    """Devuelve "classic" | "ee" | "revtech" | "unknown" según hojas presentes."""
    names_norm = {s.strip().lower() for s in wb.sheetnames}
    if "gantt_master" in names_norm:
        return "revtech"
    if "project_management" in names_norm:
        return "ee"
    if "gantt" in names_norm:
        return "classic"
    return "unknown"


# ---------------------------------------------------------------------------
# Parser CLÁSICO — RHO / TRONGKAI / DTE
# ---------------------------------------------------------------------------


_CLASSIC_HEADER_TOKENS = ("estado", "real", "proyectado", "codigo")


def _find_classic_header_row(ws: Worksheet) -> int | None:
    """Busca la fila header donde están Estado/REAL/PROYECTADO/Codigo en cols A-D."""
    for row_idx in range(1, min(15, ws.max_row + 1)):
        row = [_to_str(c.value) for c in ws[row_idx][:10]]
        norm = [(c or "").lower() for c in row]
        if all(tok in " ".join(norm) for tok in _CLASSIC_HEADER_TOKENS):
            # Confirmar que cada token está en su columna esperada
            if (
                norm[0] == "estado"
                and norm[1] == "real"
                and norm[2] == "proyectado"
                and norm[3] == "codigo"
            ):
                return row_idx
    return None


def _parse_classic_proyectos_catalog(
    ws: Worksheet, empresa_codigo: str = ""
) -> dict[str, ParsedProyecto]:
    """Hoja `Proyectos` → dict[codigo → ParsedProyecto base sin hitos].

    Filtra filas que parecen sub-tablas (equipo, roles) usando
    `_es_codigo_proyecto_valido`. Sin esto, REVTECH parseaba los nombres
    del staff (CEO, COO, etc.) como si fueran proyectos.
    """
    catalog: dict[str, ParsedProyecto] = {}
    # Header está típicamente en R2: ['', 'Proyecto', 'Codigo', 'Estado']
    header_row = None
    for row_idx in range(1, min(8, ws.max_row + 1)):
        row = [_to_str(c.value) or "" for c in ws[row_idx][:6]]
        if any("codigo" in c.lower() for c in row) and any(
            "proyecto" in c.lower() for c in row
        ):
            header_row = row_idx
            break
    if header_row is None:
        return catalog

    # Detectar columnas
    header_cells = [_to_str(c.value) or "" for c in ws[header_row][:6]]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(header_cells):
        hl = h.lower()
        if "proyecto" in hl and "descripcion" not in hl:
            col_map["nombre"] = idx
        elif "descrip" in hl:
            col_map["descripcion"] = idx
        elif "codigo" in hl:
            col_map["codigo"] = idx
        elif "estado" in hl:
            col_map["estado"] = idx

    if "codigo" not in col_map or "nombre" not in col_map:
        return catalog

    # Tracker para detectar segunda tabla (equipo) — si vemos un cambio
    # abrupto en formato de código (ej: pasa de REVTECH0007 a "Rol"),
    # parar de leer.
    consecutive_invalid = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo = _to_str(row[col_map["codigo"]]) if col_map["codigo"] < len(row) else None
        nombre = _to_str(row[col_map["nombre"]]) if col_map["nombre"] < len(row) else None
        if not codigo or not nombre:
            # Fila vacía: si veníamos leyendo proyectos y aparece un blanco,
            # podría ser separador antes de sub-tabla. Reset counter.
            continue

        # Validar que el código sea de proyecto (no rol/equipo)
        if not _es_codigo_proyecto_valido(codigo, empresa_codigo):
            consecutive_invalid += 1
            # Si vemos 2+ filas no-proyecto consecutivas, asumimos que
            # entramos en sub-tabla del equipo y paramos.
            if consecutive_invalid >= 2:
                break
            continue
        consecutive_invalid = 0

        descripcion = (
            _to_str(row[col_map["descripcion"]])
            if "descripcion" in col_map and col_map["descripcion"] < len(row)
            else None
        )
        estado_raw = (
            _to_str(row[col_map["estado"]])
            if "estado" in col_map and col_map["estado"] < len(row)
            else None
        )
        catalog[codigo] = ParsedProyecto(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            # Catálogo del Excel define estado del PROYECTO — usar enum
            # de proyecto (planificado/en_progreso/completado/cancelado/pausado),
            # NO el enum de hito.
            estado=_normalize_estado_proyecto(estado_raw),
        )
    return catalog


# ---------------------------------------------------------------------------
# Parser de hojas individuales por proyecto (fallback / enriquecimiento)
# ---------------------------------------------------------------------------


def _find_individual_sheet_for_codigo(
    wb: Workbook, codigo: str
) -> Worksheet | None:
    """Busca una hoja individual cuyo nombre referencie este codigo.

    Acepta variantes:
      - Exacto: "RHO0001"
      - Prefijo: "RHO0001 Panimávida"
      - Sufijo: "Panimávida_RHO0001"
      - Con/sin zeros: "TKAI005" matchea "TKAI005" y "TKAI5"
    """
    target_strict = _normalize_codigo_strict(codigo)
    target_loose = _normalize_codigo(codigo)
    for sname in wb.sheetnames:
        s_clean = sname.strip()
        if not s_clean:
            continue
        # Normalizar y comparar como prefijo/sufijo/contains
        s_norm = _normalize_codigo(s_clean)
        s_strict = _normalize_codigo_strict(s_clean)
        # Match si el codigo aparece como token dentro del sheet name
        if target_loose in s_norm or target_strict in s_strict:
            return wb[sname]
        # Match estricto contra cada token separado por espacio/underscore
        for token in s_clean.replace("_", " ").split():
            if (
                _normalize_codigo(token) == target_loose
                or _normalize_codigo_strict(token) == target_strict
            ):
                return wb[sname]
    return None


def _parse_individual_project_sheet(
    ws: Worksheet,
    *,
    estado_default: str = "pendiente",
) -> list[ParsedHito]:
    """Parsea una hoja individual de proyecto con layout estándar.

    Layout esperado (hojas RHO0001, TKAI005, GANTT_EE.NTR.001, etc.):
      R1-R6: header del proyecto (nombre, área, etc.) — skipear
      R7 o R8: cabecera de columnas con "Actividades", "Fecha de inicio",
        "Fecha de termino"/"Fecha de término", opcional "OBSERVACIONES",
        "ENCARGADO", "Días"/"Duración"
      R8+ o R9+: datos de actividades

    Devuelve lista de ParsedHito. Si no detecta header en filas 1-12,
    devuelve [].
    """
    # Buscar fila con cabecera "Actividades"
    header_row: int | None = None
    col_idx: dict[str, int] = {}
    for row_idx in range(1, min(15, ws.max_row + 1)):
        row = [_to_str(c.value) or "" for c in ws[row_idx][:14]]
        norm = [c.lower() for c in row]
        for idx, c in enumerate(norm):
            if c == "actividades" or c.startswith("actividad"):
                col_idx["nombre"] = idx
                header_row = row_idx
        if header_row == row_idx:
            # Mismo row: detectar otras columnas
            for idx, c in enumerate(norm):
                if c.startswith("fecha de inicio") or c == "fecha inicio":
                    col_idx["fecha_inicio"] = idx
                elif c.startswith("fecha de term") or c == "fecha término" or c == "fecha termino":
                    col_idx["fecha_termino"] = idx
                elif c == "observaciones" or c.startswith("observac"):
                    col_idx["observaciones"] = idx
                elif "encargado" in c:
                    col_idx["encargado"] = idx
                elif c == "días" or c == "dias" or "duración" in c or "duracion" in c:
                    col_idx["duracion"] = idx
            break

    if header_row is None or "nombre" not in col_idx:
        return []

    hitos: list[ParsedHito] = []
    orden = 0
    # Datos arrancan después del header (saltar 1 fila por si hay sub-header)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        nombre_idx = col_idx["nombre"]
        if nombre_idx >= len(row):
            continue
        nombre = _to_str(row[nombre_idx])
        if not nombre:
            continue
        # Skip si el nombre es claramente un sub-header (todas mayúsculas
        # cortas, palabras como "FASE", "ETAPA")
        if nombre.upper() == nombre and len(nombre) < 8:
            continue

        fecha_inicio = (
            _to_date(row[col_idx["fecha_inicio"]])
            if "fecha_inicio" in col_idx and col_idx["fecha_inicio"] < len(row)
            else None
        )
        fecha_termino = (
            _to_date(row[col_idx["fecha_termino"]])
            if "fecha_termino" in col_idx and col_idx["fecha_termino"] < len(row)
            else None
        )
        observaciones = (
            _to_str(row[col_idx["observaciones"]])
            if "observaciones" in col_idx and col_idx["observaciones"] < len(row)
            else None
        )
        encargado = (
            _to_str(row[col_idx["encargado"]])
            if "encargado" in col_idx and col_idx["encargado"] < len(row)
            else None
        )

        hitos.append(
            ParsedHito(
                nombre=nombre[:255],
                descripcion=observaciones,
                fecha_planificada=fecha_inicio,
                fecha_completado=None,
                estado=estado_default,
                progreso_pct=0,
                orden=orden,
                encargado=encargado,
                observaciones=observaciones,
            )
        )
        orden += 1

    return hitos


def _enriquecer_proyecto_con_hoja_individual(
    wb: Workbook,
    proyecto: ParsedProyecto,
) -> int:
    """Si un proyecto tiene 0 hitos en el master, intenta extraer de su
    hoja individual. Devuelve cuántos hitos se sumaron.
    """
    if proyecto.hitos:  # ya tiene hitos del master, no enriquecer
        return 0
    sheet = _find_individual_sheet_for_codigo(wb, proyecto.codigo)
    if sheet is None:
        return 0
    nuevos = _parse_individual_project_sheet(sheet, estado_default=proyecto.estado)
    if nuevos:
        proyecto.hitos.extend(nuevos)
    return len(nuevos)


def _resolve_codigo(catalog: dict[str, ParsedProyecto], codigo: str) -> str | None:
    """Busca el codigo canónico en el catálogo aplicando fuzzy match.

    Cascada de matching:
    1. Match exacto
    2. Match con _normalize_codigo (separadores normalizados)
    3. Match con _normalize_codigo_strict (también equivale leading zeros)

    Devuelve la key real del catálogo si encuentra match, o None si no.
    """
    if codigo in catalog:
        return codigo

    # Match con normalización de separadores
    target = _normalize_codigo(codigo)
    for key in catalog.keys():
        if _normalize_codigo(key) == target:
            return key

    # Match con normalización estricta (leading zeros)
    target_strict = _normalize_codigo_strict(codigo)
    for key in catalog.keys():
        if _normalize_codigo_strict(key) == target_strict:
            return key

    return None


def parse_classic(wb: Workbook, empresa_codigo: str) -> ParsedGantt:
    """Parser para RHO / TRONGKAI / DTE."""
    result = ParsedGantt(formato="classic", empresa_codigo=empresa_codigo)

    # 1. Catálogo de proyectos
    proyectos_ws = _find_sheet(wb, ["Proyectos"])
    catalog: dict[str, ParsedProyecto] = {}
    if proyectos_ws is not None:
        catalog = _parse_classic_proyectos_catalog(proyectos_ws, empresa_codigo)
    else:
        result.warnings.append("Hoja 'Proyectos' no encontrada — proyectos se inferirán del Gantt.")

    # 2. Hoja Gantt principal
    gantt_ws = _find_sheet(wb, ["Gantt", "Gantt "])
    if gantt_ws is None:
        result.warnings.append("Hoja 'Gantt' no encontrada — no se puede importar.")
        return result

    header_row = _find_classic_header_row(gantt_ws)
    if header_row is None:
        result.warnings.append("Header Estado/REAL/PROYECTADO/Codigo no detectado en hoja Gantt.")
        return result

    # Columnas fijas del formato clásico (0-indexed en row tuple):
    #   A=0 Estado, B=1 REAL, C=2 PROYECTADO, D=3 Codigo, E=4 Cliente|Tarea,
    #   F=5 Ubicación, G=6 Encargado, H=7 Estado, I=8 Semana/Fecha Crítica,
    #   J=9 Observaciones
    orden_counter: dict[str, int] = {}

    # El header del clásico ocupa 2 filas (fila header + fila "Equipo|Fecha Crítica").
    # Las filas de datos arrancan en header_row + 2.
    for row in gantt_ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row or len(row) < 10:
            continue
        codigo = _to_str(row[3])
        actividad = _to_str(row[4])
        if not codigo or not actividad:
            continue

        # Si codigo no estaba en catálogo, intentar fuzzy match antes de crear nuevo
        canonical = _resolve_codigo(catalog, codigo)
        if canonical is None:
            catalog[codigo] = ParsedProyecto(
                codigo=codigo,
                nombre=actividad,  # fallback: primera actividad como nombre
                estado="en_progreso",
            )
            result.warnings.append(f"Proyecto {codigo} no estaba en catálogo — creado desde Gantt.")
            # La primera fila del codigo es el header del proyecto, no un hito
            continue
        codigo = canonical

        estado_raw = _to_str(row[0]) or _to_str(row[7])
        encargado = _to_str(row[6])
        fecha_crit = _to_date(row[8])
        observaciones = _to_str(row[9])
        real = _to_float(row[1])
        proyectado = _to_float(row[2])
        estado_norm = _normalize_estado_texto(estado_raw)
        progreso = 100 if estado_norm == "completado" else 0

        # Heurística: si la actividad es texto MAYÚSCULA larga sin encargado ni
        # fecha, es probable un sub-header de proyecto (ej: "PANIMÁVIDA (BESS RHO)").
        # Lo skipeamos como hito pero podríamos enriquecer descripción del proyecto.
        is_subheader = (
            actividad.isupper()
            and not encargado
            and fecha_crit is None
            and real is None
            and proyectado is None
        )
        if is_subheader:
            if not catalog[codigo].descripcion:
                catalog[codigo].descripcion = actividad
            continue

        orden = orden_counter.get(codigo, 0)
        orden_counter[codigo] = orden + 1

        catalog[codigo].hitos.append(
            ParsedHito(
                nombre=actividad[:255],
                descripcion=observaciones,
                fecha_planificada=fecha_crit,
                fecha_completado=fecha_crit if estado_norm == "completado" else None,
                estado=estado_norm,
                progreso_pct=progreso,
                orden=orden,
                encargado=encargado,
                observaciones=observaciones,
                monto_real=real,
                monto_proyectado=proyectado,
            )
        )

    # 3. Fallback: si un proyecto del catálogo NO tiene hitos en el master
    # Gantt, intentar extraerlos de su hoja individual (RHO0001 Panimávida,
    # TKAI005, etc.)
    enriched_count = 0
    for proy in catalog.values():
        if not proy.hitos:
            n = _enriquecer_proyecto_con_hoja_individual(wb, proy)
            if n > 0:
                enriched_count += n
    if enriched_count > 0:
        result.warnings.append(
            f"Enriquecidos {enriched_count} hitos desde hojas individuales "
            "para proyectos sin datos en el Gantt master."
        )

    # 4. Calcular fecha_inicio / fecha_fin_estimada / progreso_pct por proyecto
    for proy in catalog.values():
        _enrich_proyecto_from_hitos(proy)

    result.proyectos = list(catalog.values())
    result.total_hitos = sum(len(p.hitos) for p in result.proyectos)
    return result


# ---------------------------------------------------------------------------
# Parser EE
# ---------------------------------------------------------------------------


def parse_ee(wb: Workbook, empresa_codigo: str) -> ParsedGantt:
    """Parser para EE — usa hoja PROJECT_MANAGEMENT con jerarquía Actividad → Sub-actividad."""
    result = ParsedGantt(formato="ee", empresa_codigo=empresa_codigo)

    # 1. Catálogo (mismo schema que clásico)
    proyectos_ws = _find_sheet(wb, ["Proyectos"])
    catalog: dict[str, ParsedProyecto] = {}
    if proyectos_ws is not None:
        catalog = _parse_classic_proyectos_catalog(proyectos_ws, empresa_codigo)

    # 2. Hoja PROJECT_MANAGEMENT
    pm_ws = _find_sheet(wb, ["PROJECT_MANAGEMENT"])
    if pm_ws is None:
        result.warnings.append("Hoja 'PROJECT_MANAGEMENT' no encontrada.")
        return result

    # El header está en R1 con columnas:
    #   B=1 ESTADO, C=2 REAL, D=3 PROYECTADO, E=4 CÓDIGO PROYECTO,
    #   F=5 ACTIVIDAD PRINCIPAL, G=6 SUB ACTIVIDAD, H=7 ENCARGADO,
    #   I=8 FECHA INICIO, J=9 FECHA TÉRMINO, K=10 PERIODO (SEMANAS),
    #   L=11 OBSERVACIONES
    orden_counter: dict[str, int] = {}

    for row in pm_ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 12:
            continue
        codigo = _to_str(row[4])
        actividad_principal = _to_str(row[5])
        sub_actividad = _to_str(row[6])
        if not codigo or not (actividad_principal or sub_actividad):
            continue

        # Asegurar proyecto en catálogo (con fuzzy match para typos)
        canonical = _resolve_codigo(catalog, codigo)
        if canonical is None:
            catalog[codigo] = ParsedProyecto(
                codigo=codigo,
                nombre=actividad_principal or codigo,
                estado="en_progreso",
            )
            result.warnings.append(
                f"Proyecto {codigo} no estaba en catálogo — creado desde PROJECT_MANAGEMENT."
            )
        else:
            codigo = canonical

        estado_raw = _to_str(row[1])
        real = _to_float(row[2])
        proyectado = _to_float(row[3])
        encargado = _to_str(row[7])
        fecha_inicio = _to_date(row[8])
        fecha_termino = _to_date(row[9])
        observaciones = _to_str(row[11])
        estado_norm = _normalize_estado_texto(estado_raw)
        progreso = 100 if estado_norm == "completado" else 0

        # Para EE, el "nombre" del hito = sub-actividad; si no hay, fallback a actividad principal
        nombre_hito = sub_actividad or actividad_principal or "Sin actividad"
        descripcion = observaciones
        if actividad_principal and sub_actividad:
            descripcion = f"[{actividad_principal}] {observaciones or ''}".strip()

        orden = orden_counter.get(codigo, 0)
        orden_counter[codigo] = orden + 1

        catalog[codigo].hitos.append(
            ParsedHito(
                nombre=nombre_hito[:255],
                descripcion=descripcion,
                fecha_planificada=fecha_inicio,
                fecha_completado=fecha_termino if estado_norm == "completado" else None,
                estado=estado_norm,
                progreso_pct=progreso,
                orden=orden,
                encargado=encargado,
                observaciones=observaciones,
                monto_real=real,
                monto_proyectado=proyectado,
                actividad_principal=actividad_principal,
            )
        )

    # Fallback: enriquecer proyectos sin hitos desde hojas individuales
    # GANTT_EE.NTR.001, GANTT_EE.CHO.001, etc.
    enriched_count = 0
    for proy in catalog.values():
        if not proy.hitos:
            n = _enriquecer_proyecto_con_hoja_individual(wb, proy)
            if n > 0:
                enriched_count += n
    if enriched_count > 0:
        result.warnings.append(
            f"Enriquecidos {enriched_count} hitos desde hojas individuales."
        )

    for proy in catalog.values():
        _enrich_proyecto_from_hitos(proy)

    result.proyectos = list(catalog.values())
    result.total_hitos = sum(len(p.hitos) for p in result.proyectos)
    return result


# ---------------------------------------------------------------------------
# Parser REVTECH
# ---------------------------------------------------------------------------


def _extract_revtech_codigo_from_sheetname(name: str) -> str | None:
    """De 'Minera Tornasol_REVTECH0002' extrae 'REVTECH0002'."""
    if "_REVTECH" not in name.upper():
        return None
    # Busca el último token con REVTECH
    parts = name.rsplit("_", 1)
    if len(parts) == 2 and "REVTECH" in parts[1].upper():
        return parts[1].strip()
    return None


def parse_revtech(wb: Workbook, empresa_codigo: str) -> ParsedGantt:
    """Parser para REVTECH — Gantt_Master con avance% + hojas por proyecto."""
    result = ParsedGantt(formato="revtech", empresa_codigo=empresa_codigo)

    # 1. Catálogo
    proyectos_ws = _find_sheet(wb, ["Proyectos"])
    catalog: dict[str, ParsedProyecto] = {}
    if proyectos_ws is not None:
        catalog = _parse_classic_proyectos_catalog(proyectos_ws, empresa_codigo)

    # 2. Recorrer hojas por proyecto (Minera Tornasol_REVTECH0002, etc.)
    # Cada una tiene su propio Gantt detallado con columnas:
    #   B=1 Actividades, C=2 Fecha inicio, D=3 Fecha término, E=4 Semanas,
    #   F=5 Encargado, G=6 Gasto Proyectado, H=7 Avance
    project_sheets = [
        s for s in wb.sheetnames if "_REVTECH" in s.upper() and s.strip() not in {"Proyectos"}
    ]

    for sname in project_sheets:
        ws = wb[sname]
        codigo = _extract_revtech_codigo_from_sheetname(sname)
        if not codigo:
            continue

        # Si no estaba en catálogo, fuzzy match primero
        canonical = _resolve_codigo(catalog, codigo)
        if canonical is None:
            nombre = sname.split("_")[0].strip() or codigo
            catalog[codigo] = ParsedProyecto(
                codigo=codigo,
                nombre=nombre,
                estado="en_progreso",
            )
        else:
            codigo = canonical

        # Buscar header row (tiene Actividades, Fecha inicio, etc.)
        header_row = None
        for row_idx in range(1, min(8, ws.max_row + 1)):
            cells = [_to_str(c.value) or "" for c in ws[row_idx][:9]]
            joined = " ".join(c.lower() for c in cells)
            if "actividades" in joined and "fecha" in joined and "avance" in joined:
                header_row = row_idx
                break
        if header_row is None:
            result.warnings.append(f"Header no encontrado en hoja '{sname}' — saltada.")
            continue

        # Datos arrancan después del header (skip 2 filas: header + sub-header con dates)
        orden = 0
        for row in ws.iter_rows(min_row=header_row + 3, values_only=True):
            if not row or len(row) < 8:
                continue
            actividad = _to_str(row[1])
            if not actividad:
                continue
            fecha_inicio = _to_date(row[2])
            fecha_termino = _to_date(row[3])
            encargado = _to_str(row[5])
            gasto = _to_float(row[6])
            avance = _to_float(row[7])
            estado, progreso = _avance_to_estado(avance)

            catalog[codigo].hitos.append(
                ParsedHito(
                    nombre=actividad[:255],
                    fecha_planificada=fecha_inicio,
                    fecha_completado=fecha_termino if estado == "completado" else None,
                    estado=estado,
                    progreso_pct=progreso,
                    orden=orden,
                    encargado=encargado,
                    monto_proyectado=gasto,
                    avance_decimal=avance,
                )
            )
            orden += 1

    # 3. Si Gantt_Master existe y los hitos están vacíos para algún proyecto,
    # caer ahí como fallback (cuando no hay hojas por proyecto).
    master_ws = _find_sheet(wb, ["Gantt_Master"])
    if master_ws is not None and not any(p.hitos for p in catalog.values()):
        result.warnings.append("Hojas por proyecto vacías — usando Gantt_Master agregado.")
        # estructura: A=ID, B=Actividades, C=fecha inicio, D=fecha termino,
        # E=semanas, F=encargado, G=gasto, H=avance
        current_codigo: str | None = None
        orden_by_codigo: dict[str, int] = {}
        for row in master_ws.iter_rows(min_row=4, values_only=True):
            if not row or len(row) < 8:
                continue
            actividad = _to_str(row[1])
            if not actividad:
                continue
            # Detectar cambio de proyecto si el nombre matchea catálogo
            for codigo, proy in catalog.items():
                if codigo in actividad or actividad.startswith(proy.nombre[:20]):
                    current_codigo = codigo
                    break
            if current_codigo is None:
                continue
            fecha_inicio = _to_date(row[2])
            fecha_termino = _to_date(row[3])
            encargado = _to_str(row[5])
            gasto = _to_float(row[6])
            avance = _to_float(row[7])
            estado, progreso = _avance_to_estado(avance)
            orden = orden_by_codigo.get(current_codigo, 0)
            orden_by_codigo[current_codigo] = orden + 1
            catalog[current_codigo].hitos.append(
                ParsedHito(
                    nombre=actividad[:255],
                    fecha_planificada=fecha_inicio,
                    fecha_completado=fecha_termino if estado == "completado" else None,
                    estado=estado,
                    progreso_pct=progreso,
                    orden=orden,
                    encargado=encargado,
                    monto_proyectado=gasto,
                    avance_decimal=avance,
                )
            )

    for proy in catalog.values():
        _enrich_proyecto_from_hitos(proy)

    result.proyectos = list(catalog.values())
    result.total_hitos = sum(len(p.hitos) for p in result.proyectos)
    return result


# ---------------------------------------------------------------------------
# Enriquecimiento agregado: derivar fechas y % desde hitos
# ---------------------------------------------------------------------------


def _enrich_proyecto_from_hitos(proy: ParsedProyecto) -> None:
    """Calcula fecha_inicio, fecha_fin_estimada y progreso_pct desde sus hitos."""
    if not proy.hitos:
        return

    # Fechas: min/max de fecha_planificada
    fechas = [h.fecha_planificada for h in proy.hitos if h.fecha_planificada]
    if fechas and proy.fecha_inicio is None:
        proy.fecha_inicio = min(fechas)
    if fechas and proy.fecha_fin_estimada is None:
        proy.fecha_fin_estimada = max(fechas)

    # Progreso: promedio de progreso_pct de hitos
    if proy.progreso_pct == 0:
        total = sum(h.progreso_pct for h in proy.hitos)
        proy.progreso_pct = total // len(proy.hitos)

    # Estado: si todos los hitos completados → completado.
    # IMPORTANTE: el enum de Proyecto NO acepta "pendiente" — usar "planificado".
    # Defensivo: si proy.estado quedó con un valor inválido del enum proyecto,
    # mapearlo a uno válido.
    valid_proyecto_estados = {
        "planificado", "en_progreso", "completado", "cancelado", "pausado"
    }
    estados = {h.estado for h in proy.hitos}
    if estados == {"completado"}:
        proy.estado = "completado"
    elif "en_progreso" in estados:
        proy.estado = "en_progreso"
    elif estados == {"pendiente"}:
        proy.estado = "planificado"
    elif estados == {"cancelado"}:
        proy.estado = "cancelado"
    # Backstop defensivo: si el estado actual no es válido para Proyecto,
    # forzar un valor válido (esto cubre cualquier dato corrupto del catálogo).
    if proy.estado not in valid_proyecto_estados:
        proy.estado = _normalize_estado_proyecto(proy.estado)


# ---------------------------------------------------------------------------
# Dispatcher público
# ---------------------------------------------------------------------------


def parse_gantt_excel(content: bytes, empresa_codigo: str) -> ParsedGantt:
    """Punto de entrada del parser.

    Detecta automáticamente la familia (clásico/EE/REVTECH) y delega al
    parser específico. Si el formato es desconocido, devuelve un
    ParsedGantt vacío con un warning.

    Raises:
        ValueError si el archivo no es un Excel válido.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"No se pudo abrir el Excel: {e}") from e

    formato = detect_format(wb)
    if formato == "classic":
        return parse_classic(wb, empresa_codigo)
    if formato == "ee":
        return parse_ee(wb, empresa_codigo)
    if formato == "revtech":
        return parse_revtech(wb, empresa_codigo)

    return ParsedGantt(
        formato="unknown",
        empresa_codigo=empresa_codigo,
        warnings=[
            f"Formato de Gantt no reconocido. Hojas: {wb.sheetnames}. "
            "Esperaba 'Gantt', 'PROJECT_MANAGEMENT' o 'Gantt_Master'."
        ],
    )
