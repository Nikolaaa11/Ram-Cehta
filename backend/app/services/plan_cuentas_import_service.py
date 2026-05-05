"""Servicio importador del Plan de Cuentas v2.

Lógica reutilizable extraída de `scripts/import_plan_cuentas.py` para
poder invocarse desde:
  1. CLI (`python -m scripts.import_plan_cuentas --apply`)
  2. Endpoint admin `POST /api/v1/admin/plan-cuentas/import` (multipart)

La lógica es la misma: leer .xlsx → parsear → UPSERT en DB. Idempotente.
"""
from __future__ import annotations

from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession


# Empresas del Excel (orden importa para iterar las columnas Hab_X)
EMPRESAS_EXCEL = ["CSL", "RHO", "DTE", "RVT", "EVQ", "TRK", "AFIS", "FIP"]

# Mapeo de prefijo de empresa a código real en core.empresas.codigo
EMPRESA_CODE_MAP = {
    "CSL": "CSL",
    "RHO": "RHO",
    "DTE": "DTE",
    "RVT": "REVTECH",
    "EVQ": "EVOQUE",
    "TRK": "TRONGKAI",
    "AFIS": "AFIS",
    "FIP": "FIP_CEHTA",
}

# CENERGY replica las habilitaciones de la empresa fuente.
# Acordado con COO: CENERGY es servicios, idéntico giro a DTE.
CENERGY_REPLICA_FROM = "DTE"


# ---------------------------------------------------------------------
# Parsers de tipos del Excel
# ---------------------------------------------------------------------


def _bool_x(value: Any) -> bool:
    """Excel marca true como 'x', false como None/empty."""
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ("x", "1", "true", "si", "sí")


def _bool_excel(value: Any) -> bool:
    """Para columnas TRUE/FALSE explícitas (Activa, Hab_X)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("true", "1", "si", "sí", "x")


def _str_or_none(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _int_or_none(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------
# Parsing del .xlsx
# ---------------------------------------------------------------------


class PlanCuentasParseError(ValueError):
    """Excel no tiene la estructura esperada."""


class ImportPayload:
    """Resultado del parseo de las 3 hojas relevantes del Excel.

    Atributos:
      cuentas: filas de `core.plan_cuentas`
      habilitaciones_cuentas: dict empresa_codigo → set(cuenta_codigo)
      proyectos: filas de `core.proyectos_contables`
      areas: filas de `core.areas` (10 áreas con descripciones del Excel)
      area_empresa: dict area_codigo → set(empresa_codigo) que aplican
    """

    __slots__ = (
        "cuentas",
        "habilitaciones_cuentas",
        "proyectos",
        "areas",
        "area_empresa",
    )

    def __init__(self) -> None:
        self.cuentas: list[dict[str, Any]] = []
        self.habilitaciones_cuentas: dict[str, set[str]] = defaultdict(set)
        self.proyectos: list[dict[str, Any]] = []
        self.areas: list[dict[str, Any]] = []
        self.area_empresa: dict[str, set[str]] = defaultdict(set)


def parse_xlsx_bytes(xlsx_bytes: bytes) -> ImportPayload:
    """Parsea el .xlsx desde bytes (uso desde upload)."""
    return _parse_workbook(
        openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    )


def parse_xlsx_path(xlsx_path: Path) -> ImportPayload:
    """Parsea desde una ruta del filesystem (uso desde CLI)."""
    return _parse_workbook(
        openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    )


def _parse_workbook(wb: Any) -> ImportPayload:
    payload = ImportPayload()
    _parse_plan_cuentas(wb, payload)
    # Las hojas Proyectos y Areas son opcionales — si faltan, se importa
    # solo el plan de cuentas (compatibilidad con Excel viejos).
    if "Proyectos" in wb.sheetnames:
        _parse_proyectos(wb, payload)
    if "Areas" in wb.sheetnames:
        _parse_areas(wb, payload)
    return payload


def _parse_plan_cuentas(wb: Any, payload: ImportPayload) -> None:
    if "PlanDeCuentas" not in wb.sheetnames:
        raise PlanCuentasParseError(
            f"El Excel no tiene hoja 'PlanDeCuentas'. Hojas: {wb.sheetnames}"
        )
    ws = wb["PlanDeCuentas"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise PlanCuentasParseError("Hoja PlanDeCuentas vacía")

    header = [str(c).strip() if c else "" for c in rows[0]]
    expected = [
        "Cuenta", "Nivel", "Tipo", "Descripcion", "CuentaPadre",
        "Imputable", "IvaTratamiento", "CorfoElegible", "TipoGastoCorfo",
        "NuboxCode", "CodigoF22", "Ajuste14D",
        "FlagPartida", "FlagConcepto", "FlagCapital", "FlagActivoFijo",
        "FlagDocumento", "FlagControlGestion", "FlagActivoNeto", "FlagCaja",
        "Flag14D", "FlagPercepcion",
        "Activa",
        "Hab_CSL", "Hab_RHO", "Hab_DTE", "Hab_RVT",
        "Hab_EVQ", "Hab_TRK", "Hab_AFIS", "Hab_FIP",
    ]
    missing = [c for c in expected if c not in header]
    if missing:
        raise PlanCuentasParseError(f"Columnas faltantes en hoja PlanDeCuentas: {missing}")

    idx = {col: header.index(col) for col in expected}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        codigo = str(row[idx["Cuenta"]]).strip()
        if not codigo:
            continue

        nivel = int(row[idx["Nivel"]])
        tipo = str(row[idx["Tipo"]]).strip()

        cuenta = {
            "codigo": codigo,
            "nivel": nivel,
            "tipo": tipo,
            "nombre": str(row[idx["Descripcion"]]).strip(),
            "descripcion": None,
            "codigo_padre": _str_or_none(row[idx["CuentaPadre"]]),
            "imputable": _bool_excel(row[idx["Imputable"]]),
            "iva_tratamiento": _str_or_none(row[idx["IvaTratamiento"]]) or "NA",
            "corfo_elegible": _bool_excel(row[idx["CorfoElegible"]]),
            "tipo_gasto_corfo": _str_or_none(row[idx["TipoGastoCorfo"]]),
            "nubox_code": _str_or_none(row[idx["NuboxCode"]]) or codigo,
            "codigo_f22": _int_or_none(row[idx["CodigoF22"]]),
            "ajuste_14d": _str_or_none(row[idx["Ajuste14D"]]),
            "flag_partida": _bool_x(row[idx["FlagPartida"]]),
            "flag_concepto": _bool_x(row[idx["FlagConcepto"]]),
            "flag_capital": _bool_x(row[idx["FlagCapital"]]),
            "flag_activo_fijo": _bool_x(row[idx["FlagActivoFijo"]]),
            "flag_documento": _bool_x(row[idx["FlagDocumento"]]),
            "flag_control_gestion": _bool_x(row[idx["FlagControlGestion"]]),
            "flag_activo_neto": _bool_x(row[idx["FlagActivoNeto"]]),
            "flag_caja": _bool_x(row[idx["FlagCaja"]]),
            "flag_marca_14d": _bool_x(row[idx["Flag14D"]]),
            "flag_percepcion": _bool_x(row[idx["FlagPercepcion"]]),
            "activa": _bool_excel(row[idx["Activa"]]),
        }
        payload.cuentas.append(cuenta)

        for empresa_excel in EMPRESAS_EXCEL:
            col = f"Hab_{empresa_excel}"
            if _bool_excel(row[idx[col]]):
                empresa_codigo = EMPRESA_CODE_MAP[empresa_excel]
                payload.habilitaciones_cuentas[empresa_codigo].add(codigo)

    # CENERGY replica las habilitaciones de la empresa fuente (DTE)
    fuente = EMPRESA_CODE_MAP[CENERGY_REPLICA_FROM]
    if fuente in payload.habilitaciones_cuentas:
        payload.habilitaciones_cuentas["CENERGY"] = set(
            payload.habilitaciones_cuentas[fuente]
        )


def _parse_proyectos(wb: Any, payload: ImportPayload) -> None:
    """Parsea hoja 'Proyectos' del Excel.

    Columnas esperadas: Codigo, Empresa, Nombre, TipoFinanciamiento,
    Programa, FechaInicio, FechaTermino, PresupuestoTotal, Moneda,
    PrimerDesembolsoCorfo, TiposGastoElegibles, Estado.
    """
    ws = wb["Proyectos"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    header = [str(c).strip() if c else "" for c in rows[0]]
    expected = [
        "Codigo", "Empresa", "Nombre", "TipoFinanciamiento", "Programa",
        "FechaInicio", "FechaTermino", "PresupuestoTotal", "Moneda",
        "PrimerDesembolsoCorfo", "TiposGastoElegibles", "Estado",
    ]
    missing = [c for c in expected if c not in header]
    if missing:
        raise PlanCuentasParseError(
            f"Columnas faltantes en hoja Proyectos: {missing}"
        )
    idx = {col: header.index(col) for col in expected}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        codigo = str(row[idx["Codigo"]]).strip()
        if not codigo:
            continue

        empresa_excel = str(row[idx["Empresa"]]).strip()
        empresa_codigo = EMPRESA_CODE_MAP.get(empresa_excel, empresa_excel)

        tipos_str = _str_or_none(row[idx["TiposGastoElegibles"]])
        tipos: list[str] = []
        if tipos_str:
            tipos = [
                t.strip().upper()
                for t in tipos_str.replace(",", ";").split(";")
                if t.strip()
            ]

        proyecto = {
            "codigo": codigo,
            "empresa_codigo": empresa_codigo,
            "nombre": str(row[idx["Nombre"]]).strip(),
            "tipo_financiamiento": str(
                row[idx["TipoFinanciamiento"]]
            ).strip().upper(),
            "programa": _str_or_none(row[idx["Programa"]]),
            "fecha_inicio": _date_or_none(row[idx["FechaInicio"]]),
            "fecha_termino": _date_or_none(row[idx["FechaTermino"]]),
            "presupuesto_total": _decimal_or_none(row[idx["PresupuestoTotal"]]),
            "moneda": _str_or_none(row[idx["Moneda"]]) or "CLP",
            "primer_desembolso_corfo": _date_or_none(
                row[idx["PrimerDesembolsoCorfo"]]
            ),
            "tipos_gasto_elegibles": tipos,
            "estado": _str_or_none(row[idx["Estado"]]) or "ACTIVE",
        }
        payload.proyectos.append(proyecto)


def _parse_areas(wb: Any, payload: ImportPayload) -> None:
    """Parsea hoja 'Areas' del Excel.

    Columnas esperadas: Codigo, Nombre, Descripcion, Activa,
    Aplica_CSL, Aplica_RHO, Aplica_DTE, Aplica_RVT, Aplica_EVQ,
    Aplica_TRK, Aplica_AFIS, Aplica_FIP.
    """
    ws = wb["Areas"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    header = [str(c).strip() if c else "" for c in rows[0]]
    expected_base = ["Codigo", "Nombre", "Descripcion", "Activa"]
    expected_aplica = [f"Aplica_{e}" for e in EMPRESAS_EXCEL]
    expected = expected_base + expected_aplica

    missing = [c for c in expected if c not in header]
    if missing:
        raise PlanCuentasParseError(
            f"Columnas faltantes en hoja Areas: {missing}"
        )
    idx = {col: header.index(col) for col in expected}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        codigo = str(row[idx["Codigo"]]).strip().upper()
        if not codigo:
            continue

        area = {
            "codigo": codigo,
            "nombre": str(row[idx["Nombre"]]).strip(),
            "descripcion": _str_or_none(row[idx["Descripcion"]]),
            "activa": _bool_excel(row[idx["Activa"]]),
        }
        payload.areas.append(area)

        for empresa_excel in EMPRESAS_EXCEL:
            col = f"Aplica_{empresa_excel}"
            if _bool_excel(row[idx[col]]):
                empresa_codigo = EMPRESA_CODE_MAP[empresa_excel]
                payload.area_empresa[codigo].add(empresa_codigo)

        # CENERGY replica DTE para áreas igual que para cuentas
        fuente = EMPRESA_CODE_MAP[CENERGY_REPLICA_FROM]
        if fuente in payload.area_empresa[codigo]:
            payload.area_empresa[codigo].add("CENERGY")


def _date_or_none(value: Any) -> Any:
    """Extrae date de un valor del Excel. None si vacío o no parseable."""
    if value is None:
        return None
    # openpyxl ya devuelve datetime.date / datetime para celdas tipo fecha
    from datetime import date, datetime
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Strings tipo '2024-08-26' o '26/08/2024'
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _decimal_or_none(value: Any) -> Any:
    """Convierte a Decimal para presupuesto. None si vacío."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    from decimal import Decimal, InvalidOperation
    try:
        return Decimal(s.replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------
# Apply a DB (idempotente)
# ---------------------------------------------------------------------


async def apply_to_db(
    db: AsyncSession,
    payload_or_cuentas: ImportPayload | list[dict[str, Any]],
    habilitaciones: dict[str, set[str]] | None = None,
) -> dict[str, int]:
    """UPSERT atómico de las 3 secciones del Excel. Caller hace commit.

    Acepta dos firmas (compatibilidad con código viejo):
      apply_to_db(db, payload)              # nueva — recibe ImportPayload
      apply_to_db(db, cuentas, habilits)    # vieja — solo cuentas
    """
    if isinstance(payload_or_cuentas, ImportPayload):
        payload = payload_or_cuentas
    else:
        # Backward compat: armar payload mínimo desde args sueltos
        payload = ImportPayload()
        payload.cuentas = payload_or_cuentas
        payload.habilitaciones_cuentas = defaultdict(set, habilitaciones or {})

    counters = {
        "cuentas_upserted": 0,
        "habilitaciones_cuentas_upserted": 0,
        "proyectos_upserted": 0,
        "areas_upserted": 0,
        "area_empresa_upserted": 0,
        "habilitaciones_omitidas_empresa_inexistente": 0,
        "proyectos_omitidos_empresa_inexistente": 0,
    }

    empresas_db = (
        (
            await db.execute(
                text("SELECT codigo FROM core.empresas WHERE activo = TRUE")
            )
        )
        .scalars()
        .all()
    )
    empresas_db_set = set(empresas_db)

    # ---------- CUENTAS ----------
    cuentas_ordenadas = sorted(
        payload.cuentas, key=lambda c: (c["nivel"], c["codigo"])
    )

    for c in cuentas_ordenadas:
        await db.execute(
            text(
                """
                INSERT INTO core.plan_cuentas (
                    codigo, nivel, tipo, nombre, descripcion, codigo_padre,
                    imputable, iva_tratamiento,
                    corfo_elegible, tipo_gasto_corfo, nubox_code,
                    codigo_f22, ajuste_14d,
                    flag_partida, flag_concepto, flag_capital,
                    flag_activo_fijo, flag_documento, flag_control_gestion,
                    flag_activo_neto, flag_caja, flag_marca_14d,
                    flag_percepcion, activa
                )
                VALUES (
                    :codigo, :nivel, :tipo, :nombre, :descripcion, :codigo_padre,
                    :imputable, :iva_tratamiento,
                    :corfo_elegible, :tipo_gasto_corfo, :nubox_code,
                    :codigo_f22, :ajuste_14d,
                    :flag_partida, :flag_concepto, :flag_capital,
                    :flag_activo_fijo, :flag_documento, :flag_control_gestion,
                    :flag_activo_neto, :flag_caja, :flag_marca_14d,
                    :flag_percepcion, :activa
                )
                ON CONFLICT (codigo) DO UPDATE SET
                    nivel = EXCLUDED.nivel,
                    tipo = EXCLUDED.tipo,
                    nombre = EXCLUDED.nombre,
                    descripcion = EXCLUDED.descripcion,
                    codigo_padre = EXCLUDED.codigo_padre,
                    imputable = EXCLUDED.imputable,
                    iva_tratamiento = EXCLUDED.iva_tratamiento,
                    corfo_elegible = EXCLUDED.corfo_elegible,
                    tipo_gasto_corfo = EXCLUDED.tipo_gasto_corfo,
                    nubox_code = EXCLUDED.nubox_code,
                    codigo_f22 = EXCLUDED.codigo_f22,
                    ajuste_14d = EXCLUDED.ajuste_14d,
                    flag_partida = EXCLUDED.flag_partida,
                    flag_concepto = EXCLUDED.flag_concepto,
                    flag_capital = EXCLUDED.flag_capital,
                    flag_activo_fijo = EXCLUDED.flag_activo_fijo,
                    flag_documento = EXCLUDED.flag_documento,
                    flag_control_gestion = EXCLUDED.flag_control_gestion,
                    flag_activo_neto = EXCLUDED.flag_activo_neto,
                    flag_caja = EXCLUDED.flag_caja,
                    flag_marca_14d = EXCLUDED.flag_marca_14d,
                    flag_percepcion = EXCLUDED.flag_percepcion,
                    activa = EXCLUDED.activa,
                    updated_at = now()
                """
            ),
            c,
        )
        counters["cuentas_upserted"] += 1

    # ---------- HABILITACIONES DE CUENTAS POR EMPRESA ----------
    for empresa_codigo, codigos_cuenta in payload.habilitaciones_cuentas.items():
        if empresa_codigo not in empresas_db_set:
            counters["habilitaciones_omitidas_empresa_inexistente"] += len(
                codigos_cuenta
            )
            continue
        for cuenta_codigo in codigos_cuenta:
            await db.execute(
                text(
                    """
                    INSERT INTO core.plan_cuenta_empresa (
                        cuenta_codigo, empresa_codigo, habilitada
                    )
                    VALUES (:cc, :ec, TRUE)
                    ON CONFLICT (cuenta_codigo, empresa_codigo) DO UPDATE
                        SET habilitada = TRUE,
                            habilitada_en = now()
                    """
                ),
                {"cc": cuenta_codigo, "ec": empresa_codigo},
            )
            counters["habilitaciones_cuentas_upserted"] += 1

    # ---------- ÁREAS ----------
    # Las 10 áreas estándar ya están en seed (migración 0034). Acá hacemos
    # UPSERT por si el COO modificó nombre/descripción/activa en el Excel.
    for area in payload.areas:
        await db.execute(
            text(
                """
                INSERT INTO core.areas (codigo, nombre, descripcion, activa)
                VALUES (:codigo, :nombre, :descripcion, :activa)
                ON CONFLICT (codigo) DO UPDATE SET
                    nombre = EXCLUDED.nombre,
                    descripcion = EXCLUDED.descripcion,
                    activa = EXCLUDED.activa,
                    updated_at = now()
                """
            ),
            area,
        )
        counters["areas_upserted"] += 1

    # ---------- ÁREA × EMPRESA ----------
    for area_codigo, empresa_codigos in payload.area_empresa.items():
        for empresa_codigo in empresa_codigos:
            if empresa_codigo not in empresas_db_set:
                continue
            await db.execute(
                text(
                    """
                    INSERT INTO core.area_empresa (area_codigo, empresa_codigo, aplica)
                    VALUES (:ac, :ec, TRUE)
                    ON CONFLICT (area_codigo, empresa_codigo) DO UPDATE
                        SET aplica = TRUE
                    """
                ),
                {"ac": area_codigo, "ec": empresa_codigo},
            )
            counters["area_empresa_upserted"] += 1

    # ---------- PROYECTOS CONTABLES ----------
    for p in payload.proyectos:
        if p["empresa_codigo"] not in empresas_db_set:
            counters["proyectos_omitidos_empresa_inexistente"] += 1
            continue
        await db.execute(
            text(
                """
                INSERT INTO core.proyectos_contables (
                    codigo, empresa_codigo, nombre, tipo_financiamiento,
                    programa, fecha_inicio, fecha_termino,
                    presupuesto_total, moneda, primer_desembolso_corfo,
                    tipos_gasto_elegibles, estado
                )
                VALUES (
                    :codigo, :empresa_codigo, :nombre, :tipo_financiamiento,
                    :programa, :fecha_inicio, :fecha_termino,
                    :presupuesto_total, :moneda, :primer_desembolso_corfo,
                    CAST(:tipos_gasto_elegibles AS TEXT[]), :estado
                )
                ON CONFLICT (codigo) DO UPDATE SET
                    empresa_codigo = EXCLUDED.empresa_codigo,
                    nombre = EXCLUDED.nombre,
                    tipo_financiamiento = EXCLUDED.tipo_financiamiento,
                    programa = EXCLUDED.programa,
                    fecha_inicio = EXCLUDED.fecha_inicio,
                    fecha_termino = EXCLUDED.fecha_termino,
                    presupuesto_total = EXCLUDED.presupuesto_total,
                    moneda = EXCLUDED.moneda,
                    primer_desembolso_corfo = EXCLUDED.primer_desembolso_corfo,
                    tipos_gasto_elegibles = EXCLUDED.tipos_gasto_elegibles,
                    estado = EXCLUDED.estado,
                    updated_at = now()
                """
            ),
            p,
        )
        counters["proyectos_upserted"] += 1

    return counters


# ---------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------


def build_summary(
    payload_or_cuentas: ImportPayload | list[dict[str, Any]],
    habilitaciones: dict[str, set[str]] | None = None,
) -> dict[str, Any]:
    """Resumen JSON-friendly de las 3 secciones del Excel.

    Acepta dos firmas (compatibilidad):
      build_summary(payload)
      build_summary(cuentas, habilitaciones)  # legacy
    """
    if isinstance(payload_or_cuentas, ImportPayload):
        payload = payload_or_cuentas
    else:
        payload = ImportPayload()
        payload.cuentas = payload_or_cuentas
        payload.habilitaciones_cuentas = defaultdict(set, habilitaciones or {})

    by_nivel = Counter(c["nivel"] for c in payload.cuentas)
    by_tipo = Counter(c["tipo"] for c in payload.cuentas)
    by_corfo_tipo = Counter(
        c["tipo_gasto_corfo"]
        for c in payload.cuentas
        if c["corfo_elegible"]
    )

    by_tipo_financiamiento = Counter(
        p["tipo_financiamiento"] for p in payload.proyectos
    )
    by_empresa_proyectos = Counter(
        p["empresa_codigo"] for p in payload.proyectos
    )

    return {
        # Plan de cuentas
        "total_cuentas": len(payload.cuentas),
        "por_nivel": dict(sorted(by_nivel.items())),
        "por_tipo": dict(sorted(by_tipo.items())),
        "imputables": sum(1 for c in payload.cuentas if c["imputable"]),
        "corfo_elegibles": sum(
            1 for c in payload.cuentas if c["corfo_elegible"]
        ),
        "por_tipo_gasto_corfo": {
            str(k): v
            for k, v in sorted(by_corfo_tipo.items(), key=lambda x: str(x[0]))
        },
        "con_codigo_f22": sum(
            1 for c in payload.cuentas if c["codigo_f22"] is not None
        ),
        "habilitaciones_por_empresa": {
            empresa: len(codigos)
            for empresa, codigos in sorted(payload.habilitaciones_cuentas.items())
        },
        # Proyectos contables
        "total_proyectos": len(payload.proyectos),
        "proyectos_por_tipo": dict(sorted(by_tipo_financiamiento.items())),
        "proyectos_por_empresa": dict(sorted(by_empresa_proyectos.items())),
        # Áreas
        "total_areas": len(payload.areas),
        "area_empresa_pares": sum(
            len(empresas) for empresas in payload.area_empresa.values()
        ),
    }
