"""Importador del Excel "Registro de Egresos" de Claudia (REVTECH / TRONGKAI).

Claudia lleva la operación real del subsidio CORFO en dos Excel (`CC
Bancos_Revtech.xlsx` y `Cuenta Bancos_trongkai.xlsx`). La hoja `Registro de
Egresos` es su verdad: una fila por documento con la "SEPARACIÓN VALORES"
(Subsidio / Cehta-Ptec / Cehta / Trewaox). Este módulo la lleva a
`core.corfo_registro_egresos` sin inventar nada:

    parsear_registro_egresos(bytes, empresa) -> ResultadoParseo   # puro, sin BD
    cargar_filas(db, empresa, filas, usuario, dry_run)            # idempotente

# POR QUÉ ES ASÍ

- Los encabezados se buscan POR NOMBRE (tolerante a acentos, mayúsculas y
  `\xa0`), no por posición: REVTECH tiene 16 columnas (`Fuente`) y TRONGKAI
  17 (`Tipo Financiamiento` + `Trewaox`). Un Excel con las columnas
  reordenadas sigue entrando.
- Las filas se importan TAL CUAL. Si el reparto no suma el total o las 4
  fuentes están vacías, la fila entra igual y la pantalla la marca en
  ámbar para que Claudia la resuelva. Cuadrarla acá sería mentirle a CORFO.
- Lo único que se salta es lo que la BD no puede guardar: fila sin fecha
  válida (el mes sale de la fecha), sin descripción o sin total. Cada
  salto lleva el número de fila del Excel y el motivo.
- Idempotencia: cada fila recibe `import_natural_key`, una huella de
  `empresa|rut|tipo|folio|fecha|total|descripción` más el ORDINAL de
  aparición dentro del archivo (la primera lleva la huella base; la n-ésima
  idéntica lleva `|#n`). Re-importar el mismo archivo crea 0 filas (índice
  único parcial + ON CONFLICT DO NOTHING) y el resumen dice cuántas ya
  existían. Las filas idénticas del mismo archivo SON pagos reales (8 en
  REVTECH y 23 en TRONGKAI: cuotas de $5.000.000 a un co-ejecutor el mismo
  día, dos peajes de $3.300): por eso entran todas, cada una con su huella,
  y la repetida queda con una observación para que Claudia confirme que no
  es un tipeo doble. `conservar_repetidas=False` (CLI `--colapsar-repetidas`)
  recupera el comportamiento viejo: sólo entra la primera y el resto va a
  `duplicadas_en_excel`.
- Neto e impuesto vacíos: mismo default que la API (§3.3): los dos vacíos →
  neto = total, impuesto = 0; uno solo → el otro es la diferencia (nunca
  negativa). Así `neto + impuesto ≠ total` queda SOLO en las filas con
  diferencia real (16 en TRONGKAI), no en las 33 boletas que vienen con el
  total y nada más.
- Plata en Decimal a 2 decimales, HALF_UP. Los totales con 4 decimales del
  Excel (conversiones desde UF) se redondean al centavo, igual que hace el
  motor de reparto. Nunca float.
- Las últimas filas de la hoja traen sólo fórmulas residuales (`Estado` =
  "✗ Pendiente" y `Fecha de Pago` = ""): no son gastos y no cuentan ni como
  leídas ni como saltadas.
"""
from __future__ import annotations

import hashlib
import io
import re
import unicodedata
import warnings
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.value_objects.reparto_corfo import (
    ESTADO_DESCUADRADO,
    ESTADO_SIN_CLASIFICAR,
    FUENTES,
    estado_reparto,
    normalizar_montos,
)

ORIGEN_IMPORT_EXCEL = "IMPORT_EXCEL"
HOJA_REGISTRO = "Registro de Egresos"

TIPOS_DOCUMENTO: tuple[str, ...] = (
    "FACTURA",
    "FACTURA_EXENTA",
    "BOLETA",
    "BOLETA_HONORARIO",
    "LIQUIDACION",
    "CO_EJECUTOR",
    "INVOICE",
    "OTRO",
)
ESTADOS_PAGO: tuple[str, ...] = ("PAGADO", "PARCIAL", "PENDIENTE")

_CENT = Decimal("0.01")
_CERO = Decimal("0.00")

#: Filas por INSERT. 426 filas entran en 3 viajes a Supabase en vez de 426.
_TAMANO_LOTE = 200

#: Encabezado normalizado (sin acentos, minúsculas, guiones pegados) → campo.
#: Las dos variantes de Claudia están acá; el resto son sinónimos razonables
#: por si un próximo Excel viene con otro rótulo.
_ALIAS_COLUMNAS: dict[str, str] = {
    "fecha": "fecha",
    "fecha documento": "fecha",
    "descripcion": "descripcion",
    "detalle": "descripcion",
    "rut emisor": "rut_emisor",
    "rut": "rut_emisor",
    "tipo de documento": "tipo_documento",
    "tipo documento": "tipo_documento",
    "tipo doc": "tipo_documento",
    "folio": "folio",
    "numero": "folio",
    "n documento": "folio",
    "monto neto/pagado": "monto_neto",
    "monto neto": "monto_neto",
    "neto": "monto_neto",
    "impuesto/patronal": "impuesto",
    "impuesto": "impuesto",
    "iva": "impuesto",
    "total": "total",
    "monto total": "total",
    "tipo de egreso": "tipo_egreso",
    "tipo egreso": "tipo_egreso",
    "fuente": "fuente",
    "tipo financiamiento": "fuente",
    "tipo de financiamiento": "fuente",
    "proyecto": "proyecto",
    "subsidio": "monto_subsidio",
    "subsidio corfo": "monto_subsidio",
    "cehta-ptec": "monto_cehta_ptec",
    "cehta ptec": "monto_cehta_ptec",
    "cehta": "monto_cehta",
    "trewaox": "monto_trewaox",
    "estado": "estado_pago",
    "estado de pago": "estado_pago",
    "estado pago": "estado_pago",
    "fecha de pago": "fecha_pago",
    "fecha pago": "fecha_pago",
}
#: Sin estas tres no hay gasto que guardar.
_COLUMNAS_OBLIGATORIAS = ("fecha", "descripcion", "total")
#: Columnas que en el Excel son fórmulas: una fila que SÓLO tiene esto no es
#: un gasto, es el residuo de la fórmula arrastrada hasta el final.
_CAMPOS_FORMULA = frozenset({"estado_pago", "fecha_pago"})
#: Cuántas filas se revisan buscando el encabezado (en el Excel real está en la 3).
_FILAS_BUSQUEDA_ENCABEZADO = 20

_FORMATOS_FECHA = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y")

_CAMPOS_MONTO_REPARTO: dict[str, str] = {
    "subsidio": "monto_subsidio",
    "cehta_ptec": "monto_cehta_ptec",
    "cehta": "monto_cehta",
    "trewaox": "monto_trewaox",
}


# ── Dataclasses (contrato §3.4) ──────────────────────────────────────


@dataclass
class FilaEgreso:
    """Una fila del Excel ya normalizada, lista para INSERT."""

    fecha: date
    descripcion: str
    rut_emisor: str | None
    tipo_documento: str
    folio: str | None
    monto_neto: Decimal
    impuesto: Decimal
    total: Decimal
    tipo_egreso: str | None
    fuente: str | None
    proyecto: str | None
    estado_pago: str
    fecha_pago: date | None
    monto_subsidio: Decimal | None
    monto_cehta_ptec: Decimal | None
    monto_cehta: Decimal | None
    monto_trewaox: Decimal | None
    observaciones: str | None
    import_natural_key: str
    fila_excel: int

    @property
    def periodo(self) -> str:
        """YYYY-MM. Lo deriva también el trigger de la BD; acá es para resúmenes."""
        return self.fecha.strftime("%Y-%m")

    @property
    def reparto(self) -> dict[str, Decimal | None]:
        return {f: getattr(self, _CAMPOS_MONTO_REPARTO[f]) for f in FUENTES}

    @property
    def reparto_estado(self) -> str:
        """SIN_CLASIFICAR / OK / DESCUADRADO, con la misma regla que la pantalla."""
        return estado_reparto(self.total, self.reparto)

    @property
    def neto_mas_impuesto_cuadra(self) -> bool:
        return self.monto_neto + self.impuesto == self.total

    def to_dict(self) -> dict[str, Any]:
        """JSON-friendly: fechas ISO, Decimal como string (es plata, no float)."""
        out: dict[str, Any] = {}
        for k, v in asdict(self).items():
            if isinstance(v, Decimal):
                out[k] = str(v)
            elif isinstance(v, date):
                out[k] = v.isoformat()
            else:
                out[k] = v
        return out

    @classmethod
    def from_dict(cls, datos: dict[str, Any]) -> FilaEgreso:
        """Inverso de `to_dict`. Acepta lo que escribió `--json-out`."""

        def _dec(v: Any) -> Decimal | None:
            return None if v is None else Decimal(str(v)).quantize(_CENT, rounding=ROUND_HALF_UP)

        def _fec(v: Any) -> date | None:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            return date.fromisoformat(str(v)[:10])

        fecha = _fec(datos["fecha"])
        total = _dec(datos["total"])
        if fecha is None or total is None:
            raise ValueError("Una FilaEgreso necesita fecha y total")
        return cls(
            fecha=fecha,
            descripcion=str(datos["descripcion"]),
            rut_emisor=datos.get("rut_emisor"),
            tipo_documento=str(datos["tipo_documento"]),
            folio=datos.get("folio"),
            monto_neto=_dec(datos.get("monto_neto")) or _CERO,
            impuesto=_dec(datos.get("impuesto")) or _CERO,
            total=total,
            tipo_egreso=datos.get("tipo_egreso"),
            fuente=datos.get("fuente"),
            proyecto=datos.get("proyecto"),
            estado_pago=str(datos.get("estado_pago") or "PENDIENTE"),
            fecha_pago=_fec(datos.get("fecha_pago")),
            monto_subsidio=_dec(datos.get("monto_subsidio")),
            monto_cehta_ptec=_dec(datos.get("monto_cehta_ptec")),
            monto_cehta=_dec(datos.get("monto_cehta")),
            monto_trewaox=_dec(datos.get("monto_trewaox")),
            observaciones=datos.get("observaciones"),
            import_natural_key=str(datos["import_natural_key"]),
            fila_excel=int(datos["fila_excel"]),
        )


@dataclass
class FilaSaltada:
    fila_excel: int
    motivo: str


@dataclass
class ResultadoParseo:
    filas: list[FilaEgreso] = field(default_factory=list)
    saltadas: list[FilaSaltada] = field(default_factory=list)
    #: Filas (Excel) que son la n-ésima aparición de una idéntica y que SE
    #: CARGAN igual, con huella propia y observación (default). Están también
    #: en `filas`; acá sólo para el resumen.
    repetidas_en_excel: list[int] = field(default_factory=list)
    #: Filas (Excel) de la n-ésima aparición que NO se cargan. Sólo se llena
    #: con `conservar_repetidas=False` (CLI `--colapsar-repetidas`).
    duplicadas_en_excel: list[int] = field(default_factory=list)
    #: Encabezados tal cual venían en el Excel (recortados).
    columnas: list[str] = field(default_factory=list)

    @property
    def leidas(self) -> int:
        """Filas con datos del Excel: cargables + saltadas + colapsadas.

        Las repetidas ya están dentro de `filas`, por eso no se suman aparte:
        273 filas de REVTECH = 271 a cargar + 2 saltadas, no 271 + 2 + 8.
        """
        return len(self.filas) + len(self.saltadas) + len(self.duplicadas_en_excel)


@dataclass
class ResumenCarga:
    """Lo que devuelve `cargar_filas`. Mismos nombres que la respuesta de
    `POST /claudia/egresos/importar`.

    `duplicadas_en_excel` y `saltadas` NO los conoce `cargar_filas` (recibe
    sólo las filas buenas): los completa quien tenga el `ResultadoParseo`.
    """

    empresa_codigo: str
    dry_run: bool
    leidas: int
    creadas: int
    omitidas_existentes: int
    descuadradas: int
    sin_clasificar: int
    duplicadas_en_excel: int = 0
    saltadas: list[FilaSaltada] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ── Normalización de celdas ──────────────────────────────────────────


def _vacio(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _limpiar_texto(v: Any) -> str | None:
    """Recorta espacios, `\xa0` y dobles espacios. Vacío → None."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    # str.split() sin argumentos parte por CUALQUIER espacio Unicode, \xa0 incluido.
    limpio = " ".join(str(v).split())
    return limpio or None


def _clave(s: str) -> str:
    """Forma canónica para comparar rótulos: sin acentos, minúsculas, un solo espacio."""
    sin_acentos = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    sin_acentos = re.sub(r"\s*-\s*", "-", sin_acentos)  # "Cehta - Ptec" → "cehta-ptec"
    return " ".join(sin_acentos.lower().split())


def _to_decimal(v: Any) -> Decimal | None:
    """Celda numérica → Decimal a 2 decimales (HALF_UP). Vacía → None.

    Acepta números de Excel y strings con formato chileno ("1.234.567,89")
    o anglo ("1234567.89"). Lo ilegible levanta ValueError: quien llama
    decide si la fila se salta o si queda una observación.
    """
    if v is None or isinstance(v, bool):
        if v is None:
            return None
        raise ValueError(f"Monto ilegible: {v!r}")
    if isinstance(v, int | Decimal):
        return Decimal(v).quantize(_CENT, rounding=ROUND_HALF_UP)
    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            raise ValueError(f"Monto ilegible: {v!r}")
        # str(float) da la representación corta ("255105.9504"), no el binario.
        return Decimal(str(v)).quantize(_CENT, rounding=ROUND_HALF_UP)
    s = str(v).replace("\xa0", "").replace("$", "").replace(" ", "").strip()
    if s in ("", "-"):
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")  # 1.234,56
        else:
            s = s.replace(",", "")  # 1,234.56
    elif "," in s:
        s = s.replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
        # "94.352" en un Excel chileno son 94.352 pesos, no 94,35: puntos
        # cada 3 dígitos son separador de miles.
        s = s.replace(".", "")
    try:
        return Decimal(s).quantize(_CENT, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError(f"Monto ilegible: {v!r}") from exc


def _to_fecha(v: Any) -> date | None:
    """datetime/date de Excel → date. Strings en formatos habituales → date. Resto → None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()[:10]
        for fmt in _FORMATOS_FECHA:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _normalizar_rut(v: Any) -> str | None:
    """'60.805.000-0' → '60805000-0' · '25543408-k' → '25543408-K' · 763335232 → '76333523-2'.

    Misma forma que deja el trigger de la BD (sin puntos, K mayúscula, con
    guión). No valida el dígito verificador: el Excel se importa tal cual y
    la pantalla lo señala si está mal.
    """
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    limpio = re.sub(r"[^0-9kK]", "", str(v)).upper()
    if not limpio:
        return None
    if len(limpio) < 2:
        return limpio
    return f"{limpio[:-1]}-{limpio[-1]}"


def _normalizar_folio(v: Any) -> str | None:
    """Folio int → str. Folio que Excel convirtió a fecha ("13-05-2026") → dd-mm-yyyy."""
    if v is None:
        return None
    if isinstance(v, datetime | date):
        return v.strftime("%d-%m-%Y")
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int | float | Decimal):
        return str(v)
    return _limpiar_texto(v)


def _normalizar_tipo_documento(v: Any) -> tuple[str, str | None]:
    """→ (código de la BD, texto original si no se reconoció).

    "Boletas"/"Boleta" → BOLETA, "liquidación" → LIQUIDACION, "Co-Ejecutor" →
    CO_EJECUTOR. Lo desconocido va a OTRO y el original a observaciones para
    que no se pierda.
    """
    crudo = _limpiar_texto(v)
    if crudo is None:
        return "OTRO", None
    k = re.sub(r"[\s\-_./]+", " ", _clave(crudo)).strip()
    compacto = k.replace(" ", "")
    if k.startswith("factura") and "exent" in k:
        return "FACTURA_EXENTA", None
    if k.startswith("factura"):
        return "FACTURA", None
    if k.startswith("boleta") and "honor" in k:
        return "BOLETA_HONORARIO", None
    if k.startswith("boleta"):
        return "BOLETA", None
    if k.startswith("liquidacion"):
        return "LIQUIDACION", None
    if compacto.startswith("coejecutor"):
        return "CO_EJECUTOR", None
    if k.startswith("invoice"):
        return "INVOICE", None
    if k in ("otro", "otros"):
        return "OTRO", None
    return "OTRO", crudo


def _normalizar_estado(v: Any) -> tuple[str, str | None]:
    """'✓ Pagado' → PAGADO · '◑ Pagado Parcial' → PARCIAL · '✗ Pendiente' → PENDIENTE.

    Vacío → PENDIENTE. Un texto que no se entiende → PENDIENTE y el original
    en observaciones. "Parcial" se mira ANTES que "Pagado" porque el rótulo
    es "Pagado Parcial".
    """
    crudo = _limpiar_texto(v)
    if crudo is None:
        return "PENDIENTE", None
    k = _clave(crudo)
    if "◑" in crudo or "parcial" in k:
        return "PARCIAL", None
    if "✓" in crudo or "✔" in crudo or "pagad" in k:
        return "PAGADO", None
    if "✗" in crudo or "✘" in crudo or "pendiente" in k or k == "x":
        return "PENDIENTE", None
    return "PENDIENTE", crudo


def _natural_key(
    empresa: str,
    rut: str | None,
    tipo: str,
    folio: str | None,
    fecha: date,
    total: Decimal,
    descripcion: str,
    ordinal: int = 1,
) -> str:
    """sha1("{empresa}|{rut}|{tipo}|{folio}|{fecha}|{total:.2f}|{descripcion.lower()}").

    `ordinal` es la aparición dentro del archivo: la primera fila idéntica
    lleva la huella base (sin sufijo) y la n-ésima lleva "|#n". Es lo que
    hace que dos cuotas iguales del mismo día sean dos gastos y que
    re-importar el mismo archivo no cree ninguna de las dos de nuevo.
    """
    base = (
        f"{empresa}|{rut or ''}|{tipo}|{folio or ''}|{fecha.isoformat()}|"
        f"{total:.2f}|{descripcion.lower()}"
    )
    if ordinal > 1:
        base += f"|#{ordinal}"
    return hashlib.sha1(base.encode("utf-8"), usedforsecurity=False).hexdigest()


def _clp(v: Decimal) -> str:
    """Pesos legibles para observaciones: 5645105.95 → '$5.645.105,95'."""
    s = f"{abs(v):,.2f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")
    if s.endswith(",00"):
        s = s[:-3]
    return f"-${s}" if v < 0 else f"${s}"


def _completar_neto_impuesto(
    total: Decimal, neto: Decimal | None, impuesto: Decimal | None
) -> tuple[Decimal, Decimal, str | None]:
    """→ (neto, impuesto, observación). Mismo default que la API (§3.3).

    - Los dos vacíos → neto = total, impuesto = 0 (boleta, honorarios,
      exenta: 33 filas reales de TRONGKAI vienen así). Queda observado.
    - Uno solo → el otro es la diferencia, lo que haría Excel. Si diera
      negativo NO se inventa nada: quedan como vinieron (el vacío en 0) y
      la observación lo dice; la pantalla lo marca como no cuadra.
    - Los dos con valor → tal cual, sumen o no el total (eso es información
      que hay que mostrar, no corregir).
    """
    if neto is None and impuesto is None:
        return total, _CERO, "Neto e impuesto vacíos en el Excel"
    if neto is not None and impuesto is not None:
        return neto, impuesto, None
    if neto is None:
        assert impuesto is not None  # para el tipo: el caso None/None ya salió
        resto = total - impuesto
        if resto < 0:
            return (
                _CERO,
                impuesto,
                f"Impuesto {_clp(impuesto)} supera el total {_clp(total)} en el Excel: "
                "el neto vacío queda en $0",
            )
        return resto, impuesto, None
    resto = total - neto
    if resto < 0:
        return (
            neto,
            _CERO,
            f"Neto {_clp(neto)} supera el total {_clp(total)} en el Excel: "
            "el impuesto vacío queda en $0",
        )
    return neto, resto, None


# ── Detección de hoja y encabezados ──────────────────────────────────


def _mapear_encabezados(celdas: tuple[Any, ...]) -> tuple[dict[str, int], list[str]]:
    """→ ({campo: índice de columna}, [rótulos originales]).

    Devuelve vacío si la fila no es la de encabezados (le faltan Fecha,
    Descripción o Total).
    """
    indices: dict[str, int] = {}
    rotulos: list[str] = []
    for i, celda in enumerate(celdas):
        rotulo = _limpiar_texto(celda)
        if rotulo is None:
            continue
        rotulos.append(rotulo)
        campo = _ALIAS_COLUMNAS.get(_clave(rotulo))
        # La primera columna que reclama un campo se lo queda (si el Excel
        # trae "Cehta" dos veces, vale la primera).
        if campo and campo not in indices:
            indices[campo] = i
    if all(c in indices for c in _COLUMNAS_OBLIGATORIAS):
        return indices, rotulos
    return {}, []


def _buscar_encabezado(ws: Any) -> tuple[int, dict[str, int], list[str]] | None:
    for n, celdas in enumerate(
        ws.iter_rows(min_row=1, max_row=_FILAS_BUSQUEDA_ENCABEZADO, values_only=True), start=1
    ):
        indices, rotulos = _mapear_encabezados(tuple(celdas))
        if indices:
            return n, indices, rotulos
    return None


def _elegir_hoja(wb: Any) -> tuple[Any, int, dict[str, int], list[str]]:
    """La hoja `Registro de Egresos` si existe y tiene encabezados; si no, la
    primera que los tenga."""
    candidatas = [ws for ws in wb.worksheets if _clave(ws.title) == _clave(HOJA_REGISTRO)]
    candidatas += [ws for ws in wb.worksheets if ws not in candidatas]
    for ws in candidatas:
        hallazgo = _buscar_encabezado(ws)
        if hallazgo:
            fila, indices, rotulos = hallazgo
            return ws, fila, indices, rotulos
    raise ValueError(
        f"No encontré la hoja '{HOJA_REGISTRO}' ni ninguna hoja con los encabezados "
        f"'Fecha', 'Descripción' y 'Total' en sus primeras {_FILAS_BUSQUEDA_ENCABEZADO} filas."
    )


# ── Parser ───────────────────────────────────────────────────────────


def _parsear_fila(
    empresa: str, fila_excel: int, valores: dict[str, Any]
) -> FilaEgreso | FilaSaltada:
    obs: list[str] = []

    fecha = _to_fecha(valores.get("fecha"))
    descripcion = _limpiar_texto(valores.get("descripcion"))
    if fecha is None:
        contexto = descripcion or "sin descripción"
        return FilaSaltada(
            fila_excel,
            f"Fecha inválida ({valores.get('fecha')!r}) — {contexto}, total "
            f"{_limpiar_texto(valores.get('total')) or 'vacío'}",
        )
    if descripcion is None:
        return FilaSaltada(fila_excel, "Sin descripción")

    try:
        total = _to_decimal(valores.get("total"))
    except ValueError as exc:
        return FilaSaltada(fila_excel, f"Total ilegible ({valores.get('total')!r}) — {exc}")
    if total is None:
        return FilaSaltada(fila_excel, f"Sin total — {descripcion}")
    if total < 0:
        return FilaSaltada(fila_excel, f"Total negativo ({total}) — {descripcion}")

    crudos_base: dict[str, Decimal | None] = {}
    ilegible = False
    for campo, rotulo in (("monto_neto", "Monto neto"), ("impuesto", "Impuesto")):
        try:
            crudos_base[campo] = _to_decimal(valores.get(campo))
        except ValueError:
            crudos_base[campo] = _CERO
            ilegible = True
            obs.append(f"{rotulo} ilegible en el Excel: {valores.get(campo)!r}")
    if ilegible:
        # No se deriva plata a partir de una celda que no se entiende: lo
        # ilegible queda en 0 (ya observado) y lo vacío también.
        neto = crudos_base["monto_neto"]
        impuesto = crudos_base["impuesto"]
        montos_base = {
            "monto_neto": neto if neto is not None else _CERO,
            "impuesto": impuesto if impuesto is not None else _CERO,
        }
    else:
        neto, impuesto, nota = _completar_neto_impuesto(
            total, crudos_base["monto_neto"], crudos_base["impuesto"]
        )
        montos_base = {"monto_neto": neto, "impuesto": impuesto}
        if nota:
            obs.append(nota)

    tipo_documento, tipo_original = _normalizar_tipo_documento(valores.get("tipo_documento"))
    if tipo_original:
        obs.append(f"Tipo de documento en el Excel: {tipo_original}")
    elif tipo_documento == "OTRO" and _vacio(valores.get("tipo_documento")):
        obs.append("Sin tipo de documento en el Excel")

    estado_pago, estado_original = _normalizar_estado(valores.get("estado_pago"))
    if estado_original:
        obs.append(f"Estado en el Excel: {estado_original}")

    fecha_pago = _to_fecha(valores.get("fecha_pago"))
    if fecha_pago is None and not _vacio(valores.get("fecha_pago")):
        obs.append(f"Fecha de pago ilegible en el Excel: {valores.get('fecha_pago')!r}")

    crudos: dict[str, Decimal | None] = {}
    for fuente, campo in _CAMPOS_MONTO_REPARTO.items():
        try:
            crudos[fuente] = _to_decimal(valores.get(campo))
        except ValueError:
            crudos[fuente] = None
            obs.append(f"{fuente} ilegible en el Excel: {valores.get(campo)!r}")
    # Todo-o-nada lo resuelve el motor: 4 vacías → None; alguna con valor → el resto 0.
    reparto = normalizar_montos(crudos)

    rut = _normalizar_rut(valores.get("rut_emisor"))
    folio = _normalizar_folio(valores.get("folio"))

    return FilaEgreso(
        fecha=fecha,
        descripcion=descripcion,
        rut_emisor=rut,
        tipo_documento=tipo_documento,
        folio=folio,
        monto_neto=montos_base["monto_neto"],
        impuesto=montos_base["impuesto"],
        total=total,
        tipo_egreso=_limpiar_texto(valores.get("tipo_egreso")),
        fuente=_limpiar_texto(valores.get("fuente")),
        proyecto=_limpiar_texto(valores.get("proyecto")),
        estado_pago=estado_pago,
        fecha_pago=fecha_pago,
        monto_subsidio=reparto["subsidio"],
        monto_cehta_ptec=reparto["cehta_ptec"],
        monto_cehta=reparto["cehta"],
        monto_trewaox=reparto["trewaox"],
        observaciones=" · ".join(obs) or None,
        import_natural_key=_natural_key(
            empresa, rut, tipo_documento, folio, fecha, total, descripcion
        ),
        fila_excel=fila_excel,
    )


def _recorrer_hoja(wb: Any, empresa: str, conservar_repetidas: bool) -> ResultadoParseo:
    ws, fila_encabezado, indices, rotulos = _elegir_hoja(wb)
    resultado = ResultadoParseo(columnas=rotulos)
    #: huella → (primera fila del Excel con esa huella, apariciones vistas)
    vistas: dict[str, tuple[int, int]] = {}
    for fila_excel, celdas in enumerate(
        ws.iter_rows(min_row=fila_encabezado + 1, values_only=True),
        start=fila_encabezado + 1,
    ):
        valores = {
            campo: (celdas[i] if i < len(celdas) else None) for campo, i in indices.items()
        }
        if all(_vacio(v) for campo, v in valores.items() if campo not in _CAMPOS_FORMULA):
            continue
        parseada = _parsear_fila(empresa, fila_excel, valores)
        if isinstance(parseada, FilaSaltada):
            resultado.saltadas.append(parseada)
            continue
        previa = vistas.get(parseada.import_natural_key)
        if previa is None:
            vistas[parseada.import_natural_key] = (fila_excel, 1)
            resultado.filas.append(parseada)
            continue
        primera, apariciones = previa
        vistas[parseada.import_natural_key] = (primera, apariciones + 1)
        if not conservar_repetidas:
            resultado.duplicadas_en_excel.append(fila_excel)
            continue
        # Entra como gasto propio (huella con "#n", estable entre corridas:
        # re-importar no la duplica) y queda marcada para que Claudia
        # confirme que es otro pago y no un tipeo repetido.
        resultado.repetidas_en_excel.append(fila_excel)
        parseada.import_natural_key = _natural_key(
            empresa,
            parseada.rut_emisor,
            parseada.tipo_documento,
            parseada.folio,
            parseada.fecha,
            parseada.total,
            parseada.descripcion,
            ordinal=apariciones + 1,
        )
        nota = (
            f"Idéntica a la fila {primera} del Excel (aparición {apariciones + 1}): "
            "verificar que sea un gasto distinto"
        )
        parseada.observaciones = (
            f"{parseada.observaciones} · {nota}" if parseada.observaciones else nota
        )
        resultado.filas.append(parseada)
    return resultado


def parsear_registro_egresos(
    contenido: bytes, empresa_codigo: str, *, conservar_repetidas: bool = True
) -> ResultadoParseo:
    """Lee el .xlsx y devuelve las filas normalizadas. Puro: no toca la BD.

    `empresa_codigo` entra en la huella de cada fila: el mismo Excel cargado
    en la empresa equivocada NO chocaría con el índice único, así que quien
    llama tiene que pasar la empresa correcta.

    `conservar_repetidas` (por defecto SÍ, es el contrato §3.4): las filas
    idénticas del mismo archivo son pagos distintos (cuotas de $5.000.000 a
    un co-ejecutor el mismo día, dos peajes de $3.300 en la misma fecha) y
    entran todas: la primera con la huella base y la n-ésima con "|#n"
    (estable entre corridas, así re-importar crea 0 filas), más una
    observación que dice a qué fila del Excel se parece; sus números van a
    `repetidas_en_excel`. Con `False` (CLI `--colapsar-repetidas`) sólo
    entra la primera y las demás van a `duplicadas_en_excel` sin cargarse.
    """
    empresa = (empresa_codigo or "").strip().upper()
    if not empresa:
        raise ValueError("Falta la empresa (REVTECH o TRONGKAI)")
    if not contenido:
        raise ValueError("El archivo está vacío")

    with warnings.catch_warnings():
        # "Data Validation extension is not supported": el Excel de Claudia
        # trae listas desplegables. En read_only la hoja se parsea recién al
        # iterar, así que el filtro tiene que cubrir también el recorrido.
        warnings.simplefilter("ignore", category=UserWarning)
        try:
            wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
        except Exception as exc:  # openpyxl levanta de todo (zip, xml, …)
            raise ValueError(f"El archivo no es un .xlsx válido: {exc}") from exc
        try:
            return _recorrer_hoja(wb, empresa, conservar_repetidas)
        finally:
            wb.close()


def contar_reparto(filas: list[FilaEgreso]) -> tuple[int, int]:
    """→ (descuadradas, sin_clasificar), con la regla del motor."""
    descuadradas = sum(1 for f in filas if f.reparto_estado == ESTADO_DESCUADRADO)
    sin_clasificar = sum(1 for f in filas if f.reparto_estado == ESTADO_SIN_CLASIFICAR)
    return descuadradas, sin_clasificar


# ── Carga ────────────────────────────────────────────────────────────


def empresas_corfo() -> frozenset[str]:
    """Las empresas que rinden CORFO, desde `corfo_rendiciones` (fuente única).

    Import perezoso: ese módulo arrastra FastAPI y `settings` (exige
    DATABASE_URL), y el parser tiene que poder correr en una máquina sin
    `.env` (`--dry-run` de la CLI). Si no se puede importar, el fallback son
    las mismas dos empresas: es la lista del subsidio 2024-265638.
    """
    try:
        from app.api.v1.corfo_rendiciones import CORFO_EMPRESAS
    except Exception:  # sin FastAPI/settings a mano: las mismas dos empresas
        return frozenset({"REVTECH", "TRONGKAI"})
    return frozenset(CORFO_EMPRESAS)


_COLUMNAS_INSERT: tuple[str, ...] = (
    "fecha",
    "descripcion",
    "rut_emisor",
    "tipo_documento",
    "folio",
    "monto_neto",
    "impuesto",
    "total",
    "tipo_egreso",
    "fuente",
    "proyecto",
    "estado_pago",
    "fecha_pago",
    "monto_subsidio",
    "monto_cehta_ptec",
    "monto_cehta",
    "monto_trewaox",
    "observaciones",
    "import_natural_key",
)
_TIPOS_SQL: dict[str, str] = {
    "fecha": "date[]",
    "fecha_pago": "date[]",
    "monto_neto": "numeric[]",
    "impuesto": "numeric[]",
    "total": "numeric[]",
    "monto_subsidio": "numeric[]",
    "monto_cehta_ptec": "numeric[]",
    "monto_cehta": "numeric[]",
    "monto_trewaox": "numeric[]",
}

#: Un INSERT por lote vía UNNEST de arrays (un viaje por cada 200 filas).
#: `periodo` lo pisa el trigger BEFORE desde `fecha`; se manda igual para
#: que el INSERT sea válido aunque el trigger no esté.
#: Los CAST van como CAST(:p AS tipo[]), NUNCA `:p::tipo[]` (SQLAlchemy no
#: reconoce el bind seguido de `::`).
SQL_INSERT_EGRESOS = f"""
INSERT INTO core.corfo_registro_egresos (
    empresa_codigo, periodo, {", ".join(_COLUMNAS_INSERT)},
    origen, created_by
)
SELECT
    CAST(:empresa AS text), to_char(f.fecha, 'YYYY-MM'),
    {", ".join(f"f.{c}" for c in _COLUMNAS_INSERT)},
    '{ORIGEN_IMPORT_EXCEL}', CAST(:usuario AS text)
FROM unnest(
    {", ".join(f"CAST(:{c} AS {_TIPOS_SQL.get(c, 'text[]')})" for c in _COLUMNAS_INSERT)}
) AS f({", ".join(_COLUMNAS_INSERT)})
ON CONFLICT (import_natural_key) WHERE import_natural_key IS NOT NULL DO NOTHING
RETURNING egreso_id
"""

SQL_CLAVES_EXISTENTES = """
SELECT import_natural_key
  FROM core.corfo_registro_egresos
 WHERE import_natural_key = ANY(CAST(:claves AS text[]))
"""


def _parametros_lote(
    empresa: str, usuario_email: str, lote: list[FilaEgreso]
) -> dict[str, Any]:
    params: dict[str, Any] = {"empresa": empresa, "usuario": usuario_email}
    for c in _COLUMNAS_INSERT:
        params[c] = [getattr(f, c) for f in lote]
    return params


async def cargar_filas(
    db: AsyncSession,
    empresa_codigo: str,
    filas: list[FilaEgreso],
    usuario_email: str,
    dry_run: bool = False,
) -> ResumenCarga:
    """INSERT idempotente de las filas parseadas.

    - `ON CONFLICT (import_natural_key) … DO NOTHING RETURNING egreso_id`:
      `creadas` son los ids que volvieron; `omitidas_existentes`, el resto.
      Una fila borrada lógicamente conserva su huella, así que también
      cuenta como existente (re-importar no la resucita).
    - `dry_run`: no escribe nada; consulta qué huellas ya existen para que
      los conteos sean reales igual.
    - Commitea al final si no es dry_run: es una carga completa, no un paso
      dentro de otra transacción. Si un lote falla, no queda nada a medias.
    """
    empresa = (empresa_codigo or "").strip().upper()
    if empresa not in empresas_corfo():
        raise ValueError(
            f"El registro de egresos CORFO es sólo para REVTECH o TRONGKAI (llegó {empresa!r})"
        )
    usuario = (usuario_email or "").strip()
    if not usuario:
        raise ValueError("Falta el email del usuario que importa (created_by)")

    descuadradas, sin_clasificar = contar_reparto(filas)
    resumen = ResumenCarga(
        empresa_codigo=empresa,
        dry_run=dry_run,
        leidas=len(filas),
        creadas=0,
        omitidas_existentes=0,
        descuadradas=descuadradas,
        sin_clasificar=sin_clasificar,
    )
    if not filas:
        return resumen

    if dry_run:
        claves = [f.import_natural_key for f in filas]
        existentes: set[str] = set()
        for i in range(0, len(claves), _TAMANO_LOTE):
            res = await db.execute(
                text(SQL_CLAVES_EXISTENTES), {"claves": claves[i : i + _TAMANO_LOTE]}
            )
            existentes.update(str(r[0]) for r in res.fetchall())
        resumen.omitidas_existentes = sum(1 for c in claves if c in existentes)
        resumen.creadas = len(claves) - resumen.omitidas_existentes
        return resumen

    for i in range(0, len(filas), _TAMANO_LOTE):
        lote = filas[i : i + _TAMANO_LOTE]
        res = await db.execute(text(SQL_INSERT_EGRESOS), _parametros_lote(empresa, usuario, lote))
        creadas = len(res.fetchall())
        resumen.creadas += creadas
        resumen.omitidas_existentes += len(lote) - creadas
    await db.commit()
    return resumen
