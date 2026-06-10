"""Round 152w — Generador de Rendiciones CORFO.

Para REVTECH + TRONGKAI (subsidio CORFO 2026). Genera 2 Excel con la
estructura OFICIAL CORFO (folio tipo 2024-265638):

  - RendicionesRRHH.xlsx   (17 cols, una fila por persona/mes)
  - RendicionesGastos.xlsx (21 cols, una fila por documento)

Pre-llena automáticamente lo que se puede desde vouchers + remuneraciones
y deja las columnas CORFO-específicas vacías con DATA VALIDATION (dropdowns)
para que el user elija valores del catálogo oficial.

Endpoints:
  GET  /admin/corfo/rendicion/preview     ?empresa=REVTECH&periodo=2026-04
  POST /admin/corfo/rendicion/excel       (json: {empresa, periodo, tipo='gastos'|'rrhh'})
  GET  /admin/corfo/catalogos             (dropdowns CORFO)
  GET  /admin/corfo/mapping/{empresa}     (mapeo cuenta_local→CORFO)
  POST /admin/corfo/mapping/{empresa}     (guardar mapeo)
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession

router = APIRouter(prefix="/admin/corfo", tags=["corfo-rendiciones"])

# Empresas autorizadas para rendir CORFO (subsidio 2026)
CORFO_EMPRESAS = {"REVTECH", "TRONGKAI"}

MESES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def _periodo_to_corfo(periodo: str) -> str:
    """'2026-04' -> 'Abr de 2026'"""
    y, m = periodo.split("-")
    return f"{MESES_ES[int(m) - 1]} de {y}"


def _require_corfo_empresa(empresa: str) -> None:
    if empresa not in CORFO_EMPRESAS:
        raise HTTPException(
            status_code=400,
            detail=f"La rendición CORFO solo aplica para REVTECH o TRONGKAI. Recibido: {empresa}",
        )


async def _require_admin(user: CurrentUser) -> None:
    if user.app_role not in {"admin", "finance"}:
        raise HTTPException(status_code=403, detail="Solo admin o finance")


# =====================================================================
# Catálogos CORFO (dropdowns oficiales)
# =====================================================================
class CatalogosResp(BaseModel):
    cuenta_gastos: list[str]
    cuenta_rrhh: list[str]
    item_gastos: list[str]
    tipo_doc_gastos: list[str]
    tipo_doc_rrhh: list[str]
    etapa: list[str]


@router.get("/catalogos", response_model=CatalogosResp)
async def get_catalogos(user: CurrentUser, db: DBSession) -> CatalogosResp:
    rows = (await db.execute(
        text("SELECT catalogo, valor FROM core.corfo_catalogos WHERE active = TRUE ORDER BY catalogo, orden")
    )).fetchall()
    cats: dict[str, list[str]] = {}
    for r in rows:
        cats.setdefault(r[0], []).append(r[1])
    return CatalogosResp(
        cuenta_gastos=cats.get("cuenta_gastos", []),
        cuenta_rrhh=cats.get("cuenta_rrhh", []),
        item_gastos=cats.get("item_gastos", []),
        tipo_doc_gastos=cats.get("tipo_doc_gastos", []),
        tipo_doc_rrhh=cats.get("tipo_doc_rrhh", []),
        etapa=cats.get("etapa", []),
    )


# =====================================================================
# Preview — qué vouchers/líneas se incluirían en la rendición
# =====================================================================
class PreviewRow(BaseModel):
    voucher_id: int
    fecha: str
    cuenta_codigo: str
    cuenta_nombre: str | None
    monto_neto: float
    monto_iva: float
    monto_total: float
    proveedor_rut: str | None
    proveedor_nombre: str | None
    folio: str | None
    glosa: str | None
    # mapeo CORFO si ya está definido
    corfo_cuenta: str | None
    corfo_item: str | None


class PreviewResp(BaseModel):
    empresa_codigo: str
    periodo: str
    periodo_corfo: str
    rows: list[PreviewRow]
    total_neto: float
    total_iva: float
    total_total: float
    sin_mapeo: int


@router.get("/rendicion/preview", response_model=PreviewResp)
async def preview_rendicion(
    user: CurrentUser,
    db: DBSession,
    empresa: str = Query(..., description="REVTECH o TRONGKAI"),
    periodo: str = Query(..., pattern=r"^\d{4}-\d{2}$"),
) -> PreviewResp:
    """Preview de vouchers EXECUTED del período que irían a la rendición de Gastos."""
    await _require_admin(user)
    _require_corfo_empresa(empresa)

    y, m = periodo.split("-")
    # Vouchers tipo COMPRA del periodo, con sus líneas + montos
    sql = """
        SELECT
            v.voucher_id, v.fecha_contable::text AS fecha,
            v.contraparte_rut, v.contraparte_nombre,
            v.doc_tributario_folio AS folio, v.glosa,
            COALESCE((
                SELECT SUM(CASE WHEN vl2.balance_treatment='GASTO' THEN vl2.debit ELSE 0 END)
                FROM core.voucher_lines vl2 WHERE vl2.voucher_id = v.voucher_id
            ), 0) AS gasto_neto,
            COALESCE((
                SELECT SUM(vl2.iva_amount)
                FROM core.voucher_lines vl2 WHERE vl2.voucher_id = v.voucher_id AND vl2.iva_amount IS NOT NULL
            ), 0) AS iva,
            (
                SELECT vl3.cuenta_codigo
                FROM core.voucher_lines vl3 WHERE vl3.voucher_id = v.voucher_id
                  AND vl3.balance_treatment = 'GASTO'
                ORDER BY vl3.line_number LIMIT 1
            ) AS gasto_cuenta,
            (
                SELECT pce.nombre
                FROM core.voucher_lines vl4
                JOIN core.plan_cuenta_empresa pce ON pce.codigo = vl4.cuenta_codigo
                                                  AND pce.empresa_codigo = v.empresa_codigo
                WHERE vl4.voucher_id = v.voucher_id AND vl4.balance_treatment = 'GASTO'
                ORDER BY vl4.line_number LIMIT 1
            ) AS gasto_cuenta_nombre
        FROM core.vouchers v
        WHERE v.empresa_codigo = :emp
          AND v.tipo = 'COMPRA'
          -- R152EEEEEE — Rango fecha en lugar de EXTRACT() para usar
          -- ix_vouchers_fecha_contable. EXTRACT() mataba el índice.
          AND v.fecha_contable >= make_date(:y, :m, 1)
          AND v.fecha_contable <  (make_date(:y, :m, 1) + INTERVAL '1 month')
          AND v.status IN ('APPROVED','EXECUTED','SYNCED','RECONCILED')
        ORDER BY v.fecha_contable, v.voucher_id
    """
    rows_db = (await db.execute(text(sql), {"emp": empresa, "y": int(y), "m": int(m)})).fetchall()

    # Mapeo CORFO existente para esta empresa
    mapping_db = (await db.execute(
        text("SELECT cuenta_codigo, corfo_cuenta, corfo_item FROM core.corfo_cuenta_mapping WHERE empresa_codigo = :e"),
        {"e": empresa},
    )).fetchall()
    mapping = {r[0]: (r[1], r[2]) for r in mapping_db}

    rows: list[PreviewRow] = []
    # R152JJJJJJ — acumular en Decimal (no float): los NUMERIC de Postgres
    # llegan como Decimal vía asyncpg; el cast a float perdía precisión en
    # montos con decimales (UF/USD) y violaba la regla "Decimal everywhere".
    total_neto = total_iva = Decimal("0")
    sin_mapeo = 0
    for r in rows_db:
        neto = Decimal(str(r[6] or 0))
        iva = Decimal(str(r[7] or 0))
        total = neto + iva
        cc = r[8] or ""
        ccname = r[9]
        corfo_cuenta, corfo_item = mapping.get(cc, (None, None))
        if not corfo_cuenta:
            sin_mapeo += 1
        rows.append(PreviewRow(
            voucher_id=r[0], fecha=r[1],
            cuenta_codigo=cc, cuenta_nombre=ccname,
            monto_neto=neto, monto_iva=iva, monto_total=total,
            proveedor_rut=r[2], proveedor_nombre=r[3],
            folio=r[4], glosa=r[5],
            corfo_cuenta=corfo_cuenta, corfo_item=corfo_item,
        ))
        total_neto += neto
        total_iva += iva

    return PreviewResp(
        empresa_codigo=empresa, periodo=periodo,
        periodo_corfo=_periodo_to_corfo(periodo),
        rows=rows,
        total_neto=total_neto, total_iva=total_iva, total_total=total_neto + total_iva,
        sin_mapeo=sin_mapeo,
    )


# =====================================================================
# Guardar mapeo cuenta_local → CORFO
# =====================================================================
class MappingItem(BaseModel):
    cuenta_codigo: str
    corfo_cuenta: str
    corfo_item: str | None = None
    corfo_cargo: str | None = None


class MappingPost(BaseModel):
    items: list[MappingItem]


# =====================================================================
# R152x — vista FULL: todas las cuentas locales usadas en COMPRAS
# históricas + mapeo CORFO actual (si existe) + usage_count para priorizar.
# =====================================================================
class CuentaUsoRow(BaseModel):
    cuenta_codigo: str
    cuenta_nombre: str | None
    uso_count: int          # cuántos vouchers usaron esta cuenta
    monto_acumulado: float  # total $ histórico
    corfo_cuenta: str | None  # ya mapeada
    corfo_item: str | None
    corfo_cargo: str | None


@router.get("/mapping/{empresa}/full", response_model=list[CuentaUsoRow])
async def mapping_full(
    empresa: str, user: CurrentUser, db: DBSession,
) -> list[CuentaUsoRow]:
    """Devuelve TODAS las cuentas contables usadas históricamente en
    vouchers COMPRA de esta empresa, junto con el mapeo CORFO actual
    si ya existe. Ordenado por uso descendente.
    """
    await _require_admin(user)
    _require_corfo_empresa(empresa)
    rows = (await db.execute(text("""
        WITH usos AS (
            SELECT
                vl.cuenta_codigo,
                COUNT(DISTINCT v.voucher_id) AS uso_count,
                SUM(vl.debit) AS monto_acumulado
            FROM core.vouchers v
            JOIN core.voucher_lines vl ON vl.voucher_id = v.voucher_id
            WHERE v.empresa_codigo = :emp
              AND v.tipo = 'COMPRA'
              AND vl.balance_treatment = 'GASTO'
            GROUP BY vl.cuenta_codigo
        )
        SELECT
            u.cuenta_codigo,
            pce.nombre AS cuenta_nombre,
            u.uso_count,
            COALESCE(u.monto_acumulado, 0) AS monto_acumulado,
            m.corfo_cuenta, m.corfo_item, m.corfo_cargo
        FROM usos u
        LEFT JOIN core.plan_cuenta_empresa pce
          ON pce.codigo = u.cuenta_codigo AND pce.empresa_codigo = :emp
        LEFT JOIN core.corfo_cuenta_mapping m
          ON m.cuenta_codigo = u.cuenta_codigo AND m.empresa_codigo = :emp
        ORDER BY u.uso_count DESC, u.cuenta_codigo
    """), {"emp": empresa})).fetchall()
    return [
        CuentaUsoRow(
            cuenta_codigo=r[0], cuenta_nombre=r[1],
            uso_count=int(r[2]),
            monto_acumulado=float(r[3] or 0),
            corfo_cuenta=r[4], corfo_item=r[5], corfo_cargo=r[6],
        )
        for r in rows
    ]


@router.get("/mapping/{empresa}")
async def get_mapping(empresa: str, user: CurrentUser, db: DBSession) -> list[MappingItem]:
    await _require_admin(user)
    _require_corfo_empresa(empresa)
    rows = (await db.execute(
        text("""SELECT cuenta_codigo, corfo_cuenta, corfo_item, corfo_cargo
                FROM core.corfo_cuenta_mapping WHERE empresa_codigo = :e
                ORDER BY cuenta_codigo"""),
        {"e": empresa},
    )).fetchall()
    return [MappingItem(cuenta_codigo=r[0], corfo_cuenta=r[1], corfo_item=r[2], corfo_cargo=r[3]) for r in rows]


@router.post("/mapping/{empresa}")
async def set_mapping(
    empresa: str, body: MappingPost, user: CurrentUser, db: DBSession,
) -> dict[str, int]:
    await _require_admin(user)
    _require_corfo_empresa(empresa)
    n = 0
    for it in body.items:
        await db.execute(
            text("""
                INSERT INTO core.corfo_cuenta_mapping
                    (empresa_codigo, cuenta_codigo, corfo_cuenta, corfo_item, corfo_cargo)
                VALUES (:e, :c, :cc, :ci, :cg)
                ON CONFLICT (empresa_codigo, cuenta_codigo) DO UPDATE SET
                    corfo_cuenta = EXCLUDED.corfo_cuenta,
                    corfo_item = EXCLUDED.corfo_item,
                    corfo_cargo = EXCLUDED.corfo_cargo,
                    updated_at = NOW()
            """),
            {"e": empresa, "c": it.cuenta_codigo, "cc": it.corfo_cuenta,
             "ci": it.corfo_item, "cg": it.corfo_cargo},
        )
        n += 1
    await db.commit()
    return {"saved": n}


# =====================================================================
# Generar Excel (Gastos o RRHH) con DATA VALIDATION
# =====================================================================
class GenExcelReq(BaseModel):
    empresa: str
    periodo: str
    tipo: str  # 'gastos' o 'rrhh'


@router.post("/rendicion/excel")
async def generar_excel(
    body: GenExcelReq, user: CurrentUser, db: DBSession,
) -> Response:
    """Genera el Excel CORFO pre-llenado + con dropdowns oficiales."""
    await _require_admin(user)
    _require_corfo_empresa(body.empresa)
    if body.tipo not in {"gastos", "rrhh"}:
        raise HTTPException(status_code=400, detail="tipo debe ser 'gastos' o 'rrhh'")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    # Catálogos
    cats_db = (await db.execute(text("SELECT catalogo, valor FROM core.corfo_catalogos WHERE active=TRUE ORDER BY catalogo, orden"))).fetchall()
    cats: dict[str, list[str]] = {}
    for r in cats_db:
        cats.setdefault(r[0], []).append(r[1])

    # Periodo CORFO (ej "Abr de 2026")
    periodo_corfo = _periodo_to_corfo(body.periodo)

    wb = Workbook()
    ws = wb.active
    thin = Side(border_style="thin", color="D2D2D7")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    H = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    H_FILL = PatternFill("solid", start_color="1D6F42")
    YELLOW = PatternFill("solid", start_color="FFF3CD")

    if body.tipo == "gastos":
        ws.title = "Carga_Gastos"
        headers = [
            "Cuenta", "Ítem", "Fuente Financiamiento", "Periodo", "Etapa",
            "Tipo Documento", "N° Documento", "Rut Proveedor",
            "Nombre Proveedor o Razón Social", "Monto Neto", "Monto IVA",
            "Monto Total", "Monto Rendir", "Fecha de Recepción",
            "Monto Cancelado", "Forma de Pago", "Fecha de Pago",
            "Fecha del documento", "Glosa / Justificación", "Receptor Rut",
            "Nombre Receptor",
        ]
    else:
        ws.title = "Carga_RRHH"
        headers = [
            "Cuenta", "Ítem", "Fuente de Financiamiento", "Periodo", "Etapa",
            "Tipo documento", "N° Documento", "RUT RRHH", "Nombre RRHH",
            "Monto Total", "Monto a Rendir", "Valor Hora", "Hora rendidas / Mes",
            "Forma de Pago", "Fecha de Pago", "Fecha del Documento", "Glosa",
        ]

    # Header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = H; cell.fill = H_FILL; cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Pre-fill datos reales
    y, m = body.periodo.split("-")
    if body.tipo == "gastos":
        # Vouchers tipo COMPRA del período + fecha_ejecucion para llenar Forma/Fecha Pago
        rows_db = (await db.execute(text("""
            SELECT
                v.voucher_id, v.fecha_contable::text, v.contraparte_rut, v.contraparte_nombre,
                v.doc_tributario_folio, v.glosa,
                COALESCE((SELECT SUM(CASE WHEN vl.balance_treatment='GASTO' THEN vl.debit ELSE 0 END)
                          FROM core.voucher_lines vl WHERE vl.voucher_id = v.voucher_id), 0) AS neto,
                COALESCE((SELECT SUM(vl.iva_amount) FROM core.voucher_lines vl
                          WHERE vl.voucher_id = v.voucher_id AND vl.iva_amount IS NOT NULL), 0) AS iva,
                (SELECT vl.cuenta_codigo FROM core.voucher_lines vl WHERE vl.voucher_id = v.voucher_id
                 AND vl.balance_treatment='GASTO' ORDER BY vl.line_number LIMIT 1) AS cuenta,
                v.fecha_ejecucion::text AS fecha_ejecucion,
                v.status
            FROM core.vouchers v
            WHERE v.empresa_codigo = :emp AND v.tipo = 'COMPRA'
              AND EXTRACT(YEAR FROM v.fecha_contable) = :y
              AND EXTRACT(MONTH FROM v.fecha_contable) = :m
              AND v.status IN ('APPROVED','EXECUTED','SYNCED','RECONCILED')
            ORDER BY v.fecha_contable, v.voucher_id
        """), {"emp": body.empresa, "y": int(y), "m": int(m)})).fetchall()
        # Mapeo
        mapping_db = (await db.execute(
            text("SELECT cuenta_codigo, corfo_cuenta, corfo_item FROM core.corfo_cuenta_mapping WHERE empresa_codigo = :e"),
            {"e": body.empresa},
        )).fetchall()
        mapping = {r[0]: (r[1], r[2]) for r in mapping_db}

        for ridx, r in enumerate(rows_db, start=2):
            cc, ci = mapping.get(r[8] or "", (None, None))
            neto = float(r[6] or 0); iva = float(r[7] or 0)
            # R152ll — si el voucher ya está EXECUTED/SYNCED/RECONCILED y tiene
            # fecha_ejecucion, prellenamos Forma de Pago + Fecha de Pago.
            fecha_ejecucion = r[9] if len(r) > 9 else None
            status_v = r[10] if len(r) > 10 else None
            ejecutado = status_v in ("EXECUTED", "SYNCED", "RECONCILED")
            forma_pago_val = "Transferencia electrónica" if ejecutado else ""
            fecha_pago_val = fecha_ejecucion if (ejecutado and fecha_ejecucion) else ""
            row_data = [
                cc, ci, "CORFO", periodo_corfo, "ETAPA 1",
                "FACTURA", r[4] or "", r[2] or "", r[3] or "",
                neto, iva, neto + iva, neto + iva,  # Monto a Rendir default = total
                r[1], neto + iva, forma_pago_val, fecha_pago_val,
                r[1], r[5], "", "",
            ]
            for cidx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ridx, column=cidx, value=val)
                cell.border = border_all
                cell.font = Font(name="Arial", size=10)
                # Marcar en amarillo lo que SE TIENE QUE COMPLETAR manualmente
                if cidx in (1, 2) and not val:  # Cuenta + Ítem sin mapeo
                    cell.fill = YELLOW
                # R152ll — solo marcamos Forma/Fecha Pago en amarillo si NO se llenaron
                if cidx == 16 and not forma_pago_val:
                    cell.fill = YELLOW
                if cidx == 17 and not fecha_pago_val:
                    cell.fill = YELLOW
                # Formato numérico
                if cidx in (10, 11, 12, 13, 15) and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
    else:
        # RRHH — desde trabajadores + nubox_remuneraciones (si hay datos)
        rows_db = (await db.execute(text("""
            SELECT t.rut, t.nombre_completo, t.cargo,
                   COALESCE(nr.sueldo_liquido, 0) AS liquido,
                   COALESCE(nr.sueldo_bruto, 0) AS bruto
            FROM core.trabajadores t
            LEFT JOIN core.nubox_remuneraciones nr
              ON nr.rut = t.rut AND nr.empresa_codigo = t.empresa_codigo
              AND EXTRACT(YEAR FROM nr.periodo) = :y
              AND EXTRACT(MONTH FROM nr.periodo) = :m
            WHERE t.empresa_codigo = :emp AND t.activo = TRUE
            ORDER BY t.nombre_completo
        """), {"emp": body.empresa, "y": int(y), "m": int(m)})).fetchall()

        for ridx, r in enumerate(rows_db, start=2):
            bruto = float(r[4] or 0)
            row_data = [
                None, None, "CORFO", periodo_corfo, "ETAPA 1",
                "LIQ. SUELDO", "", r[0], r[1],
                bruto, bruto, None, None,
                "", "", "", f"Remuneración {r[2] or 'trabajador'} {periodo_corfo}",
            ]
            for cidx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ridx, column=cidx, value=val)
                cell.border = border_all
                cell.font = Font(name="Arial", size=10)
                if cidx in (1, 2, 12, 13, 14, 15):
                    cell.fill = YELLOW
                if cidx in (10, 11, 12) and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

    # Anchos columna
    widths_gastos = [22, 28, 14, 14, 10, 14, 12, 14, 32, 12, 12, 12, 12, 14, 12, 14, 14, 14, 30, 14, 24]
    widths_rrhh = [22, 22, 14, 14, 10, 14, 12, 14, 32, 12, 12, 10, 10, 14, 14, 14, 30]
    widths = widths_gastos if body.tipo == "gastos" else widths_rrhh
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Hoja Listados (igual al original CORFO)
    ls = wb.create_sheet("Listados")
    if body.tipo == "gastos":
        listas = [
            ("A", "Cuentas (Gastos)", cats.get("cuenta_gastos", [])),
            ("B", "Ítems", cats.get("item_gastos", [])),
            ("C", "Fuente", ["CORFO", "CHUCAO TECHNOLOGY CONSULTANTS"]),
            ("D", "Periodo", [_periodo_to_corfo(f"{int(y)}-{mm:02d}") for mm in range(1, 13)]),
            ("E", "Etapa", cats.get("etapa", [])),
            ("F", "Tipo Doc", cats.get("tipo_doc_gastos", [])),
        ]
    else:
        listas = [
            ("A", "Cuentas (RRHH)", cats.get("cuenta_rrhh", [])),
            ("B", "Ítems", ["DIRECTOR 1", "PROFESIONAL 1", "TÉCNICO 1"]),
            ("C", "Fuente", ["CORFO", "CHUCAO TECHNOLOGY CONSULTANTS"]),
            ("D", "Periodo", [_periodo_to_corfo(f"{int(y)}-{mm:02d}") for mm in range(1, 13)]),
            ("E", "Etapa", cats.get("etapa", [])),
            ("F", "Tipo Doc", cats.get("tipo_doc_rrhh", [])),
        ]
    for col_letter, title, values in listas:
        col_idx = ord(col_letter) - 64
        ls.cell(row=1, column=col_idx, value=title).font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            ls.cell(row=i, column=col_idx, value=v)

    # Data validations sobre Carga_*
    def add_dv(col_letter: str, formula: str) -> None:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col_letter}2:{col_letter}1000")
        ws.add_data_validation(dv)

    if body.tipo == "gastos":
        add_dv("A", f"=Listados!$A$2:$A${1 + len(cats.get('cuenta_gastos', []))}")
        add_dv("B", f"=Listados!$B$2:$B${1 + len(cats.get('item_gastos', []))}")
        add_dv("C", "=Listados!$C$2:$C$3")
        add_dv("D", "=Listados!$D$2:$D$13")
        add_dv("E", f"=Listados!$E$2:$E${1 + len(cats.get('etapa', []))}")
        add_dv("F", f"=Listados!$F$2:$F${1 + len(cats.get('tipo_doc_gastos', []))}")
    else:
        add_dv("A", f"=Listados!$A$2:$A${1 + len(cats.get('cuenta_rrhh', []))}")
        add_dv("D", "=Listados!$D$2:$D$13")
        add_dv("E", f"=Listados!$E$2:$E${1 + len(cats.get('etapa', []))}")
        add_dv("F", f"=Listados!$F$2:$F${1 + len(cats.get('tipo_doc_rrhh', []))}")

    # Freeze headers
    ws.freeze_panes = "A2"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Rendicion_{body.tipo.title()}_{body.empresa}_{body.periodo}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
