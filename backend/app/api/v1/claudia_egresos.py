"""Registro de egresos CORFO — la sección de Claudia.

Contrato: `docs/MEGAPROMPT_REGISTRO_EGRESOS_CLAUDIA.md` §3.3 / §3.6.

Es el libro operativo de rendición de REVTECH y TRONGKAI (lo que Claudia
lleva hoy en `Registro de Egresos` de su Excel), con el reparto por fuente
(Subsidio / Cehta-Ptec / Cehta / Trewaox) y las 21 columnas oficiales de
`Carga_Gastos`. No son vouchers: un voucher es un asiento con dos firmas;
esto es lo que se rinde a CORFO.

Acceso (§3.6): el grupo ClaudIA del sidebar hoy es la única barrera. Acá el
gate vive en el backend: admin, o email de Claudia / dominios del equipo
CORFO, o rol activo en REVTECH/TRONGKAI. Encima de eso, el scope
multi-tenant (`assert_empresa_access`) y `empresa ∈ CORFO_EMPRESAS`.

Sin ORM a propósito: las tablas son nuevas y el trigger de la BD arma el
historial, deriva `periodo` desde `fecha` y normaliza el RUT. Todo `text()`
con binds; los casts van como `CAST(:x AS tipo)` — `:x::tipo` no lo entiende
SQLAlchemy y explota en runtime (lección de exports.py).
"""
from __future__ import annotations

import io
import json
from collections.abc import Mapping
from datetime import date
from decimal import Decimal
from typing import Annotated, Any, Literal

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from pydantic import ValidationError
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import CurrentUser, DBSession
from app.api.v1.corfo_rendiciones import CORFO_EMPRESAS, _periodo_to_corfo
from app.api.v1.exports import _XML_ILEGAL
from app.core.security import AuthenticatedUser
from app.domain.value_objects.reparto_corfo import FUENTES
from app.schemas.claudia_egresos import (
    ESTADOS_PAGO_LABELS,
    FUENTES_CATALOGO,
    TIPOS_DOCUMENTO_LABELS,
    CambioHistorial,
    CatalogoItem,
    ClaudiaCatalogosResponse,
    CorfoCatalogos,
    CorfoIn,
    EgresoBase,
    EgresoBatchFila,
    EgresoBatchRequest,
    EgresoBatchResponse,
    EgresoCreate,
    EgresoDeleteRequest,
    EgresoDeleteResponse,
    EgresoDetail,
    EgresoListResponse,
    EgresoRead,
    EgresoUpdate,
    EstadoResumen,
    FilaSaltada,
    HistorialItem,
    ImportarResponse,
    PeriodoItem,
    PeriodosResponse,
    PorEstadoResumen,
    PorFuenteResumen,
    ResumenResponse,
    Sugerencias,
    TipoDocumentoResumen,
    a_decimal,
    egreso_read_desde_fila,
    fmt_monto,
    fusionar_update,
)
from app.services.corfo_egresos_import_service import cargar_filas, parsear_registro_egresos
from app.services.empresa_scope_service import assert_empresa_access, get_allowed_empresa_codes

router = APIRouter(prefix="/claudia/egresos", tags=["claudia-egresos"])

# ---------------------------------------------------------------------------
# §3.6 Acceso
# ---------------------------------------------------------------------------

CLAUDIA_EMAILS = {"claudia@trongkai.com"}
CLAUDIA_DOMAINS = ("@trongkai.com", "@revtech.cl", "@revtech.com")
_MSG_403 = "Sección reservada a la coordinación CORFO (Claudia) y admins."

_SQL_ROL_CORFO = """
    SELECT 1
      FROM core.user_company_roles
     WHERE user_id = CAST(:uid AS UUID)
       AND active = TRUE
       AND empresa_codigo = ANY(CAST(:emps AS TEXT[]))
     LIMIT 1
"""


async def _check_claudia_access(user: AuthenticatedUser, db: AsyncSession) -> None:
    """Espejo backend de `canSeeClaudiaGroup` del sidebar (§3.6)."""
    if user.is_admin:
        return
    email = (user.email or "").strip().lower()
    if email and (email in CLAUDIA_EMAILS or email.endswith(CLAUDIA_DOMAINS)):
        return
    tiene_rol = (
        await db.execute(
            text(_SQL_ROL_CORFO), {"uid": str(user.sub), "emps": sorted(CORFO_EMPRESAS)}
        )
    ).scalar()
    if tiene_rol:
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=_MSG_403)


def _require_corfo_empresa(empresa: str) -> str:
    codigo = (empresa or "").strip().upper()
    if codigo not in CORFO_EMPRESAS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "El registro de egresos CORFO sólo existe para REVTECH o TRONGKAI. "
                f"Recibido: {empresa!r}"
            ),
        )
    return codigo


async def _gate(user: AuthenticatedUser, db: AsyncSession, empresa: str) -> str:
    """Los tres candados de toda ruta, en orden: grupo ClaudIA → empresa CORFO → scope."""
    await _check_claudia_access(user, db)
    codigo = _require_corfo_empresa(empresa)
    await assert_empresa_access(user, db, codigo)
    return codigo


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------

_LIMITE_LISTA = 2000
_MAX_UPLOAD_BYTES = 15 * 1024 * 1024
#: El upload se lee de a 1 MiB y se corta apenas pasa el tope: un .xlsx de
#: 200 MB no llega a materializarse entero en RAM (S2).
_CHUNK_UPLOAD = 1024 * 1024

#: Reparto en SQL sólo para FILTRAR; el estado que se devuelve lo calcula el motor.
_SQL_REPARTO_ESTADO = """
    CASE
        WHEN e.monto_subsidio IS NULL THEN 'SIN_CLASIFICAR'
        WHEN e.monto_subsidio + e.monto_cehta_ptec + e.monto_cehta + e.monto_trewaox = e.total
            THEN 'OK'
        ELSE 'DESCUADRADO'
    END
"""

# Los f-strings de abajo sólo pegan fragmentos CONSTANTES de este módulo;
# todo valor del usuario viaja como bind. S608 no distingue eso.
_SQL_SELECT_EGRESO = f"""
    SELECT e.*,
           {_SQL_REPARTO_ESTADO} AS reparto_estado,
           (SELECT COALESCE(MAX(h.version), 1)
              FROM core.corfo_registro_egresos_hist h
             WHERE h.egreso_id = e.egreso_id) AS version
      FROM core.corfo_registro_egresos e
"""  # noqa: S608

_SQL_LISTA = f"""
    WITH base AS (
        {_SQL_SELECT_EGRESO}
        WHERE e.deleted_at IS NULL
          AND e.empresa_codigo = :empresa
          AND (CAST(:periodo AS TEXT) IS NULL OR e.periodo = :periodo)
          AND (CAST(:estado_pago AS TEXT) IS NULL OR e.estado_pago = :estado_pago)
          AND (CAST(:qlike AS TEXT) IS NULL
               OR e.descripcion ILIKE :qlike
               OR COALESCE(e.rut_emisor, '') ILIKE :qlike
               OR COALESCE(e.folio, '') ILIKE :qlike)
    )
    SELECT * FROM base
     WHERE (CAST(:reparto_estado AS TEXT) IS NULL OR reparto_estado = :reparto_estado)
     ORDER BY fecha DESC, egreso_id DESC
     LIMIT :lim
"""  # noqa: S608

_SQL_POR_ID = f"{_SQL_SELECT_EGRESO} WHERE e.egreso_id = :id"

_SQL_POR_IDS = f"""
    {_SQL_SELECT_EGRESO}
    WHERE e.egreso_id = ANY(CAST(:ids AS BIGINT[]))
    ORDER BY e.egreso_id
"""

_SQL_PERIODOS = """
    SELECT periodo,
           COUNT(*) AS n,
           COALESCE(SUM(total), 0) AS total,
           COALESCE(SUM(total) FILTER (WHERE estado_pago = 'PENDIENTE'), 0) AS pendiente,
           COUNT(*) FILTER (WHERE monto_subsidio IS NULL) AS sin_clasificar,
           COUNT(*) FILTER (
               WHERE monto_subsidio IS NOT NULL
                 AND monto_subsidio + monto_cehta_ptec + monto_cehta + monto_trewaox <> total
           ) AS descuadrados
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
     GROUP BY periodo
     ORDER BY periodo DESC
"""

_SQL_RESUMEN_TOTALES = """
    SELECT COUNT(*) AS n,
           COALESCE(SUM(total), 0) AS total,
           COALESCE(SUM(monto_subsidio), 0) AS subsidio,
           COALESCE(SUM(monto_cehta_ptec), 0) AS cehta_ptec,
           COALESCE(SUM(monto_cehta), 0) AS cehta,
           COALESCE(SUM(monto_trewaox), 0) AS trewaox,
           COALESCE(SUM(total) FILTER (WHERE monto_subsidio IS NULL), 0) AS sin_clasificar_monto,
           COUNT(*) FILTER (WHERE monto_subsidio IS NULL) AS sin_clasificar,
           COUNT(*) FILTER (
               WHERE monto_subsidio IS NOT NULL
                 AND monto_subsidio + monto_cehta_ptec + monto_cehta + monto_trewaox <> total
           ) AS descuadrados
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND (CAST(:periodo AS TEXT) IS NULL OR periodo = :periodo)
"""

_SQL_RESUMEN_ESTADOS = """
    SELECT estado_pago, COUNT(*) AS n, COALESCE(SUM(total), 0) AS monto
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND (CAST(:periodo AS TEXT) IS NULL OR periodo = :periodo)
     GROUP BY estado_pago
"""

_SQL_RESUMEN_TIPOS = """
    SELECT tipo_documento, COUNT(*) AS n, COALESCE(SUM(total), 0) AS monto
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND (CAST(:periodo AS TEXT) IS NULL OR periodo = :periodo)
     GROUP BY tipo_documento
     ORDER BY monto DESC, tipo_documento
"""

_SQL_CATALOGOS_CORFO = """
    SELECT catalogo, valor
      FROM core.corfo_catalogos
     WHERE active = TRUE
     ORDER BY catalogo, orden
"""

_SQL_SUGERENCIAS = """
    SELECT 'tipo_egreso' AS campo, tipo_egreso AS valor
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND tipo_egreso IS NOT NULL AND btrim(tipo_egreso) <> ''
    UNION
    SELECT 'fuente', fuente
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND fuente IS NOT NULL AND btrim(fuente) <> ''
    UNION
    SELECT 'proyecto', proyecto
      FROM core.corfo_registro_egresos
     WHERE empresa_codigo = :empresa AND deleted_at IS NULL
       AND proyecto IS NOT NULL AND btrim(proyecto) <> ''
    ORDER BY 1, 2
"""

_SQL_INSERT = """
    INSERT INTO core.corfo_registro_egresos (
        empresa_codigo, periodo, fecha, descripcion, rut_emisor, tipo_documento, folio,
        monto_neto, impuesto, total, tipo_egreso, fuente, proyecto, estado_pago, fecha_pago,
        monto_subsidio, monto_cehta_ptec, monto_cehta, monto_trewaox,
        corfo_cuenta, corfo_item, corfo_fuente_financiamiento, corfo_etapa,
        corfo_fecha_recepcion, corfo_monto_rendir, corfo_monto_cancelado, corfo_forma_pago,
        corfo_glosa, corfo_receptor_rut, corfo_receptor_nombre,
        observaciones, adjunto_dropbox_path, origen, created_by
    ) VALUES (
        :empresa_codigo, :periodo, :fecha, :descripcion, :rut_emisor, :tipo_documento, :folio,
        :monto_neto, :impuesto, :total, :tipo_egreso, :fuente, :proyecto, :estado_pago,
        :fecha_pago,
        :monto_subsidio, :monto_cehta_ptec, :monto_cehta, :monto_trewaox,
        :corfo_cuenta, :corfo_item, :corfo_fuente_financiamiento, :corfo_etapa,
        :corfo_fecha_recepcion, :corfo_monto_rendir, :corfo_monto_cancelado, :corfo_forma_pago,
        :corfo_glosa, :corfo_receptor_rut, :corfo_receptor_nombre,
        :observaciones, :adjunto_dropbox_path, :origen, :created_by
    )
    RETURNING egreso_id
"""

_SQL_UPDATE = """
    UPDATE core.corfo_registro_egresos SET
        periodo = :periodo, fecha = :fecha, descripcion = :descripcion,
        rut_emisor = :rut_emisor, tipo_documento = :tipo_documento, folio = :folio,
        monto_neto = :monto_neto, impuesto = :impuesto, total = :total,
        tipo_egreso = :tipo_egreso, fuente = :fuente, proyecto = :proyecto,
        estado_pago = :estado_pago, fecha_pago = :fecha_pago,
        monto_subsidio = :monto_subsidio, monto_cehta_ptec = :monto_cehta_ptec,
        monto_cehta = :monto_cehta, monto_trewaox = :monto_trewaox,
        corfo_cuenta = :corfo_cuenta, corfo_item = :corfo_item,
        corfo_fuente_financiamiento = :corfo_fuente_financiamiento, corfo_etapa = :corfo_etapa,
        corfo_fecha_recepcion = :corfo_fecha_recepcion,
        corfo_monto_rendir = :corfo_monto_rendir, corfo_monto_cancelado = :corfo_monto_cancelado,
        corfo_forma_pago = :corfo_forma_pago, corfo_glosa = :corfo_glosa,
        corfo_receptor_rut = :corfo_receptor_rut, corfo_receptor_nombre = :corfo_receptor_nombre,
        observaciones = :observaciones, adjunto_dropbox_path = :adjunto_dropbox_path,
        updated_by = :updated_by
    WHERE egreso_id = :id AND deleted_at IS NULL
    RETURNING egreso_id
"""

_SQL_SOFT_DELETE = """
    UPDATE core.corfo_registro_egresos
       SET deleted_at = now(), deleted_by = :who, delete_motivo = :motivo, updated_by = :who
     WHERE egreso_id = :id AND deleted_at IS NULL
    RETURNING deleted_at
"""

_SQL_HISTORIAL = """
    SELECT version, accion, changed_by, changed_at, CAST(snapshot AS TEXT) AS snapshot
      FROM core.corfo_registro_egresos_hist
     WHERE egreso_id = :id
     ORDER BY version
"""

_SQL_EXPORT = """
    SELECT e.*
      FROM core.corfo_registro_egresos e
     WHERE e.empresa_codigo = :empresa AND e.deleted_at IS NULL
       AND (CAST(:periodo AS TEXT) IS NULL OR e.periodo = :periodo)
     ORDER BY e.fecha, e.egreso_id
"""

#: Para el test que compila todo el SQL con el dialecto asyncpg real.
SQL_TODOS: dict[str, str] = {
    "rol_corfo": _SQL_ROL_CORFO,
    "lista": _SQL_LISTA,
    "por_id": _SQL_POR_ID,
    "por_ids": _SQL_POR_IDS,
    "periodos": _SQL_PERIODOS,
    "resumen_totales": _SQL_RESUMEN_TOTALES,
    "resumen_estados": _SQL_RESUMEN_ESTADOS,
    "resumen_tipos": _SQL_RESUMEN_TIPOS,
    "catalogos_corfo": _SQL_CATALOGOS_CORFO,
    "sugerencias": _SQL_SUGERENCIAS,
    "insert": _SQL_INSERT,
    "update": _SQL_UPDATE,
    "soft_delete": _SQL_SOFT_DELETE,
    "historial": _SQL_HISTORIAL,
    "export": _SQL_EXPORT,
}


# ---------------------------------------------------------------------------
# Helpers de persistencia
# ---------------------------------------------------------------------------


def _params_insert(empresa: str, fila: EgresoBase, origen: str, quien: str) -> dict[str, Any]:
    montos = fila.montos_reparto()
    corfo = fila.corfo or CorfoIn()
    return {
        "empresa_codigo": empresa,
        # El trigger BEFORE lo recalcula desde `fecha`; se manda igual para
        # que el INSERT sea legible y no dependa del orden de los triggers.
        "periodo": fila.fecha.strftime("%Y-%m"),
        "fecha": fila.fecha,
        "descripcion": fila.descripcion,
        "rut_emisor": fila.rut_emisor,
        "tipo_documento": fila.tipo_documento,
        "folio": fila.folio,
        "monto_neto": fila.monto_neto,
        "impuesto": fila.impuesto,
        "total": fila.total,
        "tipo_egreso": fila.tipo_egreso,
        "fuente": fila.fuente,
        "proyecto": fila.proyecto,
        "estado_pago": fila.estado_pago,
        "fecha_pago": fila.fecha_pago,
        **{f"monto_{f}": montos[f] for f in FUENTES},
        **{f"corfo_{c}": getattr(corfo, c) for c in CorfoIn.model_fields},
        "observaciones": fila.observaciones,
        "adjunto_dropbox_path": fila.adjunto_dropbox_path,
        "origen": origen,
        "created_by": quien,
    }


async def _insertar(
    db: AsyncSession, empresa: str, fila: EgresoBase, origen: str, quien: str
) -> int:
    params = _params_insert(empresa, fila, origen, quien)
    nuevo_id = (await db.execute(text(_SQL_INSERT), params)).scalar()
    if nuevo_id is None:  # pragma: no cover — RETURNING siempre trae el id
        raise HTTPException(status_code=500, detail="No se pudo crear el gasto")
    return int(nuevo_id)


async def _leer_fila(db: AsyncSession, egreso_id: int) -> Mapping[str, Any] | None:
    return (await db.execute(text(_SQL_POR_ID), {"id": egreso_id})).mappings().first()


def _404_egreso(egreso_id: int) -> HTTPException:
    """Un solo 404 para "no existe", "fue borrado" y "es de otra empresa" (S3):
    si el mensaje o el código cambiaran según el caso, se podrían enumerar ids."""
    return HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"No existe el gasto #{egreso_id} (o fue borrado)",
    )


async def _leer_o_404(db: AsyncSession, egreso_id: int) -> Mapping[str, Any]:
    fila = await _leer_fila(db, egreso_id)
    if fila is None or fila.get("deleted_at") is not None:
        raise _404_egreso(egreso_id)
    return fila


async def _fila_propia_o_404(
    user: AuthenticatedUser, db: AsyncSession, egreso_id: int
) -> Mapping[str, Any]:
    """Los candados de las rutas por id, en orden: grupo ClaudIA (403 antes de
    leer nada) → la fila existe y no está borrada → la empresa de la FILA está
    en el scope del usuario.

    Una fila que existe pero es de una empresa a la que el usuario no accede
    responde el MISMO 404 que una inexistente, y sin pasar por
    `assert_empresa_access` (que loguea y persiste una violación de scope):
    con un 403 distinto, alguien de REVTECH podría enumerar los ids de
    TRONGKAI probando números. El 403 con detalle queda sólo para el ajeno al
    grupo, que se corta antes de tocar la fila.
    """
    await _check_claudia_access(user, db)
    fila = await _leer_fila(db, egreso_id)
    if fila is None or fila.get("deleted_at") is not None:
        raise _404_egreso(egreso_id)
    permitidas = await get_allowed_empresa_codes(user, db)  # None = admin
    if permitidas is not None and fila["empresa_codigo"] not in permitidas:
        raise _404_egreso(egreso_id)
    _require_corfo_empresa(fila["empresa_codigo"])
    return fila


def _like(q: str | None) -> str | None:
    if q is None or not q.strip():
        return None
    limpio = q.strip().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"%{limpio}%"


#: Los errores que pydantic escribe en inglés y Claudia ve en la grilla.
#: Los `ValueError` propios ya vienen en español; esto cubre los genéricos.
_PYDANTIC_ES: dict[str, str] = {
    "string_too_long": "máximo {max_length} caracteres",
    "string_too_short": "mínimo {min_length} caracteres",
    "missing": "falta este campo",
    "literal_error": "valor no válido (esperado: {expected})",
    "date_from_datetime_parsing": "fecha inválida (usá AAAA-MM-DD)",
    "date_parsing": "fecha inválida (usá AAAA-MM-DD)",
    "date_type": "fecha inválida (usá AAAA-MM-DD)",
    "decimal_parsing": "monto inválido",
    "decimal_type": "monto inválido",
    "int_parsing": "número inválido",
    "extra_forbidden": "campo desconocido",
}


def _mensaje_validacion(exc: ValidationError) -> str:
    """El primer error de pydantic, en español y con el campo adelante.

    'String should have at most 500 characters' → 'descripcion: máximo 500
    caracteres'. Lo que no está en la tabla sale tal cual, sin el prefijo
    'Value error, ' que pydantic le pone a los ValueError propios.
    """
    err = exc.errors()[0]
    plantilla = _PYDANTIC_ES.get(str(err.get("type", "")))
    ctx: dict[str, Any] = dict(err.get("ctx") or {})
    if plantilla is not None:
        try:
            msg = plantilla.format(**ctx)
        except (KeyError, IndexError):  # ctx sin la clave que espera la plantilla
            msg = str(err.get("msg", ""))
    else:
        msg = str(err.get("msg", "")).removeprefix("Value error, ")
    loc = ".".join(str(p) for p in err.get("loc", ()) if p != "__root__")
    return f"{loc}: {msg}" if loc else msg


# ---------------------------------------------------------------------------
# Historial: diff campo a campo entre snapshots consecutivos
# ---------------------------------------------------------------------------

_HIST_IGNORAR = {"updated_at", "updated_by", "created_at"}
_HIST_MONTOS = {
    "monto_neto", "impuesto", "total", "corfo_monto_rendir", "corfo_monto_cancelado",
    *(f"monto_{f}" for f in FUENTES),
}


def _valor_hist(campo: str, v: Any) -> str | None:
    if v is None:
        return None
    if campo in _HIST_MONTOS:
        return fmt_monto(v)
    if isinstance(v, bool):
        return "sí" if v else "no"
    return str(v)


def diff_snapshots(
    antes: Mapping[str, Any] | None, despues: Mapping[str, Any]
) -> list[CambioHistorial]:
    """Qué cambió entre dos versiones. v1 (INSERT) no tiene 'antes' → sin cambios."""
    if antes is None:
        return []
    cambios: list[CambioHistorial] = []
    for campo in sorted(set(antes) | set(despues)):
        if campo in _HIST_IGNORAR:
            continue
        a, d = _valor_hist(campo, antes.get(campo)), _valor_hist(campo, despues.get(campo))
        if a != d:
            cambios.append(CambioHistorial(campo=campo, antes=a, despues=d))
    return cambios


def _parsear_snapshot(raw: Any) -> dict[str, Any]:
    if isinstance(raw, Mapping):
        return dict(raw)
    # Se pide CAST(snapshot AS TEXT) para parsear con Decimal: un NUMERIC
    # que pasa por float pierde centavos en montos grandes.
    return json.loads(raw, parse_float=Decimal) if raw else {}


def armar_historial(filas: list[Mapping[str, Any]]) -> list[HistorialItem]:
    items: list[HistorialItem] = []
    anterior: dict[str, Any] | None = None
    for h in filas:
        snap = _parsear_snapshot(h.get("snapshot"))
        items.append(
            HistorialItem(
                version=int(h["version"]),
                accion=h["accion"],
                changed_at=h.get("changed_at"),
                changed_by=h.get("changed_by"),
                cambios=diff_snapshots(anterior, snap),
            )
        )
        anterior = snap
    return items


# ---------------------------------------------------------------------------
# Export XLSX (puro: el test le pasa filas falsas)
# ---------------------------------------------------------------------------

#: Las 17 columnas del Excel de Claudia (variante TRONGKAI, que incluye Trewaox).
REGISTRO_HEADERS: list[str] = [
    "Fecha", "Descripción", "RUT Emisor", "Tipo de Documento", "Folio",
    "Monto Neto/Pagado", "Impuesto/Patronal", "Total", "Tipo Financiamiento",
    "Tipo de Egreso", "Proyecto", "Trewaox", "Subsidio", "Cehta-Ptec", "Cehta",
    "Estado", "Fecha de Pago",
]

#: COPIA LITERAL de `corfo_rendiciones.generar_excel` (hoja Carga_Gastos).
#: `test_claudia_egresos_export.py` la compara contra el fuente de ese módulo.
CARGA_GASTOS_HEADERS: list[str] = [
    "Cuenta", "Ítem", "Fuente Financiamiento", "Periodo", "Etapa",
    "Tipo Documento", "N° Documento", "Rut Proveedor",
    "Nombre Proveedor o Razón Social", "Monto Neto", "Monto IVA",
    "Monto Total", "Monto Rendir", "Fecha de Recepción",
    "Monto Cancelado", "Forma de Pago", "Fecha de Pago",
    "Fecha del documento", "Glosa / Justificación", "Receptor Rut",
    "Nombre Receptor",
]

#: tipo_documento nuestro → vocabulario `tipo_doc_gastos` de CORFO (§3.3).
MAPEO_TIPO_DOC_CORFO: dict[str, str] = {
    "FACTURA": "FACTURA",
    "FACTURA_EXENTA": "FACTURA",
    "BOLETA": "BOLETA",
    "BOLETA_HONORARIO": "BOLETA HONORARIOS",
    "LIQUIDACION": "LIQ. SUELDO",
    "INVOICE": "INVOICE",
    "CO_EJECUTOR": "OTRO",
    "OTRO": "OTRO",
}


def _celda(v: Any) -> Any:
    """Lo que openpyxl traga: strings sin bytes de control, Decimal como float."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, str):
        return _XML_ILEGAL.sub(" ", v)
    return v


def _monto_o_none(v: Any) -> Decimal | None:
    return a_decimal(v)


def construir_export_xlsx(
    empresa: str, periodo: str | None, filas: list[Mapping[str, Any]]
) -> bytes:
    """Dos hojas: 'Registro de Egresos' (Claudia) y 'Carga_Gastos' (CORFO)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(bold=True, color="111827")

    def _hoja(ws: Any, headers: list[str], rows: list[list[Any]], montos_idx: set[int]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        anchos = [len(h) for h in headers]
        for r in rows:
            vals = [_celda(v) for v in r]
            ws.append(vals)
            for i, v in enumerate(vals):
                if v not in (None, ""):
                    anchos[i] = max(anchos[i], len(str(v)))
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(ancho + 2, 60)
        for row in ws.iter_rows(min_row=2):
            for idx in montos_idx:
                row[idx].number_format = "#,##0.00"
        ws.freeze_panes = "A2"

    registro: list[list[Any]] = []
    carga: list[list[Any]] = []
    for f in filas:
        total = _monto_o_none(f.get("total")) or Decimal("0.00")
        subsidio = _monto_o_none(f.get("monto_subsidio"))
        tipo = f.get("tipo_documento") or "OTRO"
        estado = f.get("estado_pago") or "PENDIENTE"
        registro.append([
            f.get("fecha"), f.get("descripcion"), f.get("rut_emisor"),
            TIPOS_DOCUMENTO_LABELS.get(tipo, tipo), f.get("folio"),
            _monto_o_none(f.get("monto_neto")), _monto_o_none(f.get("impuesto")), total,
            f.get("fuente"), f.get("tipo_egreso"), f.get("proyecto"),
            _monto_o_none(f.get("monto_trewaox")), subsidio,
            _monto_o_none(f.get("monto_cehta_ptec")), _monto_o_none(f.get("monto_cehta")),
            ESTADOS_PAGO_LABELS.get(estado, estado), f.get("fecha_pago"),
        ])
        # Defaults honestos (§2.9): Monto Rendir = subsidio si no se cargó otro;
        # Monto Cancelado = total sólo si está PAGADO. Nada más se inventa.
        monto_rendir = _monto_o_none(f.get("corfo_monto_rendir"))
        if monto_rendir is None:
            monto_rendir = subsidio
        monto_cancelado = _monto_o_none(f.get("corfo_monto_cancelado"))
        if monto_cancelado is None and estado == "PAGADO":
            monto_cancelado = total
        per = f.get("periodo")
        if not per and isinstance(f.get("fecha"), date):
            per = f["fecha"].strftime("%Y-%m")
        carga.append([
            f.get("corfo_cuenta"), f.get("corfo_item"), f.get("corfo_fuente_financiamiento"),
            _periodo_to_corfo(per) if per else None, f.get("corfo_etapa"),
            MAPEO_TIPO_DOC_CORFO.get(tipo, "OTRO"), f.get("folio"), f.get("rut_emisor"),
            f.get("descripcion"), _monto_o_none(f.get("monto_neto")),
            _monto_o_none(f.get("impuesto")), total, monto_rendir,
            f.get("corfo_fecha_recepcion"), monto_cancelado, f.get("corfo_forma_pago"),
            f.get("fecha_pago"), f.get("fecha"), f.get("corfo_glosa"),
            f.get("corfo_receptor_rut"), f.get("corfo_receptor_nombre"),
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Registro de Egresos"
    _hoja(ws1, REGISTRO_HEADERS, registro, {5, 6, 7, 11, 12, 13, 14})
    ws2 = wb.create_sheet("Carga_Gastos")
    _hoja(ws2, CARGA_GASTOS_HEADERS, carga, {9, 10, 11, 12, 14})
    for ws in (ws1, ws2):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, date):
                    cell.number_format = "DD-MM-YYYY"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nombre_archivo_export(empresa: str, periodo: str | None) -> str:
    return f"registro_egresos_{empresa}_{periodo or 'todos'}.xlsx"


# ---------------------------------------------------------------------------
# Rutas estáticas (ANTES de /{egreso_id})
# ---------------------------------------------------------------------------

_PERIODO_RE = r"^\d{4}-(0[1-9]|1[0-2])$"
RepartoEstadoQ = Literal["SIN_CLASIFICAR", "OK", "DESCUADRADO"]
EstadoPagoQ = Literal["PAGADO", "PARCIAL", "PENDIENTE"]


@router.get("", response_model=EgresoListResponse)
async def listar_egresos(
    user: CurrentUser,
    db: DBSession,
    empresa: Annotated[str, Query(description="REVTECH o TRONGKAI")],
    periodo: Annotated[str | None, Query(pattern=_PERIODO_RE)] = None,
    q: Annotated[str | None, Query(max_length=120)] = None,
    estado_pago: EstadoPagoQ | None = None,
    reparto_estado: RepartoEstadoQ | None = None,
) -> EgresoListResponse:
    """Lista del mes (o de todo) para la grilla. Tope 2000, `truncado` avisa."""
    codigo = await _gate(user, db, empresa)
    filas = (
        await db.execute(
            text(_SQL_LISTA),
            {
                "empresa": codigo,
                "periodo": periodo,
                "estado_pago": estado_pago,
                "qlike": _like(q),
                "reparto_estado": reparto_estado,
                "lim": _LIMITE_LISTA + 1,
            },
        )
    ).mappings().all()
    truncado = len(filas) > _LIMITE_LISTA
    items = [egreso_read_desde_fila(f) for f in filas[:_LIMITE_LISTA]]
    return EgresoListResponse(
        empresa_codigo=codigo, periodo=periodo, items=items, n=len(items), truncado=truncado
    )


@router.get("/periodos", response_model=PeriodosResponse)
async def listar_periodos(
    user: CurrentUser,
    db: DBSession,
    empresa: Annotated[str, Query()],
) -> PeriodosResponse:
    """Los chips de meses: cuántos gastos, cuánta plata y si hay algo que resolver."""
    codigo = await _gate(user, db, empresa)
    filas = (await db.execute(text(_SQL_PERIODOS), {"empresa": codigo})).mappings().all()
    items = [
        PeriodoItem(
            periodo=f["periodo"],
            n=int(f["n"]),
            total=fmt_monto(f["total"]) or "0.00",
            pendiente=fmt_monto(f["pendiente"]) or "0.00",
            sin_clasificar=int(f["sin_clasificar"]),
            descuadrados=int(f["descuadrados"]),
        )
        for f in filas
    ]
    total_general = sum((a_decimal(f["total"]) or Decimal("0") for f in filas), Decimal("0"))
    return PeriodosResponse(
        items=items,
        n_total=sum(i.n for i in items),
        total_general=fmt_monto(total_general) or "0.00",
    )


@router.get("/resumen", response_model=ResumenResponse)
async def resumen_egresos(
    user: CurrentUser,
    db: DBSession,
    empresa: Annotated[str, Query()],
    periodo: Annotated[str | None, Query(pattern=_PERIODO_RE)] = None,
) -> ResumenResponse:
    """KPIs del mes: total, por fuente, por estado, % pagado, por tipo de documento."""
    codigo = await _gate(user, db, empresa)
    params = {"empresa": codigo, "periodo": periodo}
    tot = (await db.execute(text(_SQL_RESUMEN_TOTALES), params)).mappings().first() or {}
    estados = (await db.execute(text(_SQL_RESUMEN_ESTADOS), params)).mappings().all()
    tipos = (await db.execute(text(_SQL_RESUMEN_TIPOS), params)).mappings().all()

    por_estado = {e: EstadoResumen(n=0, monto="0.00") for e in ("PAGADO", "PARCIAL", "PENDIENTE")}
    for f in estados:
        if f["estado_pago"] in por_estado:
            por_estado[f["estado_pago"]] = EstadoResumen(
                n=int(f["n"]), monto=fmt_monto(f["monto"]) or "0.00"
            )
    total = a_decimal(tot.get("total")) or Decimal("0.00")
    pagado = a_decimal(por_estado["PAGADO"].monto) or Decimal("0.00")
    pct_pagado = (pagado / total * 100).quantize(Decimal("0.01")) if total else Decimal("0.00")

    return ResumenResponse(
        empresa_codigo=codigo,
        periodo=periodo,
        n=int(tot.get("n") or 0),
        total=fmt_monto(total) or "0.00",
        por_fuente=PorFuenteResumen(
            subsidio=fmt_monto(tot.get("subsidio")) or "0.00",
            cehta_ptec=fmt_monto(tot.get("cehta_ptec")) or "0.00",
            cehta=fmt_monto(tot.get("cehta")) or "0.00",
            trewaox=fmt_monto(tot.get("trewaox")) or "0.00",
            sin_clasificar=fmt_monto(tot.get("sin_clasificar_monto")) or "0.00",
        ),
        por_estado=PorEstadoResumen(**por_estado),
        pct_pagado=f"{pct_pagado:.2f}",
        por_tipo_documento=[
            TipoDocumentoResumen(
                tipo_documento=f["tipo_documento"],
                n=int(f["n"]),
                monto=fmt_monto(f["monto"]) or "0.00",
            )
            for f in tipos
        ],
        descuadrados=int(tot.get("descuadrados") or 0),
        sin_clasificar=int(tot.get("sin_clasificar") or 0),
    )


FORMAS_PAGO: list[str] = ["Transferencia electrónica", "Cheque", "Efectivo", "Tarjeta", "Otro"]
#: CORFO no publicó el vocabulario de Fuente Financiamiento en los catálogos
#: cargados (§0): estos son sugeridos, el campo queda libre.
FUENTE_FINANCIAMIENTO_SUGERIDAS: list[str] = ["SUBSIDIO", "APORTE PECUNIARIO", "APORTE VALORIZADO"]


@router.get("/catalogos", response_model=ClaudiaCatalogosResponse)
async def catalogos_egresos(
    user: CurrentUser,
    db: DBSession,
    empresa: Annotated[str, Query()],
) -> ClaudiaCatalogosResponse:
    """Dropdowns de la ficha + autocompletar con lo que esa empresa ya usó."""
    codigo = await _gate(user, db, empresa)
    cats: dict[str, list[str]] = {}
    for f in (await db.execute(text(_SQL_CATALOGOS_CORFO))).mappings().all():
        cats.setdefault(f["catalogo"], []).append(f["valor"])
    sug: dict[str, list[str]] = {"tipo_egreso": [], "fuente": [], "proyecto": []}
    for f in (await db.execute(text(_SQL_SUGERENCIAS), {"empresa": codigo})).mappings().all():
        sug.setdefault(f["campo"], []).append(f["valor"])
    return ClaudiaCatalogosResponse(
        tipos_documento=[
            CatalogoItem(codigo=k, label=v) for k, v in TIPOS_DOCUMENTO_LABELS.items()
        ],
        estados_pago=[CatalogoItem(codigo=k, label=v) for k, v in ESTADOS_PAGO_LABELS.items()],
        fuentes=list(FUENTES_CATALOGO),
        formas_pago=[CatalogoItem(codigo=fp, label=fp) for fp in FORMAS_PAGO],
        corfo=CorfoCatalogos(
            cuenta_gastos=cats.get("cuenta_gastos", []),
            item_gastos=cats.get("item_gastos", []),
            etapa=cats.get("etapa", []),
            tipo_doc_gastos=cats.get("tipo_doc_gastos", []),
            fuente_financiamiento_sugeridas=list(FUENTE_FINANCIAMIENTO_SUGERIDAS),
        ),
        sugerencias=Sugerencias(**sug),
    )


@router.post("/batch", response_model=EgresoBatchResponse, status_code=status.HTTP_201_CREATED)
async def crear_egresos_batch(
    body: EgresoBatchRequest, user: CurrentUser, db: DBSession
) -> EgresoBatchResponse:
    """Pegado desde Excel: se validan TODAS las filas y recién después se
    insertan en una sola transacción (todo o nada). `origen='PASTE'`."""
    codigo = await _gate(user, db, body.empresa_codigo)
    validas: list[EgresoBatchFila] = []
    errores: list[dict[str, Any]] = []
    for i, cruda in enumerate(body.filas, start=1):
        try:
            validas.append(EgresoBatchFila.model_validate(cruda))
        except ValidationError as exc:
            errores.append({"fila": i, "error": _mensaje_validacion(exc)})
    if errores:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=errores)

    quien = user.email or str(user.sub)
    ids = [await _insertar(db, codigo, fila, "PASTE", quien) for fila in validas]
    await db.commit()
    filas = (await db.execute(text(_SQL_POR_IDS), {"ids": ids})).mappings().all()
    creados = [egreso_read_desde_fila(f) for f in filas]
    return EgresoBatchResponse(creados=creados, n=len(creados))


def _413_upload() -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
        detail=f"El archivo supera el máximo de {_MAX_UPLOAD_BYTES // (1024 * 1024)} MB",
    )


async def _leer_upload_acotado(archivo: UploadFile) -> bytes:
    """Lee el upload sin materializar nada por encima de `_MAX_UPLOAD_BYTES` (S2).

    Primero mira el tamaño declarado (si el cliente lo mandó) y corta ahí
    mismo sin leer un byte; después lee de a `_CHUNK_UPLOAD` y corta apenas
    lo acumulado pasa el tope. Un .xlsx de 200 MB nunca llega a estar
    entero en memoria.
    """
    declarado = getattr(archivo, "size", None)
    if declarado is not None and declarado > _MAX_UPLOAD_BYTES:
        raise _413_upload()
    partes: list[bytes] = []
    leido = 0
    while True:
        chunk = await archivo.read(_CHUNK_UPLOAD)
        if not chunk:
            break
        leido += len(chunk)
        if leido > _MAX_UPLOAD_BYTES:
            raise _413_upload()
        partes.append(chunk)
    return b"".join(partes)


@router.post("/importar", response_model=ImportarResponse)
async def importar_excel(
    user: CurrentUser,
    db: DBSession,
    archivo: Annotated[UploadFile, File(description="Excel de Claudia (.xlsx)")],
    empresa_codigo: Annotated[str, Form(min_length=2, max_length=20)],
    dry_run: Annotated[bool, Form()] = False,
) -> ImportarResponse:
    """Importa la hoja `Registro de Egresos` (idempotente por `import_natural_key`).

    Las filas idénticas del mismo Excel entran todas (son pagos distintos:
    cuotas a co-ejecutores, peajes), cada una con huella propia; el conteo
    `duplicadas_en_excel` dice cuántas entraron con observación para revisar.
    `leidas` son las filas con datos del Excel (cargables + saltadas), no
    sólo las cargables.
    """
    codigo = await _gate(user, db, empresa_codigo)
    nombre = archivo.filename or ""
    if not nombre.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Sólo se aceptan archivos .xlsx (llegó {nombre or 'sin nombre'!r})",
        )
    contenido = await _leer_upload_acotado(archivo)
    if not contenido:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="El archivo está vacío"
        )

    try:
        parseo = parsear_registro_egresos(contenido, codigo)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)
        ) from exc

    resumen = await cargar_filas(
        db, codigo, parseo.filas, user.email or str(user.sub), dry_run=dry_run
    )
    if not dry_run:
        await db.commit()

    return ImportarResponse(
        empresa_codigo=codigo,
        dry_run=dry_run,
        leidas=parseo.leidas,
        creadas=resumen.creadas,
        omitidas_existentes=resumen.omitidas_existentes,
        duplicadas_en_excel=len(parseo.repetidas_en_excel),
        saltadas=[FilaSaltada(fila_excel=s.fila_excel, motivo=s.motivo) for s in parseo.saltadas],
        descuadradas=resumen.descuadradas,
        sin_clasificar=resumen.sin_clasificar,
    )


@router.get("/exportar.xlsx")
async def exportar_xlsx(
    user: CurrentUser,
    db: DBSession,
    empresa: Annotated[str, Query()],
    periodo: Annotated[str | None, Query(pattern=_PERIODO_RE)] = None,
) -> Response:
    """Registro de Egresos (17 columnas) + Carga_Gastos (21 oficiales)."""
    codigo = await _gate(user, db, empresa)
    filas = (
        await db.execute(text(_SQL_EXPORT), {"empresa": codigo, "periodo": periodo})
    ).mappings().all()
    contenido = construir_export_xlsx(codigo, periodo, list(filas))
    nombre = nombre_archivo_export(codigo, periodo)
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nombre}"',
            "X-Total-Rows": str(len(filas)),
        },
    )


# ---------------------------------------------------------------------------
# Alta + rutas por id
# ---------------------------------------------------------------------------


@router.post("", response_model=EgresoRead, status_code=status.HTTP_201_CREATED)
async def crear_egreso(body: EgresoCreate, user: CurrentUser, db: DBSession) -> EgresoRead:
    codigo = await _gate(user, db, body.empresa_codigo)
    quien = user.email or str(user.sub)
    nuevo_id = await _insertar(db, codigo, body, body.origen, quien)
    await db.commit()
    return egreso_read_desde_fila(await _leer_o_404(db, nuevo_id))


@router.get("/{egreso_id}", response_model=EgresoDetail)
async def obtener_egreso(egreso_id: int, user: CurrentUser, db: DBSession) -> EgresoDetail:
    """La ficha: el gasto + historial con diff campo a campo."""
    fila = await _fila_propia_o_404(user, db, egreso_id)
    hist = (await db.execute(text(_SQL_HISTORIAL), {"id": egreso_id})).mappings().all()
    base = egreso_read_desde_fila(fila)
    return EgresoDetail(**base.model_dump(), historial=armar_historial(list(hist)))


@router.put("/{egreso_id}", response_model=EgresoRead)
async def actualizar_egreso(
    egreso_id: int, body: EgresoUpdate, user: CurrentUser, db: DBSession
) -> EgresoRead:
    fila = await _fila_propia_o_404(user, db, egreso_id)
    try:
        valores = fusionar_update(fila, body)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)
        ) from exc
    valores["periodo"] = valores["fecha"].strftime("%Y-%m")
    valores["id"] = egreso_id
    valores["updated_by"] = user.email or str(user.sub)
    actualizado = (await db.execute(text(_SQL_UPDATE), valores)).scalar()
    if actualizado is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No existe el gasto #{egreso_id} (o fue borrado)",
        )
    await db.commit()
    return egreso_read_desde_fila(await _leer_o_404(db, egreso_id))


@router.delete("/{egreso_id}", response_model=EgresoDeleteResponse)
async def borrar_egreso(
    egreso_id: int, body: EgresoDeleteRequest, user: CurrentUser, db: DBSession
) -> EgresoDeleteResponse:
    """Borrado lógico con motivo: la fila desaparece de la grilla, no de la BD."""
    await _fila_propia_o_404(user, db, egreso_id)
    quien = user.email or str(user.sub)
    deleted_at = (
        await db.execute(
            text(_SQL_SOFT_DELETE), {"id": egreso_id, "who": quien, "motivo": body.motivo}
        )
    ).scalar()
    if deleted_at is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No existe el gasto #{egreso_id} (o fue borrado)",
        )
    await db.commit()
    return EgresoDeleteResponse(egreso_id=egreso_id, deleted_at=deleted_at)
