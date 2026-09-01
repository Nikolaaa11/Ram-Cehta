"""Exportar listas a Excel.

Un solo endpoint `GET /exports/{entity_type}.xlsx?<filtros>` genera un workbook
con la misma data que la lista paginada del frontend, sin paginación —
hasta `_MAX_ROWS` filas. Se respeta el filtro de empresa cuando aplica.

Diseño:
- Devolvemos un `StreamingResponse` con `application/vnd.openxmlformats-...`.
- `openpyxl` ya es dep del proyecto (lo usa el ETL).
- Header row con estilos suaves (fondo gris claro, bold) — Apple-ish dentro de
  las limitaciones de Excel.
- Nombre de archivo: `{entity}_{empresa|all}_{YYYY-MM-DD}.xlsx`.
- Auto-width: usamos `column_dimensions[col].width = max(len_header, max_data_len) + 2`
  con tope 60 para evitar columnas de 200+ con observaciones largas.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession

router = APIRouter()

_MAX_ROWS = 10_000  # techo defensivo; alerta si lo cruzamos
_HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")  # ink-100
_HEADER_FONT = Font(bold=True, color="111827")  # ink-900
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

# Mapa entity_type → (sql, params, columnas)
_ENTITY_QUERIES: dict[str, dict] = {
    "ordenes_compra": {
        # Columnas al día con la tabla real. Las que faltaban dolían:
        #   · sin tipo_documento/retención/total_a_pagar, una OC de
        #     HONORARIOS exportaba "Total" bruto y el Excel del operador no
        #     cuadraba contra lo que se paga (el líquido);
        #   · sin firmas, el Excel no servía para perseguir quién falta —
        #     que es justo para lo que Nicolás lo exporta.
        "headers": [
            "OC ID",
            "Número OC",
            "Empresa",
            "Proveedor",
            "Fecha emisión",
            "Moneda",
            "Tipo documento",
            "Neto",
            "IVA",
            "Total",
            "Retención",
            "Total a pagar",
            "Estado",
            "Firmas",
            "Faltan por firmar",
            "Forma pago",
            "Plazo pago",
            "Observaciones",
        ],
        "sql": """
            SELECT oc.oc_id, oc.numero_oc, oc.empresa_codigo,
                   COALESCE(p.razon_social, '—') AS proveedor,
                   oc.fecha_emision, oc.moneda, oc.tipo_documento,
                   oc.neto, oc.iva, oc.total,
                   oc.retencion_monto,
                   COALESCE(oc.total_a_pagar, oc.total) AS total_a_pagar,
                   oc.estado,
                   COALESCE(fi.firmadas, 0) || '/' || COALESCE(fi.total_f, 0)
                       AS firmas,
                   COALESCE(fi.pendientes, '') AS faltan,
                   oc.forma_pago, oc.plazo_pago, oc.observaciones
            FROM core.ordenes_compra oc
            LEFT JOIN core.proveedores p ON p.proveedor_id = oc.proveedor_id
            LEFT JOIN (
                SELECT oc_id,
                       count(*) AS total_f,
                       count(*) FILTER (WHERE status = 'FIRMADA') AS firmadas,
                       string_agg(COALESCE(firmante_nombre, firmante_email),
                                  ', ' ORDER BY orden)
                           FILTER (WHERE status = 'PENDIENTE') AS pendientes
                  FROM core.oc_firmas GROUP BY oc_id
            ) fi ON fi.oc_id = oc.oc_id
            WHERE (CAST(:empresa AS TEXT) IS NULL OR oc.empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL OR oc.estado = :estado)
            ORDER BY oc.fecha_emision DESC NULLS LAST
            LIMIT :lim
        """,
    },
    "f29": {
        "headers": [
            "F29 ID",
            "Empresa",
            "Periodo",
            "Vencimiento",
            "Monto a pagar",
            "Fecha pago",
            "Estado",
        ],
        "sql": """
            SELECT f29_id, empresa_codigo, periodo_tributario, fecha_vencimiento,
                   monto_a_pagar, fecha_pago, estado
            FROM core.f29_obligaciones
            WHERE (CAST(:empresa AS TEXT) IS NULL OR empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL OR estado = :estado)
            ORDER BY fecha_vencimiento DESC NULLS LAST
            LIMIT :lim
        """,
    },
    "proveedores": {
        "headers": [
            "ID",
            "Razón social",
            "RUT",
            "Giro",
            "Email",
            "Teléfono",
            "Banco",
            "Tipo cuenta",
            "Nº cuenta",
            "Activo",
        ],
        "sql": """
            SELECT proveedor_id, razon_social, rut, giro, email, telefono,
                   banco, tipo_cuenta, numero_cuenta, activo
            FROM core.proveedores
            WHERE (CAST(:empresa AS TEXT) IS NULL)  -- proveedores son globales
              AND (CAST(:estado AS TEXT) IS NULL)
            ORDER BY razon_social
            LIMIT :lim
        """,
    },
    "trabajadores": {
        "headers": [
            "ID",
            "Empresa",
            "Nombre",
            "RUT",
            "Cargo",
            "Email",
            "Teléfono",
            "Fecha ingreso",
            "Fecha egreso",
            "Sueldo bruto",
            "Tipo contrato",
            "Estado",
        ],
        "sql": """
            SELECT trabajador_id, empresa_codigo, nombre_completo, rut, cargo,
                   email, telefono, fecha_ingreso, fecha_egreso, sueldo_bruto,
                   tipo_contrato, estado
            FROM core.trabajadores
            WHERE (CAST(:empresa AS TEXT) IS NULL OR empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL OR estado = :estado)
            ORDER BY empresa_codigo, nombre_completo
            LIMIT :lim
        """,
    },
    "legal_documents": {
        "headers": [
            "ID",
            "Empresa",
            "Categoría",
            "Subcategoría",
            "Nombre",
            "Contraparte",
            "Vigencia desde",
            "Vigencia hasta",
            "Monto",
            "Moneda",
            "Estado",
        ],
        "sql": """
            SELECT documento_id, empresa_codigo, categoria, subcategoria,
                   nombre, contraparte, fecha_vigencia_desde, fecha_vigencia_hasta,
                   monto, moneda, estado
            FROM core.legal_documents
            WHERE (CAST(:empresa AS TEXT) IS NULL OR empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL OR estado = :estado)
            ORDER BY empresa_codigo, uploaded_at DESC
            LIMIT :lim
        """,
    },
    "movimientos": {
        "headers": [
            "ID",
            "Empresa",
            "Fecha",
            "Descripción",
            "Concepto general",
            "Concepto detallado",
            "Abono",
            "Egreso",
            "Proyecto",
            "Fuente",
            "Banco",
            "Tipo doc",
            "Nº doc",
        ],
        "sql": """
            SELECT movimiento_id, empresa_codigo, fecha, descripcion,
                   concepto_general, concepto_detallado, abono, egreso,
                   proyecto, fuente, banco, tipo_documento, numero_documento
            FROM core.movimientos
            WHERE (CAST(:empresa AS TEXT) IS NULL OR empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL OR real_proyectado = :estado)
            ORDER BY fecha DESC NULLS LAST
            LIMIT :lim
        """,
    },
    "suscripciones": {
        "headers": [
            "ID",
            "Empresa",
            "Fecha recibo",
            "Acciones pagadas",
            "Monto UF",
            "Monto CLP",
            "Contrato ref",
            "Firmado",
            "Fecha firma",
        ],
        "sql": """
            SELECT suscripcion_id, empresa_codigo, fecha_recibo, acciones_pagadas,
                   monto_uf, monto_clp, contrato_ref, firmado, fecha_firma
            FROM core.suscripciones_acciones
            WHERE (CAST(:empresa AS TEXT) IS NULL OR empresa_codigo = :empresa)
              AND (CAST(:estado AS TEXT) IS NULL)
            ORDER BY fecha_recibo DESC NULLS LAST
            LIMIT :lim
        """,
    },
    "fondos": {
        "headers": [
            "ID",
            "Nombre",
            "Tipo",
            "País",
            "Región",
            "Ticket min USD",
            "Ticket max USD",
            "Estado outreach",
            "Próx. contacto",
            "Email",
            "Website",
        ],
        "sql": """
            SELECT fondo_id, nombre, tipo, pais, region,
                   ticket_min_usd, ticket_max_usd, estado_outreach,
                   fecha_proximo_contacto, contacto_email, website
            FROM core.fondos
            WHERE (CAST(:empresa AS TEXT) IS NULL)  -- fondos son globales
              AND (CAST(:estado AS TEXT) IS NULL OR estado_outreach = :estado)
            ORDER BY nombre
            LIMIT :lim
        """,
    },
    "entregables": {
        "headers": [
            "ID",
            "Template",
            "Nombre",
            "Categoría",
            "Subcategoría",
            "Período",
            "Fecha límite",
            "Frecuencia",
            "Prioridad",
            "Responsable",
            "Estado",
            "Fecha entrega real",
            "Días restantes",
            "Motivo no entrega",
            "Notas",
            "Adjunto URL",
            "Referencia normativa",
            "Empresa",
        ],
        "sql": """
            SELECT entregable_id, id_template, nombre, categoria, subcategoria,
                   periodo, fecha_limite, frecuencia, prioridad, responsable,
                   estado, fecha_entrega_real,
                   (fecha_limite - CURRENT_DATE) AS dias_restantes,
                   motivo_no_entrega, notas, adjunto_url, referencia_normativa,
                   COALESCE(extra->>'empresa_codigo', subcategoria) AS empresa
            FROM app.entregables_regulatorios
            WHERE (
                CAST(:empresa AS TEXT) IS NULL
                OR subcategoria = :empresa
                OR extra->>'empresa_codigo' = :empresa
            )
              AND (CAST(:estado AS TEXT) IS NULL OR estado = :estado)
            ORDER BY fecha_limite ASC
            LIMIT :lim
        """,
    },
}


#: Caracteres de control prohibidos por el XML de Excel. openpyxl los
#: rechaza con IllegalCharacterError y el export entero muere en 500 por UNA
#: celda sucia — y las observaciones vienen de PDFs/emails parseados, donde
#: estos bytes aparecen de verdad.
_XML_ILEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _serialize(value: object) -> object:
    """Convierte tipos que openpyxl no traga directamente."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value
    if isinstance(value, str):
        return _XML_ILEGAL.sub(" ", value)
    return value


def _build_workbook(headers: list[str], rows: list[tuple], sheet_name: str) -> bytes:
    """Genera el xlsx en memoria y devuelve los bytes."""
    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover — defensive
        raise RuntimeError("openpyxl no inicializó worksheet")
    ws.title = sheet_name[:31]  # límite Excel

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    max_lens = [len(h) for h in headers]
    for row in rows:
        serialized = [_serialize(v) for v in row]
        ws.append(serialized)
        for i, val in enumerate(serialized):
            ln = len(str(val)) if val not in (None, "") else 0
            if ln > max_lens[i]:
                max_lens[i] = ln

    for i, ln in enumerate(max_lens, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(ln + 2, 60)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/{entity_type}.xlsx")
async def export_entity_xlsx(
    user: CurrentUser,
    db: DBSession,
    entity_type: str,
    empresa_codigo: Annotated[str | None, Query(alias="empresa_codigo")] = None,
    estado: str | None = None,
) -> StreamingResponse:
    """Exporta una lista a Excel, respetando filtros opcionales + scope.

    `entity_type` debe ser una clave de `_ENTITY_QUERIES`. 404 si no existe.

    R152FFFFFF — SECURITY: el SQL filtra por el `empresa_codigo` del query
    param, NO por el scope del user. Sin el check de abajo, un user de la
    empresa A podía exportar data de la empresa B (sueldos, RUTs de
    trabajadores, OCs, movimientos). Ahora validamos scope.
    """
    if entity_type not in _ENTITY_QUERIES:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"entity_type '{entity_type}' no soportado para exportación",
        )

    # R152FFFFFF — Validar scope multi-tenant.
    from app.services.empresa_scope_service import get_allowed_empresa_codes

    allowed = await get_allowed_empresa_codes(user, db)
    if allowed is not None:  # None = admin global, sin restricción
        if empresa_codigo:
            # Pidió una empresa específica → debe estar en su scope.
            if empresa_codigo not in allowed:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"Sin acceso a empresa '{empresa_codigo}'.",
                )
        else:
            # Sin filtro de empresa → forzar a su empresa si solo tiene 1,
            # o rechazar si tiene varias (evita export cross-empresa).
            if len(allowed) == 1:
                empresa_codigo = next(iter(allowed))
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        "Especificá una empresa (parámetro empresa_codigo) "
                        "para exportar. No se permite exportar todas a la vez."
                    ),
                )

    cfg = _ENTITY_QUERIES[entity_type]
    rows = (
        await db.execute(
            text(cfg["sql"]),
            {"empresa": empresa_codigo, "estado": estado, "lim": _MAX_ROWS},
        )
    ).fetchall()

    sheet_name = entity_type.replace("_", " ").title()
    xlsx_bytes = _build_workbook(cfg["headers"], list(rows), sheet_name)

    today = date.today().isoformat()
    suffix = empresa_codigo or "all"
    filename = f"{entity_type}_{suffix}_{today}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Total-Rows": str(len(rows)),
            "X-Truncated": "true" if len(rows) >= _MAX_ROWS else "false",
        },
    )
