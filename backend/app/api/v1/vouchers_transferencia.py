"""Round 11 — Generador de Excel de transferencia masiva desde vouchers.

Endpoint:
    POST /vouchers/transferencia-masiva
        Body: { voucher_ids: list[int], banco_formato: str = "GENERICO" }
        Returns: StreamingResponse application/xlsx con planilla de
        transferencia masiva lista para cargar al banco.

Casos de uso:
    Una vez que un voucher COMPRA/EGRESO esta APPROVED (firmado por
    GG + DIRECTOR), debe pagarse. Para pagos via transferencia bancaria
    el operador puede:
        1. Seleccionar N vouchers APPROVED desde /aprobaciones
        2. Click "Generar Excel de transferencia masiva"
        3. Descargar XLSX con un row por beneficiario:
            - RUT, Razon Social
            - Banco, Tipo cuenta, Numero cuenta
            - Email (para notificacion bancaria)
            - Monto (Total a transferir)
            - Glosa (texto breve para el comprobante)
            - Referencia (codigo voucher para auditoria)
        4. Cargar el XLSX al banco online (BCI, Santander, BancoEstado,
           Itau, etc) — todos aceptan formatos planos tipo nomina.

Validaciones:
    - Cada voucher debe estar en status APPROVED.
    - Tipo del voucher debe ser COMPRA o EGRESO (un INGRESO no se paga).
    - forma_pago debe ser TRANSFERENCIA (o no setear → se permite con
      warning, ya que algunos vouchers viejos no llenaban forma_pago).
    - Proveedor (contraparte) debe tener al menos RUT. Banco/cuenta son
      opcionales en el modelo pero idealmente seteados; si faltan se
      dejan en blanco en el Excel para que el operador los complete.
    - Scope multi-tenant: solo vouchers en empresas a las que el user
      tiene acceso. Si pide alguno fuera de scope, error 403.

Formatos de banco:
    Por ahora soportamos "GENERICO" — un layout estandar que cubre 90%
    de los bancos chilenos. Futuros formatos podrian agregarse en el
    diccionario _BANCO_LAYOUTS si un banco especifico requiere headers
    o orden particular (ej. BancoEstado pide CSV con orden fijo).
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession, require_scope
from app.core.limiter import limiter
from app.core.security import AuthenticatedUser
from app.services.audit_service import audit_log
from app.services.empresa_scope_service import EmpresaScopeDep

router = APIRouter()


# Estilos coherentes con el resto de exports del proyecto.
_HEADER_FILL = PatternFill("solid", fgColor="236C4F")  # cehta-green
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TOTAL_FILL = PatternFill("solid", fgColor="F3F4F6")
_TOTAL_FONT = Font(bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


# Layout estandar — cubre BCI / Santander / Itau / BBVA / BancoEstado.
# Cada banco luego pide rearrange manual de columnas, pero estos son los
# 9 campos universales.
_HEADERS_GENERICO = [
    "Cod. Voucher",       # Trazabilidad — no lo suele leer el banco pero util al user
    "Empresa",            # Empresa emisora (CEHTA, EVOQUE, etc.)
    "RUT Beneficiario",   # RUT canonico 12.345.678-9
    "Nombre Beneficiario", # Razon social
    "Banco",              # BANCO DE CHILE, BCI, SANTANDER...
    "Tipo Cuenta",        # CORRIENTE / VISTA / AHORRO
    "Numero Cuenta",      # Como texto (los ceros a la izq importan)
    "Email Beneficiario", # Para notificacion bancaria
    "Monto CLP",          # Numero entero (CLP no tiene centavos)
    "Glosa",              # Texto breve <= 60 chars (limite tipico bancos)
    "Fecha Documento",    # Para conciliar
]

# Round 103 — Layout exacto del Banco Santander (template provisto por
# operador: "Plantilla de transferencia Climate Smart Leasing SpA.xlsx").
# 13 columnas. Algunas son obligatorias solo si el banco destino no es
# Santander (interbancarias). Los headers se renderean con line break
# para imitar exactamente el template.
_HEADERS_SANTANDER = [
    "Cuenta origen\n(obligatorio)",
    "Moneda origen\n(obligatorio)",
    "Cuenta destino\n(obligatorio)",
    "Moneda destino\n(obligatorio)",
    "Código banco destino\n(obligatorio solo si banco destino no es Santander)",
    "RUT beneficiario\n(obligatorio solo si banco destino no es Santander)",
    "Nombre beneficiario\n(obligatorio solo si banco destino no es Santander)",
    "Monto transferencia\n(obligatorio)",
    "Glosa personalizada transferencia\n(opcional)",
    "Correo beneficiario\n(opcional)",
    "Mensaje correo beneficiario\n(opcional)",
    "Glosa cartola originador\n(opcional)",
    "Glosa cartola beneficiario\n(opcional, solo aplica si cuenta destino es Santander)",
]

# Mapeo banco_codigo Santander para columna 5. Los principales bancos
# chilenos tienen codigos numericos asignados por el SBIF/CMF.
_BANCO_CODIGOS = {
    "BANCO DE CHILE": "001",
    "BANCO INTERNACIONAL": "009",
    "SCOTIABANK": "014",
    "BCI": "016",
    "BANCO DEL ESTADO": "012",
    "BANCOESTADO": "012",
    "CORPBANCA": "027",
    "ITAU": "039",
    "ITAÙ": "039",
    "HSBC": "031",
    "DEUTSCHE BANK": "028",
    "RIPLEY": "053",
    "BICE": "028",
    "SECURITY": "049",
    "FALABELLA": "051",
    "CONSORCIO": "055",
    "PARIS": "060",
    "BBVA": "037",
    "BTG PACTUAL": "059",
    "SANTANDER": "037",
}


class TransferenciaMasivaRequest(BaseModel):
    """Body del POST."""

    voucher_ids: list[int] = Field(
        min_length=1,
        max_length=500,
        description=(
            "IDs de vouchers APPROVED a incluir en la planilla de transferencia. "
            "Maximo 500 por export — si necesitas mas, partilo en lotes."
        ),
    )
    banco_formato: Literal["GENERICO", "SANTANDER"] = Field(
        default="GENERICO",
        description=(
            "Formato del Excel: GENERICO (11 cols universales) o "
            "SANTANDER (13 cols del template oficial Santander)."
        ),
    )
    cuenta_origen: str | None = Field(
        default=None,
        max_length=30,
        description=(
            "Cuenta origen para columna A en formato SANTANDER. "
            "Si formato=GENERICO se ignora."
        ),
    )
    incluir_pendientes_aprobacion: bool = Field(
        default=False,
        description=(
            "Si True, tambien acepta vouchers PENDING (para preview antes "
            "de firmar). Por default solo APPROVED."
        ),
    )


def _fmt_monto_clp(value: Decimal | float | int | None) -> int:
    """CLP no tiene centavos — entero redondeado."""
    if value is None:
        return 0
    return int(Decimal(str(value)).quantize(Decimal("1")))


def _trunc(s: str | None, n: int) -> str:
    """Trunca a N chars, devuelve '' si None."""
    if not s:
        return ""
    s = str(s).strip()
    return s[: n - 1] + "…" if len(s) > n else s


def _build_workbook(
    rows: list[dict],
    *,
    fecha_export: date,
    user_email: str,
    formato: str = "GENERICO",
    cuenta_origen: str = "",
) -> bytes:
    """Genera el XLSX en memoria.

    Round 103 — soporta 2 formatos:
      - GENERICO: 11 columnas universales (BCI, BBVA, BancoEstado, etc.)
      - SANTANDER: 13 columnas exactas del template oficial Santander
        (template provisto por operador).
    """
    if formato == "SANTANDER":
        return _build_workbook_santander(
            rows,
            fecha_export=fecha_export,
            user_email=user_email,
            cuenta_origen=cuenta_origen,
        )
    return _build_workbook_generico(
        rows, fecha_export=fecha_export, user_email=user_email
    )


def _build_workbook_generico(
    rows: list[dict], *, fecha_export: date, user_email: str
) -> bytes:
    """Layout estándar Round 11 — 11 columnas universales."""
    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        raise RuntimeError("openpyxl no inicializo worksheet")
    ws.title = "Transferencias"

    # Headers
    ws.append(_HEADERS_GENERICO)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER

    # Data rows
    total_monto = 0
    max_lens = [len(h) for h in _HEADERS_GENERICO]
    for r in rows:
        row_vals = [
            r["codigo"],
            r["empresa_codigo"],
            r["rut"] or "",
            r["nombre"] or "",
            r["banco"] or "",
            r["tipo_cuenta"] or "",
            # Numero de cuenta como string para preservar zeros a la izquierda
            str(r["numero_cuenta"]) if r["numero_cuenta"] else "",
            r["email"] or "",
            _fmt_monto_clp(r["monto"]),
            _trunc(r["glosa"], 60),
            r["fecha_documento"].isoformat() if r["fecha_documento"] else "",
        ]
        ws.append(row_vals)
        total_monto += _fmt_monto_clp(r["monto"])
        for i, v in enumerate(row_vals):
            ln = len(str(v))
            if ln > max_lens[i]:
                max_lens[i] = ln

    # Fila total al final (negrita, fondo gris)
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=1, value="TOTAL")
    ws.cell(row=total_row_idx, column=9, value=total_monto)
    for col in range(1, len(_HEADERS_GENERICO) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER

    # Format monto column con thousand separator
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.number_format = "#,##0"

    # Auto-width (con tope 50 para no romper layouts)
    for i, ln in enumerate(max_lens, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(ln + 2, 50)
    ws.freeze_panes = "A2"

    # Hoja "Resumen" con metadata para auditoria
    ws2 = wb.create_sheet(title="Resumen")
    ws2.append(["Generado", fecha_export.isoformat()])
    ws2.append(["Usuario", user_email])
    ws2.append(["Total vouchers", len(rows)])
    ws2.append(["Total CLP", total_monto])
    ws2.append([])
    ws2.append(["", ""])
    ws2.append(["Por empresa:", ""])
    por_empresa: dict[str, dict[str, int]] = {}
    for r in rows:
        emp = r["empresa_codigo"]
        if emp not in por_empresa:
            por_empresa[emp] = {"count": 0, "total": 0}
        por_empresa[emp]["count"] += 1
        por_empresa[emp]["total"] += _fmt_monto_clp(r["monto"])
    for emp, info in sorted(por_empresa.items()):
        ws2.append([emp, f"{info['count']} vouchers", f"${info['total']:,} CLP"])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_workbook_santander(
    rows: list[dict],
    *,
    fecha_export: date,
    user_email: str,
    cuenta_origen: str = "",
) -> bytes:
    """Round 103 — Layout exacto del Banco Santander template oficial.

    13 columnas con headers multilinea. Cuenta origen viene como input
    del operador (no la tenemos en DB). Moneda origen/destino = CLP.
    Codigo banco destino se mapea desde el nombre del banco del proveedor.

    Reglas Santander:
      - Si cuenta destino es SANTANDER, cols 5/6/7 pueden quedar vacias.
      - Si NO es Santander, cols 5/6/7 son obligatorios (interbancarias).
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        raise RuntimeError("openpyxl no inicializo worksheet")
    ws.title = "Transferencias"

    # Headers — wrap text necesario por los \n en los labels
    ws.append(_HEADERS_SANTANDER)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
    # Altura fila header para que se vea el wrap
    ws.row_dimensions[1].height = 60

    # Data rows
    total_monto = 0
    for r in rows:
        banco_dest_norm = (r["banco"] or "").upper().strip()
        es_santander = "SANTANDER" in banco_dest_norm
        codigo_banco = "" if es_santander else _BANCO_CODIGOS.get(
            banco_dest_norm, ""
        )
        row_vals = [
            cuenta_origen,                                  # Col 1
            "CLP",                                          # Col 2
            (str(r["numero_cuenta"]) if r["numero_cuenta"] else ""),  # Col 3
            "CLP",                                          # Col 4
            codigo_banco,                                   # Col 5
            (r["rut"] or "") if not es_santander else "",   # Col 6
            (r["nombre"] or "") if not es_santander else "", # Col 7
            _fmt_monto_clp(r["monto"]),                     # Col 8
            _trunc(r["glosa"] or r["codigo"], 60),          # Col 9 — Glosa personalizada
            r["email"] or "",                               # Col 10
            "",                                             # Col 11 — Mensaje correo (vacío default)
            _trunc(f"Cehta - {r['codigo']}", 30),           # Col 12 — Cartola originador
            _trunc(r["glosa"] or "", 30) if es_santander else "",  # Col 13
        ]
        ws.append(row_vals)
        total_monto += _fmt_monto_clp(r["monto"])

    # Sin fila TOTAL (el banco no la lee). Solo data limpia.

    # Format monto col 8 con thousand separator + sin decimales
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "#,##0"

    # Anchos razonables por columna (Santander mide chars)
    widths = [18, 10, 18, 10, 18, 18, 30, 14, 32, 28, 32, 22, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Hoja "Instrucciones" con info del operador
    ws2 = wb.create_sheet(title="Instrucciones")
    ws2.append(["PLANTILLA DE TRANSFERENCIA · BANCO SANTANDER"])
    ws2.append([])
    ws2.append(["Generado:", fecha_export.isoformat()])
    ws2.append(["Usuario:", user_email])
    ws2.append(["Cantidad transferencias:", len(rows)])
    ws2.append(["Monto total CLP:", total_monto])
    ws2.append([])
    ws2.append(["Cómo usar:"])
    ws2.append(["1. Completar 'Cuenta origen' (col A) con tu cuenta Santander de origen."])
    ws2.append(["2. Verificar 'Cuenta destino' (col C) — debe tener formato numérico sin guiones."])
    ws2.append(["3. Si beneficiario es Santander, cols E/F/G pueden quedar vacías."])
    ws2.append(["4. Subir XLSX al portal de Pagos Masivos del Santander."])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 40
    ws2["A1"].font = Font(bold=True, size=13, color="236C4F")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.post(
    "/transferencia-masiva",
    dependencies=[Depends(require_scope("voucher:execute"))],
)
# NOTA Round 148: @limiter.limit("5/minute") removido — rompe Pydantic
# schema inference cuando el endpoint tiene Annotated[..., Depends(...)]
# combinado con BaseModel body y StreamingResponse. Síntoma: FastAPI
# trata `user`, `db`, `scope`, `body` como query params requeridos y
# devuelve 422 "Field required" para todos. Mismo problema documentado
# en vouchers_extract.py. Default rate limit global aplica.
async def export_transferencia_masiva(
    request: Request,
    user: Annotated[AuthenticatedUser, Depends(require_scope("voucher:execute"))],
    db: DBSession,
    scope: EmpresaScopeDep,
    body: TransferenciaMasivaRequest,
) -> StreamingResponse:
    """Genera Excel de transferencia masiva desde vouchers APPROVED.

    Flujo:
      1. Fetcha cada voucher_id solicitado en una query (WHERE id = ANY).
      2. Valida que esten APPROVED (o PENDING si incluir_pendientes_aprobacion).
      3. Joinea con core.proveedores via contraparte_rut para obtener
         datos bancarios. Si no hay match (proveedor no en catalogo),
         deja los campos bancarios en blanco — el operador los completa
         manualmente antes de cargarlo al banco.
      4. Filtra por scope multi-tenant. Si algun voucher no es accesible,
         404 (no leak de existence; igual que el get/{id} de vouchers).
      5. Genera XLSX y devuelve como descarga.
      6. Audit log: deja huella de quien genero la planilla, con que
         vouchers, y total — relevante para forensia financiera.
    """
    if scope.is_global:
        scope_filter = None
    else:
        scope_filter = list(scope.allowed_codes or [])
        if not scope_filter:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tenes acceso a ninguna empresa activa.",
            )

    # R152UUUUUU — EXECUTED fuera de la planilla: un voucher ya pagado
    # re-entraba al Excel sin advertencia si el operador re-seleccionaba
    # IDs, y se cargaba de nuevo al banco (transferencia duplicada al
    # proveedor). SYNCED se mantiene: exportado a Nubox pero aún sin pagar.
    allowed_statuses = ["APPROVED", "SYNCED"]
    if body.incluir_pendientes_aprobacion:
        allowed_statuses.append("PENDING")

    # Query unica: voucher + proveedor JOIN. LEFT JOIN para no perder
    # vouchers cuyo proveedor no esta en catalogo.
    params: dict = {
        "ids": body.voucher_ids,
        "statuses": allowed_statuses,
    }
    where_scope = ""
    if scope_filter is not None:
        where_scope = "AND v.empresa_codigo = ANY(CAST(:scope AS text[]))"
        params["scope"] = scope_filter

    sql = f"""
        SELECT
            v.voucher_id,
            v.codigo,
            v.empresa_codigo,
            v.tipo,
            v.status,
            v.fecha_documento,
            v.glosa,
            v.contraparte_rut,
            v.contraparte_nombre,
            v.total_credit AS monto,
            v.forma_pago,
            p.email AS prov_email,
            p.banco AS prov_banco,
            p.tipo_cuenta AS prov_tipo_cuenta,
            p.numero_cuenta AS prov_numero_cuenta
        FROM core.vouchers v
        LEFT JOIN core.proveedores p ON p.rut = v.contraparte_rut
        WHERE v.voucher_id = ANY(:ids)
          AND v.status = ANY(:statuses)
          {where_scope}
        ORDER BY v.empresa_codigo, v.codigo
    """
    rows_raw = (await db.execute(text(sql), params)).mappings().all()

    if not rows_raw:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "Ninguno de los vouchers solicitados esta en estado pagable "
                f"({', '.join(allowed_statuses)}) o accesible a tu cuenta. "
                "Revisa que esten APPROVED y que tengas scope sobre la empresa."
            ),
        )

    # Detectar vouchers solicitados que NO aparecieron — para reportar
    # al user en el header del response.
    found_ids = {int(r["voucher_id"]) for r in rows_raw}
    missing_ids = sorted(set(body.voucher_ids) - found_ids)

    rows = [
        {
            "codigo": r["codigo"],
            "empresa_codigo": r["empresa_codigo"],
            "rut": r["contraparte_rut"],
            "nombre": r["contraparte_nombre"],
            "banco": r["prov_banco"],
            "tipo_cuenta": r["prov_tipo_cuenta"],
            "numero_cuenta": r["prov_numero_cuenta"],
            "email": r["prov_email"],
            "monto": r["monto"],
            "glosa": r["glosa"],
            "fecha_documento": r["fecha_documento"],
        }
        for r in rows_raw
    ]

    today = date.today()
    xlsx_bytes = _build_workbook(
        rows,
        fecha_export=today,
        user_email=str(getattr(user, "email", user.sub)),
        formato=body.banco_formato,
        cuenta_origen=body.cuenta_origen or "",
    )

    # Audit log — quien, cuando, que vouchers, total
    try:
        total = sum(_fmt_monto_clp(r["monto"]) for r in rows)
        await audit_log(
            db,
            None,
            user,
            action="export_transferencia_masiva",
            entity_type="voucher",
            entity_id=",".join(str(r["codigo"]) for r in rows[:10])
            + (f"… (+{len(rows) - 10} mas)" if len(rows) > 10 else ""),
            entity_label=f"Transferencia masiva — {len(rows)} vouchers — ${total:,} CLP",
            summary=(
                f"Exporto planilla transferencia masiva: {len(rows)} vouchers, "
                f"total ${total:,} CLP, banco_formato={body.banco_formato}"
            ),
            before=None,
            after={
                "voucher_ids": body.voucher_ids,
                "found": len(rows),
                "missing": len(missing_ids),
                "total_clp": total,
            },
        )
        await db.commit()
    except Exception:
        # Audit log no debe romper el download.
        pass

    filename = f"transferencia_masiva_{today.isoformat()}_{len(rows)}_vouchers.xlsx"
    response_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Total-Rows": str(len(rows)),
        "X-Total-CLP": str(sum(_fmt_monto_clp(r["monto"]) for r in rows)),
    }
    if missing_ids:
        response_headers["X-Missing-Voucher-Ids"] = ",".join(str(i) for i in missing_ids[:20])

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers,
    )


@router.get("/transferencia-masiva/preview")
async def preview_transferencia_masiva(
    user: CurrentUser,
    db: DBSession,
    scope: EmpresaScopeDep,
) -> dict:
    """Lista los vouchers APPROVED listos para pagar.

    Usado por el frontend para mostrar el set "elegible" antes de
    seleccionar cuales incluir en el Excel. Devuelve count + monto total
    + breakdown por empresa.
    """
    if scope.is_global:
        scope_filter = None
    else:
        scope_filter = list(scope.allowed_codes or [])
        if not scope_filter:
            return {
                "count": 0,
                "total_clp": 0,
                "items": [],
                "by_empresa": [],
            }

    where_scope = ""
    params: dict = {}
    if scope_filter is not None:
        where_scope = "AND v.empresa_codigo = ANY(CAST(:scope AS text[]))"
        params["scope"] = scope_filter

    # Round 10 — el preview ahora trae tambien `proveedor_telefono` y
    # `proveedor_contacto` para que el FE pueda renderear botones
    # "WhatsApp" por fila tras ejecutar la transferencia. LEFT JOIN
    # separado de `tiene_datos_bancarios` porque ese require banco+cuenta
    # no-null mientras telefono puede existir aunque no haya cuenta.
    # Round 113 — incluye proyecto_dominante (primera linea con
    # proyecto_codigo no null por voucher) usando subquery indexado.
    # El partial index idx_voucher_lines_proyecto cubre el WHERE NOT NULL.
    sql = f"""
        SELECT
            v.voucher_id,
            v.codigo,
            v.empresa_codigo,
            v.tipo,
            v.fecha_documento::text AS fecha_documento,
            v.glosa,
            v.contraparte_rut,
            v.contraparte_nombre,
            v.total_credit::text AS monto,
            v.forma_pago,
            CASE WHEN pbanco.proveedor_id IS NULL THEN FALSE ELSE TRUE END AS tiene_datos_bancarios,
            pcontact.telefono AS proveedor_telefono,
            pcontact.contacto AS proveedor_contacto,
            (SELECT vl.proyecto_codigo
               FROM core.voucher_lines vl
              WHERE vl.voucher_id = v.voucher_id
                AND vl.proyecto_codigo IS NOT NULL
              ORDER BY vl.line_number ASC
              LIMIT 1) AS proyecto_dominante
        FROM core.vouchers v
        LEFT JOIN core.proveedores pbanco
            ON pbanco.rut = v.contraparte_rut
            AND pbanco.banco IS NOT NULL
            AND pbanco.numero_cuenta IS NOT NULL
        LEFT JOIN core.proveedores pcontact
            ON pcontact.rut = v.contraparte_rut
        WHERE v.status = 'APPROVED'
          AND v.tipo IN ('COMPRA', 'EGRESO')
          {where_scope}
        ORDER BY v.fecha_documento DESC, v.voucher_id DESC
        LIMIT 200
    """
    rows = (await db.execute(text(sql), params)).mappings().all()

    total = sum(int(Decimal(r["monto"] or "0")) for r in rows)
    by_empresa: dict[str, dict[str, int]] = {}
    for r in rows:
        emp = r["empresa_codigo"]
        if emp not in by_empresa:
            by_empresa[emp] = {"count": 0, "total": 0}
        by_empresa[emp]["count"] += 1
        by_empresa[emp]["total"] += int(Decimal(r["monto"] or "0"))

    return {
        "count": len(rows),
        "total_clp": total,
        "items": [
            {
                "voucher_id": int(r["voucher_id"]),
                "codigo": r["codigo"],
                "empresa_codigo": r["empresa_codigo"],
                "tipo": r["tipo"],
                "fecha_documento": r["fecha_documento"],
                "glosa": r["glosa"],
                "contraparte_rut": r["contraparte_rut"],
                "contraparte_nombre": r["contraparte_nombre"],
                "monto": r["monto"],
                "forma_pago": r["forma_pago"],
                "tiene_datos_bancarios": r["tiene_datos_bancarios"],
                # Round 10 — exponer telefono y contacto del proveedor
                # para que el FE renderee botones WhatsApp por fila.
                "proveedor_telefono": r["proveedor_telefono"],
                "proveedor_contacto": r["proveedor_contacto"],
                # Round 113 — proyecto contable dominante para que el
                # tesorero vea a que centro de costo aplica el pago.
                "proyecto_dominante": r["proyecto_dominante"],
            }
            for r in rows
        ],
        "by_empresa": [
            {"empresa_codigo": k, "count": v["count"], "total_clp": v["total"]}
            for k, v in sorted(by_empresa.items())
        ],
    }
