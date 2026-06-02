"""R152zzz · Flujo de caja proyectado por proyecto.

Endpoints:
  GET  /flujos-caja/proyecto/{proyecto_codigo}        matriz periodo × categoría
  PUT  /flujos-caja/proyecto/{proyecto_codigo}/cell   upsert una celda
  POST /flujos-caja/proyecto/{proyecto_codigo}/upload subir Excel para
                                                      poblar de una vez
  DELETE /flujos-caja/proyecto/{proyecto_codigo}/cell/{id}
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Annotated, Any

from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession

router = APIRouter()


class FlujoCell(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    proyecto_codigo: str
    periodo: str
    categoria: str
    tipo: str
    monto_proyectado: Decimal
    monto_real: Decimal = Decimal("0")
    notas: str | None = None


class CellUpsert(BaseModel):
    periodo: str = Field(..., pattern=r"^\d{4}-\d{2}$")
    categoria: str = Field(..., min_length=1, max_length=80)
    tipo: str = Field(default="EGRESO", pattern=r"^(INGRESO|EGRESO)$")
    monto_proyectado: Decimal = Field(..., ge=0)
    notas: str | None = Field(default=None, max_length=500)


class UploadResult(BaseModel):
    proyecto_codigo: str
    celdas_creadas: int
    celdas_actualizadas: int


async def _check_proyecto_exists(db, proyecto_codigo: str) -> None:
    row = await db.scalar(
        text(
            "SELECT 1 FROM core.proyectos_contables "
            "WHERE codigo = :c AND estado = 'ACTIVE'"
        ),
        {"c": proyecto_codigo},
    )
    if not row:
        raise HTTPException(
            status_code=404,
            detail=f"Proyecto contable {proyecto_codigo} no existe o no está activo",
        )


@router.get(
    "/flujos-caja/proyecto/{proyecto_codigo}",
    response_model=list[FlujoCell],
)
async def list_flujo(
    user: CurrentUser, db: DBSession, proyecto_codigo: str
) -> list[FlujoCell]:
    """Devuelve todas las celdas del flujo de caja del proyecto,
    con monto_real calculado desde vouchers EXECUTED."""
    await _check_proyecto_exists(db, proyecto_codigo)
    rows = await db.execute(
        text(
            """SELECT id, proyecto_codigo, periodo, categoria, tipo,
                      monto_proyectado, monto_real, notas
               FROM core.v_flujos_caja_proyecto_con_real
               WHERE proyecto_codigo = :c
               ORDER BY periodo, categoria"""
        ),
        {"c": proyecto_codigo},
    )
    return [FlujoCell.model_validate(dict(r._mapping)) for r in rows]


@router.put(
    "/flujos-caja/proyecto/{proyecto_codigo}/cell",
    response_model=FlujoCell,
)
async def upsert_cell(
    user: CurrentUser,
    db: DBSession,
    proyecto_codigo: str,
    body: CellUpsert,
) -> FlujoCell:
    """Upsert idempotente de una celda. UNIQUE(proyecto, periodo, categoria)."""
    await _check_proyecto_exists(db, proyecto_codigo)
    row = (
        await db.execute(
            text(
                """INSERT INTO core.flujos_caja_proyecto
                       (proyecto_codigo, periodo, categoria, tipo,
                        monto_proyectado, notas)
                   VALUES (:c, :p, :cat, :tipo, :monto, :notas)
                   ON CONFLICT (proyecto_codigo, periodo, categoria)
                   DO UPDATE SET
                       tipo = EXCLUDED.tipo,
                       monto_proyectado = EXCLUDED.monto_proyectado,
                       notas = EXCLUDED.notas,
                       updated_at = NOW()
                   RETURNING id, proyecto_codigo, periodo, categoria, tipo,
                             monto_proyectado, notas"""
            ),
            {
                "c": proyecto_codigo,
                "p": body.periodo,
                "cat": body.categoria.strip(),
                "tipo": body.tipo,
                "monto": body.monto_proyectado,
                "notas": body.notas,
            },
        )
    ).mappings().first()
    await db.commit()
    data = dict(row)
    data["monto_real"] = Decimal("0")
    return FlujoCell.model_validate(data)


@router.delete(
    "/flujos-caja/proyecto/{proyecto_codigo}/cell/{cell_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_cell(
    user: CurrentUser, db: DBSession, proyecto_codigo: str, cell_id: int
) -> None:
    await db.execute(
        text(
            """DELETE FROM core.flujos_caja_proyecto
               WHERE id = :id AND proyecto_codigo = :c"""
        ),
        {"id": cell_id, "c": proyecto_codigo},
    )
    await db.commit()


@router.post(
    "/flujos-caja/proyecto/{proyecto_codigo}/upload",
    response_model=UploadResult,
)
async def upload_excel(
    user: CurrentUser,
    db: DBSession,
    proyecto_codigo: str,
    file: Annotated[UploadFile, File(...)],
    tipo_default: Annotated[str, Query(pattern=r"^(INGRESO|EGRESO)$")] = "EGRESO",
) -> UploadResult:
    """Sube Excel con columnas: periodo, categoria, monto, (tipo opcional, notas opcional).

    Estructura esperada:
      Fila 1: headers
      Fila 2+: una fila por celda
    Periodo válido: 'YYYY-MM'.
    Idempotente: si la celda ya existe, se sobrescribe.
    """
    await _check_proyecto_exists(db, proyecto_codigo)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Sólo archivos .xlsx/.xls")
    import openpyxl
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Excel inválido: {exc}") from exc

    ws = wb.active
    headers = [
        str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))
    ]
    required = {"periodo", "categoria", "monto"}
    if not required.issubset(headers):
        raise HTTPException(
            400,
            f"Faltan columnas obligatorias en el Excel. Requeridas: {required}. "
            f"Encontradas: {headers}",
        )
    idx_periodo = headers.index("periodo")
    idx_categoria = headers.index("categoria")
    idx_monto = headers.index("monto")
    idx_tipo = headers.index("tipo") if "tipo" in headers else None
    idx_notas = headers.index("notas") if "notas" in headers else None

    creadas = 0
    actualizadas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx_periodo] is None:
            continue
        periodo = str(row[idx_periodo]).strip()
        if len(periodo) != 7 or periodo[4] != "-":
            continue  # skip silenciosamente filas inválidas
        categoria = str(row[idx_categoria] or "").strip()
        if not categoria:
            continue
        monto = row[idx_monto]
        if monto is None:
            continue
        try:
            monto_dec = Decimal(str(monto))
        except Exception:
            continue
        tipo = (
            str(row[idx_tipo] or tipo_default).strip().upper()
            if idx_tipo is not None
            else tipo_default
        )
        if tipo not in ("INGRESO", "EGRESO"):
            tipo = tipo_default
        notas = (
            str(row[idx_notas] or "").strip() or None if idx_notas is not None else None
        )

        # Detectar si existía
        existed = await db.scalar(
            text(
                """SELECT 1 FROM core.flujos_caja_proyecto
                   WHERE proyecto_codigo = :c AND periodo = :p AND categoria = :cat"""
            ),
            {"c": proyecto_codigo, "p": periodo, "cat": categoria},
        )
        await db.execute(
            text(
                """INSERT INTO core.flujos_caja_proyecto
                       (proyecto_codigo, periodo, categoria, tipo,
                        monto_proyectado, notas)
                   VALUES (:c, :p, :cat, :tipo, :monto, :notas)
                   ON CONFLICT (proyecto_codigo, periodo, categoria)
                   DO UPDATE SET
                       tipo = EXCLUDED.tipo,
                       monto_proyectado = EXCLUDED.monto_proyectado,
                       notas = EXCLUDED.notas,
                       updated_at = NOW()"""
            ),
            {
                "c": proyecto_codigo,
                "p": periodo,
                "cat": categoria,
                "tipo": tipo,
                "monto": monto_dec,
                "notas": notas,
            },
        )
        if existed:
            actualizadas += 1
        else:
            creadas += 1
    await db.commit()

    return UploadResult(
        proyecto_codigo=proyecto_codigo,
        celdas_creadas=creadas,
        celdas_actualizadas=actualizadas,
    )
