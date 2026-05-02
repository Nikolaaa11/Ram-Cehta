"""Endpoints — Entregables Regulatorios FIP CEHTA ESG (V4 fase 6).

Diseño:
- `nivel_alerta` y `dias_restantes` se calculan server-side en cada GET
  (no se persisten — siempre frescos respecto al "hoy" actual).
- Marcar como `entregado` un recurrente (mensual/trimestral/semestral/
  anual/bienal) auto-crea la próxima instancia si no existe — así el
  pipeline operativo nunca pierde un período. Idempotente vía
  `ON CONFLICT (id_template, periodo) DO NOTHING`.
- Compliance: el campo `es_publico` queda forzado a False (restricción
  del PROMPT_MAESTRO Bloque 10).
"""
from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import PlainTextResponse, Response, StreamingResponse
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession, require_scope
from app.core.security import AuthenticatedUser
from app.schemas.entregable import (
    BulkReassignRequest,
    BulkReassignResponse,
    BulkUpdateRequest,
    BulkUpdateResponse,
    Categoria,
    ComplianceGradeEmpresa,
    ComplianceGradeReport,
    CriticalCount,
    CsvImportError,
    CsvImportResponse,
    EntregableCreate,
    EntregableEstadosCounts,
    EntregableRead,
    EntregableUpdate,
    EstadoEntregable,
    GenerarSerieRequest,
    GenerarSerieResponse,
    NivelAlerta,
    ReporteRegulatorio,
)
from app.services.audit_service import audit_log

router = APIRouter()

_FIELDS = (
    "entregable_id, id_template, nombre, descripcion, categoria, subcategoria, "
    "referencia_normativa, fecha_limite, frecuencia, prioridad, responsable, "
    "estado, fecha_entrega_real, motivo_no_entrega, notas, adjunto_url, "
    "periodo, alerta_15, alerta_10, alerta_5, generado_automaticamente, "
    "es_publico, extra, created_at, updated_at"
)


def _next_periodo_y_fecha(
    frecuencia: str, periodo: str, fecha_limite: date
) -> tuple[str, date] | None:
    """Calcula la próxima instancia de un template recurrente.

    Devuelve `None` si la frecuencia no es recurrente (`unico`/`segun_evento`)
    o si el formato del `periodo` no se reconoce.

    Convenciones de período (espejo de `_fechas_del_anio`):
        mensual:    MM-YYYY        → siguiente mes calendario, último día
        trimestral: Q{n}-YYYY      → siguiente trimestre, cierre + 30 días
        semestral:  S{n}-YYYY      → S1 julio, S2 enero+1
        anual:      YYYY           → año siguiente, 30 abril
        bienal:     YYYY-YYYY+1    → bienio siguiente, 31 diciembre

    Para mensual/trimestral preferimos derivar la fecha desde la fecha
    actual (no desde el período) para respetar feriados u offsets que el
    usuario haya editado manualmente — desplazamos +1 unidad de cadencia.
    """
    import calendar as cal
    from datetime import timedelta

    try:
        if frecuencia == "mensual":
            # MM-YYYY → siguiente mes
            mm, yyyy = periodo.split("-", 1)
            m, y = int(mm), int(yyyy)
            ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
            last = cal.monthrange(ny, nm)[1]
            return (f"{nm:02d}-{ny}", date(ny, nm, last))
        if frecuencia == "trimestral":
            # Q{n}-YYYY → siguiente trimestre
            if not periodo.startswith("Q"):
                return None
            qstr, yyyy = periodo[1:].split("-", 1)
            q, y = int(qstr), int(yyyy)
            ny, nq = (y + 1, 1) if q == 4 else (y, q + 1)
            m_close = nq * 3
            last_close = cal.monthrange(ny, m_close)[1]
            entrega = date(ny, m_close, last_close) + timedelta(days=30)
            return (f"Q{nq}-{ny}", entrega)
        if frecuencia == "semestral":
            if not periodo.startswith("S"):
                return None
            sstr, yyyy = periodo[1:].split("-", 1)
            s, y = int(sstr), int(yyyy)
            if s == 1:
                # S1-Y → S2-Y vence enero del año siguiente
                return (f"S2-{y}", date(y + 1, 1, 31))
            # S2-Y → S1-(Y+1) vence julio Y+1
            return (f"S1-{y + 1}", date(y + 1, 7, 31))
        if frecuencia == "anual":
            y = int(periodo)
            return (f"{y + 1}", date(y + 2, 4, 30))
        if frecuencia == "bienal":
            # YYYY-YYYY+1 → siguiente bienio
            y_start, y_end = periodo.split("-", 1)
            y0 = int(y_start)
            return (f"{y0 + 2}-{y0 + 3}", date(y0 + 3, 12, 31))
    except (ValueError, IndexError):
        return None
    return None


def _calcular_alerta(fecha_limite: date, estado: str) -> tuple[NivelAlerta, int]:
    """Replica la lógica del PROMPT_MAESTRO Bloque 3.1.

    Si ya está entregado, devolvemos 'normal' aunque la fecha haya pasado —
    no tiene sentido mostrar alerta de algo ya cumplido.
    """
    if estado == "entregado":
        diff = (fecha_limite - date.today()).days
        return ("normal", diff)
    diff = (fecha_limite - date.today()).days
    if diff < 0:
        return ("vencido", diff)
    if diff == 0:
        return ("hoy", 0)
    if diff <= 5:
        return ("critico", diff)
    if diff <= 10:
        return ("urgente", diff)
    if diff <= 15:
        return ("proximo", diff)
    if diff <= 30:
        return ("en_rango", diff)
    return ("normal", diff)


def _row_to_read(row: dict[str, Any]) -> EntregableRead:
    nivel, dias = _calcular_alerta(row["fecha_limite"], row["estado"])
    return EntregableRead(
        entregable_id=row["entregable_id"],
        id_template=row["id_template"],
        nombre=row["nombre"],
        descripcion=row.get("descripcion"),
        categoria=row["categoria"],
        subcategoria=row.get("subcategoria"),
        referencia_normativa=row.get("referencia_normativa"),
        fecha_limite=row["fecha_limite"],
        frecuencia=row["frecuencia"],
        prioridad=row["prioridad"],
        responsable=row["responsable"],
        estado=row["estado"],
        fecha_entrega_real=row.get("fecha_entrega_real"),
        motivo_no_entrega=row.get("motivo_no_entrega"),
        notas=row.get("notas"),
        adjunto_url=row.get("adjunto_url"),
        periodo=row["periodo"],
        alerta_15=row["alerta_15"],
        alerta_10=row["alerta_10"],
        alerta_5=row["alerta_5"],
        generado_automaticamente=row["generado_automaticamente"],
        es_publico=row["es_publico"],
        extra=row.get("extra"),
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        nivel_alerta=nivel,
        dias_restantes=dias,
    )


# ---------------------------------------------------------------------------
# GET / — listado paginado + filtros
# ---------------------------------------------------------------------------
@router.get("", response_model=list[EntregableRead])
async def list_entregables(
    user: CurrentUser,
    db: DBSession,
    categoria: Categoria | None = None,
    estado: EstadoEntregable | None = None,
    anio: int | None = Query(None, ge=2024, le=2030),
    mes: int | None = Query(None, ge=1, le=12),
    desde: date | None = None,
    hasta: date | None = None,
    responsable: str | None = Query(
        None, max_length=120, description="Match parcial case-insensitive"
    ),
    empresa: str | None = Query(
        None, max_length=120, description="Filtra por subcategoria o extra.empresa_codigo"
    ),
    q: str | None = Query(
        None, max_length=200, description="Búsqueda libre en nombre/descripcion/notas"
    ),
    only_alerta: bool = Query(False, description="Solo entregables en alerta activa"),
) -> list[EntregableRead]:
    """Lista entregables con filtros opcionales.

    Filtros V4 fase 7 añadidos: `mes`, `responsable` (parcial), `empresa`
    (match contra `subcategoria` o `extra->>empresa_codigo`), `q`
    (búsqueda full-text en `nombre || descripcion || notas`).

    Si `only_alerta=true`, solo devuelve los que estén en estado pendiente o
    en_proceso AND con días_restantes ≤ 15 (criterio del banner del frontend).
    """
    conditions: list[str] = []
    params: dict[str, Any] = {}
    if categoria:
        conditions.append("categoria = :categoria")
        params["categoria"] = categoria
    if estado:
        conditions.append("estado = :estado")
        params["estado"] = estado
    if anio:
        conditions.append("EXTRACT(year FROM fecha_limite) = :anio")
        params["anio"] = anio
    if mes:
        conditions.append("EXTRACT(month FROM fecha_limite) = :mes")
        params["mes"] = mes
    if desde:
        conditions.append("fecha_limite >= :desde")
        params["desde"] = desde
    if hasta:
        conditions.append("fecha_limite <= :hasta")
        params["hasta"] = hasta
    if responsable:
        conditions.append("responsable ILIKE :resp")
        params["resp"] = f"%{responsable}%"
    if empresa:
        # Doble match: subcategoria (legacy) o extra->>empresa_codigo (json).
        conditions.append(
            "(subcategoria ILIKE :emp OR extra->>'empresa_codigo' ILIKE :emp)"
        )
        params["emp"] = f"%{empresa}%"
    if q:
        # Búsqueda case-insensitive en los campos de texto principales.
        conditions.append(
            "(nombre ILIKE :q OR descripcion ILIKE :q OR notas ILIKE :q "
            "OR id_template ILIKE :q OR referencia_normativa ILIKE :q)"
        )
        params["q"] = f"%{q}%"
    if only_alerta:
        conditions.append(
            "estado IN ('pendiente','en_proceso') "
            "AND fecha_limite <= (CURRENT_DATE + INTERVAL '15 days')"
        )

    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    rows = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                f"{where} ORDER BY fecha_limite ASC"
            ),
            params,
        )
    ).mappings().all()
    return [_row_to_read(dict(r)) for r in rows]


# ---------------------------------------------------------------------------
# GET /facets — valores únicos para alimentar dropdowns de filtros
# ---------------------------------------------------------------------------
@router.get("/facets")
async def get_facets(
    user: CurrentUser, db: DBSession
) -> dict[str, list[str]]:
    """Devuelve valores únicos para alimentar selects de filtros.

    Útil para que el frontend muestre dropdowns de Responsable / Empresa
    sin tener que adivinar valores. Devuelve lista ordenada alfabéticamente.
    """
    rows_resp = (
        await db.execute(
            text(
                "SELECT DISTINCT responsable FROM app.entregables_regulatorios "
                "WHERE responsable IS NOT NULL ORDER BY responsable"
            )
        )
    ).all()
    rows_emp = (
        await db.execute(
            text(
                "SELECT DISTINCT COALESCE(extra->>'empresa_codigo', subcategoria) AS emp "
                "FROM app.entregables_regulatorios "
                "WHERE COALESCE(extra->>'empresa_codigo', subcategoria) IS NOT NULL "
                "ORDER BY emp"
            )
        )
    ).all()
    return {
        "responsables": [r[0] for r in rows_resp if r[0]],
        "empresas": [r[0] for r in rows_emp if r[0]],
    }


# ---------------------------------------------------------------------------
# GET /counts — contadores por estado para el header del módulo
# ---------------------------------------------------------------------------
@router.get("/counts", response_model=EntregableEstadosCounts)
async def get_counts(
    user: CurrentUser, db: DBSession
) -> EntregableEstadosCounts:
    """Conteo por estado, sin filtros — para el header KPI del módulo."""
    rows = (
        await db.execute(
            text(
                "SELECT estado, COUNT(*) AS n "
                "FROM app.entregables_regulatorios "
                "GROUP BY estado"
            )
        )
    ).mappings().all()
    counts = EntregableEstadosCounts()
    for r in rows:
        if hasattr(counts, r["estado"]):
            setattr(counts, r["estado"], r["n"])
    return counts


# ---------------------------------------------------------------------------
# GET /critical-count — conteo ligero para el badge sidebar
# ---------------------------------------------------------------------------
@router.get("/critical-count", response_model=CriticalCount)
async def critical_count(
    user: CurrentUser, db: DBSession
) -> CriticalCount:
    """Conteo agregado de entregables en alerta crítica.

    Crítico = vencido / hoy / ≤5 días, AND aún no entregado.

    Endpoint pensado para el badge del sidebar — devuelve solo enteros,
    sin payloads, con un solo SELECT agregado para que sea barato de
    llamar frecuentemente (cada vez que el usuario navega).
    """
    row = (
        await db.execute(
            text(
                """
                SELECT
                  COUNT(*) FILTER (WHERE fecha_limite < CURRENT_DATE)             AS vencidos,
                  COUNT(*) FILTER (WHERE fecha_limite = CURRENT_DATE)             AS hoy,
                  COUNT(*) FILTER (
                    WHERE fecha_limite > CURRENT_DATE
                      AND fecha_limite <= (CURRENT_DATE + INTERVAL '5 days')
                  )                                                               AS proximos_5d
                FROM app.entregables_regulatorios
                WHERE estado IN ('pendiente','en_proceso')
                """
            )
        )
    ).mappings().first()
    if row is None:
        return CriticalCount()
    vencidos = int(row["vencidos"] or 0)
    hoy = int(row["hoy"] or 0)
    proximos = int(row["proximos_5d"] or 0)
    return CriticalCount(
        critical=vencidos + hoy + proximos,
        vencidos=vencidos,
        hoy=hoy,
        proximos_5d=proximos,
    )


# ---------------------------------------------------------------------------
# GET /compliance-grade — % cumplimiento por empresa con grade A-F
# ---------------------------------------------------------------------------
def _grade_from_tasa(tasa_a_tiempo: float, tasa_cumplimiento: float) -> str:
    """Calcula la nota A-F basada en tasa de entrega a tiempo + cumplimiento.

    Filosofía: lo que se entrega a tiempo pesa más que lo que se entrega
    tarde, pero ambos cuentan para no penalizar de más.
    """
    score = tasa_a_tiempo * 0.7 + tasa_cumplimiento * 0.3
    if score >= 95:
        return "A"
    if score >= 85:
        return "B"
    if score >= 70:
        return "C"
    if score >= 50:
        return "D"
    return "F"


async def _compute_compliance_for_empresa(
    db: DBSession, empresa_codigo: str
) -> ComplianceGradeEmpresa:
    """Calcula compliance grade YTD para una empresa específica."""
    # YTD: del 1 enero del año actual hasta hoy
    rows = (
        await db.execute(
            text(
                """
                SELECT estado, fecha_limite, fecha_entrega_real
                FROM app.entregables_regulatorios
                WHERE fecha_limite >=
                      make_date(EXTRACT(year FROM CURRENT_DATE)::int, 1, 1)
                  AND (
                    subcategoria = :emp
                    OR extra->>'empresa_codigo' = :emp
                  )
                """
            ),
            {"emp": empresa_codigo},
        )
    ).mappings().all()

    total_vencidos = 0  # los que ya tenían que estar entregados (fecha < hoy)
    entregados_a_tiempo = 0
    entregados_atrasados = 0
    no_entregados = 0
    pendientes_aun = 0  # fecha futura

    today = date.today()
    for r in rows:
        fl = r["fecha_limite"]
        es = r["estado"]
        fer = r.get("fecha_entrega_real")
        if fl >= today:
            pendientes_aun += 1
            continue
        total_vencidos += 1
        if es == "entregado":
            if fer is not None and fer <= fl:
                entregados_a_tiempo += 1
            else:
                entregados_atrasados += 1
        else:
            no_entregados += 1

    if total_vencidos == 0:
        tasa_cumplimiento = 100.0
        tasa_a_tiempo = 100.0
    else:
        entregados_total = entregados_a_tiempo + entregados_atrasados
        tasa_cumplimiento = round(
            entregados_total / total_vencidos * 100.0, 1
        )
        tasa_a_tiempo = round(
            entregados_a_tiempo / total_vencidos * 100.0, 1
        )

    return ComplianceGradeEmpresa(
        empresa_codigo=empresa_codigo,
        total=total_vencidos,
        entregados_a_tiempo=entregados_a_tiempo,
        entregados_atrasados=entregados_atrasados,
        no_entregados=no_entregados,
        pendientes=pendientes_aun,
        tasa_cumplimiento=tasa_cumplimiento,
        tasa_a_tiempo=tasa_a_tiempo,
        grade=_grade_from_tasa(tasa_a_tiempo, tasa_cumplimiento),
    )


@router.get(
    "/compliance-grade/{empresa_codigo}",
    response_model=ComplianceGradeEmpresa,
)
async def compliance_grade_empresa(
    user: CurrentUser,
    db: DBSession,
    empresa_codigo: str,
) -> ComplianceGradeEmpresa:
    """Devuelve el compliance grade YTD de una empresa específica.

    Pensado para el widget de empresa (`/empresa/[codigo]`). Muestra:
        - Total de entregables que ya vencieron en YTD
        - Cuántos a tiempo / atrasados / no entregados
        - % cumplimiento + % a tiempo
        - Nota A/B/C/D/F como score consolidado.
    """
    return await _compute_compliance_for_empresa(db, empresa_codigo)


@router.get("/compliance-grade", response_model=ComplianceGradeReport)
async def compliance_grade_report(
    user: CurrentUser, db: DBSession
) -> ComplianceGradeReport:
    """Snapshot consolidado de compliance por empresa para todo el portafolio.

    Itera sobre las empresas que tienen al menos 1 entregable asignado
    y devuelve la lista ordenada por compliance score descendente.
    """
    empresas_rows = (
        await db.execute(
            text(
                """
                SELECT DISTINCT
                  COALESCE(extra->>'empresa_codigo', subcategoria) AS emp
                FROM app.entregables_regulatorios
                WHERE COALESCE(extra->>'empresa_codigo', subcategoria) IS NOT NULL
                """
            )
        )
    ).all()
    empresas_codigos = [r[0] for r in empresas_rows if r[0]]

    grades: list[ComplianceGradeEmpresa] = []
    for codigo in empresas_codigos:
        grade = await _compute_compliance_for_empresa(db, codigo)
        grades.append(grade)

    grades.sort(key=lambda g: g.tasa_a_tiempo, reverse=True)
    promedio = (
        round(
            sum(g.tasa_cumplimiento for g in grades) / len(grades),
            1,
        )
        if grades
        else 100.0
    )

    return ComplianceGradeReport(
        generado_at=datetime.now(UTC),
        empresas=grades,
        promedio_cumplimiento=promedio,
    )


# ---------------------------------------------------------------------------
# GET /reporte-cv.xlsx — Excel multi-sheet rico para acta Comité Vigilancia
# ---------------------------------------------------------------------------
@router.get("/reporte-cv.xlsx")
async def reporte_cv_xlsx(
    user: CurrentUser,
    db: DBSession,
    empresa: str | None = Query(
        None,
        description=(
            "Si se pasa, filtra el reporte a una sola empresa "
            "(match contra subcategoria o extra.empresa_codigo). "
            "Útil para acta empresa-específica."
        ),
    ),
) -> StreamingResponse:
    """Genera un Excel multi-sheet con formato profesional para Comité Vigilancia.

    5 sheets:
        1. Resumen ejecutivo — KPIs + tasa cumplimiento + meta
        2. Vencidos sin entregar — fondo rojo claro, requiere explicación
        3. Próximos 30 días — pipeline corto
        4. Distribución por estado — counts agregados
        5. Compliance por empresa — ranking con grade A/B/C/D/F (oculto si
           filtro empresa específica, ya que solo habría 1 fila)

    Apple-ish formatting: header bold + fill suave, freeze panes, auto-width,
    borders sutiles. Pensado para imprimir o entregar al CV.
    """
    # Filtro empresa — usado en todas las queries de este reporte si se pasa
    empresa_filter = ""
    empresa_params: dict[str, Any] = {}
    if empresa:
        empresa_filter = (
            " AND (subcategoria = :emp OR extra->>'empresa_codigo' = :emp)"
        )
        empresa_params["emp"] = empresa
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    # Paleta inline (no podemos usar tokens Apple en Excel)
    hdr_fill = PatternFill("solid", fgColor="111827")  # ink-900
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    neg_fill = PatternFill("solid", fgColor="FEE2E2")  # negative-50
    warn_fill = PatternFill("solid", fgColor="FEF3C7")  # warning-50
    pos_fill = PatternFill("solid", fgColor="D1FAE5")  # positive-50
    thin_side = Side(style="thin", color="E5E7EB")
    cell_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    def _autofit(ws, headers: list[str], rows_data: list[list[Any]]):
        max_lens = [len(h) for h in headers]
        for row in rows_data:
            for i, v in enumerate(row):
                ln = len(str(v)) if v not in (None, "") else 0
                if ln > max_lens[i]:
                    max_lens[i] = ln
        for i, ln in enumerate(max_lens, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(ln + 2, 50)

    def _write_header(ws, headers: list[str], row_num: int = 1):
        ws.append(headers)
        for cell in ws[row_num]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = cell_border
        ws.row_dimensions[row_num].height = 22

    # ── Cargar datos ─────────────────────────────────────────────────
    today = date.today()
    proximos = (
        await db.execute(
            text(
                f"""
                SELECT entregable_id, nombre, categoria, periodo, fecha_limite,
                       responsable, estado, referencia_normativa,
                       (fecha_limite - CURRENT_DATE) AS dias
                FROM app.entregables_regulatorios
                WHERE estado IN ('pendiente','en_proceso')
                  AND fecha_limite >= CURRENT_DATE
                  AND fecha_limite <= (CURRENT_DATE + INTERVAL '30 days')
                  {empresa_filter}
                ORDER BY fecha_limite ASC
                """
            ),
            empresa_params,
        )
    ).mappings().all()
    vencidos = (
        await db.execute(
            text(
                f"""
                SELECT entregable_id, nombre, categoria, periodo, fecha_limite,
                       responsable, estado, motivo_no_entrega,
                       (fecha_limite - CURRENT_DATE) AS dias
                FROM app.entregables_regulatorios
                WHERE estado IN ('pendiente','en_proceso','no_entregado')
                  AND fecha_limite < CURRENT_DATE
                  {empresa_filter}
                ORDER BY fecha_limite ASC
                """
            ),
            empresa_params,
        )
    ).mappings().all()
    counts_rows = (
        await db.execute(
            text(
                "SELECT estado, COUNT(*) AS n FROM app.entregables_regulatorios "
                f"WHERE 1=1 {empresa_filter} "
                "GROUP BY estado"
            ),
            empresa_params,
        )
    ).mappings().all()
    ytd_total = (
        await db.execute(
            text(
                "SELECT COUNT(*) FROM app.entregables_regulatorios "
                "WHERE fecha_limite >= make_date(EXTRACT(year FROM CURRENT_DATE)::int, 1, 1) "
                f"AND fecha_limite < CURRENT_DATE {empresa_filter}"
            ),
            empresa_params,
        )
    ).scalar() or 0
    ytd_entregados = (
        await db.execute(
            text(
                "SELECT COUNT(*) FROM app.entregables_regulatorios "
                "WHERE fecha_limite >= make_date(EXTRACT(year FROM CURRENT_DATE)::int, 1, 1) "
                "AND fecha_limite < CURRENT_DATE "
                f"AND estado = 'entregado' {empresa_filter}"
            ),
            empresa_params,
        )
    ).scalar() or 0
    tasa = round(ytd_entregados / ytd_total * 100, 1) if ytd_total > 0 else 100.0

    # Compliance per empresa
    empresas_rows = (
        await db.execute(
            text(
                """
                SELECT DISTINCT
                  COALESCE(extra->>'empresa_codigo', subcategoria) AS emp
                FROM app.entregables_regulatorios
                WHERE COALESCE(extra->>'empresa_codigo', subcategoria) IS NOT NULL
                """
            )
        )
    ).all()
    # Si filtramos por empresa, sólo computamos esa empresa (skip ranking)
    compliance_data = []
    if empresa:
        compliance_data.append(
            await _compute_compliance_for_empresa(db, empresa)
        )
    else:
        for r in empresas_rows:
            emp = r[0]
            if not emp:
                continue
            grade = await _compute_compliance_for_empresa(db, emp)
            compliance_data.append(grade)
        compliance_data.sort(key=lambda g: g.tasa_a_tiempo, reverse=True)

    # ── Build workbook ───────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Resumen ejecutivo
    ws1 = wb.active
    ws1.title = "Resumen ejecutivo"
    ws1_title_text = (
        f"FIP CEHTA ESG · {empresa} · Reporte Comité de Vigilancia"
        if empresa
        else "FIP CEHTA ESG · AFIS S.A. · Reporte Comité de Vigilancia"
    )
    ws1["A1"] = ws1_title_text
    ws1["A1"].font = Font(bold=True, size=14, color="111827")
    ws1["A1"].alignment = Alignment(horizontal="left")
    ws1.merge_cells("A1:D1")
    ws1["A2"] = (
        f"Generado el {datetime.now(UTC).strftime('%d %b %Y · %H:%M UTC')}"
        + (f" · Filtrado a empresa: {empresa}" if empresa else "")
    )
    ws1["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws1.merge_cells("A2:D2")

    # KPIs
    ws1["A4"] = "Indicador"
    ws1["B4"] = "Valor"
    ws1["A4"].fill = hdr_fill
    ws1["A4"].font = hdr_font
    ws1["B4"].fill = hdr_fill
    ws1["B4"].font = hdr_font

    counts_dict = {r["estado"]: r["n"] for r in counts_rows}
    kpi_rows = [
        ("Tasa de cumplimiento YTD", f"{tasa:.1f}%"),
        ("Total entregables YTD", str(ytd_total)),
        ("Entregados YTD", str(ytd_entregados)),
        ("Vencidos sin entregar", str(len(vencidos))),
        ("Próximos 30 días pendientes", str(len(proximos))),
        ("Pendientes (estado)", str(counts_dict.get("pendiente", 0))),
        ("En proceso (estado)", str(counts_dict.get("en_proceso", 0))),
        ("Entregados (estado)", str(counts_dict.get("entregado", 0))),
        ("No entregados (estado)", str(counts_dict.get("no_entregado", 0))),
    ]
    for i, (label, value) in enumerate(kpi_rows, start=5):
        ws1.cell(row=i, column=1, value=label).font = Font(size=10)
        c = ws1.cell(row=i, column=2, value=value)
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="right")
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 18

    # Sheet 2: Vencidos sin entregar
    ws2 = wb.create_sheet("Vencidos s entregar")
    headers2 = [
        "ID", "Categoría", "Período", "Fecha límite", "Días vencido",
        "Entregable", "Responsable", "Estado", "Motivo no entrega",
    ]
    _write_header(ws2, headers2)
    rows_data2 = []
    for r in vencidos:
        row = [
            r["entregable_id"],
            r["categoria"],
            r["periodo"],
            r["fecha_limite"],
            abs(int(r["dias"] or 0)),
            r["nombre"],
            r["responsable"],
            r["estado"],
            r.get("motivo_no_entrega") or "",
        ]
        ws2.append(row)
        rows_data2.append(row)
        # Highlight rojo claro
        for cell in ws2[ws2.max_row]:
            cell.fill = neg_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autofit(ws2, headers2, rows_data2)
    ws2.freeze_panes = "A2"

    # Sheet 3: Próximos 30 días
    ws3 = wb.create_sheet("Próximos 30 días")
    headers3 = [
        "ID", "Categoría", "Período", "Fecha límite", "Días restantes",
        "Entregable", "Responsable", "Estado", "Ref. normativa",
    ]
    _write_header(ws3, headers3)
    rows_data3 = []
    for r in proximos:
        dias = int(r["dias"] or 0)
        row = [
            r["entregable_id"],
            r["categoria"],
            r["periodo"],
            r["fecha_limite"],
            dias,
            r["nombre"],
            r["responsable"],
            r["estado"],
            r.get("referencia_normativa") or "",
        ]
        ws3.append(row)
        rows_data3.append(row)
        # Highlight según urgencia
        fill = warn_fill if dias <= 5 else None
        for cell in ws3[ws3.max_row]:
            if fill is not None:
                cell.fill = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autofit(ws3, headers3, rows_data3)
    ws3.freeze_panes = "A2"

    # Sheet 4: Distribución por estado
    ws4 = wb.create_sheet("Distribución estado")
    headers4 = ["Estado", "Cantidad"]
    _write_header(ws4, headers4)
    rows_data4 = []
    for estado_name in ["pendiente", "en_proceso", "entregado", "no_entregado"]:
        cnt = counts_dict.get(estado_name, 0)
        row = [estado_name.replace("_", " ").title(), cnt]
        ws4.append(row)
        rows_data4.append(row)
        for cell in ws4[ws4.max_row]:
            cell.border = cell_border
            if estado_name == "entregado":
                cell.fill = pos_fill
            elif estado_name == "no_entregado":
                cell.fill = neg_fill
    _autofit(ws4, headers4, rows_data4)
    ws4.freeze_panes = "A2"

    # Sheet 5: Compliance por empresa
    ws5 = wb.create_sheet("Compliance por empresa")
    headers5 = [
        "Posición", "Empresa", "Grade",
        "Tasa a tiempo (%)", "Tasa cumplimiento (%)",
        "Total vencidos", "A tiempo", "Atrasados", "No entregados", "Pendientes",
    ]
    _write_header(ws5, headers5)
    rows_data5 = []
    for i, g in enumerate(compliance_data, start=1):
        row = [
            i, g.empresa_codigo, g.grade,
            g.tasa_a_tiempo, g.tasa_cumplimiento,
            g.total, g.entregados_a_tiempo, g.entregados_atrasados,
            g.no_entregados, g.pendientes,
        ]
        ws5.append(row)
        rows_data5.append(row)
        # Highlight por grade
        fill = (
            pos_fill if g.grade in ("A", "B")
            else warn_fill if g.grade == "C"
            else neg_fill if g.grade in ("D", "F")
            else None
        )
        for cell in ws5[ws5.max_row]:
            if fill is not None:
                cell.fill = fill
            cell.border = cell_border
    _autofit(ws5, headers5, rows_data5)
    ws5.freeze_panes = "B2"

    # ── Output ───────────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = f"-{empresa.lower()}" if empresa else ""
    filename = f"reporte-cv{suffix}-{today.isoformat()}.xlsx"
    return StreamingResponse(
        _io.BytesIO(buf.read()),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Sheets": "5",
        },
    )


# ---------------------------------------------------------------------------
# GET /reporte-regulatorio — resumen para actas Comité Vigilancia
# ---------------------------------------------------------------------------
@router.get("/reporte-regulatorio", response_model=ReporteRegulatorio)
async def reporte_regulatorio(
    user: CurrentUser, db: DBSession
) -> ReporteRegulatorio:
    """Snapshot ejecutivo: counts + próximos 30d + vencidos + tasa cumplimiento."""
    counts_rows = (
        await db.execute(
            text(
                "SELECT estado, COUNT(*) AS n FROM app.entregables_regulatorios "
                "GROUP BY estado"
            )
        )
    ).mappings().all()
    counts = EntregableEstadosCounts()
    for r in counts_rows:
        if hasattr(counts, r["estado"]):
            setattr(counts, r["estado"], r["n"])

    proximos = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE estado IN ('pendiente','en_proceso') "
                "AND fecha_limite >= CURRENT_DATE "
                "AND fecha_limite <= (CURRENT_DATE + INTERVAL '30 days') "
                "ORDER BY fecha_limite ASC"
            )
        )
    ).mappings().all()

    vencidos = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE estado IN ('pendiente','en_proceso') "
                "AND fecha_limite < CURRENT_DATE "
                "ORDER BY fecha_limite ASC"
            )
        )
    ).mappings().all()

    # Tasa de cumplimiento YTD: del 1 enero del año actual hasta hoy.
    today = date.today()
    ytd_total_row = (
        await db.execute(
            text(
                "SELECT COUNT(*) AS n FROM app.entregables_regulatorios "
                "WHERE fecha_limite >= make_date(EXTRACT(year FROM CURRENT_DATE)::int, 1, 1) "
                "AND fecha_limite < CURRENT_DATE"
            )
        )
    ).first()
    ytd_entregados_row = (
        await db.execute(
            text(
                "SELECT COUNT(*) AS n FROM app.entregables_regulatorios "
                "WHERE fecha_limite >= make_date(EXTRACT(year FROM CURRENT_DATE)::int, 1, 1) "
                "AND fecha_limite < CURRENT_DATE "
                "AND estado = 'entregado'"
            )
        )
    ).first()

    total_ytd = ytd_total_row[0] if ytd_total_row else 0
    entregados_ytd = ytd_entregados_row[0] if ytd_entregados_row else 0
    tasa = (entregados_ytd / total_ytd * 100.0) if total_ytd > 0 else 100.0

    return ReporteRegulatorio(
        generado_at=datetime.now(UTC),
        estados=counts,
        proximos_30d=[_row_to_read(dict(r)) for r in proximos],
        vencidos_sin_entregar=[_row_to_read(dict(r)) for r in vencidos],
        tasa_cumplimiento_ytd=round(tasa, 1),
        total_ytd=total_ytd,
        entregados_ytd=entregados_ytd,
    )


# ---------------------------------------------------------------------------
# GET /{id} — detalle
# ---------------------------------------------------------------------------
@router.get("/{entregable_id}", response_model=EntregableRead)
async def get_entregable(
    user: CurrentUser, db: DBSession, entregable_id: int
) -> EntregableRead:
    row = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = :id"
            ),
            {"id": entregable_id},
        )
    ).mappings().first()
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Entregable no encontrado"
        )
    return _row_to_read(dict(row))


# ---------------------------------------------------------------------------
# POST / — crear entregable nuevo
# ---------------------------------------------------------------------------
@router.post("", response_model=EntregableRead, status_code=status.HTTP_201_CREATED)
async def create_entregable(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    body: EntregableCreate,
) -> EntregableRead:
    """Crea un entregable manual. Reusa scope `audit:read` (admin) — la
    creación de entregables regulatorios es operación de gobernanza."""
    res = await db.execute(
        text(
            """
            INSERT INTO app.entregables_regulatorios (
                id_template, nombre, descripcion, categoria, subcategoria,
                referencia_normativa, fecha_limite, frecuencia, prioridad,
                responsable, estado, periodo, alerta_15, alerta_10, alerta_5,
                notas, adjunto_url, generado_automaticamente, es_publico, extra,
                created_by
            ) VALUES (
                :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                :referencia_normativa, :fecha_limite, :frecuencia, :prioridad,
                :responsable, :estado, :periodo, :alerta_15, :alerta_10, :alerta_5,
                :notas, :adjunto_url, :generado_automaticamente, FALSE, :extra,
                :uid
            ) RETURNING entregable_id
            """
        ),
        {
            "id_template": body.id_template,
            "nombre": body.nombre,
            "descripcion": body.descripcion,
            "categoria": body.categoria,
            "subcategoria": body.subcategoria,
            "referencia_normativa": body.referencia_normativa,
            "fecha_limite": body.fecha_limite,
            "frecuencia": body.frecuencia,
            "prioridad": body.prioridad,
            "responsable": body.responsable,
            "estado": body.estado,
            "periodo": body.periodo,
            "alerta_15": body.alerta_15,
            "alerta_10": body.alerta_10,
            "alerta_5": body.alerta_5,
            "notas": body.notas,
            "adjunto_url": body.adjunto_url,
            "generado_automaticamente": False,
            "extra": body.extra,
            "uid": user.sub,
        },
    )
    new_id = res.scalar_one()
    await db.commit()

    row = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = :id"
            ),
            {"id": new_id},
        )
    ).mappings().one()
    created = _row_to_read(dict(row))

    await audit_log(
        db, request, user,
        action="create",
        entity_type="entregable",
        entity_id=str(new_id),
        entity_label=f"{body.categoria} · {body.periodo}",
        summary=f"Creó entregable {body.id_template} ({body.periodo})",
        before=None,
        after=created.model_dump(mode="json"),
    )
    return created


# ---------------------------------------------------------------------------
# PATCH /{id} — actualizar (estado, notas, adjunto, etc.)
# ---------------------------------------------------------------------------
@router.patch("/{entregable_id}", response_model=EntregableRead)
async def update_entregable(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    entregable_id: int,
    body: EntregableUpdate,
) -> EntregableRead:
    before_row = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = :id"
            ),
            {"id": entregable_id},
        )
    ).mappings().first()
    if not before_row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Entregable no encontrado"
        )
    before = _row_to_read(dict(before_row)).model_dump(mode="json")

    fields = body.model_dump(exclude_unset=True)
    if not fields:
        return _row_to_read(dict(before_row))

    # Si marcaron como entregado y no enviaron fecha_entrega_real, default a hoy
    if fields.get("estado") == "entregado" and "fecha_entrega_real" not in fields:
        fields["fecha_entrega_real"] = date.today()

    sets = [f"{k} = :{k}" for k in fields]
    sets.append("updated_at = now()")
    sets.append("updated_by = :uid")
    params = dict(fields)
    params["id"] = entregable_id
    params["uid"] = user.sub

    await db.execute(
        text(
            f"UPDATE app.entregables_regulatorios SET {', '.join(sets)} "  # noqa: S608
            "WHERE entregable_id = :id"
        ),
        params,
    )
    await db.commit()

    after_row = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = :id"
            ),
            {"id": entregable_id},
        )
    ).mappings().one()
    updated = _row_to_read(dict(after_row))

    await audit_log(
        db, request, user,
        action="update",
        entity_type="entregable",
        entity_id=str(entregable_id),
        entity_label=f"{updated.categoria} · {updated.periodo}",
        summary=(
            f"Actualizó {updated.id_template} → estado={updated.estado}"
            if "estado" in fields
            else f"Editó {updated.id_template}"
        ),
        before=before,
        after=updated.model_dump(mode="json"),
    )

    # ── Auto-generación próximo período ──────────────────────────────
    # Si marcamos un recurrente como entregado, intentamos crear la próxima
    # instancia para que el pipeline operativo no pierda continuidad.
    # Idempotente: ON CONFLICT (id_template, periodo) DO NOTHING.
    if (
        fields.get("estado") == "entregado"
        and updated.frecuencia in ("mensual", "trimestral", "semestral", "anual", "bienal")
    ):
        nxt = _next_periodo_y_fecha(
            updated.frecuencia, updated.periodo, updated.fecha_limite
        )
        if nxt is not None:
            next_periodo, next_fecha = nxt
            res = await db.execute(
                text(
                    """
                    INSERT INTO app.entregables_regulatorios (
                        id_template, nombre, descripcion, categoria, subcategoria,
                        referencia_normativa, fecha_limite, frecuencia, prioridad,
                        responsable, periodo, alerta_15, alerta_10, alerta_5,
                        generado_automaticamente, es_publico, created_by
                    ) VALUES (
                        :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                        :referencia_normativa, :fecha, :frecuencia, :prioridad,
                        :responsable, :periodo, :a15, :a10, :a5,
                        TRUE, FALSE, :uid
                    )
                    ON CONFLICT (id_template, periodo) DO NOTHING
                    RETURNING entregable_id
                    """
                ),
                {
                    "id_template": updated.id_template,
                    "nombre": updated.nombre,
                    "descripcion": updated.descripcion,
                    "categoria": updated.categoria,
                    "subcategoria": updated.subcategoria,
                    "referencia_normativa": updated.referencia_normativa,
                    "fecha": next_fecha,
                    "frecuencia": updated.frecuencia,
                    "prioridad": updated.prioridad,
                    "responsable": updated.responsable,
                    "periodo": next_periodo,
                    "a15": updated.alerta_15,
                    "a10": updated.alerta_10,
                    "a5": updated.alerta_5,
                    "uid": user.sub,
                },
            )
            new_row = res.first()
            await db.commit()
            if new_row is not None:
                await audit_log(
                    db, request, user,
                    action="create",
                    entity_type="entregable",
                    entity_id=str(new_row[0]),
                    entity_label=f"{updated.categoria} · {next_periodo}",
                    summary=(
                        f"Auto-generó próximo período {updated.id_template} "
                        f"({next_periodo}) tras marcar {updated.periodo} entregado"
                    ),
                    before=None,
                    after={
                        "id_template": updated.id_template,
                        "periodo": next_periodo,
                        "fecha_limite": next_fecha.isoformat(),
                    },
                )

    return updated


# ---------------------------------------------------------------------------
# POST /bulk-update — marca varios entregables en una sola transacción
# ---------------------------------------------------------------------------
@router.post("/bulk-update", response_model=BulkUpdateResponse)
async def bulk_update_entregables(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    body: BulkUpdateRequest,
) -> BulkUpdateResponse:
    """Cambia el estado de varios entregables a la vez.

    Reusa la misma lógica que el PATCH single (incluye auto-generación
    del próximo período si es recurrente y se marca entregado). Audita
    cada cambio individualmente para mantener el rastro completo.

    Si `estado='entregado'` y no se manda `fecha_entrega_real`, default
    a hoy (igual que el PATCH single).
    """
    fecha_real = body.fecha_entrega_real or (
        date.today() if body.estado == "entregado" else None
    )

    # Cargar todos los entregables solicitados de un golpe
    rows = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = ANY(:ids)"
            ),
            {"ids": body.ids},
        )
    ).mappings().all()
    by_id = {r["entregable_id"]: r for r in rows}

    requested_set = set(body.ids)
    found_set = set(by_id.keys())
    not_found = sorted(requested_set - found_set)

    updated_ids: list[int] = []
    already_target: list[int] = []
    auto_generated = 0

    for eid in body.ids:
        r = by_id.get(eid)
        if r is None:
            continue
        if r["estado"] == body.estado:
            already_target.append(eid)
            continue

        before = _row_to_read(dict(r)).model_dump(mode="json")

        await db.execute(
            text(
                """
                UPDATE app.entregables_regulatorios
                SET estado = :estado,
                    fecha_entrega_real = :fecha_real,
                    motivo_no_entrega = :motivo,
                    notas = COALESCE(:notas, notas),
                    adjunto_url = COALESCE(:adjunto, adjunto_url),
                    updated_at = now(),
                    updated_by = :uid
                WHERE entregable_id = :id
                """
            ),
            {
                "estado": body.estado,
                "fecha_real": (
                    fecha_real if body.estado == "entregado"
                    else r["fecha_entrega_real"]
                ),
                "motivo": (
                    body.motivo_no_entrega if body.estado == "no_entregado"
                    else None
                ),
                "notas": body.notas,
                "adjunto": body.adjunto_url,
                "id": eid,
                "uid": user.sub,
            },
        )
        updated_ids.append(eid)

        # Auto-generación próximo período (mismo flujo que PATCH single)
        if body.estado == "entregado" and r["frecuencia"] in (
            "mensual", "trimestral", "semestral", "anual", "bienal"
        ):
            nxt = _next_periodo_y_fecha(
                r["frecuencia"], r["periodo"], r["fecha_limite"]
            )
            if nxt is not None:
                next_periodo, next_fecha = nxt
                gen = await db.execute(
                    text(
                        """
                        INSERT INTO app.entregables_regulatorios (
                            id_template, nombre, descripcion, categoria, subcategoria,
                            referencia_normativa, fecha_limite, frecuencia, prioridad,
                            responsable, periodo, alerta_15, alerta_10, alerta_5,
                            generado_automaticamente, es_publico, created_by
                        ) VALUES (
                            :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                            :referencia_normativa, :fecha, :frecuencia, :prioridad,
                            :responsable, :periodo, :a15, :a10, :a5,
                            TRUE, FALSE, :uid
                        )
                        ON CONFLICT (id_template, periodo) DO NOTHING
                        RETURNING entregable_id
                        """
                    ),
                    {
                        "id_template": r["id_template"],
                        "nombre": r["nombre"],
                        "descripcion": r["descripcion"],
                        "categoria": r["categoria"],
                        "subcategoria": r["subcategoria"],
                        "referencia_normativa": r["referencia_normativa"],
                        "fecha": next_fecha,
                        "frecuencia": r["frecuencia"],
                        "prioridad": r["prioridad"],
                        "responsable": r["responsable"],
                        "periodo": next_periodo,
                        "a15": r["alerta_15"],
                        "a10": r["alerta_10"],
                        "a5": r["alerta_5"],
                        "uid": user.sub,
                    },
                )
                if gen.first() is not None:
                    auto_generated += 1

        # Audit por cada uno individualmente para preservar trazabilidad
        await audit_log(
            db, request, user,
            action="update",
            entity_type="entregable",
            entity_id=str(eid),
            entity_label=f"{r['categoria']} · {r['periodo']}",
            summary=(
                f"Bulk-update {r['id_template']} → estado={body.estado} "
                f"(parte de batch de {len(body.ids)})"
            ),
            before=before,
            after={"estado": body.estado},
        )

    await db.commit()

    return BulkUpdateResponse(
        requested=len(body.ids),
        updated_ids=updated_ids,
        already_target=already_target,
        not_found=not_found,
        auto_generated_next_periods=auto_generated,
    )


# ---------------------------------------------------------------------------
# POST /bulk-reassign — cambia responsable de varios entregables a la vez
# ---------------------------------------------------------------------------
@router.post("/bulk-reassign", response_model=BulkReassignResponse)
async def bulk_reassign_entregables(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    body: BulkReassignRequest,
) -> BulkReassignResponse:
    """Reassigna varios entregables a un nuevo responsable en bloque.

    Caso de uso típico: el equipo legal cambia, todos los CMF pasan a otro
    contacto, etc. Audita cada cambio individualmente para preservar
    trazabilidad por entregable.
    """
    rows = (
        await db.execute(
            text(
                "SELECT entregable_id, id_template, categoria, periodo, "
                "responsable FROM app.entregables_regulatorios "
                "WHERE entregable_id = ANY(:ids)"
            ),
            {"ids": body.ids},
        )
    ).mappings().all()
    by_id = {int(r["entregable_id"]): r for r in rows}

    requested_set = set(body.ids)
    found_set = set(by_id.keys())
    not_found = sorted(requested_set - found_set)
    updated_ids: list[int] = []

    for eid in body.ids:
        r = by_id.get(eid)
        if r is None:
            continue
        if r["responsable"] == body.responsable:
            continue  # ya estaba asignado a ese — skip
        await db.execute(
            text(
                """
                UPDATE app.entregables_regulatorios
                SET responsable = :resp,
                    updated_at = now(),
                    updated_by = :uid
                WHERE entregable_id = :id
                """
            ),
            {"resp": body.responsable, "uid": user.sub, "id": eid},
        )
        updated_ids.append(eid)

        await audit_log(
            db, request, user,
            action="update",
            entity_type="entregable",
            entity_id=str(eid),
            entity_label=f"{r['categoria']} · {r['periodo']}",
            summary=(
                f"Bulk-reassign {r['id_template']}: "
                f"{r['responsable']} → {body.responsable}"
            ),
            before={"responsable": r["responsable"]},
            after={"responsable": body.responsable},
        )

    await db.commit()
    return BulkReassignResponse(
        requested=len(body.ids),
        updated_ids=updated_ids,
        not_found=not_found,
    )


# ---------------------------------------------------------------------------
# GET /calendar.ics — feed iCalendar para Google Cal / Outlook / Apple Cal
# ---------------------------------------------------------------------------
def _ics_escape(text_val: str) -> str:
    """Escapa según RFC 5545 — comas, semicolons, backslashes, newlines."""
    return (
        text_val.replace("\\", "\\\\")
        .replace(",", "\\,")
        .replace(";", "\\;")
        .replace("\n", "\\n")
        .replace("\r", "")
    )


@router.get("/calendar.ics", response_class=PlainTextResponse)
async def entregables_ics(
    user: CurrentUser,
    db: DBSession,
    desde: date | None = Query(
        None, description="Default: hoy - 30 días (para que aparezcan vencidos)"
    ),
    hasta: date | None = Query(
        None, description="Default: hoy + 365 días"
    ),
    estado: EstadoEntregable | None = None,
    empresa: str | None = None,
) -> PlainTextResponse:
    """Devuelve un feed iCalendar (RFC 5545) con los entregables.

    Para suscribirse desde Google Calendar / Outlook / Apple Calendar:
        Settings → Add calendar by URL →
        https://api.../entregables/calendar.ics?...

    Cada entregable es un VEVENT all-day con:
        - SUMMARY: [{categoria}] {nombre}
        - DESCRIPTION: período + responsable + ref normativa + nivel alerta
        - DTSTART: fecha_limite (date-only)
        - UID estable: entregable-{id}@cehta.cl  → al re-sync, el cliente
          solo actualiza lo cambiado en lugar de duplicar.
        - CATEGORIES: la categoría regulatoria (CMF/CORFO/UAF/...)
        - STATUS: CONFIRMED por default, CANCELLED si ya entregado.

    Filtros opcionales mismo nombre que `GET /entregables`.
    """
    today = date.today()
    desde_real = desde or (today - timedelta(days=30))
    hasta_real = hasta or (today + timedelta(days=365))

    conditions = ["fecha_limite >= :desde", "fecha_limite <= :hasta"]
    params: dict[str, Any] = {"desde": desde_real, "hasta": hasta_real}
    if estado:
        conditions.append("estado = :estado")
        params["estado"] = estado
    if empresa:
        conditions.append(
            "(subcategoria ILIKE :emp OR extra->>'empresa_codigo' ILIKE :emp)"
        )
        params["emp"] = f"%{empresa}%"

    where = "WHERE " + " AND ".join(conditions)
    rows = (
        await db.execute(
            text(
                "SELECT entregable_id, nombre, categoria, subcategoria, "
                "periodo, fecha_limite, responsable, estado, "
                "referencia_normativa, descripcion "
                f"FROM app.entregables_regulatorios {where} "
                "ORDER BY fecha_limite ASC LIMIT 5000"
            ),
            params,
        )
    ).mappings().all()

    # Build ICS content
    now_utc = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    lines: list[str] = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//FIP CEHTA ESG//Entregables//ES",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Entregables FIP CEHTA",
        "X-WR-CALDESC:Calendario regulatorio CMF/CORFO/UAF/SII + Reglamento Interno",
        "X-WR-TIMEZONE:America/Santiago",
    ]

    for r in rows:
        fecha = r["fecha_limite"]
        # all-day event: DTSTART;VALUE=DATE en formato YYYYMMDD; DTEND es +1
        dt_start = fecha.strftime("%Y%m%d")
        dt_end = (fecha + timedelta(days=1)).strftime("%Y%m%d")
        uid = f"entregable-{r['entregable_id']}@cehta.cl"
        cat = str(r["categoria"])
        nombre = str(r["nombre"])
        summary = f"[{cat}] {nombre}"
        descripcion_parts: list[str] = []
        descripcion_parts.append(f"Período: {r['periodo']}")
        descripcion_parts.append(f"Responsable: {r['responsable']}")
        descripcion_parts.append(f"Estado: {r['estado']}")
        if r.get("subcategoria"):
            descripcion_parts.append(f"Empresa/Subcategoría: {r['subcategoria']}")
        if r.get("referencia_normativa"):
            descripcion_parts.append(
                f"Ref normativa: {r['referencia_normativa']}"
            )
        if r.get("descripcion"):
            descripcion_parts.append(f"\\n{r['descripcion']}")
        descripcion = " | ".join(descripcion_parts)
        ics_status = "CANCELLED" if r["estado"] == "entregado" else "CONFIRMED"

        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{now_utc}",
                f"DTSTART;VALUE=DATE:{dt_start}",
                f"DTEND;VALUE=DATE:{dt_end}",
                f"SUMMARY:{_ics_escape(summary)}",
                f"DESCRIPTION:{_ics_escape(descripcion)}",
                f"CATEGORIES:{_ics_escape(cat)}",
                f"STATUS:{ics_status}",
                "TRANSP:TRANSPARENT",  # all-day no bloquea slot del usuario
                "END:VEVENT",
            ]
        )

    lines.append("END:VCALENDAR")
    # RFC 5545 dice CRLF; ningún parser razonable se rompe por LF, pero por
    # las dudas usamos CRLF como manda la spec.
    body = "\r\n".join(lines) + "\r\n"
    return PlainTextResponse(
        content=body,
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": (
                'inline; filename="entregables-fip-cehta.ics"'
            ),
            "Cache-Control": "private, max-age=300",  # 5 min cache
        },
    )


# ---------------------------------------------------------------------------
# DELETE /{id}
# ---------------------------------------------------------------------------
@router.delete("/{entregable_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_entregable(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    entregable_id: int,
) -> Response:
    before_row = (
        await db.execute(
            text(
                f"SELECT {_FIELDS} FROM app.entregables_regulatorios "
                "WHERE entregable_id = :id"
            ),
            {"id": entregable_id},
        )
    ).mappings().first()
    if not before_row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Entregable no encontrado"
        )
    before = _row_to_read(dict(before_row)).model_dump(mode="json")

    await db.execute(
        text("DELETE FROM app.entregables_regulatorios WHERE entregable_id = :id"),
        {"id": entregable_id},
    )
    await db.commit()

    await audit_log(
        db, request, user,
        action="delete",
        entity_type="entregable",
        entity_id=str(entregable_id),
        entity_label=f"{before['categoria']} · {before['periodo']}",
        summary=f"Eliminó entregable {before['id_template']} ({before['periodo']})",
        before=before,
        after=None,
    )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# POST /serie — generar instancias de un template recurrente para un año
# ---------------------------------------------------------------------------
def _fechas_del_anio(frecuencia: str, anio: int) -> list[tuple[str, date]]:
    """Genera (periodo, fecha_limite) según frecuencia y año.

    Las fechas son aproximaciones razonables — el usuario puede luego
    ajustar instancia por instancia via PATCH si una fecha específica
    necesita corregirse (ej. feriado).
    """
    import calendar as cal

    out: list[tuple[str, date]] = []
    if frecuencia == "mensual":
        for m in range(1, 13):
            last = cal.monthrange(anio, m)[1]
            out.append((f"{m:02d}-{anio}", date(anio, m, last)))
    elif frecuencia == "trimestral":
        # 30 días después del cierre de cada trimestre
        for q, m_close in [(1, 3), (2, 6), (3, 9), (4, 12)]:
            last_close = cal.monthrange(anio, m_close)[1]
            close = date(anio, m_close, last_close)
            # 5to día hábil tras cierre — aproximamos a +30 días naturales
            from datetime import timedelta

            entrega = close + timedelta(days=30)
            out.append((f"Q{q}-{anio}", entrega))
    elif frecuencia == "semestral":
        out.append((f"S1-{anio}", date(anio, 7, 31)))
        out.append((f"S2-{anio}", date(anio + 1, 1, 31)))
    elif frecuencia == "anual":
        out.append((f"{anio}", date(anio + 1, 4, 30)))
    elif frecuencia == "bienal":
        out.append((f"{anio}-{anio + 1}", date(anio + 1, 12, 31)))
    return out


@router.post(
    "/serie",
    response_model=GenerarSerieResponse,
    status_code=status.HTTP_201_CREATED,
)
async def generar_serie(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    body: GenerarSerieRequest,
) -> GenerarSerieResponse:
    """Genera todas las instancias de un template recurrente para un año.

    Idempotente — usa ON CONFLICT (id_template, periodo) DO NOTHING. Si
    ya hay instancias para ese año, se ignoran (cuenta vuelve en `instancias_existentes`).
    """
    fechas = _fechas_del_anio(body.frecuencia, body.anio)
    creadas = 0
    existentes = 0
    fechas_creadas: list[date] = []

    for periodo, fecha in fechas:
        res = await db.execute(
            text(
                """
                INSERT INTO app.entregables_regulatorios (
                    id_template, nombre, descripcion, categoria, subcategoria,
                    referencia_normativa, fecha_limite, frecuencia, prioridad,
                    responsable, periodo, alerta_15, alerta_10, alerta_5,
                    generado_automaticamente, es_publico, created_by
                ) VALUES (
                    :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                    :referencia_normativa, :fecha, :frecuencia, :prioridad,
                    :responsable, :periodo, :a15, :a10, :a5,
                    TRUE, FALSE, :uid
                )
                ON CONFLICT (id_template, periodo) DO NOTHING
                RETURNING entregable_id
                """
            ),
            {
                "id_template": body.id_template,
                "nombre": body.nombre,
                "descripcion": body.descripcion,
                "categoria": body.categoria,
                "subcategoria": body.subcategoria,
                "referencia_normativa": body.referencia_normativa,
                "fecha": fecha,
                "frecuencia": body.frecuencia,
                "prioridad": body.prioridad,
                "responsable": body.responsable,
                "periodo": periodo,
                "a15": body.alerta_15,
                "a10": body.alerta_10,
                "a5": body.alerta_5,
                "uid": user.sub,
            },
        )
        row = res.first()
        if row is not None:
            creadas += 1
            fechas_creadas.append(fecha)
        else:
            existentes += 1
    await db.commit()

    await audit_log(
        db, request, user,
        action="create",
        entity_type="entregable_serie",
        entity_id=f"{body.id_template}:{body.anio}",
        entity_label=f"{body.categoria} · serie {body.anio}",
        summary=(
            f"Generó serie {body.id_template} {body.anio}: "
            f"{creadas} creadas, {existentes} ya existentes"
        ),
        before=None,
        after={"creadas": creadas, "existentes": existentes},
    )

    return GenerarSerieResponse(
        template=body.id_template,
        anio=body.anio,
        instancias_creadas=creadas,
        instancias_existentes=existentes,
        fechas=fechas_creadas,
    )


# ---------------------------------------------------------------------------
# POST /extend-forward — extiende todos los templates recurrentes hacia
#                        adelante si se está acercando el fin del año cubierto
# ---------------------------------------------------------------------------
@router.post("/extend-forward")
async def extend_forward(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    horizon_days: int = Query(
        90, ge=30, le=365,
        description="Si quedan ≤ horizon_days de instancias futuras, se "
                    "genera el año siguiente",
    ),
) -> dict[str, Any]:
    """Auto-extiende templates recurrentes hacia adelante.

    Para cada `id_template` recurrente, identifica la última instancia
    futura. Si está a menos de `horizon_days`, genera todas las instancias
    del año siguiente. Idempotente.

    Pensado para correr desde un cron mensual (ej. cada 1ro del mes).
    Así nunca llegamos a fin de año sin entregables 2027/2028 cargados.
    """
    # Buscar últimas instancias por id_template (solo recurrentes)
    rows = (
        await db.execute(
            text(
                """
                SELECT
                  id_template,
                  MAX(fecha_limite) AS ultima_fecha,
                  -- tomamos un row representativo para los metadatos del template
                  (ARRAY_AGG(nombre ORDER BY fecha_limite DESC))[1] AS nombre,
                  (ARRAY_AGG(descripcion ORDER BY fecha_limite DESC))[1] AS descripcion,
                  (ARRAY_AGG(categoria ORDER BY fecha_limite DESC))[1] AS categoria,
                  (ARRAY_AGG(subcategoria ORDER BY fecha_limite DESC))[1] AS subcategoria,
                  (ARRAY_AGG(referencia_normativa ORDER BY fecha_limite DESC))[1]
                      AS referencia_normativa,
                  (ARRAY_AGG(frecuencia ORDER BY fecha_limite DESC))[1] AS frecuencia,
                  (ARRAY_AGG(prioridad ORDER BY fecha_limite DESC))[1] AS prioridad,
                  (ARRAY_AGG(responsable ORDER BY fecha_limite DESC))[1] AS responsable,
                  (ARRAY_AGG(alerta_15 ORDER BY fecha_limite DESC))[1] AS alerta_15,
                  (ARRAY_AGG(alerta_10 ORDER BY fecha_limite DESC))[1] AS alerta_10,
                  (ARRAY_AGG(alerta_5 ORDER BY fecha_limite DESC))[1] AS alerta_5
                FROM app.entregables_regulatorios
                WHERE frecuencia IN ('mensual','trimestral','semestral','anual','bienal')
                GROUP BY id_template
                """
            )
        )
    ).mappings().all()

    today = date.today()
    templates_extendidos = 0
    instancias_creadas = 0
    detalle: list[dict[str, Any]] = []

    for r in rows:
        ultima: date = r["ultima_fecha"]
        dias_remaining = (ultima - today).days
        if dias_remaining > horizon_days:
            continue  # sigue cubierto, no necesitamos extender

        # Año a generar: el siguiente al de la última instancia
        anio_target = ultima.year + 1
        if anio_target > 2030:
            continue  # safeguard del CHECK constraint del schema

        fechas = _fechas_del_anio(r["frecuencia"], anio_target)
        creadas_este = 0
        for periodo, fecha in fechas:
            res = await db.execute(
                text(
                    """
                    INSERT INTO app.entregables_regulatorios (
                        id_template, nombre, descripcion, categoria, subcategoria,
                        referencia_normativa, fecha_limite, frecuencia, prioridad,
                        responsable, periodo, alerta_15, alerta_10, alerta_5,
                        generado_automaticamente, es_publico, created_by
                    ) VALUES (
                        :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                        :referencia_normativa, :fecha, :frecuencia, :prioridad,
                        :responsable, :periodo, :a15, :a10, :a5,
                        TRUE, FALSE, :uid
                    )
                    ON CONFLICT (id_template, periodo) DO NOTHING
                    RETURNING entregable_id
                    """
                ),
                {
                    "id_template": r["id_template"],
                    "nombre": r["nombre"],
                    "descripcion": r["descripcion"],
                    "categoria": r["categoria"],
                    "subcategoria": r["subcategoria"],
                    "referencia_normativa": r["referencia_normativa"],
                    "fecha": fecha,
                    "frecuencia": r["frecuencia"],
                    "prioridad": r["prioridad"],
                    "responsable": r["responsable"],
                    "periodo": periodo,
                    "a15": r["alerta_15"],
                    "a10": r["alerta_10"],
                    "a5": r["alerta_5"],
                    "uid": user.sub,
                },
            )
            if res.first() is not None:
                creadas_este += 1

        if creadas_este > 0:
            templates_extendidos += 1
            instancias_creadas += creadas_este
            detalle.append(
                {
                    "id_template": r["id_template"],
                    "anio_extendido": anio_target,
                    "instancias_creadas": creadas_este,
                }
            )

    await db.commit()

    if templates_extendidos > 0:
        await audit_log(
            db, request, user,
            action="create",
            entity_type="entregable_extend_forward",
            entity_id=f"horizon-{horizon_days}d",
            entity_label="Extensión forward de templates recurrentes",
            summary=(
                f"Auto-extensión: {templates_extendidos} templates · "
                f"{instancias_creadas} instancias nuevas"
            ),
            before=None,
            after={
                "horizon_days": horizon_days,
                "templates_extendidos": templates_extendidos,
                "instancias_creadas": instancias_creadas,
                "detalle": detalle,
            },
        )

    return {
        "horizon_days": horizon_days,
        "templates_evaluated": len(rows),
        "templates_extendidos": templates_extendidos,
        "instancias_creadas": instancias_creadas,
        "detalle": detalle,
    }


# ---------------------------------------------------------------------------
# POST /import-csv — sube un CSV con entregables (batch loads)
# ---------------------------------------------------------------------------
_CSV_REQUIRED_FIELDS = (
    "id_template", "nombre", "categoria", "fecha_limite",
    "frecuencia", "prioridad", "responsable", "periodo",
)
_CSV_OPTIONAL_FIELDS = (
    "descripcion", "subcategoria", "referencia_normativa",
    "estado", "notas", "adjunto_url", "alerta_15", "alerta_10", "alerta_5",
    "empresa_codigo",
)
_CSV_VALID_FIELDS = (*_CSV_REQUIRED_FIELDS, *_CSV_OPTIONAL_FIELDS)
_CATEGORIAS_VALIDAS = {
    "CMF", "CORFO", "UAF", "SII",
    "INTERNO", "AUDITORIA", "ASAMBLEA", "OPERACIONAL",
}
_FRECUENCIAS_VALIDAS = {
    "mensual", "trimestral", "semestral",
    "anual", "bienal", "unico", "segun_evento",
}
_PRIORIDADES_VALIDAS = {"critica", "alta", "media", "baja"}
_ESTADOS_VALIDOS = {"pendiente", "en_proceso", "entregado", "no_entregado"}


def _parse_csv_bool(val: str) -> bool:
    return val.strip().lower() in ("true", "1", "yes", "si", "sí", "y")


@router.post("/import-csv", response_model=CsvImportResponse)
async def import_entregables_csv(
    user: Annotated[AuthenticatedUser, Depends(require_scope("audit:read"))],
    db: DBSession,
    request: Request,
    file: UploadFile,
) -> CsvImportResponse:
    """Importa entregables desde un CSV.

    Headers requeridos: id_template, nombre, categoria, fecha_limite,
    frecuencia, prioridad, responsable, periodo.
    Headers opcionales: descripcion, subcategoria, referencia_normativa,
    estado (default 'pendiente'), notas, adjunto_url, alerta_15/10/5
    (default true), empresa_codigo (se guarda en `extra->>'empresa_codigo'`).

    fecha_limite: ISO `YYYY-MM-DD`.
    Idempotente vía UNIQUE (id_template, periodo) — filas duplicadas se
    cuentan como `rows_skipped`, no fallan.

    Validación por fila — si falla, sigue con las demás. Errores se
    devuelven en `errors[]` con número de fila para que el operador pueda
    corregir el CSV y reintentar.
    """
    import csv
    import io
    import json as json_module

    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Archivo debe ser .csv",
        )

    raw = await file.read()
    try:
        text_data = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text_data = raw.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se pudo decodificar (probá UTF-8)",
            ) from exc

    reader = csv.DictReader(io.StringIO(text_data))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV vacío o sin header",
        )

    headers = [h.strip() for h in reader.fieldnames]
    missing = [f for f in _CSV_REQUIRED_FIELDS if f not in headers]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan headers requeridos: {', '.join(missing)}",
        )

    rows_imported = 0
    rows_skipped = 0
    rows_failed = 0
    errors: list[CsvImportError] = []
    sample_imported: list[int] = []

    rows_data = list(reader)
    rows_received = len(rows_data)

    for idx, row in enumerate(rows_data, start=1):
        row = {(k or "").strip(): (v or "").strip() for k, v in row.items()}

        # Validar campos requeridos no vacíos
        empty_required = [f for f in _CSV_REQUIRED_FIELDS if not row.get(f)]
        if empty_required:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Campos vacíos: {', '.join(empty_required)}",
                    raw=row,
                )
            )
            continue

        # Validar enums
        cat = row["categoria"]
        if cat not in _CATEGORIAS_VALIDAS:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Categoría inválida '{cat}'. Válidas: {sorted(_CATEGORIAS_VALIDAS)}",
                    raw=row,
                )
            )
            continue
        freq = row["frecuencia"]
        if freq not in _FRECUENCIAS_VALIDAS:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Frecuencia inválida '{freq}'",
                    raw=row,
                )
            )
            continue
        prio = row["prioridad"]
        if prio not in _PRIORIDADES_VALIDAS:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Prioridad inválida '{prio}'",
                    raw=row,
                )
            )
            continue
        estado_val = row.get("estado", "pendiente") or "pendiente"
        if estado_val not in _ESTADOS_VALIDOS:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Estado inválido '{estado_val}'",
                    raw=row,
                )
            )
            continue

        # Parsear fecha
        try:
            fecha = date.fromisoformat(row["fecha_limite"])
        except ValueError:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Fecha inválida '{row['fecha_limite']}' (esperado YYYY-MM-DD)",
                    raw=row,
                )
            )
            continue

        # Construir extra JSON con empresa_codigo si viene
        extra: dict[str, Any] = {}
        if row.get("empresa_codigo"):
            extra["empresa_codigo"] = row["empresa_codigo"]

        # INSERT idempotente
        try:
            res = await db.execute(
                text(
                    """
                    INSERT INTO app.entregables_regulatorios (
                        id_template, nombre, descripcion, categoria, subcategoria,
                        referencia_normativa, fecha_limite, frecuencia, prioridad,
                        responsable, estado, periodo, alerta_15, alerta_10, alerta_5,
                        notas, adjunto_url, generado_automaticamente, es_publico,
                        extra, created_by
                    ) VALUES (
                        :id_template, :nombre, :descripcion, :categoria, :subcategoria,
                        :referencia_normativa, :fecha, :frecuencia, :prioridad,
                        :responsable, :estado, :periodo, :a15, :a10, :a5,
                        :notas, :adjunto, FALSE, FALSE,
                        CAST(:extra AS jsonb), :uid
                    )
                    ON CONFLICT (id_template, periodo) DO NOTHING
                    RETURNING entregable_id
                    """
                ),
                {
                    "id_template": row["id_template"],
                    "nombre": row["nombre"],
                    "descripcion": row.get("descripcion") or None,
                    "categoria": cat,
                    "subcategoria": row.get("subcategoria") or None,
                    "referencia_normativa": row.get("referencia_normativa") or None,
                    "fecha": fecha,
                    "frecuencia": freq,
                    "prioridad": prio,
                    "responsable": row["responsable"],
                    "estado": estado_val,
                    "periodo": row["periodo"],
                    "a15": _parse_csv_bool(row.get("alerta_15", "true")),
                    "a10": _parse_csv_bool(row.get("alerta_10", "true")),
                    "a5": _parse_csv_bool(row.get("alerta_5", "true")),
                    "notas": row.get("notas") or None,
                    "adjunto": row.get("adjunto_url") or None,
                    "extra": json_module.dumps(extra) if extra else None,
                    "uid": user.sub,
                },
            )
        except Exception as exc:
            rows_failed += 1
            errors.append(
                CsvImportError(
                    row=idx,
                    error=f"Error SQL: {type(exc).__name__}: {exc}"[:500],
                    raw=row,
                )
            )
            continue

        new_row = res.first()
        if new_row is None:
            rows_skipped += 1
        else:
            rows_imported += 1
            if len(sample_imported) < 10:
                sample_imported.append(int(new_row[0]))

    await db.commit()

    await audit_log(
        db, request, user,
        action="create",
        entity_type="entregable_csv_import",
        entity_id=f"upload-{rows_received}",
        entity_label=f"CSV import {file.filename}",
        summary=(
            f"CSV import: {rows_imported} imported, "
            f"{rows_skipped} skipped (dup), {rows_failed} failed"
        ),
        before=None,
        after={
            "filename": file.filename,
            "rows_received": rows_received,
            "rows_imported": rows_imported,
            "rows_skipped": rows_skipped,
            "rows_failed": rows_failed,
        },
    )

    return CsvImportResponse(
        rows_received=rows_received,
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
        rows_failed=rows_failed,
        errors=errors[:50],  # cap para no bombardear el frontend
        sample_imported_ids=sample_imported,
    )
