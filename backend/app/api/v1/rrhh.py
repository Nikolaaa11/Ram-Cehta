"""R152vvv · Endpoints módulo RRHH.

Owners: Benjamín Toro (Adm. y Finanzas) + Victoria (allowlist en DB).

Recursos:
  GET    /rrhh/empleados                        catálogo
  GET    /rrhh/empleados/{rut}/costos           histórico costos por periodo
  GET    /rrhh/libros                           listar libros mensuales
  GET    /rrhh/libros/{libro_id}                detalle libro + líneas
  POST   /rrhh/libros/upload                    subir Excel libro
  GET    /rrhh/costo-empresa?empresa=&periodo=  costo agregado empresa-mes
  GET    /rrhh/access                           verifica si el user puede ver RRHH
"""

from __future__ import annotations

import logging
import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Annotated

import io
from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import text

from app.api.deps import CurrentUser, DBSession
from app.core.security import AuthenticatedUser
from app.services.libro_remuneraciones_parser import (
    LibroParseado,
    parse_libro_remuneraciones,
)

router = APIRouter()


# ─────────────────────────────────────────────────────────────────────
# Permisos: allowlist en core.rrhh_allowlist + admins
# ─────────────────────────────────────────────────────────────────────


async def _check_rrhh_access(user: AuthenticatedUser, db) -> None:
    """Lanza 403 si el user no es admin ni está en allowlist RRHH."""
    if getattr(user, "is_admin", False):
        return
    email = (getattr(user, "email", "") or "").lower().strip()
    if not email:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No autorizado para RRHH (sin email)",
        )
    row = await db.execute(
        text("SELECT 1 FROM core.rrhh_allowlist WHERE email = :e"),
        {"e": email},
    )
    if not row.first():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Módulo RRHH restringido a Benjamín, Victoria y admins. "
                "Para acceder pedile a un admin que te agregue a "
                "core.rrhh_allowlist."
            ),
        )


# ─────────────────────────────────────────────────────────────────────
# Schemas
# ─────────────────────────────────────────────────────────────────────


class EmpleadoRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    rut: str
    nombre: str
    empresa_codigo: str
    area: str | None
    cargo: str | None
    fecha_ingreso: str | None
    activo: bool
    sueldo_base_actual: Decimal | None


class LineaLibroRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    empleado_rut: str
    nombre: str
    area: str | None
    dias_trabajados: Decimal
    total_haberes: Decimal
    liquido_pagado: Decimal
    total_descuentos_legales: Decimal
    total_aportes_patronales: Decimal
    costo_total_empresa: Decimal
    # Detalles haberes
    sueldo_base: Decimal
    horas_extras: Decimal
    gratificacion_legal: Decimal
    asignacion_familiar: Decimal
    # Aportes patronales detallados
    aporte_afp_empleador: Decimal
    sis: Decimal
    seguro_cesantia_empleador: Decimal
    seguro_social: Decimal
    mutual: Decimal
    base_tributable: Decimal
    impuesto_unico: Decimal


class LibroRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    empresa_codigo: str
    periodo: str
    total_haberes: Decimal
    total_liquido: Decimal
    total_descuentos_legales: Decimal
    total_aportes_patronales: Decimal
    total_costo_empresa: Decimal
    cantidad_empleados: int
    archivo_origen: str | None
    uploaded_at: datetime


class LibroDetalle(LibroRead):
    lineas: list[LineaLibroRead]


class CostoHistorico(BaseModel):
    rut: str
    nombre: str
    periodo: str
    empresa_codigo: str
    area: str | None
    total_haberes: Decimal
    liquido_pagado: Decimal
    total_aportes_patronales: Decimal
    costo_total_empresa: Decimal
    multiplicador_costo: Decimal | None


class UploadResult(BaseModel):
    libro_id: int
    empresa_codigo: str
    periodo: str
    cantidad_empleados: int
    total_costo_empresa: Decimal
    reemplazo: bool  # True si había uno previo y se sobreescribió


class AccessResponse(BaseModel):
    allowed: bool
    reason: str | None = None


# ─────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────


@router.get("/access", response_model=AccessResponse)
async def check_access(user: CurrentUser, db: DBSession) -> AccessResponse:
    """Endpoint público (autenticado) que devuelve si el user puede ver /rrhh."""
    try:
        await _check_rrhh_access(user, db)
        return AccessResponse(allowed=True)
    except HTTPException as e:
        return AccessResponse(allowed=False, reason=e.detail)


@router.get("/empleados", response_model=list[EmpleadoRead])
async def list_empleados(
    user: CurrentUser,
    db: DBSession,
    empresa_codigo: str | None = Query(default=None),
    incluir_inactivos: bool = Query(default=False),
) -> list[EmpleadoRead]:
    await _check_rrhh_access(user, db)
    wheres = []
    params: dict = {}
    if empresa_codigo:
        wheres.append("empresa_codigo = :e")
        params["e"] = empresa_codigo
    if not incluir_inactivos:
        wheres.append("activo = TRUE")
    where_sql = (" WHERE " + " AND ".join(wheres)) if wheres else ""
    rows = await db.execute(
        text(
            f"""
            SELECT rut, nombre, empresa_codigo, area, cargo,
                   to_char(fecha_ingreso, 'YYYY-MM-DD') AS fecha_ingreso,
                   activo, sueldo_base_actual
            FROM core.empleados
            {where_sql}
            ORDER BY empresa_codigo, area NULLS LAST, nombre
            """
        ),
        params,
    )
    return [EmpleadoRead.model_validate(dict(r._mapping)) for r in rows]


@router.get("/empleados/{rut}/costos", response_model=list[CostoHistorico])
async def costos_empleado(
    user: CurrentUser, db: DBSession, rut: str
) -> list[CostoHistorico]:
    await _check_rrhh_access(user, db)
    rows = await db.execute(
        text(
            """
            SELECT rut, nombre, periodo, empresa_codigo, area,
                   total_haberes, liquido_pagado, total_aportes_patronales,
                   costo_total_empresa, multiplicador_costo
            FROM core.v_costo_empleado_mensual
            WHERE rut = :r
            ORDER BY periodo DESC
            """
        ),
        {"r": rut},
    )
    return [CostoHistorico.model_validate(dict(r._mapping)) for r in rows]


@router.get("/libros", response_model=list[LibroRead])
async def list_libros(
    user: CurrentUser,
    db: DBSession,
    empresa_codigo: str | None = Query(default=None),
    periodo: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}$"),
) -> list[LibroRead]:
    await _check_rrhh_access(user, db)
    wheres = []
    params: dict = {}
    if empresa_codigo:
        wheres.append("empresa_codigo = :e")
        params["e"] = empresa_codigo
    if periodo:
        wheres.append("periodo = :p")
        params["p"] = periodo
    where_sql = (" WHERE " + " AND ".join(wheres)) if wheres else ""
    rows = await db.execute(
        text(
            f"""
            SELECT id, empresa_codigo, periodo, total_haberes, total_liquido,
                   total_descuentos_legales, total_aportes_patronales,
                   total_costo_empresa, cantidad_empleados,
                   archivo_origen, uploaded_at
            FROM core.libros_remuneraciones
            {where_sql}
            ORDER BY periodo DESC, empresa_codigo
            """
        ),
        params,
    )
    return [LibroRead.model_validate(dict(r._mapping)) for r in rows]


@router.get("/libros/{libro_id}", response_model=LibroDetalle)
async def get_libro(
    user: CurrentUser, db: DBSession, libro_id: int
) -> LibroDetalle:
    await _check_rrhh_access(user, db)
    libro_row = (
        await db.execute(
            text(
                """
                SELECT id, empresa_codigo, periodo, total_haberes, total_liquido,
                       total_descuentos_legales, total_aportes_patronales,
                       total_costo_empresa, cantidad_empleados,
                       archivo_origen, uploaded_at
                FROM core.libros_remuneraciones
                WHERE id = :id
                """
            ),
            {"id": libro_id},
        )
    ).first()
    if not libro_row:
        raise HTTPException(status_code=404, detail="Libro no encontrado")
    lineas_rows = await db.execute(
        text(
            """
            SELECT id, empleado_rut, nombre, area, dias_trabajados,
                   total_haberes, liquido_pagado, total_descuentos_legales,
                   total_aportes_patronales, costo_total_empresa,
                   sueldo_base, horas_extras, gratificacion_legal,
                   asignacion_familiar,
                   aporte_afp_empleador, sis, seguro_cesantia_empleador,
                   seguro_social, mutual,
                   base_tributable, impuesto_unico
            FROM core.libro_remuneraciones_lineas
            WHERE libro_id = :id
            ORDER BY area NULLS LAST, nombre
            """
        ),
        {"id": libro_id},
    )
    lineas = [LineaLibroRead.model_validate(dict(r._mapping)) for r in lineas_rows]
    base = dict(libro_row._mapping)
    base["lineas"] = lineas
    return LibroDetalle.model_validate(base)


@router.post("/libros/upload", response_model=UploadResult)
async def upload_libro(
    user: CurrentUser,
    db: DBSession,
    file: Annotated[UploadFile, File(...)],
    empresa_codigo: Annotated[
        str, Query(description="Código empresa destino (RVT, AFIS, etc.)")
    ],
) -> UploadResult:
    """Sube un Excel libro de remuneraciones. Reemplaza el del periodo si existe."""
    await _check_rrhh_access(user, db)

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Sólo se aceptan archivos .xlsx / .xls"
        )

    # Verificar empresa
    e = await db.scalar(
        text("SELECT 1 FROM core.empresas WHERE codigo = :c AND activo = TRUE"),
        {"c": empresa_codigo},
    )
    if not e:
        raise HTTPException(
            status_code=400,
            detail=f"Empresa {empresa_codigo} no existe o está inactiva",
        )

    # R152EEEEEE — Cap defensivo de tamaño. Sin esto, un user con
    # permisos RRHH puede subir N×10GB hasta llenar el FS de la VM Fly
    # (512MB). Excel real de libro remuneraciones rara vez pasa 5MB.
    contents = await file.read()
    MAX_LIBRO_BYTES = 20 * 1024 * 1024  # 20MB
    if len(contents) > MAX_LIBRO_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"Archivo demasiado grande ({len(contents) / 1024 / 1024:.1f}MB). "
                f"Máximo: 20MB."
            ),
        )
    # Guardar tmp + parsear. delete=True con context manager para evitar
    # tmp leak si parse_libro_remuneraciones falla.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = Path(tmp.name)
    try:
        parsed = parse_libro_remuneraciones(tmp_path)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo parsear el Excel: {exc}",
        ) from exc
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    if not parsed.lineas:
        raise HTTPException(
            status_code=400,
            detail="El Excel no contiene líneas de empleados parseables",
        )

    # Reemplazo si ya existe libro de este periodo-empresa
    existing = await db.scalar(
        text(
            """SELECT id FROM core.libros_remuneraciones
               WHERE empresa_codigo = :c AND periodo = :p"""
        ),
        {"c": empresa_codigo, "p": parsed.periodo},
    )
    reemplazo = bool(existing)
    if existing:
        await db.execute(
            text("DELETE FROM core.libros_remuneraciones WHERE id = :id"),
            {"id": existing},
        )

    # Insert libro cabecera
    libro_id_row = (
        await db.execute(
            text(
                """
                INSERT INTO core.libros_remuneraciones
                    (empresa_codigo, periodo,
                     total_haberes, total_liquido, total_descuentos_legales,
                     total_aportes_patronales, total_costo_empresa,
                     archivo_origen, archivo_hash, uploaded_by,
                     cantidad_empleados)
                VALUES
                    (:c, :p, :th, :tl, :tdl, :tap, :tce,
                     :origen, :hash, CAST(:u AS UUID), :cant)
                RETURNING id
                """
            ),
            {
                "c": empresa_codigo,
                "p": parsed.periodo,
                "th": parsed.total_haberes,
                "tl": parsed.total_liquido,
                "tdl": parsed.total_descuentos_legales,
                "tap": parsed.total_aportes_patronales,
                "tce": parsed.total_costo_empresa,
                "origen": parsed.archivo_origen,
                "hash": parsed.archivo_hash,
                "u": str(getattr(user, "sub", "") or ""),
                "cant": len(parsed.lineas),
            },
        )
    ).fetchone()
    libro_id = libro_id_row[0]

    # Insert líneas + upsert empleados
    for l in parsed.lineas:
        # Auto-crear empleado si no existe (operativo: el upload trae el catálogo)
        await db.execute(
            text(
                """
                INSERT INTO core.empleados
                    (rut, nombre, empresa_codigo, area, sueldo_base_actual, activo)
                VALUES (:rut, :nombre, :emp, :area, :sb, TRUE)
                ON CONFLICT (rut) DO UPDATE SET
                    nombre = EXCLUDED.nombre,
                    empresa_codigo = EXCLUDED.empresa_codigo,
                    area = COALESCE(EXCLUDED.area, core.empleados.area),
                    sueldo_base_actual = EXCLUDED.sueldo_base_actual,
                    updated_at = NOW()
                """
            ),
            {
                "rut": l.rut,
                "nombre": l.nombre,
                "emp": empresa_codigo,
                "area": l.area,
                "sb": l.sueldo_base,
            },
        )

        await db.execute(
            text(
                """
                INSERT INTO core.libro_remuneraciones_lineas (
                    libro_id, empleado_rut, nombre, area, dias_trabajados,
                    sueldo_base, horas_extras, gratificacion_legal,
                    otros_imponibles, total_imponibles,
                    asignacion_familiar, otros_no_imponibles,
                    total_no_imponibles, total_haberes,
                    prevision, salud, seguro_cesantia_trab,
                    otros_descuentos_legales, total_descuentos_legales,
                    descuentos_varios, total_descuentos, liquido_pagado,
                    aporte_afp_empleador, sis, seguro_cesantia_empleador,
                    seguro_social, mutual, total_aportes_patronales,
                    base_tributable, impuesto_unico,
                    costo_total_empresa
                ) VALUES (
                    :libro_id, :rut, :nombre, :area, :dt,
                    :sb, :he, :gl, :oi, :ti,
                    :af, :oni, :tni, :th,
                    :prev, :sal, :sct, :odl, :tdl,
                    :dv, :td, :liq,
                    :aae, :sis, :sce, :ss, :mut, :tap,
                    :bt, :iu, :cte
                )
                """
            ),
            {
                "libro_id": libro_id,
                "rut": l.rut,
                "nombre": l.nombre,
                "area": l.area,
                "dt": l.dias_trabajados,
                "sb": l.sueldo_base,
                "he": l.horas_extras,
                "gl": l.gratificacion_legal,
                "oi": l.otros_imponibles,
                "ti": l.total_imponibles,
                "af": l.asignacion_familiar,
                "oni": l.otros_no_imponibles,
                "tni": l.total_no_imponibles,
                "th": l.total_haberes,
                "prev": l.prevision,
                "sal": l.salud,
                "sct": l.seguro_cesantia_trab,
                "odl": l.otros_descuentos_legales,
                "tdl": l.total_descuentos_legales,
                "dv": l.descuentos_varios,
                "td": l.total_descuentos,
                "liq": l.liquido_pagado,
                "aae": l.aporte_afp_empleador,
                "sis": l.sis,
                "sce": l.seguro_cesantia_empleador,
                "ss": l.seguro_social,
                "mut": l.mutual,
                "tap": l.total_aportes_patronales,
                "bt": l.base_tributable,
                "iu": l.impuesto_unico,
                "cte": l.costo_total_empresa,
            },
        )

    await db.commit()

    return UploadResult(
        libro_id=libro_id,
        empresa_codigo=empresa_codigo,
        periodo=parsed.periodo,
        cantidad_empleados=len(parsed.lineas),
        total_costo_empresa=parsed.total_costo_empresa,
        reemplazo=reemplazo,
    )


@router.get("/costo-empresa")
async def costo_empresa(
    user: CurrentUser,
    db: DBSession,
    empresa_codigo: str = Query(...),
    desde: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}$"),
    hasta: str | None = Query(default=None, pattern=r"^\d{4}-\d{2}$"),
) -> dict:
    """Costo agregado por empresa en un rango de periodos."""
    await _check_rrhh_access(user, db)
    wheres = ["empresa_codigo = :c"]
    params: dict = {"c": empresa_codigo}
    if desde:
        wheres.append("periodo >= :d")
        params["d"] = desde
    if hasta:
        wheres.append("periodo <= :h")
        params["h"] = hasta
    where_sql = " WHERE " + " AND ".join(wheres)
    rows = await db.execute(
        text(
            f"""
            SELECT periodo,
                   SUM(total_haberes) AS total_haberes,
                   SUM(total_liquido) AS total_liquido,
                   SUM(total_aportes_patronales) AS total_aportes_patronales,
                   SUM(total_costo_empresa) AS total_costo_empresa,
                   SUM(cantidad_empleados) AS cantidad_empleados
            FROM core.libros_remuneraciones
            {where_sql}
            GROUP BY periodo
            ORDER BY periodo
            """
        ),
        params,
    )
    series = [dict(r._mapping) for r in rows]
    total_costo = sum((Decimal(s["total_costo_empresa"] or 0) for s in series), Decimal("0"))
    return {
        "empresa_codigo": empresa_codigo,
        "periodos": series,
        "total_costo_acumulado": total_costo,
    }


# ─────────────────────────────────────────────────────────────────────
# R152CCCC — CRUD empleados + edición de líneas del libro
# ─────────────────────────────────────────────────────────────────────


class EmpleadoCreate(BaseModel):
    rut: str = Field(..., min_length=8, max_length=14)
    nombre: str = Field(..., min_length=2, max_length=200)
    empresa_codigo: str = Field(..., min_length=2, max_length=20)
    area: str | None = Field(default=None, max_length=120)
    cargo: str | None = Field(default=None, max_length=120)
    fecha_ingreso: str | None = None  # YYYY-MM-DD
    afp: str | None = None
    salud: str | None = None
    sueldo_base_actual: Decimal | None = Field(default=None, ge=0)
    notas: str | None = None


class EmpleadoUpdate(BaseModel):
    nombre: str | None = Field(default=None, min_length=2, max_length=200)
    area: str | None = Field(default=None, max_length=120)
    cargo: str | None = Field(default=None, max_length=120)
    fecha_ingreso: str | None = None
    fecha_salida: str | None = None
    activo: bool | None = None
    afp: str | None = None
    salud: str | None = None
    sueldo_base_actual: Decimal | None = Field(default=None, ge=0)
    notas: str | None = None


class LineaUpdate(BaseModel):
    """Permite editar campos básicos de una línea del libro."""

    sueldo_base: Decimal | None = Field(default=None, ge=0)
    horas_extras: Decimal | None = Field(default=None, ge=0)
    gratificacion_legal: Decimal | None = Field(default=None, ge=0)
    otros_imponibles: Decimal | None = Field(default=None, ge=0)
    asignacion_familiar: Decimal | None = Field(default=None, ge=0)
    otros_no_imponibles: Decimal | None = Field(default=None, ge=0)
    prevision: Decimal | None = Field(default=None, ge=0)
    salud: Decimal | None = Field(default=None, ge=0)
    seguro_cesantia_trab: Decimal | None = Field(default=None, ge=0)
    otros_descuentos_legales: Decimal | None = Field(default=None, ge=0)
    descuentos_varios: Decimal | None = Field(default=None, ge=0)
    aporte_afp_empleador: Decimal | None = Field(default=None, ge=0)
    sis: Decimal | None = Field(default=None, ge=0)
    seguro_cesantia_empleador: Decimal | None = Field(default=None, ge=0)
    seguro_social: Decimal | None = Field(default=None, ge=0)
    mutual: Decimal | None = Field(default=None, ge=0)


@router.post("/empleados", response_model=EmpleadoRead, status_code=201)
async def create_empleado(
    user: CurrentUser, db: DBSession, body: EmpleadoCreate
) -> EmpleadoRead:
    """Crea un empleado manualmente (sin pasar por upload de libro)."""
    await _check_rrhh_access(user, db)
    e = await db.scalar(
        text("SELECT 1 FROM core.empresas WHERE codigo = :c AND activo = TRUE"),
        {"c": body.empresa_codigo},
    )
    if not e:
        raise HTTPException(400, f"Empresa {body.empresa_codigo} no existe")

    try:
        await db.execute(
            text(
                """INSERT INTO core.empleados
                       (rut, nombre, empresa_codigo, area, cargo,
                        fecha_ingreso, afp, salud, sueldo_base_actual, notas, activo)
                   VALUES (:rut, :nombre, :empresa_codigo, :area, :cargo,
                           CAST(:fecha_ingreso AS DATE), :afp, :salud,
                           :sueldo_base_actual, :notas, TRUE)"""
            ),
            body.model_dump(),
        )
        await db.commit()
    except Exception as exc:
        await db.rollback()
        # R152JJJJJJ — no filtrar detalles del constraint DB al frontend.
        # El caso típico es RUT duplicado → mensaje claro; el resto genérico.
        logging.getLogger(__name__).warning(
            "rrhh.crear_empleado_failed", extra={"err": str(exc)[:200]}
        )
        if "unique" in str(exc).lower() or "duplicate" in str(exc).lower():
            raise HTTPException(
                409, "Ya existe un empleado con ese RUT en esa empresa."
            ) from exc
        raise HTTPException(
            409,
            "No se pudo crear el empleado. Revisá que los datos sean válidos "
            "(RUT, fechas y empresa).",
        ) from exc

    row = (
        await db.execute(
            text(
                """SELECT rut, nombre, empresa_codigo, area, cargo,
                          to_char(fecha_ingreso, 'YYYY-MM-DD') AS fecha_ingreso,
                          activo, sueldo_base_actual
                   FROM core.empleados WHERE rut = :r"""
            ),
            {"r": body.rut},
        )
    ).first()
    return EmpleadoRead.model_validate(dict(row._mapping))


@router.patch("/empleados/{rut}", response_model=EmpleadoRead)
async def update_empleado(
    user: CurrentUser, db: DBSession, rut: str, body: EmpleadoUpdate
) -> EmpleadoRead:
    """Edita campos del empleado. Solo los provistos cambian."""
    await _check_rrhh_access(user, db)
    # R152UUUUUU - exclude_unset (no exclude_none): el modal de edicion
    # manda `area: null` para LIMPIAR el campo, pero exclude_none lo
    # descartaba: el usuario veia "Empleado actualizado" y el valor viejo
    # reaparecia. Con exclude_unset, null explicito = borrar; campo no
    # enviado = no tocar (mismo patron que el PATCH de OCs).
    fields = body.model_dump(exclude_unset=True)
    if not fields:
        existing = (
            await db.execute(
                text(
                    """SELECT rut, nombre, empresa_codigo, area, cargo,
                              to_char(fecha_ingreso, 'YYYY-MM-DD') AS fecha_ingreso,
                              activo, sueldo_base_actual
                       FROM core.empleados WHERE rut = :r"""
                ),
                {"r": rut},
            )
        ).first()
        if not existing:
            raise HTTPException(404, "Empleado no encontrado")
        return EmpleadoRead.model_validate(dict(existing._mapping))

    set_clauses = []
    params: dict = {"r": rut}
    for k, v in fields.items():
        if k in ("fecha_ingreso", "fecha_salida") and v is not None:
            set_clauses.append(f"{k} = CAST(:{k} AS DATE)")
        else:
            set_clauses.append(f"{k} = :{k}")
        params[k] = v
    set_clauses.append("updated_at = NOW()")

    sql = "UPDATE core.empleados SET " + ", ".join(set_clauses) + " WHERE rut = :r"
    result = await db.execute(text(sql), params)
    if result.rowcount == 0:
        raise HTTPException(404, "Empleado no encontrado")
    await db.commit()

    row = (
        await db.execute(
            text(
                """SELECT rut, nombre, empresa_codigo, area, cargo,
                          to_char(fecha_ingreso, 'YYYY-MM-DD') AS fecha_ingreso,
                          activo, sueldo_base_actual
                   FROM core.empleados WHERE rut = :r"""
            ),
            {"r": rut},
        )
    ).first()
    return EmpleadoRead.model_validate(dict(row._mapping))


@router.delete(
    "/empleados/{rut}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
)
async def delete_empleado(
    user: CurrentUser, db: DBSession, rut: str
) -> Response:
    """Soft-delete: activo=FALSE + fecha_salida=hoy."""
    await _check_rrhh_access(user, db)
    result = await db.execute(
        text(
            "UPDATE core.empleados SET activo=FALSE, "
            "fecha_salida=COALESCE(fecha_salida, CURRENT_DATE), "
            "updated_at=NOW() WHERE rut = :r"
        ),
        {"r": rut},
    )
    if result.rowcount == 0:
        raise HTTPException(404, "Empleado no encontrado")
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.patch(
    "/libros/{libro_id}/lineas/{linea_id}",
    response_model=LineaLibroRead,
)
async def update_linea(
    user: CurrentUser,
    db: DBSession,
    libro_id: int,
    linea_id: int,
    body: LineaUpdate,
) -> LineaLibroRead:
    """Edita campos de una línea + recalcula totales derivados.

    Después de cambiar un campo base (ej: sueldo_base), recalcula:
      - total_imponibles, total_no_imponibles, total_haberes
      - total_descuentos_legales, total_descuentos, liquido_pagado
      - total_aportes_patronales, costo_total_empresa
    Y re-suma los totales del libro cabecera.
    """
    await _check_rrhh_access(user, db)
    fields = body.model_dump(exclude_none=True)
    if not fields:
        raise HTTPException(400, "Sin cambios")

    set_clauses = [f"{k} = :{k}" for k in fields.keys()]
    params = {**fields, "libro_id": libro_id, "linea_id": linea_id}
    sql = (
        "UPDATE core.libro_remuneraciones_lineas SET "
        + ", ".join(set_clauses)
        + " WHERE id = :linea_id AND libro_id = :libro_id"
    )
    result = await db.execute(text(sql), params)
    if result.rowcount == 0:
        raise HTTPException(404, "Línea no encontrada")

    # Recalcular totales derivados
    await db.execute(
        text(
            """UPDATE core.libro_remuneraciones_lineas SET
                 total_imponibles = sueldo_base + horas_extras + gratificacion_legal + otros_imponibles,
                 total_no_imponibles = asignacion_familiar + otros_no_imponibles,
                 total_haberes = (sueldo_base + horas_extras + gratificacion_legal + otros_imponibles)
                                 + (asignacion_familiar + otros_no_imponibles),
                 total_descuentos_legales = prevision + salud + seguro_cesantia_trab + otros_descuentos_legales,
                 total_descuentos = (prevision + salud + seguro_cesantia_trab + otros_descuentos_legales) + descuentos_varios,
                 liquido_pagado = (
                     (sueldo_base + horas_extras + gratificacion_legal + otros_imponibles)
                     + (asignacion_familiar + otros_no_imponibles)
                 ) - (
                     (prevision + salud + seguro_cesantia_trab + otros_descuentos_legales) + descuentos_varios
                 ),
                 total_aportes_patronales = aporte_afp_empleador + sis + seguro_cesantia_empleador + seguro_social + mutual,
                 costo_total_empresa = (
                     (sueldo_base + horas_extras + gratificacion_legal + otros_imponibles)
                     + (asignacion_familiar + otros_no_imponibles)
                 ) + (
                     aporte_afp_empleador + sis + seguro_cesantia_empleador + seguro_social + mutual
                 )
               WHERE id = :linea_id"""
        ),
        {"linea_id": linea_id},
    )

    # Re-sumar libro cabecera
    await db.execute(
        text(
            """UPDATE core.libros_remuneraciones lr SET
                 total_haberes = (
                     SELECT COALESCE(SUM(total_haberes),0)
                     FROM core.libro_remuneraciones_lineas WHERE libro_id = lr.id),
                 total_liquido = (
                     SELECT COALESCE(SUM(liquido_pagado),0)
                     FROM core.libro_remuneraciones_lineas WHERE libro_id = lr.id),
                 total_descuentos_legales = (
                     SELECT COALESCE(SUM(total_descuentos_legales),0)
                     FROM core.libro_remuneraciones_lineas WHERE libro_id = lr.id),
                 total_aportes_patronales = (
                     SELECT COALESCE(SUM(total_aportes_patronales),0)
                     FROM core.libro_remuneraciones_lineas WHERE libro_id = lr.id),
                 total_costo_empresa = (
                     SELECT COALESCE(SUM(costo_total_empresa),0)
                     FROM core.libro_remuneraciones_lineas WHERE libro_id = lr.id)
               WHERE id = :libro_id"""
        ),
        {"libro_id": libro_id},
    )
    await db.commit()

    row = (
        await db.execute(
            text(
                """SELECT id, empleado_rut, nombre, area, dias_trabajados,
                          total_haberes, liquido_pagado, total_descuentos_legales,
                          total_aportes_patronales, costo_total_empresa,
                          sueldo_base, horas_extras, gratificacion_legal,
                          asignacion_familiar,
                          aporte_afp_empleador, sis, seguro_cesantia_empleador,
                          seguro_social, mutual,
                          base_tributable, impuesto_unico
                   FROM core.libro_remuneraciones_lineas
                   WHERE id = :linea_id"""
            ),
            {"linea_id": linea_id},
        )
    ).first()
    return LineaLibroRead.model_validate(dict(row._mapping))


# ─────────────────────────────────────────────────────────────────────
# R152DDDD — Excel export del libro (formato Nubox/SII estándar)
# ─────────────────────────────────────────────────────────────────────


@router.get("/libros/{libro_id}/export-excel")
async def export_libro_excel(
    user: CurrentUser, db: DBSession, libro_id: int
) -> StreamingResponse:
    """Genera el libro de remuneraciones en Excel formato Nubox/SII.

    El layout matchea exactamente el del Excel original que el operador
    puede subir, lo que facilita envío al contador externo o re-importar.

    Estructura:
      Filas 1-7: header con datos de la empresa
      Fila 9: headers de tabla 1 (HABERES y DESCUENTOS)
      Filas 10+: una por empleado
      Fila X: TOTAL GENERAL
      Fila X+2: header "Patronales" + "Calculo Imp Unico"
      Fila X+3: headers tabla 2
      Filas X+4+: una por empleado (aportes patronales)
    """
    await _check_rrhh_access(user, db)

    # Cargar libro + empresa + líneas
    libro_row = (
        await db.execute(
            text(
                """SELECT lr.id, lr.empresa_codigo, lr.periodo,
                          lr.total_haberes, lr.total_liquido,
                          lr.total_descuentos_legales,
                          lr.total_aportes_patronales,
                          lr.total_costo_empresa, lr.cantidad_empleados,
                          e.razon_social, e.rut AS empresa_rut,
                          e.giro, e.direccion
                   FROM core.libros_remuneraciones lr
                   JOIN core.empresas e ON e.codigo = lr.empresa_codigo
                   WHERE lr.id = :id"""
            ),
            {"id": libro_id},
        )
    ).first()
    if not libro_row:
        raise HTTPException(404, "Libro no encontrado")
    lr = dict(libro_row._mapping)

    lineas_rows = await db.execute(
        text(
            """SELECT * FROM core.libro_remuneraciones_lineas
               WHERE libro_id = :id
               ORDER BY area NULLS LAST, nombre"""
        ),
        {"id": libro_id},
    )
    lineas = [dict(r._mapping) for r in lineas_rows]

    # Generar Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIBRO DE REMUNERACIONES"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", start_color="DCFCE7")
    total_fill = PatternFill("solid", start_color="FEF3C7")
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    meses_es = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    año, mes = lr["periodo"].split("-")
    mes_label = f"{meses_es[int(mes)]} DEL {año}"

    # Header empresa
    ws["A1"] = (lr.get("razon_social") or "").upper()
    ws["A1"].font = title_font
    ws["A2"] = f"Rut: {lr.get('empresa_rut') or '-'}"
    ws["A3"] = lr.get("giro") or ""
    ws["A4"] = lr.get("direccion") or ""
    ws["A6"] = "LIBRO DE REMUNERACIONES"
    ws["A6"].font = bold
    ws["A7"] = f"MES: {mes_label}"

    # Tabla 1 headers (fila 9)
    HEADERS_1 = [
        "Cód", "R.U.T", "Nombre", "DT",
        "S. Base", "H. Extras", "Grat. Legal", "Otros Imp.",
        "Total Imp.", "Asig. Fam.", "Otr. No Imp.", "Tot. No Imp.",
        "Tot. Haberes", "Previsión", "Salud", "Seg. Ces.",
        "Otros D.Leg.", "Tot. D.Leg.", "Desc. Varios", "Tot. Desc.", "Líquido",
    ]
    for i, h in enumerate(HEADERS_1, start=1):
        c = ws.cell(row=9, column=i, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border_thin

    # Datos tabla 1
    row_idx = 10
    for i, l in enumerate(lineas, start=1):
        cells = [
            i, l["empleado_rut"], l["nombre"], int(l["dias_trabajados"] or 0),
            int(l["sueldo_base"] or 0), int(l["horas_extras"] or 0),
            int(l["gratificacion_legal"] or 0), int(l["otros_imponibles"] or 0),
            int(l["total_imponibles"] or 0), int(l["asignacion_familiar"] or 0),
            int(l["otros_no_imponibles"] or 0), int(l["total_no_imponibles"] or 0),
            int(l["total_haberes"] or 0), int(l["prevision"] or 0),
            int(l["salud"] or 0), int(l["seguro_cesantia_trab"] or 0),
            int(l["otros_descuentos_legales"] or 0),
            int(l["total_descuentos_legales"] or 0),
            int(l["descuentos_varios"] or 0), int(l["total_descuentos"] or 0),
            int(l["liquido_pagado"] or 0),
        ]
        for j, v in enumerate(cells, start=1):
            c = ws.cell(row=row_idx, column=j, value=v)
            c.border = border_thin
            if j >= 4:
                c.number_format = "#,##0"
        row_idx += 1

    # Total general tabla 1
    total_row = row_idx
    ws.cell(row=total_row, column=3, value="TOTAL GENERAL").font = bold
    totals_1 = [
        sum(int(l["dias_trabajados"] or 0) for l in lineas),
        sum(int(l["sueldo_base"] or 0) for l in lineas),
        sum(int(l["horas_extras"] or 0) for l in lineas),
        sum(int(l["gratificacion_legal"] or 0) for l in lineas),
        sum(int(l["otros_imponibles"] or 0) for l in lineas),
        sum(int(l["total_imponibles"] or 0) for l in lineas),
        sum(int(l["asignacion_familiar"] or 0) for l in lineas),
        sum(int(l["otros_no_imponibles"] or 0) for l in lineas),
        sum(int(l["total_no_imponibles"] or 0) for l in lineas),
        sum(int(l["total_haberes"] or 0) for l in lineas),
        sum(int(l["prevision"] or 0) for l in lineas),
        sum(int(l["salud"] or 0) for l in lineas),
        sum(int(l["seguro_cesantia_trab"] or 0) for l in lineas),
        sum(int(l["otros_descuentos_legales"] or 0) for l in lineas),
        sum(int(l["total_descuentos_legales"] or 0) for l in lineas),
        sum(int(l["descuentos_varios"] or 0) for l in lineas),
        sum(int(l["total_descuentos"] or 0) for l in lineas),
        sum(int(l["liquido_pagado"] or 0) for l in lineas),
    ]
    for j, v in enumerate(totals_1, start=4):
        c = ws.cell(row=total_row, column=j, value=v)
        c.font = bold
        c.fill = total_fill
        c.number_format = "#,##0"
        c.border = border_thin

    # Tabla 2: Aportes patronales + Cálculo Imp. Único
    row_idx = total_row + 2
    ws.cell(row=row_idx, column=1, value="Patronales").font = bold
    ws.cell(row=row_idx, column=11, value="Calculo Imp Unico").font = bold
    row_idx += 1
    HEADERS_2 = [
        "Cód", "R.U.T", "Nombre", "DT",
        "AFP EMP", "SIS", "Seg Cesantia", "Seg Social", "Mutual",
        "", "Tributable", "Calculo",
    ]
    for i, h in enumerate(HEADERS_2, start=1):
        if not h:
            continue
        c = ws.cell(row=row_idx, column=i, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border_thin
    row_idx += 1

    for i, l in enumerate(lineas, start=1):
        cells = [
            i, l["empleado_rut"], l["nombre"], int(l["dias_trabajados"] or 0),
            float(l["aporte_afp_empleador"] or 0),
            float(l["sis"] or 0),
            float(l["seguro_cesantia_empleador"] or 0),
            float(l["seguro_social"] or 0),
            float(l["mutual"] or 0),
            None,
            int(l["base_tributable"] or 0),
            float(l["impuesto_unico"] or 0),
        ]
        for j, v in enumerate(cells, start=1):
            if v is None:
                continue
            c = ws.cell(row=row_idx, column=j, value=v)
            c.border = border_thin
            if j >= 4:
                c.number_format = "#,##0" if j in (4, 11) else "#,##0.00"
        row_idx += 1

    # Auto-width columnas
    for col_idx in range(1, 22):
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
            min(max_len + 2, 30)
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"libro_remuneraciones_{lr['empresa_codigo']}_{lr['periodo']}.xlsx"

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
