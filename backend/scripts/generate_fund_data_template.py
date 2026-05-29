"""Genera template Excel para cargar la DATA REAL del fondo (Dashboard Institucional).

Output: C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx con hojas:
  1. INSTRUCCIONES   — cómo llenar y cargar
  2. Fondo & LPs     — committed capital del fondo + commitments de cada LP
  3. Cashflows       — capital calls / distributions / fees (J-Curve)
  4. Valuations      — valuación trimestral por empresa portfolio (MOIC/IRR)
  5. KPIs Operativos — revenue/ebitda/headcount/MW por empresa por mes
  6. Impact (IRIS+)  — métricas de impacto verificadas

Cada hoja tiene los headers EXACTOS que el import script espera + 1-2 filas
de ejemplo (marcadas en gris claro para borrar).

Después de llenar: correr
    python scripts/import_fund_data.py "C:/.../Template_Data_Fondo_REAL.xlsx"
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path("C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx")

GREEN = "1D6F42"
GREEN_LIGHT = "E7F3EC"
GRAY = "F0F0F0"
thin = Side(border_style="thin", color="D2D2D7")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
hf = Font(name="Arial", size=10, bold=True, color="FFFFFF")
hfill = PatternFill("solid", start_color=GREEN)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
example_fill = PatternFill("solid", start_color=GRAY)
example_font = Font(name="Arial", size=10, italic=True, color="6E6E73")
data_font = Font(name="Arial", size=10)

PORTFOLIO = ["CSL", "RHO", "DTE", "REVTECH", "EVOQUE", "TRONGKAI"]


def _write_sheet(ws, headers, examples, widths, notes=""):
    # Header
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = center
        c.border = border
    # Examples
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = example_font
            c.fill = example_fill
            c.border = border
            if isinstance(val, (int, float)) and ci > 2:
                c.number_format = "#,##0.00;(#,##0.00);-"
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    if notes:
        nr = len(examples) + 3
        nc = ws.cell(row=nr, column=1, value=notes)
        nc.font = Font(name="Arial", size=9, italic=True, color="6E6E73")
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=len(headers))
        ws.row_dimensions[nr].height = 45


def main() -> None:
    wb = Workbook()

    # ===== Hoja 1: INSTRUCCIONES =====
    ws0 = wb.active
    ws0.title = "INSTRUCCIONES"
    inst = [
        ("CEHTA CAPITAL — CARGA DE DATA REAL DEL FONDO", "title"),
        ("", ""),
        ("Este Excel reemplaza los datos DEMO que se borraron del Dashboard Institucional.", ""),
        ("Llená cada hoja con los números REALES del fondo y sus empresas.", ""),
        ("", ""),
        ("ORDEN DE LLENADO", "h2"),
        ("1. Fondo & LPs    → committed capital del fondo + de cada Limited Partner", ""),
        ("2. Cashflows      → cada capital call / distribución / fee con su fecha", ""),
        ("3. Valuations     → valuación de cada empresa portfolio (trimestral)", ""),
        ("4. KPIs Operativos→ revenue/ebitda/headcount por empresa por mes (opcional)", ""),
        ("5. Impact         → métricas IRIS+ verificadas (opcional)", ""),
        ("", ""),
        ("REGLAS", "h2"),
        ("• Fechas: formato YYYY-MM-DD (ej. 2025-03-31)", ""),
        ("• Montos: en USD, solo números (sin $ ni puntos de miles). Ej: 1500000", ""),
        ("• Empresas válidas: CSL, RHO, DTE, REVTECH, EVOQUE, TRONGKAI", ""),
        ("• Borrá las filas de ejemplo (gris) antes de cargar.", ""),
        ("• Si no tenés un dato, dejá la celda vacía (excepto las obligatorias marcadas *).", ""),
        ("", ""),
        ("EL DASHBOARD CALCULA SOLO", "h2"),
        ("• TVPI / DPI / RVPI / MOIC del fondo → desde los Cashflows", ""),
        ("• % Called, NAV, J-Curve → desde Cashflows + Valuations", ""),
        ("• No tenés que calcular ratios; solo cargá los movimientos y valuaciones.", ""),
        ("", ""),
        ("CÓMO CARGAR", "h2"),
        ("Cuando termines, avisale a Claude (o a tu equipo técnico) y corré:", ""),
        ("   python backend/scripts/import_fund_data.py \"<ruta del Excel>\"", "code"),
        ("El script valida todo y carga en la DB. Si hay un error, te dice qué fila.", ""),
        ("", ""),
        (f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "foot"),
    ]
    for i, (txt, st) in enumerate(inst, 1):
        c = ws0.cell(row=i, column=1, value=txt)
        if st == "title":
            c.font = Font(name="Arial", size=16, bold=True, color=GREEN)
        elif st == "h2":
            c.font = Font(name="Arial", size=12, bold=True)
            c.fill = PatternFill("solid", start_color=GREEN_LIGHT)
        elif st == "code":
            c.font = Font(name="Consolas", size=10, color="1E40AF")
        elif st == "foot":
            c.font = Font(name="Arial", size=9, italic=True, color="6E6E73")
        else:
            c.font = Font(name="Arial", size=11)
    ws0.column_dimensions["A"].width = 95

    # ===== Hoja 2: Fondo & LPs =====
    ws_fund = wb.create_sheet("Fondo & LPs")
    # Sección fondo
    ws_fund.cell(row=1, column=1, value="FONDO").font = Font(bold=True, size=12, color=GREEN)
    _write_sheet_at(ws_fund, 2,
        ["fund_codigo*", "fund_size_committed_usd*", "vintage_year", "moneda"],
        [["FIP_CEHTA_ESG", 22500000, 2024, "USD"]],
        [18, 26, 14, 10])
    # Sección LPs (más abajo)
    ws_fund.cell(row=6, column=1, value="LIMITED PARTNERS (aportantes)").font = Font(bold=True, size=12, color=GREEN)
    _write_sheet_at(ws_fund, 7,
        ["legal_name*", "lp_type*", "rut", "commitment_usd*", "ownership_pct"],
        [
            ["CORFO - Corporación de Fomento", "publico_corfo", "60.706.000-2", 11250000, 50],
            ["Aportante Privado #1", "privado", "76.xxx.xxx-x", 11250000, 50],
        ],
        [32, 16, 16, 18, 14])
    ws_fund.cell(row=11, column=1,
        value="lp_type válidos: publico_corfo, privado, family_office, institucional, otro. "
              "La suma de ownership_pct debe dar 100.").font = Font(size=9, italic=True, color="6E6E73")

    # ===== Hoja 3: Cashflows =====
    ws_cf = wb.create_sheet("Cashflows")
    _write_sheet(ws_cf,
        ["fund_codigo*", "lp (opcional)", "cashflow_type*", "amount_usd*", "effective_date*", "descripcion"],
        [
            ["FIP_CEHTA_ESG", "CORFO", "capital_call", 1500000, "2024-03-15", "1er llamado de capital"],
            ["FIP_CEHTA_ESG", "", "management_fee", 281250, "2024-03-31", "Mgmt fee Q1 (2.5%/4)"],
            ["FIP_CEHTA_ESG", "CORFO", "distribution", 500000, "2025-12-20", "Distribución parcial CSL"],
        ],
        [18, 16, 20, 16, 16, 32],
        notes="cashflow_type válidos: capital_call, distribution, management_fee, expense, "
              "carried_interest, subscription_line_draw. Una fila por movimiento. "
              "El signo lo maneja el sistema (calls negativos, distribuciones positivos en J-Curve).")

    # ===== Hoja 4: Valuations =====
    ws_val = wb.create_sheet("Valuations")
    _write_sheet(ws_val,
        ["empresa*", "as_of_date*", "invested_usd", "realized_usd", "fair_value_usd", "moic_net", "irr_net", "metodo"],
        [
            ["CSL", "2025-12-31", 2000000, 0, 3200000, 1.60, 0.18, "DCF"],
            ["RHO", "2025-12-31", 1500000, 500000, 2100000, 1.73, 0.22, "Comparables"],
        ],
        [10, 14, 16, 14, 16, 10, 10, 14],
        notes="Una fila por empresa por fecha de valuación (típico: cierre de cada trimestre). "
              "moic_net e irr_net son opcionales — si los dejás vacíos, el dashboard los calcula "
              "desde invested vs fair_value. metodo: DCF, Comparables, Costo, Última_ronda, etc.")

    # ===== Hoja 5: KPIs Operativos =====
    ws_kpi = wb.create_sheet("KPIs Operativos")
    _write_sheet(ws_kpi,
        ["empresa*", "period*", "revenue_usd", "ebitda_usd", "cash_balance_usd", "headcount", "mw_installed", "notes"],
        [
            ["RHO", "2025-12-31", 450000, 120000, 800000, 12, 5.2, "Q4 2025"],
            ["CSL", "2025-12-31", 380000, 95000, 600000, 8, 0, "Q4 2025"],
        ],
        [10, 14, 14, 14, 16, 11, 12, 20],
        notes="period = fin de mes (YYYY-MM-31). Opcional pero alimenta las sparklines de la "
              "pestaña KPIs Op. de cada empresa. mw_installed solo aplica a empresas de energía (RHO).")

    # ===== Hoja 6: Impact (IRIS+) =====
    ws_imp = wb.create_sheet("Impact")
    _write_sheet(ws_imp,
        ["empresa*", "period*", "iris_metric_id*", "metric_name*", "metric_value*", "unit*", "verified"],
        [
            ["RHO", "2025-12-31", "PI2764", "CO2 emissions avoided", 1250, "tCO2e", "SI"],
            ["CSL", "2025-12-31", "OI2535", "Jobs created", 8, "empleos", "NO"],
        ],
        [10, 14, 14, 30, 14, 12, 10],
        notes="IRIS+ codes comunes: PI2764 (CO2 avoided/tCO2e), PI5842 (Energy generated/MWh), "
              "OI2535 (Jobs/empleos), PI4060 (Waste diverted/ton). verified: SI/NO. "
              "Ver catálogo completo en iris.thegiin.org")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"[OK] Template data real generado: {OUTPUT}")


def _write_sheet_at(ws, start_row, headers, examples, widths):
    """Escribe headers+examples empezando en start_row (para hojas multi-sección)."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = center
        c.border = border
    for ri, row in enumerate(examples, start_row + 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = example_font
            c.fill = example_fill
            c.border = border
            if isinstance(val, (int, float)) and ci > 1:
                c.number_format = "#,##0;(#,##0);-"
    for ci, w in enumerate(widths, 1):
        cur = ws.column_dimensions[get_column_letter(ci)].width or 0
        if w > cur:
            ws.column_dimensions[get_column_letter(ci)].width = w


if __name__ == "__main__":
    main()
