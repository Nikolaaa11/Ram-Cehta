"""Importa Template_Data_Fondo_REAL.xlsx → core.* en Supabase Brasil.

Uso:
    python scripts/import_fondo_real.py "C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx"

Tablas que actualiza:
  - core.funds                     UPDATE (committed + aum)
  - core.limited_partners          UPDATE (commitment, paid_in, distributed, ownership)
  - core.company_valuations        INSERT (idempotente por empresa_codigo + as_of_date)
  - core.fund_cashflows            INSERT
  - core.company_operational_kpis  INSERT (idempotente por empresa_codigo + period)
  - core.impact_metrics            INSERT

Idempotente: si volvés a correrlo con los mismos datos, no duplica
(valuations y KPIs tienen unique en (empresa, fecha); usa ON CONFLICT
DO UPDATE).
"""
from __future__ import annotations

import asyncio
import os
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import asyncpg
import openpyxl


def _load_db_url() -> str:
    url = os.getenv("DATABASE_URL")
    if not url:
        env = Path(__file__).resolve().parent.parent / ".env"
        if env.exists():
            for line in env.read_text(encoding="utf-8").splitlines():
                if line.startswith("DATABASE_URL="):
                    url = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
    if not url:
        print("ERROR: DATABASE_URL no encontrado", file=sys.stderr)
        sys.exit(1)
    return url.replace("postgresql+asyncpg://", "postgresql://")


def _row_dict(ws, headers_row: int = 1) -> list[dict[str, Any]]:
    """Devuelve la hoja como lista de dicts {header: valor}."""
    headers = [c.value for c in ws[headers_row]]
    rows = []
    for r in range(headers_row + 1, ws.max_row + 1):
        row = ws[r]
        values = [c.value for c in row]
        if all(v is None or v == "" for v in values):
            continue
        rec = dict(zip(headers, values))
        rows.append(rec)
    return rows


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_dec(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _to_bool(v) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().upper() in ("TRUE", "T", "SI", "SÍ", "1", "YES")
    return bool(v)


async def main(xlsx_path: str) -> None:
    db_url = _load_db_url()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Leyendo: {xlsx_path}\n")

    db = await asyncpg.connect(db_url, timeout=30, statement_cache_size=0)
    fund_id = await db.fetchval("SELECT fund_id FROM core.funds WHERE codigo='FIP_CEHTA_ESG'")
    if not fund_id:
        print("ERROR: no encuentro core.funds con codigo FIP_CEHTA_ESG")
        await db.close()
        return

    # Mapa lp_name → lp_id
    lp_map: dict[str, Any] = {}
    for r in await db.fetch("SELECT lp_id, legal_name FROM core.limited_partners WHERE fund_id=$1", fund_id):
        lp_map[r["legal_name"]] = r["lp_id"]

    async with db.transaction():
        # ──── HOJA 1: FONDO ──────────────────────────────────────────
        if "Fondo" in wb.sheetnames:
            rows = _row_dict(wb["Fondo"])
            kv = {r["Campo"]: r["Valor a confirmar"] for r in rows if r.get("Campo")}
            committed = _to_dec(kv.get("Tamaño committed (USD)"))
            aum = _to_dec(kv.get("AUM actual (USD)"))
            if committed is not None:
                await db.execute(
                    "UPDATE core.funds SET fund_size_committed_usd=$1, updated_at=NOW() WHERE codigo='FIP_CEHTA_ESG'",
                    committed,
                )
                print(f"[OK] funds.fund_size_committed_usd = ${committed:,.0f}")
            if aum is not None:
                await db.execute(
                    "UPDATE core.funds SET aum_current_usd=$1, updated_at=NOW() WHERE codigo='FIP_CEHTA_ESG'",
                    aum,
                )
                print(f"[OK] funds.aum_current_usd = ${aum:,.0f}")

        # ──── HOJA 2: INVERSIONISTAS ────────────────────────────────
        if "Inversionistas" in wb.sheetnames:
            for r in _row_dict(wb["Inversionistas"]):
                name = r.get("LP Name (existente)")
                if not name:
                    continue
                lp_id = lp_map.get(name)
                if not lp_id:
                    print(f"[WARN] LP no encontrado: {name}")
                    continue
                commit = _to_dec(r.get("Commitment USD"))
                paid = _to_dec(r.get("Paid-in USD"))
                dist = _to_dec(r.get("Distributed USD"))
                own = _to_dec(r.get("Ownership %"))
                await db.execute(
                    """
                    UPDATE core.limited_partners
                    SET commitment_usd = COALESCE($2, commitment_usd),
                        paid_in_usd = COALESCE($3, paid_in_usd),
                        distributed_usd = COALESCE($4, distributed_usd),
                        ownership_pct = COALESCE($5, ownership_pct),
                        updated_at = NOW()
                    WHERE lp_id = $1
                    """,
                    lp_id, commit, paid, dist, own,
                )
                print(f"[OK] LP {name[:30]:<30}: commit=${commit or 0:,.0f} paid=${paid or 0:,.0f}")

        # ──── HOJA 3: VALUACIONES ───────────────────────────────────
        if "Valuaciones" in wb.sheetnames:
            count = 0
            for r in _row_dict(wb["Valuaciones"]):
                emp = r.get("empresa_codigo")
                d = _to_date(r.get("as_of_date (YYYY-MM-DD)"))
                if not emp or not d:
                    continue
                await db.execute(
                    """
                    INSERT INTO core.company_valuations
                        (empresa_codigo, as_of_date, invested_amount_usd, realized_value_usd,
                         unrealized_fv_usd, moic_gross, moic_net, irr_gross, irr_net,
                         valuation_method, notes)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)
                    ON CONFLICT (empresa_codigo, as_of_date)
                    DO UPDATE SET
                        invested_amount_usd = EXCLUDED.invested_amount_usd,
                        realized_value_usd = EXCLUDED.realized_value_usd,
                        unrealized_fv_usd = EXCLUDED.unrealized_fv_usd,
                        moic_gross = EXCLUDED.moic_gross, moic_net = EXCLUDED.moic_net,
                        irr_gross = EXCLUDED.irr_gross, irr_net = EXCLUDED.irr_net,
                        valuation_method = EXCLUDED.valuation_method, notes = EXCLUDED.notes
                    """,
                    emp, d,
                    _to_dec(r.get("invested_amount_usd")), _to_dec(r.get("realized_value_usd")),
                    _to_dec(r.get("unrealized_fv_usd")), _to_dec(r.get("moic_gross")),
                    _to_dec(r.get("moic_net")), _to_dec(r.get("irr_gross")),
                    _to_dec(r.get("irr_net")), r.get("valuation_method"), r.get("notes"),
                )
                count += 1
            print(f"[OK] Valuaciones upserted: {count}")

        # ──── HOJA 4: CASHFLOWS ─────────────────────────────────────
        if "Cashflows" in wb.sheetnames:
            count = 0
            for r in _row_dict(wb["Cashflows"]):
                d = _to_date(r.get("effective_date (YYYY-MM-DD)"))
                cf_type = r.get("cashflow_type")
                amount = _to_dec(r.get("amount_usd"))
                if not d or not cf_type or amount is None:
                    continue
                lp_name = r.get("lp_name (opcional)")
                lp_id = lp_map.get(lp_name) if lp_name else None
                await db.execute(
                    """
                    INSERT INTO core.fund_cashflows
                        (fund_id, lp_id, cashflow_type, amount_usd, effective_date,
                         descripcion, ilpa_category, recallable)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                    """,
                    fund_id, lp_id, cf_type, amount, d,
                    r.get("descripcion"), r.get("ilpa_category"),
                    _to_bool(r.get("recallable (TRUE/FALSE)")),
                )
                count += 1
            print(f"[OK] Cashflows inserted: {count}")

        # ──── HOJA 5: KPIs OPERATIVOS ───────────────────────────────
        if "KPIs Operativos" in wb.sheetnames:
            count = 0
            for r in _row_dict(wb["KPIs Operativos"]):
                emp = r.get("empresa_codigo")
                p = _to_date(r.get("period (YYYY-MM-01)"))
                if not emp or not p:
                    continue
                await db.execute(
                    """
                    INSERT INTO core.company_operational_kpis
                        (empresa_codigo, period, revenue_usd, ebitda_usd, ebitda_margin,
                         gross_margin, cash_balance_usd, burn_rate_usd, cash_runway_months,
                         headcount, mw_installed, capacity_factor, notes)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
                    ON CONFLICT (empresa_codigo, period)
                    DO UPDATE SET
                        revenue_usd=EXCLUDED.revenue_usd, ebitda_usd=EXCLUDED.ebitda_usd,
                        ebitda_margin=EXCLUDED.ebitda_margin, gross_margin=EXCLUDED.gross_margin,
                        cash_balance_usd=EXCLUDED.cash_balance_usd, burn_rate_usd=EXCLUDED.burn_rate_usd,
                        cash_runway_months=EXCLUDED.cash_runway_months, headcount=EXCLUDED.headcount,
                        mw_installed=EXCLUDED.mw_installed, capacity_factor=EXCLUDED.capacity_factor,
                        notes=EXCLUDED.notes
                    """,
                    emp, p,
                    _to_dec(r.get("revenue_usd")), _to_dec(r.get("ebitda_usd")),
                    _to_dec(r.get("ebitda_margin")), _to_dec(r.get("gross_margin")),
                    _to_dec(r.get("cash_balance_usd")), _to_dec(r.get("burn_rate_usd")),
                    _to_dec(r.get("cash_runway_months")),
                    int(r.get("headcount")) if r.get("headcount") else None,
                    _to_dec(r.get("mw_installed")), _to_dec(r.get("capacity_factor")),
                    r.get("notes"),
                )
                count += 1
            print(f"[OK] KPIs upserted: {count}")

        # ──── HOJA 6: IMPACT METRICS ────────────────────────────────
        if "Impact Metrics" in wb.sheetnames:
            count = 0
            for r in _row_dict(wb["Impact Metrics"]):
                emp = r.get("empresa_codigo")
                p = _to_date(r.get("period (YYYY-MM-DD)"))
                metric_id = r.get("iris_metric_id")
                value = _to_dec(r.get("metric_value"))
                if not emp or not p or not metric_id or value is None:
                    continue
                await db.execute(
                    """
                    INSERT INTO core.impact_metrics
                        (empresa_codigo, fund_id, period, iris_metric_id, metric_name,
                         metric_value, unit, framework, verified, verifier)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                    """,
                    emp, fund_id, p, metric_id, r.get("metric_name"),
                    value, r.get("unit"), r.get("framework"),
                    _to_bool(r.get("verified (TRUE/FALSE)")) or False,
                    r.get("verifier"),
                )
                count += 1
            print(f"[OK] Impact metrics inserted: {count}")

    print("\n[OK] Import completo. El Dashboard Institucional ya muestra estos datos.")
    await db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/import_fondo_real.py <path_al_xlsx>")
        sys.exit(1)
    asyncio.run(main(sys.argv[1]))
