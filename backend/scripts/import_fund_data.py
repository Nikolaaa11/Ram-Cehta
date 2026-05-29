"""Importa la DATA REAL del fondo desde Template_Data_Fondo_REAL.xlsx a la DB.

Lee las hojas: Fondo & LPs, Cashflows, Valuations, KPIs Operativos, Impact
y hace UPSERT en las tablas core.* correspondientes.

Idempotente: borra los datos previos de cada tabla antes de cargar (re-corrible).

Uso:
    python scripts/import_fund_data.py "C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx"
"""
from __future__ import annotations

import asyncio
import os
import sys
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl

FUND_CODE = "FIP_CEHTA_ESG"


def _db_url() -> str:
    url = os.getenv("DATABASE_URL")
    if not url:
        env = Path(__file__).resolve().parent.parent / ".env"
        for line in env.read_text(encoding="utf-8").splitlines():
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip().strip('"').strip("'")
                break
    return url.replace("postgresql+asyncpg://", "postgresql://")


def _d(v):
    """Parse fecha → date."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()


def _n(v):
    """Parse número → float o None."""
    if v is None or str(v).strip() == "":
        return None
    return float(str(v).replace(",", "").replace("$", "").strip())


def _rows(ws, header_row: int):
    """Itera filas como dicts usando header_row como encabezado. Salta filas vacías y de ejemplo (gris)."""
    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    out = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=False):
        vals = [c.value for c in r]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        # Saltar filas de ejemplo (fill gris F0F0F0)
        first_fill = r[0].fill.start_color.rgb if r[0].fill and r[0].fill.start_color else None
        if first_fill and str(first_fill).endswith("F0F0F0"):
            continue
        rec = dict(zip(headers, vals))
        out.append(rec)
    return out


def _clean_key(d: dict) -> dict:
    """Normaliza keys quitando * y espacios."""
    return {k.replace("*", "").strip(): v for k, v in d.items()}


async def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    db = await asyncpg.connect(_db_url(), timeout=30, statement_cache_size=0)

    fund_id = await db.fetchval("SELECT fund_id FROM core.funds WHERE codigo=$1", FUND_CODE)
    if not fund_id:
        print(f"ERROR: no existe fondo {FUND_CODE}")
        return

    # mapping LP name → lp_id (para cashflows)
    async def lp_id_for(name: str):
        if not name or str(name).strip() == "":
            return None
        return await db.fetchval(
            "SELECT lp_id FROM core.limited_partners WHERE legal_name ILIKE $1 LIMIT 1",
            f"%{str(name).strip()[:20]}%",
        )

    summary = {}

    async with db.transaction():
        # ---- 1. Fondo & LPs ----
        if "Fondo & LPs" in wb.sheetnames:
            ws = wb["Fondo & LPs"]
            # Fondo: header en row 2, data row 3
            fund_rows = _rows(ws, 2)
            for fr in fund_rows[:1]:
                fr = _clean_key(fr)
                if fr.get("fund_size_committed_usd") is not None:
                    await db.execute(
                        "UPDATE core.funds SET fund_size_committed_usd=$1, vintage_year=$2, updated_at=NOW() WHERE codigo=$3",
                        _n(fr["fund_size_committed_usd"]),
                        int(fr.get("vintage_year") or 2024),
                        FUND_CODE,
                    )
                    summary["fondo"] = "actualizado"
            # LPs: header en row 7
            lp_rows = _rows(ws, 7)
            if lp_rows:
                # Reset commitments, luego upsert por legal_name
                for lr in lp_rows:
                    lr = _clean_key(lr)
                    name = str(lr.get("legal_name") or "").strip()
                    if not name:
                        continue
                    existing = await db.fetchval(
                        "SELECT lp_id FROM core.limited_partners WHERE legal_name ILIKE $1",
                        f"%{name[:20]}%",
                    )
                    commit = _n(lr.get("commitment_usd")) or 0
                    pct = _n(lr.get("ownership_pct")) or 0
                    if existing:
                        await db.execute(
                            "UPDATE core.limited_partners SET commitment_usd=$1, ownership_pct=$2, lp_type=$3, rut=COALESCE($4,rut), updated_at=NOW() WHERE lp_id=$5",
                            commit, pct, str(lr.get("lp_type") or "privado"), lr.get("rut"), existing,
                        )
                    else:
                        await db.execute(
                            "INSERT INTO core.limited_partners (fund_id, legal_name, lp_type, rut, commitment_usd, ownership_pct, active) VALUES ($1,$2,$3,$4,$5,$6,TRUE)",
                            fund_id, name, str(lr.get("lp_type") or "privado"), lr.get("rut"), commit, pct,
                        )
                summary["lps"] = len(lp_rows)

        # ---- 2. Cashflows ----
        if "Cashflows" in wb.sheetnames:
            await db.execute("DELETE FROM core.fund_cashflows WHERE fund_id=$1", fund_id)
            cf = _rows(wb["Cashflows"], 1)
            n = 0
            for r in cf:
                r = _clean_key(r)
                if not r.get("cashflow_type") or _n(r.get("amount_usd")) is None:
                    continue
                await db.execute(
                    """INSERT INTO core.fund_cashflows
                       (fund_id, lp_id, cashflow_type, amount_usd, effective_date, descripcion)
                       VALUES ($1,$2,$3,$4,$5,$6)""",
                    fund_id, await lp_id_for(r.get("lp (opcional)") or r.get("lp")),
                    str(r["cashflow_type"]).strip(), _n(r["amount_usd"]),
                    _d(r["effective_date"]), r.get("descripcion"),
                )
                n += 1
            summary["cashflows"] = n

        # ---- 3. Valuations ----
        if "Valuations" in wb.sheetnames:
            await db.execute("DELETE FROM core.company_valuations")
            v = _rows(wb["Valuations"], 1)
            n = 0
            for r in v:
                r = _clean_key(r)
                emp = str(r.get("empresa") or "").strip().upper()
                if not emp or not r.get("as_of_date"):
                    continue
                await db.execute(
                    """INSERT INTO core.company_valuations
                       (empresa_codigo, as_of_date, invested_amount_usd, realized_value_usd,
                        unrealized_fv_usd, moic_net, irr_net, valuation_method)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8)""",
                    emp, _d(r["as_of_date"]), _n(r.get("invested_usd")), _n(r.get("realized_usd")),
                    _n(r.get("fair_value_usd")), _n(r.get("moic_net")), _n(r.get("irr_net")),
                    r.get("metodo"),
                )
                n += 1
            summary["valuations"] = n

        # ---- 4. KPIs ----
        if "KPIs Operativos" in wb.sheetnames:
            await db.execute("DELETE FROM core.company_operational_kpis")
            k = _rows(wb["KPIs Operativos"], 1)
            n = 0
            for r in k:
                r = _clean_key(r)
                emp = str(r.get("empresa") or "").strip().upper()
                if not emp or not r.get("period"):
                    continue
                await db.execute(
                    """INSERT INTO core.company_operational_kpis
                       (empresa_codigo, period, revenue_usd, ebitda_usd, cash_balance_usd,
                        headcount, mw_installed, notes)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8)""",
                    emp, _d(r["period"]), _n(r.get("revenue_usd")), _n(r.get("ebitda_usd")),
                    _n(r.get("cash_balance_usd")),
                    int(r["headcount"]) if r.get("headcount") not in (None, "") else None,
                    _n(r.get("mw_installed")), r.get("notes"),
                )
                n += 1
            summary["kpis"] = n

        # ---- 5. Impact ----
        if "Impact" in wb.sheetnames:
            await db.execute("DELETE FROM core.impact_metrics")
            im = _rows(wb["Impact"], 1)
            n = 0
            for r in im:
                r = _clean_key(r)
                emp = str(r.get("empresa") or "").strip().upper()
                if not emp or not r.get("iris_metric_id") or _n(r.get("metric_value")) is None:
                    continue
                ver = str(r.get("verified") or "").strip().upper() in ("SI", "SÍ", "YES", "TRUE", "1")
                await db.execute(
                    """INSERT INTO core.impact_metrics
                       (empresa_codigo, fund_id, period, iris_metric_id, metric_name,
                        metric_value, unit, framework, verified)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,'IRIS+ v5.3',$8)""",
                    emp, fund_id, _d(r["period"]), str(r["iris_metric_id"]).strip(),
                    str(r["metric_name"]).strip(), _n(r["metric_value"]),
                    str(r.get("unit") or "").strip(), ver,
                )
                n += 1
            summary["impact"] = n

    print("[OK] Import completado:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
    await db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print('Uso: python scripts/import_fund_data.py "<ruta Excel>"')
        sys.exit(1)
    asyncio.run(main(sys.argv[1]))
