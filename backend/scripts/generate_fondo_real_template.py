"""Genera Template_Data_Fondo_REAL.xlsx para cargar la data REAL del fondo.

Output: C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx con 6 hojas:
  1. Fondo            — confirmar fund_size_committed + AUM
  2. Inversionistas   — montos REALES de CORFO y Privado (commitment, paid_in)
  3. Valuaciones      — invested vs FV + MOIC + IRR por empresa portfolio × Q
  4. Cashflows        — capital calls + management fees + distributions
  5. KPIs Operativos  — revenue, EBITDA, headcount, MW, cash runway por empresa × mes
  6. Impact Metrics   — IRIS+ codes con cantidad + unidad

El script de import lee estas hojas y popula:
  - core.funds (UPDATE)
  - core.limited_partners (UPDATE)
  - core.company_valuations (INSERT)
  - core.fund_cashflows (INSERT)
  - core.company_operational_kpis (INSERT)
  - core.impact_metrics (INSERT)

Antes de cargar real, las tablas transaccionales están VACÍAS (limpieza R152j).
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path("C:/Users/DELL/Documents/Template_Data_Fondo_REAL.xlsx")

# ─── Estilos ──────────────────────────────────────────────────────────
GREEN = "1D6F42"
GREEN_LIGHT = "E7F3EC"
RED_LIGHT = "FEE2E2"
AMBER_LIGHT = "FEF3C7"
BLUE_LIGHT = "DBEAFE"

thin = Side(border_style="thin", color="D2D2D7")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
H = Font(name="Arial", size=10, bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", start_color=GREEN)
INST_FILL = PatternFill("solid", start_color=GREEN_LIGHT)
EXAMPLE_FILL = PatternFill("solid", start_color=BLUE_LIGHT)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


PORTFOLIO_EMPRESAS = ["CSL", "RHO", "DTE", "REVTECH", "EVOQUE", "TRONGKAI"]


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.font = H
        cell.fill = H_FILL
        cell.alignment = center
        cell.border = border


def _set_widths(ws, widths: list[int]) -> None:
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _note(ws, row: int, cols: int, text: str) -> None:
    """Fila de nota merge-celled debajo de los datos."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", size=10, italic=True, color="6E6E73")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = INST_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 30


# ─── HOJA 1: FONDO ────────────────────────────────────────────────────
def build_fondo(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Fondo"
    _write_header(ws, ["Campo", "Valor a confirmar", "Default actual"])
    fondo_rows = [
        ("Código del fondo (no cambiar)", "FIP_CEHTA_ESG", "FIP_CEHTA_ESG"),
        ("Nombre completo", "FIP CEHTA ESG", "FIP CEHTA ESG"),
        ("Año vintage", 2024, 2024),
        ("Tamaño committed (USD)", None, "vacío — completar"),
        ("AUM actual (USD)", None, "vacío — recalcula del cashflow"),
        ("Fecha inception", "2024-01-15", "2024-01-15"),
        ("Fin período de inversión", "2027-12-31", "2027-12-31"),
        ("Fin de vida del fondo", "2034-12-31", "2034-12-31"),
        ("Moneda base", "USD", "USD"),
        ("Administradora", "AFIS S.A.", "AFIS S.A."),
        ("Regulador", "CMF Chile · NCG 532 + NCG 554", "CMF Chile · NCG 532 + NCG 554"),
    ]
    for i, (label, val, default) in enumerate(fondo_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=i, column=1).border = border
        c = ws.cell(row=i, column=2, value=val)
        c.font = Font(name="Arial", size=10, color="0000FF")
        c.border = border
        c.fill = EXAMPLE_FILL
        if i in (5, 6):  # números USD
            c.number_format = "$#,##0"
        d = ws.cell(row=i, column=3, value=default)
        d.font = Font(name="Arial", size=10, color="6E6E73")
        d.border = border
    _set_widths(ws, [38, 30, 35])
    _note(ws, len(fondo_rows) + 3, 3,
          "Completar la columna B (azul) con valores reales. El script de import "
          "actualizará core.funds. Los campos que dejes vacíos no se tocan.")


# ─── HOJA 2: INVERSIONISTAS ───────────────────────────────────────────
def build_lps(wb: Workbook) -> None:
    ws = wb.create_sheet("Inversionistas")
    _write_header(ws, ["LP Name (existente)", "Tipo", "Commitment USD", "Paid-in USD", "Distributed USD", "Ownership %"])
    lp_rows = [
        ("CORFO - Corporacion de Fomento de la Produccion", "publico_corfo", None, 0, 0, None),
        ("Aportante Privado #1 (placeholder)", "privado", None, 0, 0, None),
    ]
    for i, row in enumerate(lp_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10, color="0000FF" if c in (3, 4, 5, 6) else "1D1D1F")
            if c in (3, 4, 5):
                cell.number_format = "$#,##0"
            if c == 6:
                cell.number_format = "0.00%"
            if c >= 3:
                cell.fill = EXAMPLE_FILL
    _set_widths(ws, [50, 20, 18, 18, 18, 14])
    _note(ws, len(lp_rows) + 3, 6,
          "Renombrá el 'Aportante Privado #1' con el nombre real cuando lo tengas. "
          "Completá los montos. La suma de ownership % debe ser 100%.")


# ─── HOJA 3: VALUACIONES ──────────────────────────────────────────────
def build_valuations(wb: Workbook) -> None:
    ws = wb.create_sheet("Valuaciones")
    headers = [
        "empresa_codigo", "as_of_date (YYYY-MM-DD)", "invested_amount_usd",
        "realized_value_usd", "unrealized_fv_usd", "moic_gross", "moic_net",
        "irr_gross", "irr_net", "valuation_method", "notes",
    ]
    _write_header(ws, headers)
    # Ejemplos: 1 por empresa, último trimestre cerrado
    today = datetime.now().strftime("%Y-%m-%d")
    example_rows = [
        ("CSL", "2025-12-31", 1_500_000, 0, 0, None, None, None, None, "DCF / cost", "Ejemplo — borrar y completar"),
        ("RHO", "2025-12-31", 2_500_000, 0, 0, None, None, None, None, "Comparables", "Ejemplo — borrar y completar"),
    ]
    for i, row in enumerate(example_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.fill = AMBER_LIGHT_FILL = PatternFill("solid", start_color=AMBER_LIGHT)
            cell.fill = AMBER_LIGHT_FILL
            if c in (3, 4, 5):
                cell.number_format = "$#,##0"
            if c in (6, 7):
                cell.number_format = "0.00\"x\""
            if c in (8, 9):
                cell.number_format = "0.00%"
    _set_widths(ws, [14, 22, 16, 16, 16, 12, 12, 12, 12, 22, 35])
    _note(ws, len(example_rows) + 3, 11,
          "Una fila por empresa portfolio por trimestre. Códigos válidos: " +
          ", ".join(PORTFOLIO_EMPRESAS) + ". Cargá tantas filas como trimestres tengás.")


# ─── HOJA 4: CASHFLOWS ────────────────────────────────────────────────
def build_cashflows(wb: Workbook) -> None:
    ws = wb.create_sheet("Cashflows")
    headers = [
        "effective_date (YYYY-MM-DD)", "lp_name (opcional)", "cashflow_type",
        "amount_usd", "descripcion", "ilpa_category", "recallable (TRUE/FALSE)",
    ]
    _write_header(ws, headers)
    example_rows = [
        ("2024-04-15", "CORFO - Corporacion de Fomento de la Produccion",
         "capital_call", 2_000_000, "Capital Call #1 CORFO", "Capital Contribution", True),
        ("2024-04-15", "Aportante Privado #1 (placeholder)",
         "capital_call", 2_000_000, "Capital Call #1 Privado", "Capital Contribution", True),
        ("2024-06-30", "", "management_fee", -56_250, "Mgmt fee Q2-2024 (2.5%/yr × $9M called)", "Mgmt Fee", False),
        ("2025-12-15", "CORFO - Corporacion de Fomento de la Produccion",
         "distribution", 500_000, "Distribución parcial CSL exit", "Distribution Return of Capital", False),
    ]
    for i, row in enumerate(example_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=BLUE_LIGHT)
            if c == 4:
                cell.number_format = "$#,##0;($#,##0)"
    _set_widths(ws, [22, 48, 18, 16, 38, 26, 14])
    _note(ws, len(example_rows) + 3, 7,
          "cashflow_type: capital_call | management_fee | distribution | expense | other. "
          "Para fees del fondo dejá lp_name vacío. Borrá las filas de ejemplo antes de subir.")


# ─── HOJA 5: KPIs OPERATIVOS ──────────────────────────────────────────
def build_kpis(wb: Workbook) -> None:
    ws = wb.create_sheet("KPIs Operativos")
    headers = [
        "empresa_codigo", "period (YYYY-MM-01)", "revenue_usd", "ebitda_usd",
        "ebitda_margin", "gross_margin", "cash_balance_usd", "burn_rate_usd",
        "cash_runway_months", "headcount", "mw_installed", "capacity_factor", "notes",
    ]
    _write_header(ws, headers)
    example_rows = [
        ("CSL", "2025-12-01", 480_000, 95_000, 0.198, 0.42, 250_000, None, None, 12, None, None, "Q4 2025"),
        ("RHO", "2025-12-01", 1_200_000, 380_000, 0.317, 0.55, 800_000, None, None, 18, 24.5, 0.31, "Q4 2025"),
    ]
    for i, row in enumerate(example_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=GREEN_LIGHT)
            if c in (3, 4, 7, 8):
                cell.number_format = "$#,##0"
            if c in (5, 6, 12):
                cell.number_format = "0.0%"
    _set_widths(ws, [14, 18, 14, 14, 12, 12, 14, 14, 12, 10, 12, 12, 30])
    _note(ws, len(example_rows) + 3, 13,
          "Una fila por empresa por mes. mw_installed + capacity_factor solo aplican "
          "para RHO (generación). Margins en decimal: 0.20 = 20%.")


# ─── HOJA 6: IMPACT METRICS ───────────────────────────────────────────
def build_impact(wb: Workbook) -> None:
    ws = wb.create_sheet("Impact Metrics")
    headers = [
        "empresa_codigo", "period (YYYY-MM-DD)", "iris_metric_id",
        "metric_name", "metric_value", "unit", "framework", "verified (TRUE/FALSE)", "verifier",
    ]
    _write_header(ws, headers)
    example_rows = [
        ("RHO", "2025-12-31", "PI2764", "CO2 emissions avoided", 1250, "tCO2e/year", "IRIS+ v5.3", False, ""),
        ("CSL", "2025-12-31", "PI5842", "Energy generated from renewable sources", 4200, "MWh", "IRIS+ v5.3", False, ""),
        ("TRONGKAI", "2025-12-31", "OI2535", "Jobs created", 8, "FTE", "IRIS+ v5.3", False, ""),
        ("REVTECH", "2025-12-31", "PI4060", "Waste recycled", 320, "tons/year", "IRIS+ v5.3", False, ""),
    ]
    for i, row in enumerate(example_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=GREEN_LIGHT)
            if c == 5:
                cell.number_format = "#,##0.00"
    _set_widths(ws, [14, 18, 14, 36, 14, 16, 14, 14, 20])
    _note(ws, len(example_rows) + 3, 9,
          "Códigos IRIS+ habituales: PI2764 (CO2 evitado), PI5842 (energía renovable), "
          "OI2535 (jobs), PI4060 (residuos reciclados). verified=TRUE si fue auditado por tercero.")


# ─── HOJA 7: Instrucciones ────────────────────────────────────────────
def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instrucciones", 0)  # primera hoja
    inst = [
        ("Cehta Capital — Template Data REAL del Fondo", "title"),
        ("", ""),
        ("CÓMO USAR ESTE EXCEL", "h2"),
        ("1. Hay 6 hojas a llenar (de izquierda a derecha): Fondo, Inversionistas, Valuaciones, Cashflows, KPIs, Impact.", ""),
        ("2. En cada hoja, las celdas azules/coloreadas son las que DEBES llenar con tus datos reales.", ""),
        ("3. Las filas son EJEMPLOS — borralas antes de subir. Agregá tantas filas como necesites debajo.", ""),
        ("4. Guardá el Excel (Ctrl+S) y mandame el archivo, o subilo a Dropbox y avisame.", ""),
        ("5. Yo (o un admin) corre el script `import_fondo_real.py` que lee el Excel y popula la base.", ""),
        ("", ""),
        ("REGLAS DE LLENADO", "h2"),
        ("• Códigos de empresa portfolio válidos: CSL, RHO, DTE, REVTECH, EVOQUE, TRONGKAI.", ""),
        ("• Fechas: SIEMPRE en formato YYYY-MM-DD (ej. 2025-12-31).", ""),
        ("• Montos en USD (no CLP). Si tenés CLP, convertí al tipo de cambio del período.", ""),
        ("• Decimales con punto: 0.198 = 19.8%.", ""),
        ("• Si no tenés un dato, dejá la celda vacía — NO inventes números.", ""),
        ("", ""),
        ("EFECTO EN LA PLATAFORMA", "h2"),
        ("Una vez cargado, vas a ver en Dashboard Institucional:", ""),
        ("• AUM real, TVPI, DPI, RVPI, MOIC del fondo.", ""),
        ("• J-Curve real con cashflow acumulado.", ""),
        ("• Tabla de portfolio companies con MOIC reales.", ""),
        ("• Sparklines de KPIs operativos en cada empresa.", ""),
        ("• Métricas IRIS+ agregadas para reportar a CORFO.", ""),
        ("", ""),
        (f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "footnote"),
    ]
    for i, (text, style) in enumerate(inst, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Arial", size=16, bold=True, color=GREEN)
        elif style == "h2":
            cell.font = Font(name="Arial", size=12, bold=True, color="1D1D1F")
            cell.fill = INST_FILL
        elif style == "footnote":
            cell.font = Font(name="Arial", size=9, italic=True, color="6E6E73")
        else:
            cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions["A"].width = 100


# ─── MAIN ─────────────────────────────────────────────────────────────
def main() -> None:
    wb = Workbook()
    build_fondo(wb)
    build_lps(wb)
    build_valuations(wb)
    build_cashflows(wb)
    build_kpis(wb)
    build_impact(wb)
    build_instructions(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"[OK] Generado: {OUTPUT}")
    print(f"     Hojas: Instrucciones, Fondo, Inversionistas, Valuaciones, Cashflows, KPIs Operativos, Impact Metrics")


if __name__ == "__main__":
    main()
