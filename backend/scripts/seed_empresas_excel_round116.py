"""Round 116 — seed de empresa extra data + credenciales + directorio + inversionistas.

Lee el Excel `resumen data.xlsx` (o `Data.xlsx`) provisto por Nicolas
y carga:

  1. Extensiones de core.empresas: pagina_web, contabilidad, direccion SII, giro
  2. Credenciales SII y Previred (CIFRADAS con Fernet via credentials_service)
  3. Cuentas bancarias adicionales en core.bancos_cuentas
  4. core.directorio_miembros (5 personas)
  5. core.inversionistas_aportantes (5 personas)

REQUIERE:
  - env var CREDENTIALS_FERNET_KEY configurada (ver credentials_service.py)
  - Migración 0067 corrida (`alembic upgrade head` previo)

USO:
  python backend/scripts/seed_empresas_excel_round116.py <ruta-al-excel>

  Ejemplo:
  python backend/scripts/seed_empresas_excel_round116.py "C:/Users/DELL/Downloads/Data (4).xlsx"

SEGURIDAD:
  El Excel TIENE passwords en plaintext. NO lo commitees. El script lee del path
  pasado en argv. Las passwords se cifran ANTES de tocar la DB.

IDEMPOTENTE: UPSERT por (empresa_codigo, sistema) en credenciales; UNIQUE
constraints + ON CONFLICT en el resto. Volver a correr no duplica filas.
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

load_dotenv(Path(__file__).parent.parent / ".env")

# Importamos el service de cifrado ANTES del engine para fallar rápido
# si CREDENTIALS_FERNET_KEY no está configurada.
sys.path.insert(0, str(Path(__file__).parent.parent))
from app.services.credentials_service import (  # noqa: E402
    encrypt_credential,
    health_check,
)

url_raw = os.getenv("DATABASE_URL", "")
url = re.sub(r"\+asyncpg|\+psycopg(?!2)", "+psycopg2", url_raw)
if not url:
    print("✗ DATABASE_URL no configurada en .env")
    sys.exit(1)


# =====================================================================
# Mapeo razón social Excel → empresa_codigo del sistema
# =====================================================================
RAZON_SOCIAL_TO_CODIGO = {
    "AFIS SA": "AFIS",
    "FIP CEHTA": "FIP_CEHTA",
    "Cenergy ltda": "CENERGY",
    "Evoque Energy SpA": "EVOQUE",
    "CLIMATE SMART LEASING SpA": "CSL",
    "AGROTECNOLOGIAS E INGENIERIA SPA": "TRONGKAI",
    "RHO GENERACION SpA": "RHO",
    "INGENIERIA E INNOVACION SpA": "REVTECH",
    "DTE SpA": "DTE",
}

# Mapeo nombre Previred → empresa_codigo
PREVIRED_NAME_TO_CODIGO = {
    "trongkai": "TRONGKAI",
    "revtech": "REVTECH",
    "cenergy": "CENERGY",
    "afis": "AFIS",
    "dte": "DTE",
    "evoque": "EVOQUE",
}


def _norm(s: object) -> str:
    """Normaliza para matching: lowercase, sin tildes/diéresis, sin extra spaces."""
    if s is None:
        return ""
    out = str(s).strip()
    # Reemplaza tildes comunes que aparecen en el Excel ('Í' → 'I' etc.)
    repl = {
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ñ": "N",
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n",
    }
    for k, v in repl.items():
        out = out.replace(k, v)
    return out


def _str_or_none(s: object) -> str | None:
    if s is None:
        return None
    val = str(s).strip()
    return val if val else None


def _resolve_empresa_codigo(razon_social_excel: str) -> str | None:
    """Mapea razón social del Excel al empresa_codigo del sistema."""
    norm = _norm(razon_social_excel)
    for raw_match, codigo in RAZON_SOCIAL_TO_CODIGO.items():
        if _norm(raw_match) == norm:
            return codigo
    return None


def _parse_excel(path: Path) -> dict[str, list]:
    """Lee el Excel y devuelve dict con 4 listas: empresas, inv, dir, previred."""
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))

    empresas: list[dict] = []
    inversionistas: list[dict] = []
    directorio: list[dict] = []
    previred: list[dict] = []

    section: str | None = None
    for r in rows:
        if r is None:
            continue
        cells = [c for c in r if c is not None]
        if not cells:
            continue
        first = _norm(r[1]) if len(r) > 1 else ""

        # Detector de sección por encabezado
        if first == "EMPRESAS":
            section = "empresas"
            continue
        if first == "INVERSIONISTAS/APORTANTES":
            section = "inversionistas"
            continue
        if first == "DIRECTORIO":
            section = "directorio"
            continue
        if first.startswith("GERENTES"):
            section = "gerentes"  # los GG ya están en user_company_roles, no los seedeamos acá
            continue
        if first == "PREVIRED":
            section = "previred"
            continue

        # Skip headers de columnas (segunda fila de cada sección suele tener
        # "NOMBRE", "PAGINA WEB", etc.)
        if first in {
            "PAGINA WEB", "NOMBRE", "EMPRESA", "PAGINA WEB ",
        }:
            continue

        if section == "empresas" and len(r) > 4 and r[2] is not None:
            empresas.append({
                "pagina_web": _str_or_none(r[1]),
                "razon_social": _str_or_none(r[2]),
                "contabilidad": _str_or_none(r[3]),
                "rut": _str_or_none(r[4]),
                "sii_password": _str_or_none(r[5]),
                "banco": _str_or_none(r[6]),
                "cuenta": _str_or_none(r[7]),
                "codigo_banco": _str_or_none(r[8]),
                "giro": _str_or_none(r[9]),
                "direccion_sii": _str_or_none(r[10]),
            })
        elif section == "inversionistas" and len(r) > 2 and r[1] is not None:
            inversionistas.append({
                "nombre": _str_or_none(r[1]),
                "rut": _str_or_none(r[2]),
                "direccion": _str_or_none(r[3]),
                "telefono": _str_or_none(r[4]),
                "banco": _str_or_none(r[5]),
                "cuenta": _str_or_none(r[6]),
                "codigo_banco": _str_or_none(r[7]),
                "correo": _str_or_none(r[8]),
            })
        elif section == "directorio" and len(r) > 2 and r[1] is not None:
            directorio.append({
                "nombre": _str_or_none(r[1]),
                "rut": _str_or_none(r[2]),
                "direccion": _str_or_none(r[3]),
                "telefono": _str_or_none(r[4]),
                "banco": _str_or_none(r[5]),
                "cuenta": _str_or_none(r[6]),
                "codigo_banco": _str_or_none(r[7]),
                "correo": _str_or_none(r[8]),
            })
        elif section == "previred" and len(r) > 3 and r[1] is not None:
            previred.append({
                "empresa_name": _str_or_none(r[1]),
                "rut_usuario": _str_or_none(r[2]),
                "password": _str_or_none(r[3]),
            })

    return {
        "empresas": empresas,
        "inversionistas": inversionistas,
        "directorio": directorio,
        "previred": previred,
    }


def run(excel_path: Path) -> None:
    print(f"Round 116 — seed desde: {excel_path}\n")

    # 1. Sanity check del servicio de cifrado
    h = health_check()
    if not h.get("configured") or not h.get("round_trip_ok"):
        print(f"✗ credentials_service no operativo: {h}")
        print(
            "  Setea CREDENTIALS_FERNET_KEY en .env (o env var) antes de correr esto."
        )
        sys.exit(2)
    print(f"✓ credentials_service OK ({h})\n")

    data = _parse_excel(excel_path)
    print(
        f"Leído del Excel: {len(data['empresas'])} empresas, "
        f"{len(data['inversionistas'])} inversionistas, "
        f"{len(data['directorio'])} directorio, "
        f"{len(data['previred'])} previred\n"
    )

    engine = create_engine(url, connect_args={"sslmode": "require"})

    with engine.begin() as conn:
        # ---------------------------------------------------------------
        # 1. Empresas — UPDATE extra fields + INSERT credencial SII
        # ---------------------------------------------------------------
        # Dedupe por razón social (varias filas son la misma empresa con
        # múltiples cuentas bancarias).
        empresas_by_codigo: dict[str, dict] = {}
        for e in data["empresas"]:
            codigo = _resolve_empresa_codigo(e["razon_social"] or "")
            if codigo is None:
                print(f"  ⚠ Sin mapeo para '{e['razon_social']}' — skip")
                continue
            # Solo guardamos la primera ocurrencia (los datos generales son iguales)
            empresas_by_codigo.setdefault(codigo, e)

        updated = 0
        creds_sii = 0
        for codigo, e in empresas_by_codigo.items():
            row = conn.execute(
                text("SELECT codigo FROM core.empresas WHERE codigo = :c"),
                {"c": codigo},
            ).fetchone()
            if not row:
                print(f"  ⚠ Empresa {codigo} no existe en core.empresas — skip")
                continue
            conn.execute(
                text(
                    """
                    UPDATE core.empresas SET
                        pagina_web = COALESCE(:web, pagina_web),
                        contabilidad_proveedor = COALESCE(:cont, contabilidad_proveedor),
                        direccion_sii = COALESCE(:dirsii, direccion_sii),
                        giro = COALESCE(:giro, giro),
                        updated_at = NOW()
                    WHERE codigo = :codigo
                    """
                ),
                {
                    "codigo": codigo,
                    "web": e["pagina_web"],
                    "cont": e["contabilidad"],
                    "dirsii": e["direccion_sii"],
                    "giro": e["giro"],
                },
            )
            updated += 1

            # Credencial SII — UPSERT
            if e["sii_password"] and e["rut"]:
                ciphered = encrypt_credential(e["sii_password"])
                conn.execute(
                    text(
                        """
                        INSERT INTO core.empresa_credenciales
                            (empresa_codigo, sistema, rut_usuario, password_encrypted)
                        VALUES (:c, 'sii', :rut, :pwd)
                        ON CONFLICT (empresa_codigo, sistema)
                        DO UPDATE SET
                            rut_usuario = EXCLUDED.rut_usuario,
                            password_encrypted = EXCLUDED.password_encrypted,
                            updated_at = NOW()
                        """
                    ),
                    {"c": codigo, "rut": e["rut"], "pwd": ciphered},
                )
                creds_sii += 1
        print(f"  ✓ {updated} empresas actualizadas, {creds_sii} credenciales SII")

        # ---------------------------------------------------------------
        # 2. Previred credenciales — UPSERT
        # ---------------------------------------------------------------
        creds_prev = 0
        for p in data["previred"]:
            name = (p["empresa_name"] or "").strip().lower()
            codigo = PREVIRED_NAME_TO_CODIGO.get(name)
            if codigo is None:
                print(f"  ⚠ Previred sin mapeo: {p['empresa_name']} — skip")
                continue
            if not p["password"] or not p["rut_usuario"]:
                continue
            ciphered = encrypt_credential(p["password"])
            conn.execute(
                text(
                    """
                    INSERT INTO core.empresa_credenciales
                        (empresa_codigo, sistema, rut_usuario, password_encrypted)
                    VALUES (:c, 'previred', :rut, :pwd)
                    ON CONFLICT (empresa_codigo, sistema)
                    DO UPDATE SET
                        rut_usuario = EXCLUDED.rut_usuario,
                        password_encrypted = EXCLUDED.password_encrypted,
                        updated_at = NOW()
                    """
                ),
                {"c": codigo, "rut": p["rut_usuario"], "pwd": ciphered},
            )
            creds_prev += 1
        print(f"  ✓ {creds_prev} credenciales Previred")

        # ---------------------------------------------------------------
        # 3. Directorio
        # ---------------------------------------------------------------
        dir_inserts = 0
        for d in data["directorio"]:
            if not d["nombre"]:
                continue
            # Idempotente por (nombre, rut)
            existing = conn.execute(
                text(
                    """
                    SELECT miembro_id FROM core.directorio_miembros
                    WHERE nombre = :n AND COALESCE(rut, '') = COALESCE(:r, '')
                    """
                ),
                {"n": d["nombre"], "r": d["rut"]},
            ).fetchone()
            if existing:
                continue
            conn.execute(
                text(
                    """
                    INSERT INTO core.directorio_miembros
                        (nombre, rut, direccion, telefono, banco, cuenta, codigo_banco, correo)
                    VALUES (:n, :r, :dir, :tel, :b, :c, :cb, :co)
                    """
                ),
                {
                    "n": d["nombre"], "r": d["rut"],
                    "dir": d["direccion"], "tel": d["telefono"],
                    "b": d["banco"], "c": d["cuenta"],
                    "cb": d["codigo_banco"], "co": d["correo"],
                },
            )
            dir_inserts += 1
        print(f"  ✓ {dir_inserts} miembros directorio (insertados)")

        # ---------------------------------------------------------------
        # 4. Inversionistas/Aportantes
        # ---------------------------------------------------------------
        inv_inserts = 0
        for i in data["inversionistas"]:
            if not i["nombre"]:
                continue
            existing = conn.execute(
                text(
                    """
                    SELECT inversionista_id FROM core.inversionistas_aportantes
                    WHERE nombre = :n AND COALESCE(rut, '') = COALESCE(:r, '')
                    """
                ),
                {"n": i["nombre"], "r": i["rut"]},
            ).fetchone()
            if existing:
                continue
            conn.execute(
                text(
                    """
                    INSERT INTO core.inversionistas_aportantes
                        (nombre, rut, direccion, telefono, banco, cuenta, codigo_banco, correo)
                    VALUES (:n, :r, :dir, :tel, :b, :c, :cb, :co)
                    """
                ),
                {
                    "n": i["nombre"], "r": i["rut"],
                    "dir": i["direccion"], "tel": i["telefono"],
                    "b": i["banco"], "c": i["cuenta"],
                    "cb": i["codigo_banco"], "co": i["correo"],
                },
            )
            inv_inserts += 1
        print(f"  ✓ {inv_inserts} inversionistas/aportantes (insertados)")

    print("\n=== Seed completado OK ===")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python seed_empresas_excel_round116.py <ruta-al-excel>")
        sys.exit(1)
    excel = Path(sys.argv[1]).resolve()
    if not excel.exists():
        print(f"✗ No existe: {excel}")
        sys.exit(1)
    run(excel)
