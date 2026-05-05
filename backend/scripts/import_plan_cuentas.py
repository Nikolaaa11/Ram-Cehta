"""Importer del Plan de Cuentas v2 (V5).

Lee `Plan_de_cuentas_v2.xlsx` y carga las 469 cuentas + matriz de
habilitación por empresa. Idempotente (UPSERT por `codigo`), seguro
para re-correr cuando el COO actualiza el Excel.

Uso:

    # Dry-run — solo imprime estadísticas y primeras filas, no toca DB
    python -m scripts.import_plan_cuentas

    # Apply — escribe a la DB
    python -m scripts.import_plan_cuentas --apply

    # Excel en otra ruta
    python -m scripts.import_plan_cuentas --xlsx /ruta/a/Plan_de_cuentas_v3.xlsx --apply

Reglas:
- Inserción en orden por nivel (1 → 4) para respetar FK self-referencial.
- Habilitación por empresa: para cada (cuenta × empresa) con Hab_X = TRUE,
  inserta row en core.plan_cuenta_empresa con habilitada=true.
- CENERGY: replica las habilitaciones de DTE (no está en el Excel pero
  el COO la mantiene activa; CENERGY es servicios, idéntico giro a DTE).
- Mapeo flags Excel → DB: el Excel usa 'x' para true, NULL/None para false.

Salida: tabla resumen al final con counts por nivel/tipo/CORFO + total
de habilitaciones creadas.
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: falta openpyxl. Instalá con: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from sqlalchemy import text

from app.core.database import SessionLocal


# Empresas del Excel (orden importa para iterar las columnas Hab_X)
EMPRESAS_EXCEL = ["CSL", "RHO", "DTE", "RVT", "EVQ", "TRK", "AFIS", "FIP"]

# Mapeo de prefijo de empresa a código real en core.empresas.codigo
EMPRESA_CODE_MAP = {
    "CSL": "CSL",
    "RHO": "RHO",
    "DTE": "DTE",
    "RVT": "REVTECH",
    "EVQ": "EVOQUE",
    "TRK": "TRONGKAI",
    "AFIS": "AFIS",
    "FIP": "FIP_CEHTA",
}

# CENERGY replica las habilitaciones de DTE (acordado con COO).
# Si querés cambiar la fuente, modificá esta constante.
CENERGY_REPLICA_FROM = "DTE"


# ---------------------------------------------------------------------
# Helpers de parseo
# ---------------------------------------------------------------------


def _bool_x(value: Any) -> bool:
    """Excel marca true como 'x', false como None/empty. Robust parser."""
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ("x", "1", "true", "si", "sí")


def _bool_excel(value: Any) -> bool:
    """Para columnas TRUE/FALSE explícitas (Activa, Hab_X)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("true", "1", "si", "sí", "x")


def _str_or_none(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _int_or_none(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------
# Parse Excel → estructuras de datos
# ---------------------------------------------------------------------


def parse_xlsx(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    """Devuelve (cuentas, habilitaciones_por_empresa).

    cuentas: lista de dicts con todos los campos para INSERT en core.plan_cuentas.
    habilitaciones_por_empresa: dict empresa_codigo → set de códigos de cuenta habilitados.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "PlanDeCuentas" not in wb.sheetnames:
        raise ValueError(
            f"El Excel no tiene hoja 'PlanDeCuentas'. Hojas: {wb.sheetnames}"
        )
    ws = wb["PlanDeCuentas"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Hoja PlanDeCuentas vacía")

    header = [str(c).strip() if c else "" for c in rows[0]]
    expected = [
        "Cuenta", "Nivel", "Tipo", "Descripcion", "CuentaPadre",
        "Imputable", "IvaTratamiento", "CorfoElegible", "TipoGastoCorfo",
        "NuboxCode", "CodigoF22", "Ajuste14D",
        "FlagPartida", "FlagConcepto", "FlagCapital", "FlagActivoFijo",
        "FlagDocumento", "FlagControlGestion", "FlagActivoNeto", "FlagCaja",
        "Flag14D", "FlagPercepcion",
        "Activa",
        "Hab_CSL", "Hab_RHO", "Hab_DTE", "Hab_RVT",
        "Hab_EVQ", "Hab_TRK", "Hab_AFIS", "Hab_FIP",
    ]
    missing = [c for c in expected if c not in header]
    if missing:
        raise ValueError(f"Columnas faltantes en Excel: {missing}")

    idx = {col: header.index(col) for col in expected}

    cuentas: list[dict[str, Any]] = []
    habilitaciones: dict[str, set[str]] = defaultdict(set)

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        codigo = str(row[idx["Cuenta"]]).strip()
        if not codigo:
            continue

        nivel = int(row[idx["Nivel"]])
        tipo = str(row[idx["Tipo"]]).strip()

        cuenta = {
            "codigo": codigo,
            "nivel": nivel,
            "tipo": tipo,
            "nombre": str(row[idx["Descripcion"]]).strip(),
            "descripcion": None,  # el Excel no diferencia nombre vs desc detallada
            "codigo_padre": _str_or_none(row[idx["CuentaPadre"]]),
            "imputable": _bool_excel(row[idx["Imputable"]]),
            "iva_tratamiento": _str_or_none(row[idx["IvaTratamiento"]]) or "NA",
            "corfo_elegible": _bool_excel(row[idx["CorfoElegible"]]),
            "tipo_gasto_corfo": _str_or_none(row[idx["TipoGastoCorfo"]]),
            "nubox_code": _str_or_none(row[idx["NuboxCode"]]) or codigo,
            "codigo_f22": _int_or_none(row[idx["CodigoF22"]]),
            "ajuste_14d": _str_or_none(row[idx["Ajuste14D"]]),
            "flag_partida": _bool_x(row[idx["FlagPartida"]]),
            "flag_concepto": _bool_x(row[idx["FlagConcepto"]]),
            "flag_capital": _bool_x(row[idx["FlagCapital"]]),
            "flag_activo_fijo": _bool_x(row[idx["FlagActivoFijo"]]),
            "flag_documento": _bool_x(row[idx["FlagDocumento"]]),
            "flag_control_gestion": _bool_x(row[idx["FlagControlGestion"]]),
            "flag_activo_neto": _bool_x(row[idx["FlagActivoNeto"]]),
            "flag_caja": _bool_x(row[idx["FlagCaja"]]),
            "flag_marca_14d": _bool_x(row[idx["Flag14D"]]),
            "flag_percepcion": _bool_x(row[idx["FlagPercepcion"]]),
            "activa": _bool_excel(row[idx["Activa"]]),
        }
        cuentas.append(cuenta)

        # Habilitaciones por empresa
        for empresa_excel in EMPRESAS_EXCEL:
            col = f"Hab_{empresa_excel}"
            if _bool_excel(row[idx[col]]):
                empresa_codigo = EMPRESA_CODE_MAP[empresa_excel]
                habilitaciones[empresa_codigo].add(codigo)

    # CENERGY replica las habilitaciones de la empresa fuente (default DTE)
    fuente_codigo = EMPRESA_CODE_MAP[CENERGY_REPLICA_FROM]
    if fuente_codigo in habilitaciones:
        habilitaciones["CENERGY"] = set(habilitaciones[fuente_codigo])

    return cuentas, dict(habilitaciones)


# ---------------------------------------------------------------------
# Apply a DB
# ---------------------------------------------------------------------


async def apply_to_db(
    cuentas: list[dict[str, Any]],
    habilitaciones: dict[str, set[str]],
) -> dict[str, int]:
    """Ejecuta el UPSERT contra la DB. Devuelve counters."""
    counters = {
        "cuentas_upserted": 0,
        "habilitaciones_upserted": 0,
        "habilitaciones_omitidas_empresa_inexistente": 0,
    }

    async with SessionLocal() as db:
        # Validar que todas las empresas referenciadas existen
        empresas_db = (
            (
                await db.execute(
                    text("SELECT codigo FROM core.empresas WHERE activo = TRUE")
                )
            )
            .scalars()
            .all()
        )
        empresas_db_set = set(empresas_db)

        # Inserción en orden por nivel (1 → 4) para FK self
        cuentas_ordenadas = sorted(cuentas, key=lambda c: (c["nivel"], c["codigo"]))

        for c in cuentas_ordenadas:
            await db.execute(
                text(
                    """
                    INSERT INTO core.plan_cuentas (
                        codigo, nivel, tipo, nombre, descripcion, codigo_padre,
                        imputable, iva_tratamiento,
                        corfo_elegible, tipo_gasto_corfo, nubox_code,
                        codigo_f22, ajuste_14d,
                        flag_partida, flag_concepto, flag_capital,
                        flag_activo_fijo, flag_documento, flag_control_gestion,
                        flag_activo_neto, flag_caja, flag_marca_14d,
                        flag_percepcion, activa
                    )
                    VALUES (
                        :codigo, :nivel, :tipo, :nombre, :descripcion, :codigo_padre,
                        :imputable, :iva_tratamiento,
                        :corfo_elegible, :tipo_gasto_corfo, :nubox_code,
                        :codigo_f22, :ajuste_14d,
                        :flag_partida, :flag_concepto, :flag_capital,
                        :flag_activo_fijo, :flag_documento, :flag_control_gestion,
                        :flag_activo_neto, :flag_caja, :flag_marca_14d,
                        :flag_percepcion, :activa
                    )
                    ON CONFLICT (codigo) DO UPDATE SET
                        nivel = EXCLUDED.nivel,
                        tipo = EXCLUDED.tipo,
                        nombre = EXCLUDED.nombre,
                        descripcion = EXCLUDED.descripcion,
                        codigo_padre = EXCLUDED.codigo_padre,
                        imputable = EXCLUDED.imputable,
                        iva_tratamiento = EXCLUDED.iva_tratamiento,
                        corfo_elegible = EXCLUDED.corfo_elegible,
                        tipo_gasto_corfo = EXCLUDED.tipo_gasto_corfo,
                        nubox_code = EXCLUDED.nubox_code,
                        codigo_f22 = EXCLUDED.codigo_f22,
                        ajuste_14d = EXCLUDED.ajuste_14d,
                        flag_partida = EXCLUDED.flag_partida,
                        flag_concepto = EXCLUDED.flag_concepto,
                        flag_capital = EXCLUDED.flag_capital,
                        flag_activo_fijo = EXCLUDED.flag_activo_fijo,
                        flag_documento = EXCLUDED.flag_documento,
                        flag_control_gestion = EXCLUDED.flag_control_gestion,
                        flag_activo_neto = EXCLUDED.flag_activo_neto,
                        flag_caja = EXCLUDED.flag_caja,
                        flag_marca_14d = EXCLUDED.flag_marca_14d,
                        flag_percepcion = EXCLUDED.flag_percepcion,
                        activa = EXCLUDED.activa,
                        updated_at = now()
                    """
                ),
                c,
            )
            counters["cuentas_upserted"] += 1

        # Habilitaciones
        for empresa_codigo, codigos_cuenta in habilitaciones.items():
            if empresa_codigo not in empresas_db_set:
                counters["habilitaciones_omitidas_empresa_inexistente"] += len(
                    codigos_cuenta
                )
                continue
            for cuenta_codigo in codigos_cuenta:
                await db.execute(
                    text(
                        """
                        INSERT INTO core.plan_cuenta_empresa (
                            cuenta_codigo, empresa_codigo, habilitada
                        )
                        VALUES (:cc, :ec, TRUE)
                        ON CONFLICT (cuenta_codigo, empresa_codigo) DO UPDATE
                            SET habilitada = TRUE,
                                habilitada_en = now()
                        """
                    ),
                    {"cc": cuenta_codigo, "ec": empresa_codigo},
                )
                counters["habilitaciones_upserted"] += 1

        await db.commit()

    return counters


# ---------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------


def report(
    cuentas: list[dict[str, Any]],
    habilitaciones: dict[str, set[str]],
) -> None:
    print()
    print("=" * 70)
    print("PLAN DE CUENTAS — REPORTE DE IMPORTACIÓN")
    print("=" * 70)
    print(f"Total cuentas: {len(cuentas)}")

    by_nivel = Counter(c["nivel"] for c in cuentas)
    print("\nPor nivel:")
    for n in sorted(by_nivel):
        print(f"  Nivel {n}: {by_nivel[n]}")

    by_tipo = Counter(c["tipo"] for c in cuentas)
    print("\nPor tipo:")
    for t, n in sorted(by_tipo.items()):
        print(f"  {t}: {n}")

    imputables = sum(1 for c in cuentas if c["imputable"])
    print(f"\nImputables (nivel 4): {imputables}")

    corfo = sum(1 for c in cuentas if c["corfo_elegible"])
    print(f"Marcadas CORFO elegible: {corfo}")
    by_corfo_tipo = Counter(
        c["tipo_gasto_corfo"] for c in cuentas if c["corfo_elegible"]
    )
    if by_corfo_tipo:
        print("  Por tipo de gasto CORFO:")
        for t, n in sorted(by_corfo_tipo.items(), key=lambda x: str(x[0])):
            print(f"    {t}: {n}")

    f22 = sum(1 for c in cuentas if c["codigo_f22"] is not None)
    print(f"\nCon código F22 (declaración renta): {f22}")

    print("\nHabilitaciones por empresa:")
    for empresa, codigos in sorted(habilitaciones.items()):
        print(f"  {empresa}: {len(codigos)} cuentas habilitadas")

    print("\nPrimeras 5 cuentas (verificación):")
    for c in cuentas[:5]:
        print(
            f"  {c['codigo']} L{c['nivel']} {c['tipo']:8s} "
            f"imp={c['imputable']:1} corfo={c['corfo_elegible']:1} "
            f"{c['nombre'][:50]}"
        )

    print()


# ---------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Importer del Plan de Cuentas v2 a core.plan_cuentas"
    )
    default_xlsx = Path.home() / "Downloads" / "Plan_de_cuentas_v2.xlsx"
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=default_xlsx,
        help=f"Ruta al .xlsx (default: {default_xlsx})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Si se omite, solo imprime el reporte sin tocar la DB",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: no se encuentra {args.xlsx}", file=sys.stderr)
        return 1

    print(f"Leyendo {args.xlsx}...")
    cuentas, habilitaciones = parse_xlsx(args.xlsx)
    report(cuentas, habilitaciones)

    if not args.apply:
        print("Modo DRY-RUN — no se tocó la DB.")
        print("Para aplicar: python -m scripts.import_plan_cuentas --apply")
        return 0

    print("Aplicando a la DB (puede tomar 30-60s para 469 cuentas)...")
    counters = asyncio.run(apply_to_db(cuentas, habilitaciones))
    print()
    print(f"  Cuentas upserted:           {counters['cuentas_upserted']}")
    print(f"  Habilitaciones upserted:    {counters['habilitaciones_upserted']}")
    if counters["habilitaciones_omitidas_empresa_inexistente"] > 0:
        print(
            f"  Habilitaciones omitidas (empresa no encontrada): "
            f"{counters['habilitaciones_omitidas_empresa_inexistente']}"
        )
    print()
    print("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
