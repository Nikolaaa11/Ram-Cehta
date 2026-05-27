"""Genera template Excel para bulk-import de vouchers via /vouchers/import-csv.

Output: C:/Users/DELL/Documents/Template_Vouchers_Import.xlsx con 3 hojas:
  1. Vouchers       — la hoja a llenar con datos (con 4 ejemplos pre-cargados)
  2. Valores válidos — listas de empresa_codigo / tipo / doc_tributario_tipo
  3. Instrucciones   — reglas, alias de columnas, validaciones

NOTA: el endpoint backend acepta CSV separado por `;`. Cuando termines de
llenar la hoja Vouchers, exporta a CSV (UTF-8 BOM, separador ;) y subí
ese .csv en /vouchers/import.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path("C:/Users/DELL/Documents/Template_Vouchers_Import.xlsx")

# =====================================================================
# Configuración estilos
# =====================================================================
GREEN = "1D6F42"
GREEN_LIGHT = "E7F3EC"
RED_LIGHT = "FEE2E2"
AMBER_LIGHT = "FEF3C7"
BLUE_LIGHT = "DBEAFE"
GRAY_LIGHT = "F5F5F7"

thin = Side(border_style="thin", color="D2D2D7")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color=GREEN)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =====================================================================
# Columnas — tomadas del backend voucher_csv_import_service.py
# =====================================================================
COLUMNS = [
    ("voucher_ref",         "OBL", "STRING",  "Agrupa filas del mismo voucher",            "V001"),
    ("empresa_codigo",      "OBL", "STRING",  "Código empresa (ver hoja Valores)",         "CSL"),
    ("tipo",                "OBL", "ENUM",    "INGRESO/EGRESO/COMPRA/VENTA/TRASPASO",      "COMPRA"),
    ("fecha_documento",     "OBL", "DATE",    "YYYY-MM-DD o DD-MM-YYYY",                    "2025-01-15"),
    ("fecha_contable",      "OBL", "DATE",    "Período contable (puede diferir del doc)",  "2025-01-15"),
    ("glosa",               "OBL", "STRING",  "Descripción general del voucher",           "Compra insumos enero"),
    ("contraparte_rut",     "OPC", "RUT",     "RUT proveedor/cliente (formato chileno)",   "76.123.456-7"),
    ("contraparte_nombre",  "OPC", "STRING",  "Razón social proveedor/cliente",            "Office Depot SpA"),
    ("doc_tributario_tipo", "OPC", "ENUM",    "FACTURA/BOLETA/NOTA_CREDITO/...",           "FACTURA"),
    ("doc_tributario_folio","OPC", "STRING",  "Número de folio del documento",             "12345"),
    ("line_number",         "OBL", "INT",     "Correlativo línea dentro del voucher",      "1"),
    ("cuenta_codigo",       "OBL", "STRING",  "Código del plan de cuentas",                "5-01-01-001"),
    ("proyecto_codigo",     "OPC", "STRING",  "Código proyecto (si aplica)",               ""),
    ("area_codigo",         "OPC", "STRING",  "3 letras: ADM/COM/OPE/IDA",                 "ADM"),
    ("debit",               "OPC*","DECIMAL", "Monto debe (CLP, ej: 100000)",              "100000"),
    ("credit",              "OPC*","DECIMAL", "Monto haber (CLP, ej: 119000)",             "0"),
    ("descripcion",         "OPC", "STRING",  "Descripción específica de la línea",        "Insumos oficina"),
]

# *OPC: Una línea debe tener debit XOR credit (uno de los dos > 0).


# =====================================================================
# Ejemplos: 4 vouchers completos (Compra + Venta + Pago + Traspaso)
# =====================================================================
EXAMPLES = [
    # Voucher 1: COMPRA con IVA — 3 líneas
    ["V001","CSL","COMPRA","2025-01-15","2025-01-15","Compra insumos oficina enero",
     "76.123.456-7","Office Depot SpA","FACTURA","12345",
     1,"5-01-01-001","","ADM",100000,0,"Insumos consumibles"],
    ["V001","CSL","COMPRA","2025-01-15","2025-01-15","Compra insumos oficina enero",
     "76.123.456-7","Office Depot SpA","FACTURA","12345",
     2,"1-01-02-001","","",19000,0,"IVA crédito fiscal"],
    ["V001","CSL","COMPRA","2025-01-15","2025-01-15","Compra insumos oficina enero",
     "76.123.456-7","Office Depot SpA","FACTURA","12345",
     3,"2-02-01-001","","",0,119000,"Cuenta por pagar Office Depot"],

    # Voucher 2: VENTA — 3 líneas
    ["V002","RHO","VENTA","2025-02-10","2025-02-10","Venta consultoría energía",
     "77.999.888-K","Cliente Demo Ltda","FACTURA_ELECTRONICA","2001",
     1,"1-01-03-001","P-2025-001","COM",2380000,0,"Cuenta por cobrar cliente"],
    ["V002","RHO","VENTA","2025-02-10","2025-02-10","Venta consultoría energía",
     "77.999.888-K","Cliente Demo Ltda","FACTURA_ELECTRONICA","2001",
     2,"4-01-01-001","P-2025-001","COM",0,2000000,"Ingreso por servicios"],
    ["V002","RHO","VENTA","2025-02-10","2025-02-10","Venta consultoría energía",
     "77.999.888-K","Cliente Demo Ltda","FACTURA_ELECTRONICA","2001",
     3,"2-03-01-001","","",0,380000,"IVA débito fiscal"],

    # Voucher 3: PAGO (EGRESO) — 2 líneas
    ["V003","CSL","EGRESO","2025-01-25","2025-01-25","Pago factura Office Depot V001",
     "76.123.456-7","Office Depot SpA","NA","",
     1,"2-02-01-001","","",119000,0,"Cancela Cta x Pagar"],
    ["V003","CSL","EGRESO","2025-01-25","2025-01-25","Pago factura Office Depot V001",
     "76.123.456-7","Office Depot SpA","NA","",
     2,"1-01-01-002","","",0,119000,"Egreso Banco BCI"],

    # Voucher 4: TRASPASO entre cuentas internas — 2 líneas
    ["V004","CEHTA","TRASPASO","2025-03-01","2025-03-01","Traspaso BCI → Itaú",
     "","","NA","",
     1,"1-01-01-003","","",5000000,0,"Ingreso Banco Itaú"],
    ["V004","CEHTA","TRASPASO","2025-03-01","2025-03-01","Traspaso BCI → Itaú",
     "","","NA","",
     2,"1-01-01-002","","",0,5000000,"Egreso Banco BCI"],
]


# =====================================================================
# Enums (tomados de app/schemas/voucher.py + seed empresas)
# =====================================================================
EMPRESAS = [
    ("CEHTA",      "Cehta Capital SpA (Holding)"),
    ("FIP_CEHTA",  "FIP Cehta ESG (Fondo)"),
    ("AFIS",       "AFIS"),
    ("CENERGY",    "Cenergy Ltda"),
    ("CSL",        "Climate Smart Leasing SpA"),
    ("DTE",        "DTE SpA"),
    ("EVOQUE",     "Evoque Energy SpA"),
    ("REVTECH",    "Ingeniería e Innovación SpA"),
    ("RHO",        "RHO Generación SpA"),
    ("TRONGKAI",   "Agrotecnologías e Ingeniería SpA"),
]

TIPOS_VOUCHER = [
    ("INGRESO",   "Ingreso de caja (cobro, depósito, abono)"),
    ("EGRESO",    "Egreso de caja (pago, transferencia salida)"),
    ("COMPRA",    "Factura de compra recibida"),
    ("VENTA",     "Factura de venta emitida"),
    ("TRASPASO",  "Movimiento entre cuentas propias"),
    ("APERTURA",  "Asiento de apertura ejercicio"),
    ("CIERRE",    "Asiento de cierre ejercicio"),
    ("REVERSO",   "Anulación de voucher previo"),
]

TIPOS_DOC_TRIBUTARIO = [
    "FACTURA",
    "FACTURA_ELECTRONICA",
    "FACTURA_ELECTRONICA_EXENTA",
    "FACTURA_EXENTA",
    "FACTURA_COMPRA",
    "FACTURA_COMPRA_ELECTRONICA",
    "FACTURA_IMPORTACION",
    "FACTURA_EXPORTACION",
    "FACTURA_INICIO",
    "BOLETA",
    "NOTA_CREDITO",
    "NOTA_CREDITO_ELECTRONICA",
    "NOTA_DEBITO",
    "NOTA_DEBITO_ELECTRONICA",
    "HONORARIOS",
    "LIQUIDACION_FACTURA",
    "LIQUIDACION_FACTURA_ELECTRONICA",
    "DECLARACION_INGRESO",
    "SOLICITUD_REGISTRO_FACTURA",
    "INVOICE",
    "NA",
]

ALIASES_HEADER = [
    ("voucher_ref",         "ref, referencia, numero, n"),
    ("empresa_codigo",      "empresa"),
    ("fecha_documento",     "fecha_doc"),
    ("fecha_contable",      "fecha"),
    ("glosa",               "descripcion_voucher"),
    ("contraparte_rut",     "rut"),
    ("contraparte_nombre",  "contraparte"),
    ("doc_tributario_tipo", "tipo_documento"),
    ("doc_tributario_folio","folio"),
    ("line_number",         "linea"),
    ("cuenta_codigo",       "cuenta"),
    ("proyecto_codigo",     "proyecto"),
    ("area_codigo",         "area"),
    ("debit",               "debe"),
    ("credit",              "haber"),
    ("descripcion",         "detalle"),
]


# =====================================================================
# Build workbook
# =====================================================================
def main() -> None:
    wb = Workbook()

    # ===== Hoja 1: Vouchers (datos editables) =====
    ws1 = wb.active
    ws1.title = "Vouchers"

    headers = [c[0] for c in COLUMNS]
    obligs = [c[1] for c in COLUMNS]

    # Header row 1: nombre columna
    for col_idx, name in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Header row 2: OBL / OPC marker
    oblig_font = Font(name="Arial", size=8, bold=True, italic=True, color="FFFFFF")
    oblig_fill_obl = PatternFill("solid", start_color="B91C1C")
    oblig_fill_opc = PatternFill("solid", start_color="6E6E73")
    for col_idx, ob in enumerate(obligs, start=1):
        cell = ws1.cell(row=2, column=col_idx, value=ob)
        cell.font = oblig_font
        cell.fill = oblig_fill_obl if ob.startswith("OBL") else oblig_fill_opc
        cell.alignment = center
        cell.border = border

    # Filas de ejemplos
    example_font = Font(name="Arial", size=10)
    voucher_colors = [GREEN_LIGHT, BLUE_LIGHT, AMBER_LIGHT, RED_LIGHT]
    prev_ref = None
    voucher_idx = -1
    for row_idx, row in enumerate(EXAMPLES, start=3):
        ref = row[0]
        if ref != prev_ref:
            voucher_idx += 1
            prev_ref = ref
        fill_color = voucher_colors[voucher_idx % len(voucher_colors)]
        fill = PatternFill("solid", start_color=fill_color)
        for col_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = example_font
            cell.border = border
            cell.fill = fill
            # debit / credit format
            if col_idx in (15, 16) and isinstance(val, (int, float)):
                cell.number_format = "#,##0;(#,##0);-"
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    widths = [12, 14, 10, 13, 13, 32, 14, 28, 22, 8, 5, 14, 12, 6, 12, 12, 28]
    for col_idx, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze headers
    ws1.freeze_panes = "A3"

    # Nota debajo de los ejemplos
    note_row = len(EXAMPLES) + 4
    note_cell = ws1.cell(
        row=note_row, column=1,
        value=(
            "▲ INSTRUCCIONES: borrá las 4 filas de ejemplo (3-11) cuando vayas a cargar "
            "tu data real. Recordá: una FILA = una LÍNEA contable; mismo voucher_ref agrupa "
            "líneas en un voucher. La suma debit = suma credit dentro de un voucher_ref."
        )
    )
    note_cell.font = Font(name="Arial", size=10, italic=True, color="6E6E73")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=17)
    ws1.row_dimensions[note_row].height = 40

    # ===== Hoja 2: Valores Válidos =====
    ws2 = wb.create_sheet("Valores Válidos")

    def write_section(title: str, headers_row: list[str], rows: list, start_row: int) -> int:
        # Title
        cell = ws2.cell(row=start_row, column=1, value=title)
        cell.font = Font(name="Arial", size=12, bold=True, color=GREEN)
        start_row += 1
        # Headers
        for col_idx, h in enumerate(headers_row, start=1):
            c = ws2.cell(row=start_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center
        start_row += 1
        # Data
        for r in rows:
            if isinstance(r, tuple):
                for col_idx, val in enumerate(r, start=1):
                    c = ws2.cell(row=start_row, column=col_idx, value=val)
                    c.font = Font(name="Arial", size=10)
                    c.border = border
                    c.alignment = left_wrap
            else:
                c = ws2.cell(row=start_row, column=1, value=r)
                c.font = Font(name="Consolas", size=10)
                c.border = border
            start_row += 1
        return start_row + 2

    next_row = 1
    next_row = write_section(
        "Empresas (empresa_codigo)",
        ["codigo", "razón social"],
        EMPRESAS, next_row
    )
    next_row = write_section(
        "Tipos de voucher (tipo)",
        ["valor", "descripción"],
        TIPOS_VOUCHER, next_row
    )
    next_row = write_section(
        "Tipos de documento tributario (doc_tributario_tipo)",
        ["valor"],
        TIPOS_DOC_TRIBUTARIO, next_row
    )
    next_row = write_section(
        "Aliases aceptados en headers (case-insensitive)",
        ["nombre canónico", "aliases también aceptados"],
        ALIASES_HEADER, next_row
    )

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 50

    # ===== Hoja 3: Instrucciones =====
    ws3 = wb.create_sheet("Instrucciones")
    inst = [
        ("CEHTA CAPITAL — TEMPLATE BULK IMPORT VOUCHERS", "title"),
        ("", ""),
        ("Endpoint: POST /vouchers/import-csv", "code"),
        ("UI: https://cehta-capital.vercel.app/vouchers/import", "code"),
        ("", ""),
        ("REGLAS GENERALES", "h2"),
        ("• Una FILA = una LÍNEA contable. NO una fila = un voucher.", ""),
        ("• Mismo voucher_ref en N filas → se agrupan como UN voucher con N líneas.", ""),
        ("• La suma de debit DEBE igualar la suma de credit dentro de cada voucher_ref.", ""),
        ("• Cada línea: debit > 0 XOR credit > 0 (NO los dos, NO ninguno).", ""),
        ("• Estado inicial: todos los vouchers se crean en DRAFT (editables).", ""),
        ("", ""),
        ("FORMATO DEL ARCHIVO", "h2"),
        ("• Separador: PUNTO Y COMA (;) — Excel chileno default", ""),
        ("• Encoding: UTF-8 (con o sin BOM)", ""),
        ("• Decimales: usar `,` o `.` (ambos aceptados)", ""),
        ("• Fechas: YYYY-MM-DD (ISO) o DD-MM-YYYY o DD/MM/YYYY", ""),
        ("", ""),
        ("CÓMO EXPORTAR DESDE EXCEL", "h2"),
        ("1. Llenar la hoja 'Vouchers' con tus datos reales (borrar ejemplos).", ""),
        ("2. Archivo → Guardar como → CSV UTF-8 (delimitado por comas) (*.csv)", ""),
        ("3. Verificar que el separador sea `;` (en regiones chilenas viene así por default)", ""),
        ("4. Si tu Excel exporta con `,` como separador, abrir el .csv con un editor de texto", ""),
        ("   y reemplazar `,` por `;` (ojo de no tocar los decimales).", ""),
        ("", ""),
        ("CÓMO IMPORTAR EN EL SISTEMA", "h2"),
        ("1. Login en https://cehta-capital.vercel.app/", ""),
        ("2. Ir a /vouchers/import", ""),
        ("3. Arrastrar el .csv o seleccionarlo", ""),
        ("4. Clic en 'Validar (dry-run)' — el backend revisa sin escribir nada", ""),
        ("5. Si hay errores, los muestra agrupados por voucher_ref + fila + mensaje", ""),
        ("6. Si todo OK, clic en 'Importar' — crea los vouchers en estado DRAFT", ""),
        ("7. Después: ir a /vouchers, filtrar por DRAFT, revisar y enviar a firma", ""),
        ("", ""),
        ("LÍMITES", "h2"),
        ("• Tamaño máx del archivo: 10 MB", ""),
        ("• Sin límite explícito de filas, pero recomendado: < 5,000 filas por archivo", ""),
        ("", ""),
        ("VALIDACIONES QUE HACE EL BACKEND", "h2"),
        ("• empresa_codigo existe en core.empresas", ""),
        ("• tipo está en el enum válido (ver hoja Valores Válidos)", ""),
        ("• cuenta_codigo existe en core.plan_cuenta_empresa para esa empresa", ""),
        ("• proyecto_codigo (si presente) existe en core.proyectos_contables", ""),
        ("• line_number único dentro del mismo voucher_ref", ""),
        ("• debit XOR credit > 0 por línea", ""),
        ("• Suma debit = suma credit por voucher_ref (descuadre solo permitido en DRAFT)", ""),
        ("• Fechas válidas y parseables", ""),
        ("", ""),
        (f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "footnote"),
    ]

    for i, (text, style) in enumerate(inst, start=1):
        cell = ws3.cell(row=i, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Arial", size=16, bold=True, color=GREEN)
        elif style == "h2":
            cell.font = Font(name="Arial", size=12, bold=True, color="1D1D1F")
            cell.fill = PatternFill("solid", start_color=GREEN_LIGHT)
        elif style == "code":
            cell.font = Font(name="Consolas", size=10, color="1E40AF")
        elif style == "footnote":
            cell.font = Font(name="Arial", size=9, italic=True, color="6E6E73")
        else:
            cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws3.column_dimensions["A"].width = 95

    # Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"[OK] Template generado: {OUTPUT}")


if __name__ == "__main__":
    main()
