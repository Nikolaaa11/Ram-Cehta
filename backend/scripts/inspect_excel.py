"""Inspecciona un Excel y dumpea su estructura (sheets, columnas, sample rows).

Uso:
    python -m scripts.inspect_excel "C:/Users/DELL/Downloads/Consolida bbdd 2.xlsx"

Output:
    - Lista de sheets con row count + columnas
    - Primeras 5 filas de cada sheet como JSON
    - Útil para diseñar el parser ETL contra la estructura real.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def inspect(path: Path) -> None:
    # Configurar UTF-8 en Windows
    if sys.stdout.encoding != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    if not path.exists():
        print(f"[ERROR] Archivo no encontrado: {path}")
        sys.exit(1)

    print(f"[FILE] Archivo: {path.name}")
    print(f"[SIZE] Tamano: {path.stat().st_size / 1024:.1f} KB")
    print()

    wb = load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True, max_row=10))

        if not rows:
            print(f"[SHEET] {sheet_name!r} (vacia)")
            print()
            continue

        # Detect header row (first non-empty row)
        header = None
        header_row_idx = -1
        for idx, row in enumerate(rows):
            non_empty = sum(1 for c in row if c not in (None, ""))
            if non_empty >= 3:
                header = list(row)
                header_row_idx = idx
                break

        # Total row count (limit to avoid loading whole sheet)
        # Re-iter with full sheet to count
        total_rows = ws.max_row or 0

        print(f"[SHEET] {sheet_name!r}")
        print(f"   filas totales: {total_rows}")
        print(f"   columnas: {ws.max_column}")
        if header is not None:
            print(f"   header detectado en fila #{header_row_idx + 1}:")
            for i, col in enumerate(header):
                if col is not None:
                    print(f"     [{i}] {col!r}")

            # Show first 3 data rows
            data_rows = rows[header_row_idx + 1 : header_row_idx + 4]
            print(f"   primeras 3 filas de datos:")
            for j, row in enumerate(data_rows, start=1):
                # Build dict
                d = {}
                for i, val in enumerate(row):
                    if header[i] is not None:
                        d[str(header[i])] = (
                            str(val)[:80] if val is not None else None
                        )
                print(f"     [fila {j}] {json.dumps(d, default=str, ensure_ascii=False)}")
        else:
            print(f"   [WARN] No se detecto header (sheet vacia o muy esparsa)")
        print()

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python -m scripts.inspect_excel <path>")
        sys.exit(1)
    inspect(Path(sys.argv[1]))
