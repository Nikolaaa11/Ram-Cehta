"""Importador del Excel "Registro de Egresos" de Claudia.

Los workbooks se arman acá con openpyxl reproduciendo las DOS variantes de
columnas (REVTECH: `Fuente`; TRONGKAI: `Tipo Financiamiento` + `Trewaox`) y
todos los detalles sucios medidos en los archivos reales: estados con
símbolo, `\xa0` al final del nombre, folio numérico y folio que Excel
convirtió a fecha, fila sin fecha, "Boletas"/"Boleta"/"liquidación",
duplicado exacto, 4 fuentes vacías vs una con valor, total con 4 decimales.

`cargar_filas` no se prueba contra Postgres (no hay en CI): se prueba el
SQL como string (compila con asyncpg, sin `:param::cast`) y la aritmética
del resumen con una sesión falsa.
"""
from __future__ import annotations

import hashlib
import importlib.util
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.dialects import postgresql

from app.domain.value_objects.reparto_corfo import (
    ESTADO_DESCUADRADO,
    ESTADO_OK,
    ESTADO_SIN_CLASIFICAR,
)
from app.services.corfo_egresos_import_service import (
    SQL_CLAVES_EXISTENTES,
    SQL_INSERT_EGRESOS,
    FilaEgreso,
    FilaSaltada,
    ResumenCarga,
    _to_decimal,
    cargar_filas,
    contar_reparto,
    parsear_registro_egresos,
)

D = Decimal

HEADERS_REVTECH = [
    "Fecha",
    "Descripción",
    "RUT Emisor",
    "Tipo de Documento",
    "Folio",
    "Monto Neto/Pagado",
    "Impuesto/Patronal",
    "Total",
    "Tipo de Egreso",
    "Fuente",
    "Proyecto",
    "Subsidio",
    "Cehta-Ptec",
    "Cehta",
    "Estado",
    "Fecha de Pago",
]
HEADERS_TRONGKAI = [
    "Fecha",
    "Descripción",
    "RUT Emisor",
    "Tipo de Documento",
    "Folio",
    "Monto Neto/Pagado",
    "Impuesto/Patronal",
    "Total",
    "Tipo Financiamiento",
    "Tipo de Egreso",
    "Proyecto",
    "Trewaox",
    "Subsidio",
    "Cehta-Ptec",
    "Cehta",
    "Estado",
    "Fecha de Pago",
]

# Fila real de REVTECH (MCG, ago-2026) como base; cada test pisa lo que necesita.
_BASE: dict[str, Any] = {
    "Fecha": datetime(2026, 8, 27),
    "Descripción": "MCG AUDITORES CONSULTORES SPA",
    "RUT Emisor": "76642280-2",
    "Tipo de Documento": "Factura",
    "Folio": 10540,
    "Monto Neto/Pagado": 79287,
    "Impuesto/Patronal": 15065,
    "Total": 94352,
    "Tipo de Egreso": "Cehta",
    "Fuente": "Cehta",
    "Tipo Financiamiento": "Cehta",
    "Proyecto": "Cehta",
    "Trewaox": None,
    "Subsidio": None,
    "Cehta-Ptec": None,
    "Cehta": 94352,
    "Estado": "✓ Pagado",
    "Fecha de Pago": datetime(2026, 1, 8),
}
#: Reparto vacío (las 4 fuentes) para pisar `_BASE`.
_SIN_REPARTO: dict[str, Any] = dict.fromkeys(("Trewaox", "Subsidio", "Cehta-Ptec", "Cehta"))

# filas 59-61 reales: Climate Smart Leasing, Co-Ejecutor folio 1, $5.000.000, mismo día
_CLIMATE: dict[str, Any] = {
    "Descripción": "Climate Smart Leasing",
    "RUT Emisor": "77868887-5",
    "Tipo de Documento": "Co-Ejecutor",
    "Folio": 1,
    "Total": 5000000,
    "Monto Neto/Pagado": 5000000,
    "Impuesto/Patronal": None,
    "Subsidio": 5000000,
    "Cehta": 0,
}


def _fila(headers: list[str], **cambios: Any) -> list[Any]:
    datos = {**_BASE, **cambios}
    return [datos.get(h) for h in headers]


def _xlsx(
    headers: list[str],
    filas: list[list[Any]],
    *,
    hoja: str = "Registro de Egresos",
    con_titulo: bool = True,
    hojas_previas: tuple[str, ...] = ("Dashboard",),
) -> bytes:
    """Workbook con la misma forma que el de Claudia: título en la fila 1,
    'SEPARACION VALORES' en la 2, encabezados en la 3, datos desde la 4."""
    wb = Workbook()
    ws = wb.active
    if hojas_previas:
        ws.title = hojas_previas[0]
        ws.append(["Dashboard de otra cosa"])
        for extra in hojas_previas[1:]:
            wb.create_sheet(extra)
        ws = wb.create_sheet(hoja)
    else:
        ws.title = hoja
    if con_titulo:
        ws.append(["REGISTRO DE EGRESOS"])
        ws.append([None] * 11 + ["SEPARACION VALORES"])
    ws.append(headers)
    for f in filas:
        ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Variantes de columnas ────────────────────────────────────────────


def test_variante_revtech_mapea_todos_los_campos():
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH)]), "REVTECH")

    assert res.columnas == HEADERS_REVTECH
    assert res.saltadas == [] and res.duplicadas_en_excel == []
    assert len(res.filas) == 1
    f = res.filas[0]
    assert f.fila_excel == 4  # encabezado en la fila 3
    assert f.fecha == date(2026, 8, 27)
    assert f.periodo == "2026-08"
    assert f.descripcion == "MCG AUDITORES CONSULTORES SPA"
    assert f.rut_emisor == "76642280-2"
    assert f.tipo_documento == "FACTURA"
    assert f.folio == "10540"
    assert f.monto_neto == D("79287.00")
    assert f.impuesto == D("15065.00")
    assert f.total == D("94352.00")
    assert f.neto_mas_impuesto_cuadra is True
    assert (f.tipo_egreso, f.fuente, f.proyecto) == ("Cehta", "Cehta", "Cehta")
    assert f.estado_pago == "PAGADO"
    assert f.fecha_pago == date(2026, 1, 8)
    # una fuente con valor → las otras en 0 (todo-o-nada), Trewaox incluida
    assert f.reparto == {
        "subsidio": D("0.00"),
        "cehta_ptec": D("0.00"),
        "cehta": D("94352.00"),
        "trewaox": D("0.00"),
    }
    assert f.reparto_estado == ESTADO_OK
    assert f.observaciones is None


def test_variante_trongkai_con_trewaox_y_tipo_financiamiento():
    fila = _fila(
        HEADERS_TRONGKAI,
        **{
            "Descripción": "Bustorf Larsen y Compañía Ltda",
            "RUT Emisor": 763335232,  # el Excel real lo trae como int
            "Tipo de Documento": "Boleta",
            "Folio": "3714",
            "Monto Neto/Pagado": 11090,
            "Impuesto/Patronal": None,
            "Total": 11090,
            "Tipo Financiamiento": "InnovaRegion",
            "Tipo de Egreso": "InnovaRegion",
            "Proyecto": "Trewaox",
            "Trewaox": 11090,
            "Cehta": None,
        },
    )
    res = parsear_registro_egresos(_xlsx(HEADERS_TRONGKAI, [fila]), "TRONGKAI")

    assert res.columnas == HEADERS_TRONGKAI
    f = res.filas[0]
    assert f.fuente == "InnovaRegion"  # "Tipo Financiamiento" cae en `fuente`
    assert f.tipo_egreso == "InnovaRegion"
    assert f.proyecto == "Trewaox"
    assert f.rut_emisor == "76333523-2"
    assert f.tipo_documento == "BOLETA"
    assert f.monto_trewaox == D("11090.00")
    assert (f.monto_subsidio, f.monto_cehta_ptec, f.monto_cehta) == (D("0.00"),) * 3
    assert f.reparto_estado == ESTADO_OK
    assert f.impuesto == D("0.00")


# ── Detalles sucios de §1.2 ──────────────────────────────────────────


@pytest.mark.parametrize(
    ("crudo", "esperado"),
    [
        ("✓ Pagado", "PAGADO"),
        ("◑ Pagado Parcial", "PARCIAL"),
        ("✗ Pendiente", "PENDIENTE"),
        ("Pagado", "PAGADO"),
        ("pendiente", "PENDIENTE"),
        ("PARCIAL", "PARCIAL"),
        (None, "PENDIENTE"),
        ("", "PENDIENTE"),
    ],
)
def test_estado_por_simbolo_o_palabra(crudo, esperado):
    res = parsear_registro_egresos(
        _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH, Estado=crudo)]), "REVTECH"
    )
    assert res.filas[0].estado_pago == esperado
    assert res.filas[0].observaciones is None


def test_estado_desconocido_queda_pendiente_y_en_observaciones():
    res = parsear_registro_egresos(
        _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH, Estado="en revisión")]), "REVTECH"
    )
    f = res.filas[0]
    assert f.estado_pago == "PENDIENTE"
    assert f.observaciones == "Estado en el Excel: en revisión"


def test_descripcion_con_nbsp_y_espacios_se_recorta():
    filas = [
        _fila(HEADERS_REVTECH, **{"Descripción": "CAMILO IVAN SALAZAR ORTIZ \xa0"}),
        _fila(HEADERS_REVTECH, **{"Descripción": " Rendición  Uber\xa01_26-05-2026 ", "Folio": 2}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert res.filas[0].descripcion == "CAMILO IVAN SALAZAR ORTIZ"
    assert res.filas[1].descripcion == "Rendición Uber 1_26-05-2026"


def test_folio_int_float_y_fecha():
    filas = [
        _fila(HEADERS_REVTECH, Folio=126),
        _fila(HEADERS_REVTECH, Folio=127.0),
        # fila 252 real: Claudia escribió "13-05-2026" y Excel lo hizo fecha
        _fila(HEADERS_REVTECH, Folio=datetime(2026, 5, 13)),
        _fila(HEADERS_REVTECH, Folio=" 040326-2 "),
        _fila(HEADERS_REVTECH, Folio=None, **{"Descripción": "sin folio"}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert [f.folio for f in res.filas] == ["126", "127", "13-05-2026", "040326-2", None]


def test_fecha_invalida_salta_la_fila_con_motivo_y_sigue():
    filas = [
        _fila(HEADERS_REVTECH),
        # filas 29/30 reales: "Nc Doing Spa" sin fecha
        _fila(HEADERS_REVTECH, Fecha=None, **{"Descripción": "Nc Doing Spa", "Total": 950000}),
        _fila(HEADERS_REVTECH, Fecha="pendiente", Folio=99),
        _fila(HEADERS_REVTECH, Folio=100),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")

    assert [f.fila_excel for f in res.filas] == [4, 7]
    assert [s.fila_excel for s in res.saltadas] == [5, 6]
    assert "Fecha inválida" in res.saltadas[0].motivo
    assert "Nc Doing Spa" in res.saltadas[0].motivo and "950000" in res.saltadas[0].motivo
    assert "'pendiente'" in res.saltadas[1].motivo
    assert res.leidas == 4


def test_fecha_como_texto_legible_se_acepta():
    filas = [
        _fila(HEADERS_REVTECH, Fecha="2026-03-04"),
        _fila(HEADERS_REVTECH, Fecha="05-03-2026", Folio=2),
        _fila(HEADERS_REVTECH, **{"Fecha de Pago": "07/03/2026", "Folio": 3}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert res.filas[0].fecha == date(2026, 3, 4)
    assert res.filas[1].fecha == date(2026, 3, 5)
    assert res.filas[2].fecha_pago == date(2026, 3, 7)


def test_sin_descripcion_o_sin_total_se_salta():
    filas = [
        _fila(HEADERS_REVTECH, **{"Descripción": "  \xa0 "}),
        _fila(HEADERS_REVTECH, Total=None, Folio=2),
        _fila(HEADERS_REVTECH, Total="abc", Folio=3),
        _fila(HEADERS_REVTECH, Total=-100, Folio=4),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert res.filas == []
    motivos = [s.motivo for s in res.saltadas]
    assert motivos[0] == "Sin descripción"
    assert motivos[1].startswith("Sin total")
    assert motivos[2].startswith("Total ilegible")
    assert motivos[3].startswith("Total negativo")


@pytest.mark.parametrize(
    ("crudo", "esperado"),
    [
        ("Factura", "FACTURA"),
        ("Factura Exenta", "FACTURA_EXENTA"),
        ("FACTURA EXENTA ", "FACTURA_EXENTA"),
        ("Boletas", "BOLETA"),
        ("Boleta", "BOLETA"),
        ("Boleta Honorario", "BOLETA_HONORARIO"),
        ("Boleta de Honorarios", "BOLETA_HONORARIO"),
        ("liquidación", "LIQUIDACION"),
        ("Liquidación", "LIQUIDACION"),
        ("Co-Ejecutor", "CO_EJECUTOR"),
        ("COEJECUTOR", "CO_EJECUTOR"),
        ("Invoice", "INVOICE"),
    ],
)
def test_tipo_de_documento_normalizado(crudo, esperado):
    res = parsear_registro_egresos(
        _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH, **{"Tipo de Documento": crudo})]),
        "REVTECH",
    )
    assert res.filas[0].tipo_documento == esperado
    assert res.filas[0].observaciones is None


def test_tipo_de_documento_desconocido_va_a_otro_con_el_original_en_observaciones():
    filas = [
        _fila(HEADERS_REVTECH, **{"Tipo de Documento": "Recibo"}),
        _fila(HEADERS_REVTECH, **{"Tipo de Documento": None, "Folio": 2}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert res.filas[0].tipo_documento == "OTRO"
    assert res.filas[0].observaciones == "Tipo de documento en el Excel: Recibo"
    assert res.filas[1].tipo_documento == "OTRO"
    assert res.filas[1].observaciones == "Sin tipo de documento en el Excel"


def _sha1(base: bytes) -> str:
    return hashlib.sha1(base, usedforsecurity=False).hexdigest()


_HUELLA_CLIMATE = b"REVTECH|77868887-5|CO_EJECUTOR|1|2026-08-27|5000000.00|climate smart leasing"


def test_repetidas_identicas_entran_todas_con_huella_propia_y_observacion():
    # Filas 59-61 reales: tres cuotas de $5.000.000 al co-ejecutor el mismo día.
    # Son tres pagos, no un tipeo: entran las tres (default del contrato §3.4).
    contenido = _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH, **_CLIMATE)] * 3)

    res = parsear_registro_egresos(contenido, "REVTECH")

    assert [f.fila_excel for f in res.filas] == [4, 5, 6]
    assert res.repetidas_en_excel == [5, 6]
    assert res.duplicadas_en_excel == [] and res.leidas == 3
    claves = [f.import_natural_key for f in res.filas]
    assert len(set(claves)) == 3
    # la primera lleva la huella base del contrato; la n-ésima, base + "|#n"
    assert claves == [
        _sha1(_HUELLA_CLIMATE),
        _sha1(_HUELLA_CLIMATE + b"|#2"),
        _sha1(_HUELLA_CLIMATE + b"|#3"),
    ]
    assert res.filas[0].observaciones is None
    assert res.filas[1].observaciones == (
        "Idéntica a la fila 4 del Excel (aparición 2): verificar que sea un gasto distinto"
    )
    assert res.filas[2].observaciones == (
        "Idéntica a la fila 4 del Excel (aparición 3): verificar que sea un gasto distinto"
    )
    # estable entre corridas: la segunda pasada produce las mismas huellas
    otra = parsear_registro_egresos(contenido, "REVTECH")
    assert [f.import_natural_key for f in otra.filas] == claves


def test_repetida_con_otro_estado_o_folio():
    filas = [
        _fila(HEADERS_REVTECH, **_CLIMATE),
        _fila(HEADERS_REVTECH, **_CLIMATE, Estado="✗ Pendiente"),  # el estado NO entra en la huella
        _fila(HEADERS_REVTECH, **{**_CLIMATE, "Folio": 2}),  # otro folio: otro gasto, no repetida
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert [f.fila_excel for f in res.filas] == [4, 5, 6]
    assert res.repetidas_en_excel == [5]
    assert res.filas[1].estado_pago == "PENDIENTE"
    assert res.filas[2].observaciones is None
    assert res.leidas == 3


def test_colapsar_repetidas_manda_la_segunda_a_duplicadas_en_excel():
    # Comportamiento viejo, opt-in (CLI --colapsar-repetidas): sólo entra la primera.
    filas = [
        _fila(HEADERS_REVTECH, **_CLIMATE),
        _fila(HEADERS_REVTECH, **_CLIMATE),
        _fila(HEADERS_REVTECH, **_CLIMATE, Estado="✗ Pendiente"),
        _fila(HEADERS_REVTECH, **{**_CLIMATE, "Folio": 2}),
    ]
    res = parsear_registro_egresos(
        _xlsx(HEADERS_REVTECH, filas), "REVTECH", conservar_repetidas=False
    )
    assert [f.fila_excel for f in res.filas] == [4, 7]
    assert res.duplicadas_en_excel == [5, 6]
    assert res.repetidas_en_excel == []
    assert res.leidas == 4
    assert res.filas[0].import_natural_key == _sha1(_HUELLA_CLIMATE)


def test_cuatro_fuentes_vacias_vs_una_con_valor():
    filas = [
        _fila(HEADERS_TRONGKAI, **_SIN_REPARTO),
        _fila(HEADERS_TRONGKAI, Folio=2, **{**_SIN_REPARTO, "Subsidio": 94352}),
        # fila 77 real: 35424 + 6731 = 42155 ≠ 42154
        _fila(
            HEADERS_TRONGKAI,
            Folio=3,
            Total=42154,
            **{**_SIN_REPARTO, "Subsidio": 35424, "Cehta": 6731},
        ),
        # un 0 explícito cuenta como valor: ya está clasificada (en 0)
        _fila(HEADERS_TRONGKAI, Folio=4, **{**_SIN_REPARTO, "Trewaox": 0}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_TRONGKAI, filas), "TRONGKAI")
    vacia, una, descuadrada, cero = res.filas

    assert vacia.reparto == dict.fromkeys(("subsidio", "cehta_ptec", "cehta", "trewaox"))
    assert vacia.reparto_estado == ESTADO_SIN_CLASIFICAR

    assert una.reparto == {
        "subsidio": D("94352.00"),
        "cehta_ptec": D("0.00"),
        "cehta": D("0.00"),
        "trewaox": D("0.00"),
    }
    assert una.reparto_estado == ESTADO_OK

    assert descuadrada.reparto_estado == ESTADO_DESCUADRADO
    assert descuadrada.neto_mas_impuesto_cuadra is False

    assert cero.reparto_estado == ESTADO_DESCUADRADO
    assert cero.monto_trewaox == D("0.00") and cero.monto_subsidio == D("0.00")

    assert contar_reparto(res.filas) == (2, 1)


def test_total_con_cuatro_decimales_se_redondea_al_centavo():
    # fila 44 real de REVTECH (UF): neto 5.390.000 + impuesto 255.105,9504
    filas = [
        _fila(
            HEADERS_REVTECH,
            **{
                "Monto Neto/Pagado": 5390000,
                "Impuesto/Patronal": 255105.9504,
                "Total": 5645105.9504,
                "Subsidio": 2000000,
                "Cehta-Ptec": 3390000,
                # residuo binario de Excel, como viene en las filas UF de TRONGKAI
                "Cehta": 255105.95040000003,
            },
        )
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    f = res.filas[0]
    assert f.total == D("5645105.95")
    assert f.impuesto == D("255105.95")
    assert f.monto_cehta == D("255105.95")
    assert f.neto_mas_impuesto_cuadra is True
    assert f.reparto_estado == ESTADO_OK
    assert f"{f.total:.2f}" == "5645105.95"


@pytest.mark.parametrize(
    ("crudo", "esperado"),
    [
        ("60.805.000-0", "60805000-0"),
        ("25543408-k", "25543408-K"),
        (763335232, "76333523-2"),
        (" 76.642.280 - 2 ", "76642280-2"),
        (None, None),
        ("", None),
    ],
)
def test_rut_normalizado(crudo, esperado):
    res = parsear_registro_egresos(
        _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH, **{"RUT Emisor": crudo})]), "REVTECH"
    )
    assert res.filas[0].rut_emisor == esperado


def _una(headers: list[str], **cambios: Any) -> FilaEgreso:
    empresa = "TRONGKAI" if headers is HEADERS_TRONGKAI else "REVTECH"
    return parsear_registro_egresos(_xlsx(headers, [_fila(headers, **cambios)]), empresa).filas[0]


def test_neto_e_impuesto_vacios_toman_el_default_de_la_api():
    # 33 boletas reales de TRONGKAI vienen sólo con el total: neto = total,
    # impuesto = 0 (el mismo default que EgresoCreate), con observación.
    f = _una(
        HEADERS_TRONGKAI,
        **{"Monto Neto/Pagado": None, "Impuesto/Patronal": None, "Total": 35398, "Cehta": 35398},
    )
    assert (f.monto_neto, f.impuesto, f.total) == (D("35398.00"), D("0.00"), D("35398.00"))
    assert f.neto_mas_impuesto_cuadra is True
    assert f.observaciones == "Neto e impuesto vacíos en el Excel"


def test_neto_o_impuesto_solo_deriva_el_otro_como_diferencia():
    solo_neto = _una(HEADERS_TRONGKAI, **{"Monto Neto/Pagado": 79287, "Impuesto/Patronal": None})
    solo_imp = _una(HEADERS_TRONGKAI, **{"Monto Neto/Pagado": None, "Impuesto/Patronal": 15065})
    assert (solo_neto.monto_neto, solo_neto.impuesto) == (D("79287.00"), D("15065.00"))
    assert (solo_imp.monto_neto, solo_imp.impuesto) == (D("79287.00"), D("15065.00"))
    assert solo_neto.neto_mas_impuesto_cuadra and solo_imp.neto_mas_impuesto_cuadra
    assert solo_neto.observaciones is None and solo_imp.observaciones is None


def test_neto_o_impuesto_mayor_al_total_no_inventa_un_negativo():
    # Nunca una diferencia negativa: quedan como vinieron (el vacío en 0) y se observa.
    neto_grande = _una(
        HEADERS_TRONGKAI, **{"Monto Neto/Pagado": 100000, "Impuesto/Patronal": None, "Total": 94352}
    )
    assert (neto_grande.monto_neto, neto_grande.impuesto) == (D("100000.00"), D("0.00"))
    assert neto_grande.neto_mas_impuesto_cuadra is False
    assert neto_grande.observaciones == (
        "Neto $100.000 supera el total $94.352 en el Excel: el impuesto vacío queda en $0"
    )
    imp_grande = _una(
        HEADERS_TRONGKAI, **{"Monto Neto/Pagado": None, "Impuesto/Patronal": 100000, "Total": 94352}
    )
    assert (imp_grande.monto_neto, imp_grande.impuesto) == (D("0.00"), D("100000.00"))
    assert imp_grande.observaciones == (
        "Impuesto $100.000 supera el total $94.352 en el Excel: el neto vacío queda en $0"
    )


def test_neto_e_impuesto_con_valor_que_no_suman_quedan_tal_cual():
    # fila 77 real: 35.424 + 6.731 = 42.155 ≠ 42.154. La diferencia es real y
    # se muestra; corregirla acá sería inventar.
    f = _una(
        HEADERS_TRONGKAI,
        **{"Monto Neto/Pagado": 35424, "Impuesto/Patronal": 6731, "Total": 42154, "Cehta": 42154},
    )
    assert (f.monto_neto, f.impuesto) == (D("35424.00"), D("6731.00"))
    assert f.neto_mas_impuesto_cuadra is False
    assert f.observaciones is None


def test_neto_ilegible_no_deriva_nada():
    f = _una(HEADERS_TRONGKAI, **{"Monto Neto/Pagado": "abc", "Impuesto/Patronal": None})
    assert (f.monto_neto, f.impuesto) == (D("0.00"), D("0.00"))
    assert f.observaciones == "Monto neto ilegible en el Excel: 'abc'"


def test_montos_como_texto_chileno():
    assert _to_decimal("1.234.567,89") == D("1234567.89")
    assert _to_decimal("1,234,567.89") == D("1234567.89")
    assert _to_decimal("$ 94.352") == D("94352.00")  # puntos de miles, no decimal
    assert _to_decimal("94.35") == D("94.35")
    assert _to_decimal("") is None
    assert _to_decimal(255105.9504) == D("255105.95")
    with pytest.raises(ValueError, match="ilegible"):
        _to_decimal("abc")


def test_filas_residuales_de_formulas_no_cuentan():
    # Las últimas ~25/350 filas del Excel real traen sólo '✗ Pendiente' (fórmula
    # arrastrada) y '' en Fecha de Pago. No son gastos.
    residual = [None] * len(HEADERS_REVTECH)
    residual[HEADERS_REVTECH.index("Estado")] = "✗ Pendiente"
    residual[HEADERS_REVTECH.index("Fecha de Pago")] = ""
    filas = [_fila(HEADERS_REVTECH), residual, residual, [None] * len(HEADERS_REVTECH)]
    res = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH")
    assert len(res.filas) == 1
    assert res.saltadas == []
    assert res.leidas == 1


# ── Hoja y encabezados ───────────────────────────────────────────────


def test_encabezados_por_nombre_tolerantes_y_reordenados():
    headers = [
        "TOTAL",
        "fecha",
        "Descripcion",
        "Rut emisor",
        "TIPO DOCUMENTO",
        "folio",
        "Cehta - Ptec",
        "SUBSIDIO",
        "cehta\xa0",
        "ESTADO DE PAGO",
    ]
    fila = [94352, datetime(2026, 8, 27), "MCG", "76642280-2", "Factura", 1, 20000, 70000, 4352]
    fila.append("✓")
    contenido = _xlsx(headers, [fila], hoja="Egresos 2026", con_titulo=False, hojas_previas=())
    res = parsear_registro_egresos(contenido, "REVTECH")

    assert res.columnas == [h.strip() for h in headers]
    f = res.filas[0]
    assert f.fila_excel == 2  # encabezado en la fila 1
    assert f.total == D("94352.00")
    assert f.fecha == date(2026, 8, 27)
    assert f.monto_cehta_ptec == D("20000.00")
    assert f.monto_subsidio == D("70000.00")
    assert f.monto_cehta == D("4352.00")
    assert f.monto_trewaox == D("0.00")  # sin columna Trewaox: 0, no None (todo-o-nada)
    assert f.estado_pago == "PAGADO"
    # sin columnas de neto/impuesto: default de la API (neto = total, impuesto = 0)
    assert (f.monto_neto, f.impuesto) == (D("94352.00"), D("0.00"))
    assert f.observaciones == "Neto e impuesto vacíos en el Excel"


def test_prefiere_la_hoja_registro_de_egresos_aunque_no_sea_la_primera():
    contenido = _xlsx(
        HEADERS_REVTECH,
        [_fila(HEADERS_REVTECH)],
        hojas_previas=("Dashboard", "CC_Santander", "Flujo"),
    )
    res = parsear_registro_egresos(contenido, "REVTECH")
    assert len(res.filas) == 1


def test_sin_encabezados_lanza_error_en_espanol():
    wb = Workbook()
    wb.active.append(["Nombre", "Monto"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Registro de Egresos"):
        parsear_registro_egresos(buf.getvalue(), "REVTECH")


def test_archivo_que_no_es_xlsx_lanza_error():
    with pytest.raises(ValueError, match="no es un .xlsx"):
        parsear_registro_egresos(b"esto no es un zip", "REVTECH")
    with pytest.raises(ValueError, match="vacío"):
        parsear_registro_egresos(b"", "REVTECH")
    with pytest.raises(ValueError, match="empresa"):
        parsear_registro_egresos(_xlsx(HEADERS_REVTECH, []), "  ")


# ── Huella e idempotencia ────────────────────────────────────────────


def test_import_natural_key_estable_entre_corridas_y_por_empresa():
    contenido = _xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH)])
    a = parsear_registro_egresos(contenido, "REVTECH").filas[0]
    b = parsear_registro_egresos(contenido, "revtech ").filas[0]
    assert a.import_natural_key == b.import_natural_key

    esperado = hashlib.sha1(
        b"REVTECH|76642280-2|FACTURA|10540|2026-08-27|94352.00|mcg auditores consultores spa",
        usedforsecurity=False,
    ).hexdigest()
    assert a.import_natural_key == esperado

    otra = parsear_registro_egresos(contenido, "TRONGKAI").filas[0]
    assert otra.import_natural_key != a.import_natural_key


def test_la_huella_usa_los_valores_normalizados():
    # Mismo gasto escrito "sucio" y "limpio" → misma huella (así el re-import no duplica)
    sucio = _fila(
        HEADERS_REVTECH,
        **{
            "Descripción": "MCG Auditores Consultores SpA\xa0",
            "RUT Emisor": "76.642.280-2",
            "Folio": 10540.0,
        },
    )
    limpio = _fila(HEADERS_REVTECH)
    k1 = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, [sucio]), "REVTECH").filas[0]
    k2 = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, [limpio]), "REVTECH").filas[0]
    assert k1.import_natural_key == k2.import_natural_key


def test_round_trip_json():
    filas = [
        _fila(HEADERS_TRONGKAI, **{"Tipo de Documento": "Recibo", "Total": 5645105.9504}),
        _fila(HEADERS_TRONGKAI, Folio=2, **{**_SIN_REPARTO, "Fecha de Pago": None, "Estado": None}),
    ]
    res = parsear_registro_egresos(_xlsx(HEADERS_TRONGKAI, filas), "TRONGKAI")
    for original in res.filas:
        d = original.to_dict()
        assert d["fecha"] == "2026-08-27"
        assert isinstance(d["total"], str) and d["total"] == str(original.total)
        assert d["monto_subsidio"] is None or isinstance(d["monto_subsidio"], str)
        copia = FilaEgreso.from_dict(json.loads(json.dumps(d, ensure_ascii=False)))
        assert copia == original
        assert copia.reparto_estado == original.reparto_estado


# ── cargar_filas: SQL como string + aritmética con sesión falsa ─────


def _sin_binds_sobrantes(sql: str) -> None:
    render = str(text(sql).compile(dialect=postgresql.asyncpg.dialect()))
    sobrantes = [s for s in re.findall(r":\w+", render) if not s.startswith("::")]
    assert not sobrantes, f"binds sin sustituir {sobrantes}: ¿alguien escribió :param::cast?"


def test_sql_insert_es_idempotente_y_compila_para_asyncpg():
    assert (
        "ON CONFLICT (import_natural_key) WHERE import_natural_key IS NOT NULL DO NOTHING"
        in SQL_INSERT_EGRESOS
    )
    assert "RETURNING egreso_id" in SQL_INSERT_EGRESOS
    assert "'IMPORT_EXCEL'" in SQL_INSERT_EGRESOS
    assert "core.corfo_registro_egresos" in SQL_INSERT_EGRESOS
    assert "::" not in SQL_INSERT_EGRESOS and "::" not in SQL_CLAVES_EXISTENTES
    _sin_binds_sobrantes(SQL_INSERT_EGRESOS)
    _sin_binds_sobrantes(SQL_CLAVES_EXISTENTES)
    # todas las columnas del contrato viajan en el INSERT
    for col in (
        "monto_subsidio",
        "monto_cehta_ptec",
        "monto_cehta",
        "monto_trewaox",
        "observaciones",
        "created_by",
        "periodo",
    ):
        assert col in SQL_INSERT_EGRESOS


class _ResultadoFalso:
    def __init__(self, filas: list[tuple[Any, ...]]) -> None:
        self._filas = filas

    def fetchall(self) -> list[tuple[Any, ...]]:
        return self._filas


class _SesionFalsa:
    """Registra cada execute; devuelve las respuestas en orden."""

    def __init__(self, respuestas: list[list[tuple[Any, ...]]]) -> None:
        self.llamadas: list[tuple[str, dict[str, Any]]] = []
        self._respuestas = list(respuestas)
        self.commits = 0

    async def execute(self, stmt: Any, params: dict[str, Any] | None = None) -> _ResultadoFalso:
        self.llamadas.append((str(stmt), params or {}))
        return _ResultadoFalso(self._respuestas.pop(0) if self._respuestas else [])

    async def commit(self) -> None:
        self.commits += 1


def _tres_filas() -> list[FilaEgreso]:
    filas = [_fila(HEADERS_REVTECH, Folio=n) for n in (1, 2, 3)]
    return parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH").filas


async def test_cargar_filas_cuenta_creadas_por_los_returning_y_commitea():
    filas = _tres_filas()
    db: Any = _SesionFalsa([[(101,), (102,)]])  # la BD devolvió 2 ids: 1 ya existía

    r = await cargar_filas(db, "revtech", filas, "claudia@trongkai.com")

    assert isinstance(r, ResumenCarga)
    assert (r.creadas, r.omitidas_existentes, r.leidas) == (2, 1, 3)
    assert r.dry_run is False and r.empresa_codigo == "REVTECH"
    assert db.commits == 1
    sql, params = db.llamadas[0]
    assert "ON CONFLICT (import_natural_key)" in sql and "RETURNING egreso_id" in sql
    assert params["empresa"] == "REVTECH"
    assert params["usuario"] == "claudia@trongkai.com"
    assert params["fecha"] == [date(2026, 8, 27)] * 3
    assert params["total"] == [D("94352.00")] * 3
    assert params["folio"] == ["1", "2", "3"]
    assert params["import_natural_key"] == [f.import_natural_key for f in filas]
    assert params["monto_trewaox"] == [D("0.00")] * 3
    assert r.to_dict()["saltadas"] == []


async def test_cargar_filas_dry_run_no_inserta_ni_commitea_pero_cuenta_de_verdad():
    filas = _tres_filas()
    db: Any = _SesionFalsa([[(filas[1].import_natural_key,)]])

    r = await cargar_filas(db, "REVTECH", filas, "claudia@trongkai.com", dry_run=True)

    assert (r.creadas, r.omitidas_existentes, r.leidas) == (2, 1, 3)
    assert r.dry_run is True
    assert db.commits == 0
    assert len(db.llamadas) == 1
    sql, params = db.llamadas[0]
    assert sql.lstrip().upper().startswith("SELECT")
    assert "INSERT" not in sql.upper()
    assert params["claves"] == [f.import_natural_key for f in filas]


async def test_cargar_filas_sin_filas_no_toca_la_bd():
    db: Any = _SesionFalsa([])
    r = await cargar_filas(db, "TRONGKAI", [], "claudia@trongkai.com")
    assert (r.creadas, r.omitidas_existentes, r.leidas) == (0, 0, 0)
    assert db.llamadas == [] and db.commits == 0


async def test_cargar_filas_rechaza_empresa_ajena_y_usuario_vacio():
    db: Any = _SesionFalsa([])
    with pytest.raises(ValueError, match="REVTECH o TRONGKAI"):
        await cargar_filas(db, "AFIS", _tres_filas(), "x@y.z")
    with pytest.raises(ValueError, match="email"):
        await cargar_filas(db, "REVTECH", _tres_filas(), "  ")
    assert db.llamadas == []


async def test_reimportar_el_mismo_archivo_con_repetidas_crea_cero_filas():
    """Criterio de aceptación §6: re-importar el mismo Excel crea 0 filas,
    también cuando trae repetidas (huella con ordinal, estable entre corridas)."""
    contenido = _xlsx(
        HEADERS_REVTECH,
        [_fila(HEADERS_REVTECH, **_CLIMATE)] * 2 + [_fila(HEADERS_REVTECH, Folio=9)],
    )
    primera = parsear_registro_egresos(contenido, "REVTECH").filas
    claves = [f.import_natural_key for f in primera]
    assert len(primera) == 3 and len(set(claves)) == 3

    # 1ª carga: la BD devuelve los tres ids (nada existía)
    db1: Any = _SesionFalsa([[(1,), (2,), (3,)]])
    r1 = await cargar_filas(db1, "REVTECH", primera, "claudia@trongkai.com")
    assert (r1.creadas, r1.omitidas_existentes) == (3, 0)
    assert db1.llamadas[0][1]["import_natural_key"] == claves

    # 2ª corrida sobre el mismo archivo: mismas huellas, en el mismo orden
    segunda = parsear_registro_egresos(contenido, "REVTECH").filas
    assert [f.import_natural_key for f in segunda] == claves

    # dry-run: la BD ya las tiene todas → se crearían 0
    db2: Any = _SesionFalsa([[(k,) for k in claves]])
    r2 = await cargar_filas(db2, "REVTECH", segunda, "claudia@trongkai.com", dry_run=True)
    assert (r2.creadas, r2.omitidas_existentes, r2.leidas) == (0, 3, 3)

    # carga real: ON CONFLICT DO NOTHING no devuelve ningún id → 0 creadas
    db3: Any = _SesionFalsa([[]])
    r3 = await cargar_filas(db3, "REVTECH", segunda, "claudia@trongkai.com")
    assert (r3.creadas, r3.omitidas_existentes) == (0, 3)


async def test_cargar_filas_reporta_descuadradas_y_sin_clasificar():
    filas = [
        _fila(HEADERS_REVTECH, Folio=1, **_SIN_REPARTO),
        _fila(HEADERS_REVTECH, Folio=2, **{**_SIN_REPARTO, "Subsidio": 1}),
        _fila(HEADERS_REVTECH, Folio=3),
    ]
    parseadas = parsear_registro_egresos(_xlsx(HEADERS_REVTECH, filas), "REVTECH").filas
    db: Any = _SesionFalsa([[(1,), (2,), (3,)]])
    r = await cargar_filas(db, "REVTECH", parseadas, "claudia@trongkai.com")
    assert (r.sin_clasificar, r.descuadradas, r.creadas) == (1, 1, 3)


# ── CLI ──────────────────────────────────────────────────────────────


def _cli() -> Any:
    ruta = Path(__file__).resolve().parents[2] / "scripts" / "importar_registro_egresos_excel.py"
    spec = importlib.util.spec_from_file_location("importar_registro_egresos_excel", ruta)
    assert spec and spec.loader
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


def test_cli_dry_run_y_json_out_sin_base_de_datos(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
):
    filas = [
        _fila(HEADERS_TRONGKAI, Folio=1),
        _fila(HEADERS_TRONGKAI, Folio=1),  # repetida: entra igual (default)
        _fila(HEADERS_TRONGKAI, Fecha=None, Folio=2),  # saltada
        _fila(HEADERS_TRONGKAI, Folio=3, **_SIN_REPARTO),
    ]
    xlsx = tmp_path / "Cuenta Bancos_trongkai.xlsx"
    xlsx.write_bytes(_xlsx(HEADERS_TRONGKAI, filas))
    salida = tmp_path / "trongkai.json"
    cli = _cli()

    assert cli.main(["--empresa", "trongkai", "--archivo", str(xlsx), "--dry-run"]) == 0
    texto = capsys.readouterr().out
    assert "Filas con datos (leídas): 4" in texto
    assert "a cargar:              3" in texto
    assert "saltadas:              1" in texto
    assert "repetidas (cargadas con observación): 1" in texto
    assert "filas repetidas (n-ésima aparición de una fila idéntica, entran igual): 5" in texto
    assert "colapsadas" not in texto
    assert "Sin clasificar (4 fuentes vacías): 1" in texto
    assert "PAGADO: 3" in texto
    assert "2026-08     3 filas" in texto
    assert "TOTAL a cargar: $283.056" in texto

    args = ["--empresa", "TRONGKAI", "--archivo", str(xlsx), "--json-out", str(salida)]
    assert cli.main(args) == 0
    payload = json.loads(salida.read_text(encoding="utf-8"))
    assert payload["empresa_codigo"] == "TRONGKAI"
    assert len(payload["filas"]) == 3
    assert payload["saltadas"][0]["fila_excel"] == 6
    assert payload["repetidas_en_excel"] == [5]
    assert payload["duplicadas_en_excel"] == []
    assert payload["filas"][0]["total"] == "94352.00"
    assert "Idéntica a la fila 4" in payload["filas"][1]["observaciones"]
    recuperadas = [FilaEgreso.from_dict(d) for d in payload["filas"]]
    assert recuperadas == parsear_registro_egresos(xlsx.read_bytes(), "TRONGKAI").filas

    # --conservar-repetidas sigue aceptándose, pero es un no-op (ya es el default)
    capsys.readouterr()
    args = ["--empresa", "TRONGKAI", "--archivo", str(xlsx), "--dry-run", "--conservar-repetidas"]
    assert cli.main(args) == 0
    assert "a cargar:              3" in capsys.readouterr().out

    # --colapsar-repetidas: comportamiento viejo, sólo entra la primera
    args = ["--empresa", "TRONGKAI", "--archivo", str(xlsx), "--dry-run", "--colapsar-repetidas"]
    assert cli.main(args) == 0
    texto = capsys.readouterr().out
    assert "a cargar:              2" in texto
    assert "repetidas (cargadas con observación): 0" in texto
    assert "colapsadas (--colapsar-repetidas): 1" in texto
    assert "filas colapsadas (no se cargan, sólo entra la primera): 5" in texto
    assert "TOTAL a cargar: $188.704" in texto


def test_cli_rechaza_empresa_ajena_y_json_de_otra_empresa(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
):
    cli = _cli()
    xlsx = tmp_path / "x.xlsx"
    xlsx.write_bytes(_xlsx(HEADERS_REVTECH, [_fila(HEADERS_REVTECH)]))
    assert cli.main(["--empresa", "AFIS", "--archivo", str(xlsx), "--dry-run"]) == 1
    assert "REVTECH o TRONGKAI" in capsys.readouterr().out

    salida = tmp_path / "revtech.json"
    args = ["--empresa", "REVTECH", "--archivo", str(xlsx), "--json-out", str(salida)]
    assert cli.main(args) == 0
    capsys.readouterr()
    # cargar el JSON de REVTECH como TRONGKAI se corta ANTES de tocar la BD
    assert cli.main(["--empresa", "TRONGKAI", "--json-in", str(salida), "--usuario", "a@b.c"]) == 1
    assert "el JSON es de REVTECH" in capsys.readouterr().out
    # sin --usuario tampoco se carga
    assert cli.main(["--empresa", "REVTECH", "--json-in", str(salida)]) == 1
    assert "--usuario" in capsys.readouterr().out


def test_fila_saltada_es_serializable():
    s = FilaSaltada(29, "Fecha inválida (None)")
    assert ResumenCarga("REVTECH", True, 1, 0, 0, 0, 0, saltadas=[s]).to_dict()["saltadas"] == [
        {"fila_excel": 29, "motivo": "Fecha inválida (None)"}
    ]
