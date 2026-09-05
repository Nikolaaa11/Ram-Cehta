"""Export del registro de egresos: Registro de Egresos (Claudia) + Carga_Gastos (CORFO).

La hoja Carga_Gastos tiene que salir con los 21 encabezados EXACTOS de
`corfo_rendiciones.py` (la planilla oficial del folio 2024-265638). Como esa
lista vive dentro de una función, se la lee del fuente con `ast`: si alguien
la toca allá, este test lo ve.
"""
from __future__ import annotations

import ast
import io
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

import app.api.v1.corfo_rendiciones as corfo_rendiciones
from app.api.v1.claudia_egresos import (
    CARGA_GASTOS_HEADERS,
    MAPEO_TIPO_DOC_CORFO,
    REGISTRO_HEADERS,
    construir_export_xlsx,
    nombre_archivo_export,
)


def _headers_oficiales_carga_gastos() -> list[str]:
    fuente = Path(corfo_rendiciones.__file__).read_text(encoding="utf-8")
    for nodo in ast.walk(ast.parse(fuente)):
        if not isinstance(nodo, ast.Assign):
            continue
        if not any(isinstance(t, ast.Name) and t.id == "headers" for t in nodo.targets):
            continue
        if isinstance(nodo.value, ast.List):
            vals = [e.value for e in nodo.value.elts if isinstance(e, ast.Constant)]
            if len(vals) == 21:  # la de RRHH tiene 17
                return vals
    raise AssertionError("no encontré los 21 headers de Carga_Gastos en corfo_rendiciones.py")


def _fila(**over: Any) -> dict[str, Any]:
    f: dict[str, Any] = {
        "egreso_id": 1,
        "empresa_codigo": "TRONGKAI",
        "periodo": "2026-08",
        "fecha": date(2026, 8, 27),
        "descripcion": "MCG AUDITORES CONSULTORES SPA",
        "rut_emisor": "76642280-2",
        "tipo_documento": "FACTURA",
        "folio": "10540",
        "monto_neto": Decimal("79287.00"),
        "impuesto": Decimal("15065.00"),
        "total": Decimal("94352.00"),
        "tipo_egreso": "Servicios",
        "fuente": "Corfo",
        "proyecto": "Trongkai",
        "estado_pago": "PAGADO",
        "fecha_pago": date(2026, 9, 1),
        "monto_subsidio": Decimal("47176.00"),
        "monto_cehta_ptec": Decimal("18870.00"),
        "monto_cehta": Decimal("28306.00"),
        "monto_trewaox": Decimal("0.00"),
        "corfo_cuenta": None,
        "corfo_item": None,
        "corfo_fuente_financiamiento": None,
        "corfo_etapa": None,
        "corfo_fecha_recepcion": None,
        "corfo_monto_rendir": None,
        "corfo_monto_cancelado": None,
        "corfo_forma_pago": None,
        "corfo_glosa": None,
        "corfo_receptor_rut": None,
        "corfo_receptor_nombre": None,
    }
    f.update(over)
    return f


def _abrir(filas: list[dict[str, Any]], periodo: str | None = "2026-08"):
    return load_workbook(io.BytesIO(construir_export_xlsx("TRONGKAI", periodo, filas)))


def _fila_hoja(ws: Any, n: int) -> list[Any]:
    return [c.value for c in ws[n]]


def test_carga_gastos_tiene_los_21_encabezados_oficiales_en_orden():
    oficiales = _headers_oficiales_carga_gastos()
    assert len(oficiales) == 21
    assert oficiales == CARGA_GASTOS_HEADERS
    wb = _abrir([_fila()])
    assert wb.sheetnames == ["Registro de Egresos", "Carga_Gastos"]
    assert _fila_hoja(wb["Carga_Gastos"], 1) == oficiales


def test_registro_de_egresos_tiene_las_17_columnas_de_claudia():
    wb = _abrir([_fila()])
    encabezados = _fila_hoja(wb["Registro de Egresos"], 1)
    assert encabezados == REGISTRO_HEADERS
    assert len(encabezados) == 17
    assert "Trewaox" in encabezados
    fila = dict(zip(encabezados, _fila_hoja(wb["Registro de Egresos"], 2), strict=True))
    assert fila["Tipo de Documento"] == "Factura"
    assert fila["Tipo Financiamiento"] == "Corfo"
    assert fila["Estado"] == "Pagado"
    assert fila["Subsidio"] == 47176.0
    assert fila["Total"] == 94352.0


def test_mapeo_tipo_documento_es_el_del_contrato():
    assert MAPEO_TIPO_DOC_CORFO == {
        "FACTURA": "FACTURA",
        "FACTURA_EXENTA": "FACTURA",
        "BOLETA": "BOLETA",
        "BOLETA_HONORARIO": "BOLETA HONORARIOS",
        "LIQUIDACION": "LIQ. SUELDO",
        "INVOICE": "INVOICE",
        "CO_EJECUTOR": "OTRO",
        "OTRO": "OTRO",
    }


@pytest.mark.parametrize(("tipo", "esperado"), sorted(MAPEO_TIPO_DOC_CORFO.items()))
def test_carga_gastos_mapea_cada_tipo_de_documento(tipo: str, esperado: str):
    wb = _abrir([_fila(tipo_documento=tipo)])
    fila = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(wb["Carga_Gastos"], 2), strict=True))
    assert fila["Tipo Documento"] == esperado


def test_carga_gastos_periodo_corfo_y_defaults_honestos():
    wb = _abrir([_fila(), _fila(egreso_id=2, estado_pago="PENDIENTE", fecha_pago=None)])
    ws = wb["Carga_Gastos"]
    pagada = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(ws, 2), strict=True))
    pendiente = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(ws, 3), strict=True))
    assert pagada["Periodo"] == "Ago de 2026"
    assert pagada["Monto Rendir"] == 47176.0  # = subsidio si no se cargó otro
    assert pagada["Monto Cancelado"] == 94352.0  # = total porque está PAGADO
    assert pendiente["Monto Cancelado"] is None  # no se inventa un pago
    assert pagada["Nombre Proveedor o Razón Social"] == "MCG AUDITORES CONSULTORES SPA"
    assert pagada["Rut Proveedor"] == "76642280-2"
    fecha_doc = pagada["Fecha del documento"]
    assert fecha_doc == date(2026, 8, 27) or str(fecha_doc).startswith("2026-08-27")


def test_corfo_monto_rendir_cargado_le_gana_al_default():
    wb = _abrir(
        [_fila(corfo_monto_rendir=Decimal("12345.00"), corfo_monto_cancelado=Decimal("1.00"))]
    )
    fila = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(wb["Carga_Gastos"], 2), strict=True))
    assert fila["Monto Rendir"] == 12345.0
    assert fila["Monto Cancelado"] == 1.0


def test_sin_clasificar_deja_monto_rendir_vacio():
    sin = _fila(monto_subsidio=None, monto_cehta_ptec=None, monto_cehta=None, monto_trewaox=None)
    wb = _abrir([sin])
    fila = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(wb["Carga_Gastos"], 2), strict=True))
    assert fila["Monto Rendir"] is None


def test_caracter_de_control_no_rompe_el_export():
    wb = _abrir([_fila(descripcion="PROVEEDOR\x00SUCIO\x1f", corfo_glosa="glosa\x0bfea")])
    registro = _fila_hoja(wb["Registro de Egresos"], 2)
    carga = dict(zip(CARGA_GASTOS_HEADERS, _fila_hoja(wb["Carga_Gastos"], 2), strict=True))
    assert registro[1] == "PROVEEDOR SUCIO "
    assert carga["Glosa / Justificación"] == "glosa fea"


def test_export_vacio_igual_trae_las_dos_hojas_con_encabezados():
    wb = _abrir([], periodo=None)
    assert wb.sheetnames == ["Registro de Egresos", "Carga_Gastos"]
    assert _fila_hoja(wb["Carga_Gastos"], 1) == CARGA_GASTOS_HEADERS
    assert wb["Carga_Gastos"].max_row == 1


def test_nombre_del_archivo():
    assert nombre_archivo_export("REVTECH", "2026-08") == "registro_egresos_REVTECH_2026-08.xlsx"
    assert nombre_archivo_export("TRONGKAI", None) == "registro_egresos_TRONGKAI_todos.xlsx"
