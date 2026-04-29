"""Tests para los helpers puros de search y exports.

Las queries SQL se prueban en integración (otra capa); acá nos enfocamos en
las decisiones de borde: longitud mínima de query, serialización de tipos,
construcción de workbook con headers y freeze panes.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.api.v1.exports import (
    _ENTITY_QUERIES,
    _build_workbook,
    _serialize,
)
from app.schemas.search import SearchHit, SearchResponse


# ---------------------------------------------------------------------
# exports._serialize
# ---------------------------------------------------------------------
class TestSerialize:
    def test_none_to_empty_string(self) -> None:
        assert _serialize(None) == ""

    def test_decimal_to_float(self) -> None:
        out = _serialize(Decimal("123.45"))
        assert out == 123.45
        assert isinstance(out, float)

    def test_date_pasa_por_referencia(self) -> None:
        d = date(2025, 1, 1)
        assert _serialize(d) is d

    def test_datetime_pasa_por_referencia(self) -> None:
        dt = datetime(2025, 1, 1, 10, 30)
        assert _serialize(dt) is dt

    def test_int_y_str_pasan(self) -> None:
        assert _serialize(42) == 42
        assert _serialize("hola") == "hola"


# ---------------------------------------------------------------------
# exports._build_workbook
# ---------------------------------------------------------------------
class TestBuildWorkbook:
    def test_workbook_tiene_header_bold(self) -> None:
        headers = ["A", "B", "C"]
        rows = [(1, "x", date(2025, 1, 1))]
        xlsx = _build_workbook(headers, rows, "Test")

        from io import BytesIO

        wb = load_workbook(BytesIO(xlsx))
        ws = wb.active
        assert ws is not None
        assert ws.title == "Test"
        # row 1 = headers, en bold
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 1).font.bold is True
        # row 2 = data
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "x"
        # freeze panes en A2 (header pegado al hacer scroll)
        assert ws.freeze_panes == "A2"

    def test_sheet_name_truncado_a_31_chars(self) -> None:
        nombre_largo = "A" * 50
        xlsx = _build_workbook(["X"], [], nombre_largo)
        from io import BytesIO

        wb = load_workbook(BytesIO(xlsx))
        assert wb.active is not None
        assert len(wb.active.title) == 31

    def test_filas_vacias_genera_solo_header(self) -> None:
        xlsx = _build_workbook(["A", "B"], [], "Vacio")
        from io import BytesIO

        wb = load_workbook(BytesIO(xlsx))
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 1  # solo header

    def test_decimales_convertidos_a_float_en_excel(self) -> None:
        xlsx = _build_workbook(
            ["Monto"], [(Decimal("9999.99"),)], "M"
        )
        from io import BytesIO

        wb = load_workbook(BytesIO(xlsx))
        ws = wb.active
        assert ws is not None
        assert ws.cell(2, 1).value == pytest.approx(9999.99)


# ---------------------------------------------------------------------
# exports._ENTITY_QUERIES
# ---------------------------------------------------------------------
class TestEntityQueries:
    def test_todas_las_entities_tienen_headers_y_sql(self) -> None:
        for name, cfg in _ENTITY_QUERIES.items():
            assert cfg.get("headers"), f"{name}: headers vacío"
            assert cfg.get("sql"), f"{name}: sql vacío"
            # toda query incluye :empresa, :estado y :lim como params
            sql = cfg["sql"]
            assert ":empresa" in sql, f"{name}: falta :empresa param"
            assert ":estado" in sql, f"{name}: falta :estado param"
            assert ":lim" in sql, f"{name}: falta :lim param"

    def test_entities_esperadas_existen(self) -> None:
        expected = {
            "ordenes_compra",
            "f29",
            "proveedores",
            "trabajadores",
            "legal_documents",
            "movimientos",
            "suscripciones",
            "fondos",
        }
        assert set(_ENTITY_QUERIES.keys()) >= expected


# ---------------------------------------------------------------------
# search.SearchHit / SearchResponse
# ---------------------------------------------------------------------
class TestSearchSchemas:
    def test_hit_minimo_valido(self) -> None:
        hit = SearchHit(
            entity_type="empresa",
            entity_id="CENERGY",
            title="Cenergy SpA",
            link="/empresa/CENERGY",
        )
        assert hit.entity_type == "empresa"
        assert hit.subtitle is None
        assert hit.score == 0.0

    def test_response_vacio_tiene_total_cero(self) -> None:
        r = SearchResponse(query="ab", total=0, by_entity={})
        assert r.total == 0
        assert r.by_entity == {}

    def test_response_con_hits_se_serializa(self) -> None:
        hit = SearchHit(
            entity_type="empresa",
            entity_id="CENERGY",
            title="Cenergy",
            link="/empresa/CENERGY",
        )
        r = SearchResponse(
            query="cen", total=1, by_entity={"empresa": [hit]}
        )
        dumped = r.model_dump()
        assert dumped["total"] == 1
        assert "empresa" in dumped["by_entity"]
        assert dumped["by_entity"]["empresa"][0]["title"] == "Cenergy"


# ---------------------------------------------------------------------
# search query length
# ---------------------------------------------------------------------
class TestSearchMinLength:
    def test_min_query_len_es_dos(self) -> None:
        from app.api.v1.search import _MIN_QUERY_LEN

        assert _MIN_QUERY_LEN == 2

    def test_per_entity_limit_es_cinco(self) -> None:
        from app.api.v1.search import _PER_ENTITY_LIMIT

        assert _PER_ENTITY_LIMIT == 5
